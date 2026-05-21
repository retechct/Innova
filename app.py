import os
import json
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
import psycopg2
from psycopg2 import pool as pg_pool
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from dotenv import load_dotenv
import cloudinary
import cloudinary.uploader
import cloudinary.api
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import db
from flask_migrate import Migrate
from flask_jwt_extended import JWTManager
import pandas as pd
from io import BytesIO
from flask import send_file
from datetime import datetime
import csv
from io import StringIO
from flask import Response
# Carga las variables del archivo .env (en producción, Railway las inyecta directo)
load_dotenv()

cloudinary.config(
    cloud_name = os.getenv('CLOUDINARY_CLOUD_NAME'),
    api_key    = os.getenv('CLOUDINARY_API_KEY'),
    api_secret = os.getenv('CLOUDINARY_API_SECRET')
)

app = Flask(__name__, static_folder='Vendedor', static_url_path='')
CORS(app)
# --- CONFIGURACIÓN E INICIALIZACIÓN DE BASE DE DATOS Y SEGURIDAD ---
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'clave-secreta-de-innova-mobili')

# Inicializar las extensiones con tu app actual sin alterar tus rutas existentes
db.init_app(app)
migrate = Migrate(app, db)
jwt = JWTManager(app)

# --- IMPORTAR Y REGISTRAR LAS RUTAS DEL KARDEX ---
from routes_kardex import kardex_bp
app.register_blueprint(kardex_bp, url_prefix='/api/kardex')
# ------------------------------------------------------------------
# Rate limiter — protege el login contra ataques de fuerza bruta
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://",
)

# URL base del backend
BACKEND_URL = os.getenv("BACKEND_URL", "https://innova-4cnn.onrender.com")

# ==========================================
# POOL DE CONEXIONES A LA BASE DE DATOS
# ==========================================
# Reutiliza conexiones TCP en lugar de abrir una nueva en cada request.
# minconn=2 conexiones siempre listas, maxconn=10 para picos de tráfico.
_db_pool = pg_pool.ThreadedConnectionPool(
    minconn=2,
    maxconn=10,
    host     = os.getenv("DB_HOST"),
    database = os.getenv("DB_NAME"),
    user     = os.getenv("DB_USER"),
    password = os.getenv("DB_PASSWORD"),
)

def get_db_connection():
    """Obtiene una conexión del pool (no abre una TCP nueva cada vez)."""
    return _db_pool.getconn()

def release_db_connection(conn):
    """Devuelve la conexión al pool para que otro request la reutilice."""
    if conn:
        _db_pool.putconn(conn)


# ==========================================
# FUNCIONES AUXILIARES
# ==========================================


def limpiar_foto(url):
    """Evita placeholders o URLs vacías en campos de foto."""
    if not url or 'via.placeholder.com' in url:
        return "imagenes/sin_foto.jpg"
    # Las URLs de Cloudinary ya son https completas; las locales antiguas se devuelven con la URL base
    if url.startswith('http'):
        return url
    return f"{BACKEND_URL}/uploads/{url}"


def enviar_notificacion_venta(correo_destino, codigo_venta, cliente):
    """Envía un correo automático al vendedor tras registrar una venta."""
    try:
        remitente  = os.getenv("EMAIL_USER")
        password   = os.getenv("EMAIL_PASS")
        smtp_server = os.getenv("EMAIL_SMTP", "smtp.gmail.com")
        smtp_port  = int(os.getenv("EMAIL_PORT", 587))

        if not remitente or not password:
            print("⚠️ Advertencia: Credenciales de correo no configuradas.")
            return

        mensaje = MIMEText(
            f"Se ha registrado una nueva venta.\n\nCódigo: {codigo_venta}\nCliente: {cliente}"
        )
        mensaje['Subject'] = f"Nueva Venta Registrada - {codigo_venta}"
        mensaje['From']    = remitente
        mensaje['To']      = correo_destino

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(remitente, password)
            server.send_message(mensaje)
    except Exception as e:
        print(f"Error al enviar correo de notificación: {e}")


# ==========================================
# 0. SERVIDOR DE IMÁGENES Y BIENVENIDA
# ==========================================

@app.route('/uploads/<filename>')
def mostrar_foto(filename):
    """Ruta de compatibilidad para fotos antiguas guardadas localmente."""
    return send_from_directory('uploads', filename)


@app.route('/', methods=['GET'])
def bienvenida():
    return send_from_directory('Vendedor', 'index.html')


# ==========================================
# 1. MÓDULO: CATÁLOGO E INSUMOS
# ==========================================

@app.route('/api/catalogo', methods=['GET'])
def obtener_catalogo():
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT id, nombre_modelo, precio_base, foto_url, es_plantilla, en_stock
            FROM catalogo_productos
        """)
        productos = cursor.fetchall()

        lista_productos = []
        for p in productos:
            lista_productos.append({
                "id":          p[0],
                "nombre":      p[1],
                "precio":      float(p[2]),
                "foto":        limpiar_foto(p[3]),
                "es_plantilla": bool(p[4]),
                "en_stock":    bool(p[5])
            })
        return jsonify(lista_productos)
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/catalogo/nuevo', methods=['POST'])
def agregar_producto_directo():
    """Agrega un producto ya terminado (en stock) directo al catálogo."""
    try:
        nombre   = request.form.get('nombre')
        precio   = float(request.form.get('precio', 0))
        cantidad = int(request.form.get('cantidad', 1))
        origen   = request.form.get('origen', 'Externo')

        if 'foto' not in request.files or request.files['foto'].filename == '':
            return jsonify({'error': 'La foto del producto es obligatoria'}), 400

        foto_file = request.files['foto']
        respuesta_nube = cloudinary.uploader.upload(foto_file, folder="catalogo")
        foto_ruta = respuesta_nube.get('secure_url')

        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            INSERT INTO catalogo_productos
                (nombre_modelo, precio_base, foto_url, es_plantilla, en_stock, origen_produccion, stock_cantidad)
            VALUES (%s, %s, %s, False, %s, %s, %s)
        """, (nombre, precio, foto_ruta, cantidad > 0, origen, cantidad))
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Producto añadido al catálogo'}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/insumos', methods=['GET'])
def obtener_insumos():
    """
    Devuelve el inventario de insumos con nombre, cantidad actual.
    Tabla: inventario_insumos (id, nombre_insumo, cantidad_actual)
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT id, nombre_insumo, cantidad_actual FROM inventario_insumos ORDER BY nombre_insumo")
        insumos = [
            {"id": r[0], "nombre": r[1], "cantidad_actual": r[2]}
            for r in cursor.fetchall()
        ]
        return jsonify(insumos)
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/upload-voucher', methods=['POST'])
def upload_voucher():
    """
    Endpoint dedicado para subir fotos de vouchers/recibos a Cloudinary.
    El frontend llama esto ANTES de agregar el pago a la lista,
    y recibe la URL real de Cloudinary para guardarla en la venta.
    """
    if 'archivo' not in request.files or request.files['archivo'].filename == '':
        return jsonify({'error': 'No se recibió ningún archivo'}), 400
    try:
        archivo = request.files['archivo']
        respuesta_nube = cloudinary.uploader.upload(archivo, folder="vouchers_pagos")
        url = respuesta_nube.get('secure_url')
        return jsonify({'url': url}), 200
    except Exception as e:
        print(f"Error al subir voucher: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ventas', methods=['GET', 'POST'])
def guardar_venta():
    if request.method == 'GET':
        return listar_ventas()
    datos = request.json
    try:
        conexion = get_db_connection()
        conexion.autocommit = False # Iniciamos transacción manual
        cursor   = conexion.cursor()

        # 1. Insertar cabecera de la venta (ACTUALIZADO MAYO 2026)
        # Extraer empresa_pago del primer pago registrado (para la columna resumen de ventas)
        lista_pagos_raw = datos.get('pagos', [])
        empresa_pago_resumen = lista_pagos_raw[0].get('empresa', '') if lista_pagos_raw else ''

        cursor.execute("""
            INSERT INTO ventas (
                codigo_venta, nombre_cliente, dni_cliente, celular_cliente,
                direccion_cliente, vendedor_id, fecha_emision, fecha_entrega,
                monto_total, vendedor_nombre, moneda, tipo_cambio, tipo_comprobante,
                empresa_ruc, empresa_pago, sede, nombre_empresa_cliente
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;
        """, (
            datos['codigo'],          datos['cliente'],
            datos.get('dni'),         datos.get('celular'),
            datos.get('direccion'),   datos.get('vendedor_id'),
            datos['fecha_emision'],   datos.get('fecha_entrega'),
            datos.get('monto_total', 0),
            datos.get('vendedor_nombre'),
            datos.get('moneda', 'PEN'),
            datos.get('tipo_cambio', 1.00),
            datos.get('tipo_comprobante', 'Boleta'),
            datos.get('empresa_ruc'),        # RUC de la empresa del vendedor (viene del login)
            empresa_pago_resumen,             # Empresa destino del primer pago (resumen)
            datos.get('sede', 'Sede Central'),
            datos.get('empresa_cliente', 'Particular')
        ))
        venta_id = cursor.fetchone()[0]
        
        _crear_tabla_historial_precios(cursor)
        cursor.execute("""
                        INSERT INTO historial_precios (
                            venta_id, codigo_venta, precio_original, precio_nuevo, 
                            motivo, vendedor_id, vendedor_nombre, estado, fecha_solicitud, fecha_resolucion
                        ) VALUES (%s, %s, %s, %s, 'Precio base de creación del contrato', %s, %s, 'Aprobado', NOW(), NOW());
                    """, (
                        venta_id, 
                        datos['codigo'], 
                        float(datos.get('monto_total', 0)), 
                        float(datos.get('monto_total', 0)),
                        datos.get('vendedor_id'),
                        datos.get('vendedor_nombre')
                    ))

        # 2. Registrar Múltiples Pagos en la tabla 'pagos' (NUEVO)
        lista_pagos = lista_pagos_raw
        total_adelanto = 0
        for p in lista_pagos:
            monto_bruto = p.get('monto', 0)
            total_adelanto += monto_bruto
            cursor.execute("""
                INSERT INTO pagos (
                    venta_id, tipo_pago, entidad, numero_operacion,
                    monto_bruto, comision_pos, monto_neto, empresa_destino, comprobante_url
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s);
            """, (
                venta_id,
                p.get('tipo'),
                p.get('entidad'),
                p.get('operacion'),
                monto_bruto,
                p.get('comision', 0),
                p.get('monto_neto', 0),
                p.get('empresa'),
                p.get('comprobante_url', 'Sin imagen')
            ))

        # Actualizar monto_adelanto en ventas con la suma real de pagos
        if total_adelanto > 0:
            cursor.execute("""
                UPDATE ventas SET monto_adelanto = %s WHERE id = %s
            """, (total_adelanto, venta_id))

        # ============================================================
        # MOTOR MAKE-VS-BUY — Mapa de componentes a áreas de taller
        # ============================================================
        TIPOS_TELA   = {'tela', 'tela-cojin', 'tela-silla', 'tela-butaca'}
        TIPOS_COJIN  = {'cojin-entero', 'cojin-diseno'}
        SUFIJO_TELA  = {
            'tela':        ' - Para Sofá/Silla',
            'tela-silla':  ' - Para Silla Comedor',
            'tela-butaca': ' - Para Butaca',
            'tela-cojin':  ' - Para Cojines',
            'cojin-entero':'Cojines Enteros',
            'cojin-diseno':'Cojines c/Diseño',
        }

        mapeo_erp = {
            'tela':              ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'tela-cojin':        ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'tela-silla':        ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'tela-butaca':       ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'cojin-entero':      ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'cojin-diseno':      ('maestro_disenos_cojin', 'CORTE_Y_CONTROL_TELAS'),
            'base':              ('maestro_bases',          'PREPARACION_PATAS_ZOCALO'),
            'tablero':           ('maestro_tableros',       'TABLEROS_Y_PIEDRAS'),
            'tablero-centro':    ('maestro_tableros',       'TABLEROS_Y_PIEDRAS'),
            'silla':             ('maestro_sillas',         'ESTRUCTURAS_SILLAS'),
            'estructura-butaca': ('maestro_butacas',        'ESTRUCTURAS_SILLAS'),
            'base-mesa':         ('maestro_bases_comedor',  'TABLEROS_Y_PIEDRAS'),
            'base-centro':       ('maestro_bases_comedor',  'TABLEROS_Y_PIEDRAS'),
        }

        for m in datos['muebles']:
            # Insertar ítem de venta
            cursor.execute("""
                INSERT INTO items_venta (venta_id, producto, color_tela, foto_url)
                VALUES (%s, %s, %s, %s) RETURNING id;
            """, (venta_id, m['tipo'], m['tela'], m['foto']))
            item_id = cursor.fetchone()[0]

            componentes = m.get('componentes', {})
            areas_internas_creadas = set()

            nombre_lower = m['tipo'].lower()
            area_estructura = None
            if any(p in nombre_lower for p in ['sofá', 'sofa', 'seccional', 'modular', 'multi', 'curvado', 'plantilla']):
                area_estructura = 'ESTRUCTURAS_MUEBLES'
            elif any(p in nombre_lower for p in ['silla', 'butaca', 'sitial', 'puff']):
                area_estructura = 'ESTRUCTURAS_SILLAS'

            if area_estructura:
                cursor.execute("""
                    INSERT INTO tickets_produccion (item_id, area_trabajo, estado_ticket, etapa)
                    VALUES (%s, %s, 'Pendiente', 1)
                """, (item_id, area_estructura))
                areas_internas_creadas.add(area_estructura)

                if area_estructura == 'ESTRUCTURAS_MUEBLES':
                    area_tap = 'TAPICERIA_SOFAS'
                elif area_estructura == 'ESTRUCTURAS_SILLAS':
                    area_tap = 'TAPICERIA_SILLAS'
                else:
                    area_tap = None

                if area_tap:
                    cursor.execute("""
                        INSERT INTO tickets_produccion (item_id, area_trabajo, estado_ticket, etapa)
                        VALUES (%s, %s, 'Bloqueado', 2)
                    """, (item_id, area_tap))
                    areas_internas_creadas.add(area_tap)

            for key, sku in componentes.items():
                if not sku or key not in mapeo_erp:
                    continue

                tabla, area_destino = mapeo_erp[key]

                if key == 'silla':
                    cursor.execute("SELECT material, origen_produccion FROM maestro_sillas WHERE sku = %s", (sku,))
                    res_silla = cursor.fetchone()
                    if res_silla:
                        material_silla = (res_silla[0] or '').lower()
                        if any(m_word in material_silla for m_word in ['metal', 'acero', 'fierro', 'aluminio']):
                            cursor.execute("""
                                INSERT INTO logistica_externa (venta_id, insumo_nombre, sku, estado)
                                VALUES (%s, %s, %s, 'POR_PEDIR')
                            """, (venta_id, key, sku))
                            continue
                        if res_silla[1] == 'Externo':
                            cursor.execute("""
                                INSERT INTO logistica_externa (venta_id, insumo_nombre, sku, estado)
                                VALUES (%s, %s, %s, 'POR_PEDIR')
                            """, (venta_id, key, sku))
                            continue

                try:
                    cursor.execute(f"SELECT origen_produccion FROM {tabla} WHERE sku = %s", (sku,))
                    res = cursor.fetchone()
                except Exception:
                    res = None

                if res and res[0] == 'Interno':
                    if area_destino not in areas_internas_creadas:
                        nombre_ticket = m['tipo']
                        if key in SUFIJO_TELA:
                            nombre_ticket = m['tipo'] + SUFIJO_TELA[key]

                        cursor.execute("""
                            INSERT INTO tickets_produccion (item_id, area_trabajo, estado_ticket, etapa)
                            VALUES (%s, %s, 'Pendiente', 1)
                        """, (item_id, area_destino))
                        areas_internas_creadas.add(area_destino)
                elif res and res[0] == 'Externo':
                    cursor.execute("""
                        INSERT INTO logistica_externa (venta_id, insumo_nombre, sku, estado)
                        VALUES (%s, %s, %s, 'POR_PEDIR')
                    """, (venta_id, key, sku))
                elif not res:
                    cursor.execute("""
                        INSERT INTO logistica_externa (venta_id, insumo_nombre, sku, estado)
                        VALUES (%s, %s, %s, 'POR_PEDIR')
                    """, (venta_id, key, sku))

            # Ticket de Despacho
            ha_generado_tickets = len(areas_internas_creadas) > 0
            estado_despacho = 'Bloqueado' if ha_generado_tickets else 'Pendiente'
            cursor.execute("""
                INSERT INTO tickets_produccion (item_id, area_trabajo, estado_ticket, etapa)
                VALUES (%s, 'DESPACHO_CENTRAL', %s, 99)
            """, (item_id, estado_despacho))

            # Descuento de stock
            cursor.execute("""
                SELECT cp.id FROM catalogo_productos cp
                WHERE cp.nombre_modelo = %s LIMIT 1;
            """, (m['tipo'],))
            prod_row = cursor.fetchone()
            
            if prod_row:
                producto_id = prod_row[0]
                cursor.execute("""
                    UPDATE inventario_insumos i
                    SET cantidad_actual = GREATEST(0, i.cantidad_actual - r.cantidad_necesaria)
                    FROM recetas_muebles r
                    WHERE r.insumo_id = i.id AND r.producto_id = %s;
                """, (producto_id,))

        conexion.commit()

        cursor.execute("SELECT email FROM usuarios WHERE id = %s", (datos['vendedor_id'],))
        v_correo = cursor.fetchone()
        if v_correo:
            enviar_notificacion_venta(v_correo[0], datos['codigo'], datos['cliente'])

        return jsonify({"mensaje": "Venta procesada exitosamente", "id": venta_id}), 201

    except Exception as ex:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print(f"ERROR SQL EXACTO: {str(ex)}")
        error_msg = str(ex)
        if "llave duplicada" in error_msg or "UniqueViolation" in error_msg:
            return jsonify({"error": "El N° de Contrato ya fue registrado."}), 400
        return jsonify({"error": error_msg}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)
        
# ==========================================
# 3. MÓDULO: SEGUIMIENTO Y PRODUCCIÓN
# ==========================================

@app.route('/api/mis-ventas/<int:vendedor_id>', methods=['GET'])
def obtener_mis_ventas(vendedor_id):
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT
                v.codigo_venta,
                v.nombre_cliente,
                v.fecha_entrega,
                COALESCE(COUNT(t.id), 0)                                              AS total,
                COALESCE(SUM(CASE WHEN t.estado_ticket = 'Terminado' THEN 1 ELSE 0 END), 0) AS terminados,
                v.monto_total,
                COALESCE(v.estado_general, 'Pendiente') AS estado
            FROM ventas v
            LEFT JOIN items_venta i        ON v.id    = i.venta_id
            LEFT JOIN tickets_produccion t ON i.id    = t.item_id
            WHERE v.vendedor_id = %s
            GROUP BY v.id, v.codigo_venta, v.nombre_cliente, v.fecha_entrega, v.monto_total, v.estado_general
            ORDER BY v.id DESC;
        """, (vendedor_id,))

        res = []
        for v in cursor.fetchall():
            total     = v[3]
            terminados = v[4]
            porcentaje = round((terminados / total * 100), 0) if total > 0 else 0
            res.append({
                "codigo":   v[0],
                "cliente":  v[1],
                "entrega":  v[2].strftime('%d/%m/%Y') if v[2] else "S/F",
                "progreso": porcentaje,
                "monto_total": float(v[5]),
                "estado":   v[6]
            })
        return jsonify(res)
    except Exception as ex:
        print("Error en seguimiento:", ex)
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/produccion/<area>', methods=['GET'])
def ver_tickets_area(area):
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT t.id, i.producto, i.color_tela, v.codigo_venta,
                   t.estado_ticket, i.foto_url, v.fecha_entrega
            FROM tickets_produccion t
            JOIN items_venta i ON t.item_id   = i.id
            JOIN ventas v      ON i.venta_id  = v.id
            WHERE t.area_trabajo = %s AND t.estado_ticket = 'Pendiente'
            ORDER BY v.fecha_entrega ASC;
        """, (area,))

        lista = []
        for t in cursor.fetchall():
            lista.append({
                "ticket_id":    t[0],
                "producto":     t[1],
                "color":        t[2],
                "codigo_venta": t[3],
                "estado":       t[4],
                "foto":         t[5],
                "fecha_entrega": t[6].strftime('%d/%m/%Y') if t[6] else "S/F"
            })
        return jsonify({"area": area, "tareas": lista})
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/pedido/detalle/<codigo>', methods=['GET'])
def obtener_detalle_pedido(codigo):
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute("""
            SELECT id, codigo_venta, nombre_cliente, fecha_entrega, empresa_ruc, vendedor_nombre
            FROM ventas WHERE codigo_venta = %s;
        """, (codigo,))
        venta = cursor.fetchone()
        if not venta:
            return jsonify({"error": "Pedido no encontrado"}), 404

        cursor.execute("""
            SELECT producto, color_tela, foto_url
            FROM items_venta WHERE venta_id = %s;
        """, (venta[0],))
        items = [{"producto": i[0], "detalles": i[1], "foto": i[2]} for i in cursor.fetchall()]

        return jsonify({
            "codigo":  venta[1],
            "cliente": venta[2],
            "entrega": venta[3].strftime('%d/%m/%Y') if venta[3] else "S/F",
            "items":   items
        }), 200
    except Exception as ex:
        print("Error al obtener detalle del pedido:", ex)
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# 4. MÓDULO: MAESTRO DE MATERIALES (SKU Y FOTOS)
# ==========================================

@app.route('/api/materiales/nuevo', methods=['POST'])
def agregar_nuevo_material():
    try:
        tipo_material = request.form.get('tipo_material')
        origen        = request.form.get('origen_produccion', 'Externo')

        foto_ruta = ""
        if 'foto' in request.files:
            foto_file = request.files['foto']
            if foto_file.filename != '':
                respuesta_nube = cloudinary.uploader.upload(foto_file, folder="materiales")
                foto_ruta = respuesta_nube.get('secure_url')

        conexion  = get_db_connection()
        cursor    = conexion.cursor()
        nuevo_sku = ""

        if tipo_material == 'tela':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_telas")
            nuevo_sku = f"TEL-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_telas
                    (sku, proveedor, coleccion, color, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,'Disponible') RETURNING id;
            """, (nuevo_sku, request.form.get('proveedor'), request.form.get('coleccion'),
                  request.form.get('color'), foto_ruta, origen))

        elif tipo_material == 'cojin':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_disenos_cojin")
            nuevo_sku = f"COJ-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_disenos_cojin
                    (sku, nombre_diseno, tipo_tela, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,'Disponible') RETURNING id;
            """, (nuevo_sku, request.form.get('nombre_diseno'), request.form.get('tipo_tela'),
                  foto_ruta, origen))

        elif tipo_material == 'base':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_bases")
            nuevo_sku = f"BAS-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_bases
                    (sku, tipo, material, modelo, color, medida_altura, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'Disponible') RETURNING id;
            """, (nuevo_sku, request.form.get('tipo'), request.form.get('material'),
                  request.form.get('modelo'), request.form.get('color'),
                  request.form.get('medida_altura'), foto_ruta, origen))

        elif tipo_material == 'tablero':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_tableros")
            nuevo_sku = f"TAB-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_tableros
                    (sku, material_base, nombre_modelo, color_veta, acabado, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,%s,'Disponible') RETURNING id;
            """, (nuevo_sku, request.form.get('material_base'), request.form.get('nombre_modelo'),
                  request.form.get('color_veta'), request.form.get('acabado'), foto_ruta, origen))

        elif tipo_material == 'base-comedor':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_bases_comedor")
            nuevo_sku = f"BAC-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_bases_comedor
                    (sku, material, modelo, color, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,'Disponible') RETURNING id;
            """, (nuevo_sku, request.form.get('material'), request.form.get('modelo'),
                  request.form.get('color'), foto_ruta, origen))

        elif tipo_material == 'silla':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_sillas")
            nuevo_sku = f"SIL-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_sillas
                    (sku, material, modelo, color_estructura, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,'Disponible') RETURNING id;
            """, (nuevo_sku, request.form.get('material'), request.form.get('modelo'),
                  request.form.get('color_estructura'), foto_ruta, origen))

        elif tipo_material == 'butaca':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_butacas")
            nuevo_sku = f"BUT-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_butacas
                    (sku, material, modelo, color_estructura, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,'Disponible') RETURNING id;
            """, (nuevo_sku, request.form.get('material'), request.form.get('modelo'),
                  request.form.get('color_estructura'), foto_ruta, origen))

        else:
            return jsonify({"error": "Tipo de material no válido"}), 400

        conexion.commit()
        return jsonify({
            "mensaje":  f"{tipo_material.capitalize()} registrada con éxito",
            "sku":      nuevo_sku,
            "foto_url": foto_ruta
        }), 201

    except Exception as ex:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print("Error al guardar material:", ex)
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/materiales/listas', methods=['GET'])
def obtener_listas_materiales():
    try:
        conexion = get_db_connection()
        # Usamos cursores nombrados para pipeline: enviamos las 7 queries
        # en secuencia sin esperar resultado entre ellas, reduciendo round-trips.
        cur_telas    = conexion.cursor(); cur_telas.execute("SELECT sku, proveedor, coleccion, color, foto_url, COALESCE(estado,'Disponible') FROM maestro_telas")
        cur_cojines  = conexion.cursor(); cur_cojines.execute("SELECT sku, nombre_diseno, tipo_tela, foto_url, COALESCE(estado,'Disponible') FROM maestro_disenos_cojin")
        cur_bases    = conexion.cursor(); cur_bases.execute("SELECT sku, tipo, material, modelo, color, medida_altura, foto_url, COALESCE(estado,'Disponible') FROM maestro_bases")
        cur_tableros = conexion.cursor(); cur_tableros.execute("SELECT sku, material_base, nombre_modelo, color_veta, acabado, foto_url, COALESCE(estado,'Disponible') FROM maestro_tableros")
        cur_bcom     = conexion.cursor(); cur_bcom.execute("SELECT sku, material, modelo, color, foto_url, COALESCE(estado,'Disponible') FROM maestro_bases_comedor")
        cur_sillas   = conexion.cursor(); cur_sillas.execute("SELECT sku, material, modelo, color_estructura, foto_url, COALESCE(estado,'Disponible') FROM maestro_sillas")
        cur_butacas  = conexion.cursor(); cur_butacas.execute("SELECT sku, material, modelo, color_estructura, foto_url, COALESCE(estado,'Disponible') FROM maestro_butacas")

        telas        = [{"sku":r[0],"proveedor":r[1],"coleccion":r[2],"color":r[3],"foto_url":limpiar_foto(r[4]),"estado":r[5]} for r in cur_telas.fetchall()]
        cojines      = [{"sku":r[0],"nombre_diseno":r[1],"tipo_tela":r[2],"foto_url":limpiar_foto(r[3]),"estado":r[4]} for r in cur_cojines.fetchall()]
        bases        = [{"sku":r[0],"tipo":r[1],"material":r[2],"modelo":r[3],"color":r[4],"medida":r[5],"foto_url":limpiar_foto(r[6]),"estado":r[7]} for r in cur_bases.fetchall()]
        tableros     = [{"sku":r[0],"material_base":r[1],"nombre":r[2],"color":r[3],"acabado":r[4],"foto_url":limpiar_foto(r[5]),"estado":r[6]} for r in cur_tableros.fetchall()]
        bases_comedor= [{"sku":r[0],"material":r[1],"modelo":r[2],"color":r[3],"foto_url":limpiar_foto(r[4]),"estado":r[5]} for r in cur_bcom.fetchall()]
        sillas       = [{"sku":r[0],"material":r[1],"modelo":r[2],"color":r[3],"foto_url":limpiar_foto(r[4]),"estado":r[5]} for r in cur_sillas.fetchall()]
        butacas      = [{"sku":r[0],"material":r[1],"modelo":r[2],"color":r[3],"foto_url":limpiar_foto(r[4]),"estado":r[5]} for r in cur_butacas.fetchall()]

        for c in (cur_telas, cur_cojines, cur_bases, cur_tableros, cur_bcom, cur_sillas, cur_butacas):
            c.close()

        return jsonify({
            "telas": telas, "cojines": cojines, "bases": bases,
            "tableros": tableros, "bases_comedor": bases_comedor,
            "sillas": sillas, "butacas": butacas
        })
    except Exception as ex:
        print("Error en obtener_listas_materiales:", ex)
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            release_db_connection(conexion)


# ==========================================
# 5. MÓDULO: GESTOR DE APROBACIÓN (CREACIONES)
# ==========================================

@app.route('/api/creaciones', methods=['POST'])
def guardar_creacion():
    """
    Vendedor sube un diseño personalizado (fotos de referencia).
    Tabla creaciones_vendedores: (id, vendedor_id, nombre_modelo, categoria,
        detalles_tecnicos, notas_casqueria, config_json, estado, fecha_creacion)
    Tabla fotos_creaciones: (id, creacion_id, foto_url)
    """
    try:
        vendedor_id      = request.form.get('vendedor_id', 1)
        nombre_modelo    = request.form.get('nombre_modelo')
        categoria        = request.form.get('categoria', 'Personalizado')
        detalles_tecnicos = request.form.get('detalles_tecnicos', '')
        notas_casqueria  = request.form.get('notas_casqueria', '')
        config_json      = request.form.get('config_json')

        if not nombre_modelo:
            return jsonify({'error': 'El nombre del modelo es obligatorio'}), 400

        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute("""
            INSERT INTO creaciones_vendedores
                (vendedor_id, nombre_modelo, categoria, detalles_tecnicos, notas_casqueria, config_json)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING id;
        """, (vendedor_id, nombre_modelo, categoria, detalles_tecnicos, notas_casqueria, config_json))
        creacion_id = cursor.fetchone()[0]

        for archivo in request.files.getlist('fotos'):
            if archivo and archivo.filename != '':
                respuesta_nube = cloudinary.uploader.upload(archivo, folder="creaciones")
                foto_url_nube = respuesta_nube.get('secure_url')
                cursor.execute(
                    "INSERT INTO fotos_creaciones (creacion_id, foto_url) VALUES (%s,%s);",
                    (creacion_id, foto_url_nube)
                )

        conexion.commit()
        return jsonify({'mensaje': '¡Creación guardada con éxito!', 'creacion_id': creacion_id}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print("Error al guardar la creación:", e)
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/creaciones', methods=['GET'])
def obtener_creaciones():
    """Devuelve las creaciones pendientes de aprobación (panel Admin)."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT c.id, c.nombre_modelo, c.categoria, c.detalles_tecnicos, c.notas_casqueria,
                   f.foto_url AS foto,
                   c.config_json, c.estado
            FROM creaciones_vendedores c
            LEFT JOIN LATERAL (
                SELECT foto_url FROM fotos_creaciones
                WHERE creacion_id = c.id
                ORDER BY id LIMIT 1
            ) f ON true
            WHERE c.estado = 'Pendiente'
            ORDER BY c.fecha_creacion DESC;
        """)
        creaciones = []
        for r in cursor.fetchall():
            creaciones.append({
                "id":       r[0], "nombre":   r[1], "categoria": r[2],
                "detalles": r[3], "notas":    r[4],
                "foto_url": limpiar_foto(r[5]),
                "config_json": r[6], "estado": r[7]
            })
        return jsonify(creaciones), 200
    except Exception as e:
        print("Error al obtener creaciones:", e)
        return jsonify({'error': 'Error al cargar creaciones'}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/creaciones/aprobar', methods=['POST'])
def aprobar_creacion():
    """Admin aprueba un diseño → pasa al catálogo oficial."""
    data        = request.json
    creacion_id = data.get('creacion_id')
    origen      = data.get('origen', 'Interno')
    precio_base = data.get('precio_base', 0)

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute("""
            SELECT c.nombre_modelo,
                   (SELECT foto_url FROM fotos_creaciones WHERE creacion_id = c.id LIMIT 1)
            FROM creaciones_vendedores c WHERE c.id = %s;
        """, (creacion_id,))
        creacion = cursor.fetchone()

        if not creacion:
            return jsonify({'error': 'Creación no encontrada'}), 404

        nombre   = creacion[0]
        foto_url = limpiar_foto(creacion[1])

        cursor.execute("""
            INSERT INTO catalogo_productos
                (nombre_modelo, precio_base, foto_url, es_plantilla, en_stock, origen_produccion)
            VALUES (%s,%s,%s,False,False,%s);
        """, (nombre, precio_base, foto_url, origen))

        cursor.execute(
            "UPDATE creaciones_vendedores SET estado = 'Aprobado' WHERE id = %s;",
            (creacion_id,)
        )
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Modelo aprobado y enviado al catálogo principal.'}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print("Error en aprobación:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/creaciones/rechazar', methods=['POST'])
def rechazar_creacion():
    """Admin rechaza un diseño con motivo opcional."""
    data        = request.json
    creacion_id = data.get('creacion_id')
    motivo      = data.get('motivo', '')

    if not creacion_id:
        return jsonify({'error': 'creacion_id es obligatorio'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute(
            "UPDATE creaciones_vendedores SET estado = 'Rechazado' WHERE id = %s;",
            (creacion_id,)
        )
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': f'Diseño rechazado. Motivo: {motivo}'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# 6. MÓDULO: USUARIOS Y LOGIN
# ==========================================

@app.route('/api/usuarios', methods=['GET'])
def obtener_usuarios():
    """Lista pública de usuarios (sin PIN ni datos sensibles)."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT id, nombre, rol, area_asignada FROM usuarios ORDER BY nombre;")
        usuarios = [
            {"id": r[0], "nombre": r[1], "rol": r[2], "area_asignada": r[3]} 
            for r in cursor.fetchall()
        ]
        return jsonify(usuarios), 200
    except Exception as e:
        return jsonify({'error': 'Error al cargar usuarios'}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/usuarios/detalle', methods=['GET'])
def obtener_usuarios_detalle():
    """Lista completa para panel Admin (sin PIN)."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT id, nombre, rol, email, area_asignada, empresa_nombre, empresa_ruc FROM usuarios ORDER BY nombre;")
        usuarios = [
            {"id": r[0], "nombre": r[1], "rol": r[2], "email": r[3], "area": r[4], "empresa": r[5], "ruc": r[6]}
            for r in cursor.fetchall()
        ]
        return jsonify(usuarios), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/usuarios/nuevo', methods=['POST'])
def crear_usuario():
    data = request.json
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            INSERT INTO usuarios (nombre, email, pin_acceso, contrasena, rol, area_asignada, empresa_nombre, empresa_ruc)
            VALUES (%s, %s, %s, '123456', %s, %s, %s, %s);
        """, (data['nombre'], data['correo'], data['pin'], data['rol'], data['area'], data['empresa_nombre'], data['empresa_ruc']))
        conexion.commit()
        return jsonify({'exito': True}), 201
    except Exception as e:
        print(f"❌ Error en crear_usuario: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/login', methods=['POST'])
@limiter.limit("10 per minute")   # Máximo 10 intentos por minuto por IP
def verificar_pin():
    try:
        usuario_id   = request.json.get('usuario_id')
        pin_ingresado = request.json.get('pin')

        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT id, nombre, rol, empresa_nombre, empresa_ruc, email, area_asignada, pin_acceso
            FROM usuarios
            WHERE id = %s AND pin_acceso = %s;
        """, (usuario_id, pin_ingresado))
        usuario = cursor.fetchone()

        if usuario:
            res = jsonify({
                "exito": True,
                "usuario": {
                    "id":            usuario[0],
                    "nombre":        usuario[1],
                    "rol":           usuario[2],
                    "empresa":       usuario[3],
                    "ruc":           usuario[4],
                    "email":         usuario[5],
                    "area_asignada": usuario[6]
                }
            })
            cursor.close(); release_db_connection(conexion)
            return res, 200

        cursor.close(); release_db_connection(conexion)
        return jsonify({"exito": False, "error": "PIN incorrecto"}), 401

    except Exception as e:
        print(f"❌ Error en login: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500


# ==========================================
# 7. MÓDULO: PROVEEDORES
# ==========================================

@app.route('/api/proveedores', methods=['GET'])
def obtener_proveedores():
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT id, nombre, especialidad, correo, telefono FROM proveedores ORDER BY nombre;")
        provs = [
            {"id": r[0], "nombre": r[1], "especialidad": r[2], "correo": r[3], "telefono": r[4]}
            for r in cursor.fetchall()
        ]
        return jsonify(provs), 200
    except Exception as e:
        print(f"❌ Error en obtener_proveedores: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/proveedores/nuevo', methods=['POST'])
def crear_proveedor():
    data = request.json
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            INSERT INTO proveedores (nombre, especialidad, correo, telefono)
            VALUES (%s,%s,%s,%s);
        """, (data['nombre'], data['especialidad'], data['correo'], data['telefono']))
        conexion.commit()
        return jsonify({'exito': True}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/taller/ticket/<int:id>/finalizar', methods=['POST'])
def finalizar_ticket(id):
    """Marca un ticket como terminado, guarda la foto de evidencia,
    y desbloquea automáticamente los tickets de tapicería si sus prerrequisitos están completos."""
    try:
        if 'foto' not in request.files:
            return jsonify({'error': 'La foto de evidencia es obligatoria'}), 400
        
        foto = request.files['foto']
        if foto.filename == '':
            return jsonify({'error': 'No se seleccionó ninguna foto'}), 400

        # Subir foto de evidencia a Cloudinary
        respuesta_nube = cloudinary.uploader.upload(foto, folder="evidencias")
        foto_url_final = respuesta_nube.get('secure_url')

        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        # 1. Marcar este ticket como Terminado
        cursor.execute("""
            UPDATE tickets_produccion
            SET estado_ticket = 'Terminado',
                foto_evidencia = %s,
                fecha_fin      = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING item_id, area_trabajo;
        """, (foto_url_final, id))
        row = cursor.fetchone()

        desbloqueados = 0
        if row:
            item_id   = row[0]
            area_term = row[1]

            # 2. Verificar si hay tickets BLOQUEADOS en Tapicería para este mismo item
            #    Prerrequisito: ESTRUCTURAS_MUEBLES/SILLAS + CORTE_Y_CONTROL_TELAS deben estar Terminados
            cursor.execute("""
                SELECT t.id, t.area_trabajo
                FROM tickets_produccion t
                WHERE t.item_id = %s
                  AND t.estado_ticket = 'Bloqueado'
                  AND t.area_trabajo IN ('TAPICERIA_SOFAS','TAPICERIA_SILLAS','ARMADO_COJINES')
            """, (item_id,))
            tickets_bloqueados = cursor.fetchall()

            for tb_id, tb_area in tickets_bloqueados:
                # Definir qué áreas deben estar Terminadas para desbloquear este ticket
                if tb_area in ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS'):
                    areas_req = ['ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS_SILLAS', 'CORTE_Y_CONTROL_TELAS', 'TELAS']
                elif tb_area == 'ARMADO_COJINES':
                    areas_req = ['CORTE_Y_CONTROL_TELAS', 'TELAS']
                else:
                    continue

                # Verificar cuántos prerrequisitos están Terminados para este item
                placeholders_req = ','.join(['%s'] * len(areas_req))
                cursor.execute(f"""
                    SELECT COUNT(*) FROM tickets_produccion
                    WHERE item_id = %s
                      AND area_trabajo IN ({placeholders_req})
                      AND estado_ticket = 'Terminado'
                """, (item_id, *areas_req))
                terminados_req = cursor.fetchone()[0]

                # También verificamos cuántos prerrequisitos EXISTEN (para no desbloquear si faltan)
                cursor.execute(f"""
                    SELECT COUNT(*) FROM tickets_produccion
                    WHERE item_id = %s
                      AND area_trabajo IN ({placeholders_req})
                """, (item_id, *areas_req))
                total_req = cursor.fetchone()[0]

                # Desbloquear solo si todos los prerrequisitos que EXISTEN están Terminados
                # (al menos debe haber 1 prerrequisito existente)
                if total_req > 0 and terminados_req >= total_req:
                    nuevo_estado_tb = 'En Proceso' if True else 'Pendiente'
                    cursor.execute("""
                        UPDATE tickets_produccion
                        SET estado_ticket = 'En Proceso',
                            fecha_inicio  = CURRENT_TIMESTAMP
                        WHERE id = %s AND estado_ticket = 'Bloqueado'
                    """, (tb_id,))
                    desbloqueados += cursor.rowcount

        conexion.commit()
        msg = 'Ticket finalizado correctamente'
        if desbloqueados > 0:
            msg += f'. {desbloqueados} ticket(s) de tapicería desbloqueado(s) automáticamente.'
        return jsonify({'exito': True, 'mensaje': msg, 'desbloqueados': desbloqueados}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/taller/cola-recojo', methods=['GET'])
def obtener_cola_recojo():
    """
    Retorna los tickets de ESTRUCTURAS que ya están Terminados
    y cuyos tickets de Tapicería aún están Bloqueados (esperan ser recogidos y entregados al tapicero).
    Esto permite al Jefe/Admin ver qué estructuras están listas para que el chofer las recoja.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT
                t.id                AS ticket_id,
                t.area_trabajo,
                i.producto,
                v.codigo_venta,
                v.nombre_cliente,
                COALESCE(u.nombre, 'Sin asignar') AS operario,
                t.fecha_fin,
                COALESCE(i.foto_url, '')           AS foto_url,
                COALESCE(t.ticket_details_override, i.color_tela, '') AS especificaciones,
                t.foto_evidencia,
                v.direccion_cliente,
                v.fecha_entrega,
                t.item_id,
                -- Tapicero pre-asignado (si ya fue asignado al ticket bloqueado)
                COALESCE(ut.nombre, 'Sin asignar') AS tapicero_nombre
            FROM tickets_produccion t
            JOIN items_venta i    ON t.item_id   = i.id
            JOIN ventas v         ON i.venta_id  = v.id
            LEFT JOIN usuarios u  ON t.trabajador_asignado_id = u.id
            -- Buscar el ticket de tapicería bloqueado para este item
            LEFT JOIN tickets_produccion tap
                ON tap.item_id = t.item_id
               AND tap.area_trabajo IN ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS')
               AND tap.estado_ticket = 'Bloqueado'
            LEFT JOIN usuarios ut ON tap.trabajador_asignado_id = ut.id
            WHERE t.area_trabajo IN ('ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS_SILLAS')
              AND t.estado_ticket = 'Terminado'
              AND tap.id IS NOT NULL
            ORDER BY t.fecha_fin DESC;
        """)
        resultado = [
            {
                "ticket_id":       r[0],
                "area":            r[1],
                "producto":        r[2],
                "codigo_venta":    r[3],
                "cliente":         r[4],
                "operario":        r[5],
                "fecha_fin":       r[6].strftime('%d/%m/%Y %H:%M') if r[6] else 'S/F',
                "foto_url":        limpiar_foto(r[7]),
                "especificaciones": r[8] or '',
                "foto_evidencia":  r[9] if r[9] else '',  # foto_evidencia_url de Cloudinary
                "direccion":       r[10] or '',
                "fecha_entrega":   r[11].strftime('%d/%m/%Y') if r[11] else 'S/F',
                "item_id":         r[12],
                "tapicero":        r[13],
            }
            for r in cursor.fetchall()
        ]
        return jsonify(resultado), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# 8. MÓDULO: TALLER Y KANBAN
# ==========================================

@app.route('/api/taller/tickets', methods=['GET'])
def obtener_tickets_taller():
    area_filtro = request.args.get('area')
    operario_id = request.args.get('operario_id')
    AREA_ALIASES = {
        'TELAS':                  ['TELAS', 'CORTE_Y_CONTROL_TELAS'],
        'CORTE_Y_CONTROL_TELAS':  ['CORTE_Y_CONTROL_TELAS', 'TELAS'],
        'TAPICERIA':              ['TAPICERIA', 'TAPICERIA_SOFAS', 'TAPICERIA_SILLAS'],
        'TAPICERIA_SOFAS':        ['TAPICERIA_SOFAS', 'TAPICERIA'],
        'TAPICERIA_SILLAS':       ['TAPICERIA_SILLAS', 'TAPICERIA'],
        'ESTRUCTURAS':            ['ESTRUCTURAS', 'ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS_SILLAS', 'CARPINTERIA'],
        'ESTRUCTURAS_MUEBLES':    ['ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS', 'CARPINTERIA'],
        'ESTRUCTURAS_SILLAS':     ['ESTRUCTURAS_SILLAS', 'ESTRUCTURAS', 'CARPINTERIA'],
        'CARPINTERIA':            ['CARPINTERIA', 'ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS_SILLAS'],
    }
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        query = """
            SELECT t.id, i.producto, t.estado_ticket, t.area_trabajo, t.ticket_details_override,
                   t.trabajador_asignado_id, v.codigo_venta, i.color_tela, t.item_id,
                   i.foto_url, t.foto_evidencia, u.nombre
            FROM tickets_produccion t
            JOIN items_venta i ON t.item_id  = i.id
            JOIN ventas v      ON i.venta_id = v.id
            LEFT JOIN usuarios u ON t.trabajador_asignado_id = u.id
            WHERE 1=1
        """
        params = []
        if area_filtro:
            areas_buscar = AREA_ALIASES.get(area_filtro, [area_filtro])
            placeholders = ','.join(['%s'] * len(areas_buscar))
            query += f' AND t.area_trabajo IN ({placeholders})'
            params.extend(areas_buscar)
        if operario_id:
            query += ' AND t.trabajador_asignado_id = %s'
            params.append(int(operario_id))
        query += " ORDER BY t.etapa ASC, t.id DESC;"

        cursor.execute(query, params)
        tickets = []
        raw_rows = cursor.fetchall()

        # Calculamos items_incompletos solo de los items que aparecen en los resultados,
        # no de toda la tabla. Esto evita el full scan masivo.
        item_ids_en_resultado = {row[8] for row in raw_rows if row[3] == 'DESPACHO_CENTRAL'}
        items_incompletos = set()
        if item_ids_en_resultado:
            placeholders_items = ','.join(['%s'] * len(item_ids_en_resultado))
            cursor.execute(f"""
                SELECT DISTINCT item_id FROM tickets_produccion
                WHERE item_id IN ({placeholders_items})
                  AND estado_ticket != 'Terminado'
                  AND area_trabajo != 'DESPACHO_CENTRAL'
            """, tuple(item_ids_en_resultado))
            items_incompletos = {r[0] for r in cursor.fetchall()}

        for row in raw_rows:
            estado = row[2] # t.estado_ticket
            # Si es despacho y el item tiene cosas pendientes en taller, forzamos BLOQUEADO
            if row[3] == 'DESPACHO_CENTRAL' and row[8] in items_incompletos: # row[3] is t.area_trabajo, row[8] is t.item_id
                estado = 'Bloqueado'

            tickets.append({
                "id":              row[0],
                "producto":        f"{row[1]} (Ref: {row[6]})",
                "estado":          estado,
                "area":            row[3],
                "trabajador":      row[5],  # trabajador_asignado_id
                "especificaciones": row[4] if row[4] else (row[7] if row[7] else "Sin notas técnicas"),
                "foto":            limpiar_foto(row[9]),
                "trabajador_nombre": row[11] if row[11] else 'Sin asignar',
                "item_id":         row[8]
            })
        return jsonify(tickets), 200
    except Exception as e:
        print("Error en tickets taller:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/taller/tickets_pendientes', methods=['GET'])
def obtener_tickets_pendientes():
    """Tickets sin maestro asignado (incluye Bloqueados de Tapicería para pre-asignación)."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT t.id, v.codigo_venta, v.nombre_cliente, i.producto,
                   t.area_trabajo, t.estado_ticket, v.fecha_entrega, i.color_tela
            FROM tickets_produccion t
            JOIN items_venta i ON t.item_id  = i.id
            JOIN ventas v      ON i.venta_id = v.id
            WHERE t.trabajador_asignado_id IS NULL
              AND t.estado_ticket != 'Terminado'
            ORDER BY v.fecha_entrega ASC;
        """)
        res = []
        for t in cursor.fetchall():
            res.append({
                "ticket_id":      t[0], "codigo":   t[1], "cliente": t[2],
                "producto":       t[3], "area":     t[4], "estado":  t[5],
                "entrega":        t[6].strftime('%d/%m/%Y') if t[6] else "S/F",
                "especificaciones": t[7]
            })
        return jsonify(res), 200
    except Exception as e:
        print("Error en tickets pendientes:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/taller/asignar', methods=['POST'])
def asignar_maestro_ticket():
    data        = request.json
    ticket_id   = data.get('ticket_id')
    trabajador_id = data.get('trabajador_id')
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        # Obtener estado actual del ticket
        cursor.execute("SELECT estado_ticket FROM tickets_produccion WHERE id = %s", (ticket_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        estado_actual = row[0]
        # Si está Bloqueado, solo asignamos el trabajador pero NO cambiamos el estado
        # (se desbloqueará automáticamente cuando los prerrequisitos terminen)
        # Si está Pendiente, pasamos a En Proceso para que el operario lo vea activo
        nuevo_estado = estado_actual if estado_actual == 'Bloqueado' else 'En Proceso'
        cursor.execute("""
            UPDATE tickets_produccion
            SET trabajador_asignado_id = %s,
                estado_ticket          = %s,
                fecha_inicio           = CASE WHEN %s = 'En Proceso' THEN CURRENT_TIMESTAMP ELSE fecha_inicio END
            WHERE id = %s;
        """, (trabajador_id, nuevo_estado, nuevo_estado, ticket_id))
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Maestro asignado correctamente', 'estado': nuevo_estado}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# 9. MÓDULO: INVENTARIO DE MATERIALES
# ==========================================

@app.route('/api/taller/fichatecnica-skus', methods=['GET'])
def obtener_fotos_skus():
    """
    Recibe ?skus=TEL-0003,COJ-0001 y devuelve foto+nombre de cada SKU
    buscando en maestro_telas y maestro_disenos_cojin.
    """
    skus_param = request.args.get('skus', '')
    if not skus_param:
        return jsonify([]), 200

    skus = [s.strip().upper() for s in skus_param.split(',') if s.strip()]
    if not skus:
        return jsonify([]), 200

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        placeholders = ','.join(['%s'] * len(skus))

        # Buscar en maestro_telas
        cursor.execute(f"""
            SELECT sku, CONCAT(coleccion, ' - ', color) AS nombre, foto_url, 'tela' AS tipo
            FROM maestro_telas
            WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados = [{'sku': r[0], 'nombre': r[1], 'foto_url': r[2], 'tipo': r[3]} for r in cursor.fetchall()]

        # Buscar en maestro_disenos_cojin
        cursor.execute(f"""
            SELECT sku, nombre_diseno AS nombre, foto_url, 'cojin' AS tipo
            FROM maestro_disenos_cojin
            WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados += [{'sku': r[0], 'nombre': r[1], 'foto_url': r[2], 'tipo': r[3]} for r in cursor.fetchall()]

        return jsonify(resultados), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/taller/inventario', methods=['GET'])
def obtener_inventario():
    """Vista unificada de todos los materiales maestros."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT id, nombre_modelo AS nombre, COALESCE(estado,'Disponible'), 'TABLERO'    AS cat, origen_produccion FROM maestro_tableros
            UNION ALL
            SELECT id, color,                   COALESCE(estado,'Disponible'), 'TELA'       AS cat, origen_produccion FROM maestro_telas
            UNION ALL
            SELECT id, nombre_diseno,            COALESCE(estado,'Disponible'), 'COJIN'      AS cat, origen_produccion FROM maestro_disenos_cojin
            UNION ALL
            SELECT id, modelo,                   COALESCE(estado,'Disponible'), 'BASE'       AS cat, origen_produccion FROM maestro_bases
            UNION ALL
            SELECT id, modelo,                   COALESCE(estado,'Disponible'), 'BASE-COMEDOR' AS cat, origen_produccion FROM maestro_bases_comedor
            UNION ALL
            SELECT id, modelo,                   COALESCE(estado,'Disponible'), 'SILLA'      AS cat, origen_produccion FROM maestro_sillas
            UNION ALL
            SELECT id, modelo,                   COALESCE(estado,'Disponible'), 'BUTACA'     AS cat, origen_produccion FROM maestro_butacas
            ORDER BY cat, nombre;
        """)
        insumos = [
            {"id": r[0], "nombre": r[1], "estado": r[2], "categoria": r[3]}
            for r in cursor.fetchall()
        ]
        return jsonify(insumos), 200
    except Exception as e:
        print("Error en inventario:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/inventario/actualizar', methods=['POST'])
def actualizar_estado_inventario():
    data       = request.json
    item_id    = data.get('id')
    categoria  = data.get('categoria')
    nuevo_estado = data.get('estado')

    tablas_permitidas = {
        'TELA':         'maestro_telas',
        'COJIN':        'maestro_disenos_cojin',
        'BASE':         'maestro_bases',
        'BASE-COMEDOR': 'maestro_bases_comedor',
        'TABLERO':      'maestro_tableros',
        'SILLA':        'maestro_sillas',
        'BUTACA':       'maestro_butacas',
    }
    tabla = tablas_permitidas.get(categoria)
    if not tabla:
        return jsonify({'error': 'Categoría no válida'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute(f"UPDATE {tabla} SET estado = %s WHERE id = %s;", (nuevo_estado, item_id))
        conexion.commit()
        return jsonify({'exito': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# 10. MÓDULO: LOGÍSTICA EXTERNA
# ==========================================

@app.route('/api/logistica', methods=['GET'])
def obtener_logistica():
    """
    Lista todos los ítems de logística externa.
    Tabla: logistica_externa
      (id, venta_id, insumo_nombre, sku, proveedor_id, precio_cotizado,
       fecha_entrega_proveedor, estado)
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT
                l.id,
                v.codigo_venta,
                l.insumo_nombre,
                l.sku,
                COALESCE(p.nombre, 'Sin asignar') AS proveedor,
                l.precio_cotizado,
                l.fecha_entrega_proveedor,
                l.estado
            FROM logistica_externa l
            JOIN ventas v           ON l.venta_id    = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            ORDER BY l.estado ASC, l.id DESC;
        """)
        items = []
        for r in cursor.fetchall():
            items.append({
                "id":                    r[0],
                "codigo_venta":          r[1],
                "insumo":                r[2],
                "sku":                   r[3],
                "proveedor":             r[4],
                "precio_cotizado":       float(r[5]) if r[5] else None,
                "fecha_entrega_proveedor": r[6].strftime('%d/%m/%Y') if r[6] else None,
                "estado":                r[7]
            })
        return jsonify(items), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/logistica/actualizar', methods=['POST'])
def actualizar_logistica():
    """
    Permite al Admin actualizar proveedor, precio cotizado,
    fecha de entrega estimada y estado de un ítem de logística.
    """
    data = request.json
    logistica_id           = data.get('id')
    proveedor_id           = data.get('proveedor_id')
    precio_cotizado        = data.get('precio_cotizado')
    fecha_entrega_proveedor = data.get('fecha_entrega_proveedor')   # 'YYYY-MM-DD'
    estado                 = data.get('estado')

    if not logistica_id:
        return jsonify({'error': 'id es obligatorio'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            UPDATE logistica_externa
            SET proveedor_id            = COALESCE(%s, proveedor_id),
                precio_cotizado         = COALESCE(%s, precio_cotizado),
                fecha_entrega_proveedor = COALESCE(%s::date, fecha_entrega_proveedor),
                estado                  = COALESCE(%s, estado)
            WHERE id = %s;
        """, (proveedor_id, precio_cotizado, fecha_entrega_proveedor, estado, logistica_id))
        conexion.commit()
        return jsonify({'exito': True}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# 11. MÓDULO: RECETAS DE MUEBLES (BOM)
# ==========================================

@app.route('/api/recetas/<int:producto_id>', methods=['GET'])
def obtener_receta(producto_id):
    """
    Devuelve la lista de materiales (BOM) de un producto.
    Tabla: recetas_muebles (id, producto_id, insumo_id, cantidad_necesaria)
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT r.id, i.nombre_insumo, r.insumo_id, r.cantidad_necesaria, i.cantidad_actual
            FROM recetas_muebles r
            JOIN inventario_insumos i ON r.insumo_id = i.id
            WHERE r.producto_id = %s
            ORDER BY i.nombre_insumo;
        """, (producto_id,))
        receta = [
            {
                "id":                 r[0],
                "nombre_insumo":      r[1],
                "insumo_id":          r[2],
                "cantidad_necesaria": r[3],
                "stock_actual":       r[4]
            }
            for r in cursor.fetchall()
        ]
        return jsonify(receta), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/recetas/nueva', methods=['POST'])
def agregar_ingrediente_receta():
    """Agrega un insumo a la receta de un producto."""
    data              = request.json
    producto_id       = data.get('producto_id')
    insumo_id         = data.get('insumo_id')
    cantidad_necesaria = data.get('cantidad_necesaria', 1)

    if not producto_id or not insumo_id:
        return jsonify({'error': 'producto_id e insumo_id son obligatorios'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            INSERT INTO recetas_muebles (producto_id, insumo_id, cantidad_necesaria)
            VALUES (%s,%s,%s) RETURNING id;
        """, (producto_id, insumo_id, cantidad_necesaria))
        nuevo_id = cursor.fetchone()[0]
        conexion.commit()
        return jsonify({'exito': True, 'id': nuevo_id}), 201
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# 12. MÓDULO: SUGERENCIAS DE INSUMOS (OPTIMIZADO)
# ==========================================

@app.route('/api/sugerencias', methods=['POST'])
def guardar_sugerencia():
    """Vendedor sugiere un insumo usando los formularios detallados."""
    try:
        nombre     = request.form.get('nombre')
        tipo       = request.form.get('tipo') # 'tela', 'tablero', 'base', etc.
        usuario_id = request.form.get('usuario_id')
        datos_json = request.form.get('datos_json')

        if not nombre or not tipo:
            return jsonify({'error': 'El nombre y tipo de insumo son obligatorios'}), 400

        foto_ruta = "imagenes/sin_foto.jpg"
        if 'foto' in request.files:
            foto_file = request.files['foto']
            if foto_file.filename != '':
                respuesta_nube = cloudinary.uploader.upload(foto_file, folder="sugerencias")
                foto_ruta = respuesta_nube.get('secure_url')

        conexion = get_db_connection()
        cursor   = conexion.cursor()
        
        cursor.execute("""
            INSERT INTO sugerencias_insumos (nombre, tipo, foto_ref, usuario_id, datos_json, estado)
            VALUES (%s, %s, %s, %s, %s, 'Pendiente') RETURNING id;
        """, (nombre, tipo, foto_ruta, usuario_id, datos_json))
        
        sugerencia_id = cursor.fetchone()[0]
        conexion.commit()
        return jsonify({'exito': True, 'id': sugerencia_id, 'mensaje': 'Sugerencia enviada al Gestor de Aprobación'}), 201

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)


@app.route('/api/sugerencias', methods=['GET'])
def obtener_sugerencias():
    """Trae todas las sugerencias de insumos pendientes para el panel Admin."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT s.id, s.nombre, s.tipo, s.foto_ref, 
                   COALESCE(u.nombre, 'Vendedor') AS vendedor, 
                   s.datos_json, s.estado
            FROM sugerencias_insumos s
            LEFT JOIN usuarios u ON s.usuario_id = u.id
            WHERE s.estado = 'Pendiente'
            ORDER BY s.fecha_registro DESC;
        """)
        resultado = [{
            "id": r[0], "nombre": r[1], "tipo": r[2],
            "foto_url": r[3], "vendedor": r[4], 
            "datos_json": r[5], "estado": r[6]
        } for r in cursor.fetchall()]
        return jsonify(resultado), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)


@app.route('/api/sugerencias/aprobar', methods=['POST'])
def aprobar_sugerencia_insumo():
    """Admin aprueba la sugerencia del vendedor, define Origen (Make vs Buy) y genera SKU."""
    data          = request.json
    sugerencia_id = data.get('sugerencia_id')
    origen        = data.get('origen', 'Value') # 'Interno' o 'Externo'
    
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Buscar la sugerencia
        cursor.execute("SELECT tipo, foto_ref, datos_json FROM sugerencias_insumos WHERE id = %s;", (sugerencia_id,))
        sug = cursor.fetchone()
        if not sug:
            return jsonify({'error': 'Sugerencia no encontrada'}), 404

        tipo_material, foto_ruta, datos_raw = sug
        datos = json.loads(datos_raw) if datos_raw else {}

        nuevo_sku = ""

        if tipo_material == 'tela':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_telas")
            nuevo_sku = f"TEL-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_telas (sku, proveedor, coleccion, color, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,'Disponible');
            """, (nuevo_sku, datos.get('proveedor'), datos.get('coleccion'), datos.get('color'), foto_ruta, origen))

        elif tipo_material == 'cojin':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_disenos_cojin")
            nuevo_sku = f"COJ-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_disenos_cojin (sku, nombre_diseno, tipo_tela, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,'Disponible');
            """, (nuevo_sku, datos.get('nombre_diseno'), datos.get('tipo_tela'), foto_ruta, origen))

        elif tipo_material == 'base':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_bases")
            nuevo_sku = f"BAS-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_bases (sku, tipo, material, modelo, color, medida_altura, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'Disponible');
            """, (nuevo_sku, datos.get('tipo'), datos.get('material'), datos.get('modelo'), datos.get('color'), datos.get('medida_altura'), foto_ruta, origen))

        elif tipo_material == 'tablero':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_tableros")
            nuevo_sku = f"TAB-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_tableros (sku, material_base, nombre_modelo, color_veta, acabado, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,%s,'Disponible');
            """, (nuevo_sku, datos.get('material_base'), datos.get('nombre_modelo'), datos.get('color_veta'), datos.get('acabado'), foto_ruta, origen))

        elif tipo_material == 'base-comedor':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_bases_comedor")
            nuevo_sku = f"BAC-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_bases_comedor (sku, material, modelo, color, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,'Disponible');
            """, (nuevo_sku, datos.get('material'), datos.get('modelo'), datos.get('color'), foto_ruta, origen))

        elif tipo_material == 'silla':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_sillas")
            nuevo_sku = f"SIL-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_sillas (sku, material, modelo, color_estructura, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,'Disponible');
            """, (nuevo_sku, datos.get('material'), datos.get('modelo'), datos.get('color_estructura'), foto_ruta, origen))

        elif tipo_material == 'butaca':
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM maestro_butacas")
            nuevo_sku = f"BUT-{str(cursor.fetchone()[0]+1).zfill(4)}"
            cursor.execute("""
                INSERT INTO maestro_butacas (sku, material, modelo, color_estructura, foto_url, origen_produccion, estado)
                VALUES (%s,%s,%s,%s,%s,%s,'Disponible');
            """, (nuevo_sku, datos.get('material'), datos.get('modelo'), datos.get('color_estructura'), foto_ruta, origen))

        # Marcar sugerencia como aprobada
        cursor.execute("UPDATE sugerencias_insumos SET estado = 'Aprobado' WHERE id = %s;", (sugerencia_id,))
        conexion.commit()

        return jsonify({'exito': True, 'mensaje': f'Insumo oficializado con éxito. Nuevo SKU: {nuevo_sku}'}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)

@app.route('/api/taller/ticket/<int:ticket_id>/derivar', methods=['POST'])
def derivar_ticket_con_foto(ticket_id):
    """
    Pase de Posta — Área de Telas deriva el material al área final.
    Recibe: foto (archivo), nueva_area (str), nuevo_trabajador_id (int).
    Guarda la foto como evidencia, cambia el área y el trabajador,
    y devuelve el estado a 'Pendiente' para que el tapicero/cojinero pueda tomarlo.
    """
    nueva_area         = request.form.get('nueva_area')
    nuevo_trabajador_id = request.form.get('nuevo_trabajador_id')

    if not nueva_area or not nuevo_trabajador_id:
        return jsonify({'error': 'nueva_area y nuevo_trabajador_id son obligatorios'}), 400

    # Subir foto de evidencia a Cloudinary
    foto_ruta = None
    if 'foto' in request.files and request.files['foto'].filename != '':
        foto_file = request.files['foto']
        respuesta_nube = cloudinary.uploader.upload(foto_file, folder="derivaciones")
        foto_ruta = respuesta_nube.get('secure_url')

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Verificar que el ticket existe
        cursor.execute("SELECT id FROM tickets_produccion WHERE id = %s", (ticket_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Ticket no encontrado'}), 404

        # Actualizar: cambiar área, trabajador, foto y volver a Pendiente
        cursor.execute("""
            UPDATE tickets_produccion
            SET area_trabajo            = %s,
                trabajador_asignado_id  = %s,
                foto_evidencia          = COALESCE(%s, foto_evidencia),
                estado_ticket           = 'Pendiente',
                fecha_inicio            = NULL,
                fecha_fin               = NULL
            WHERE id = %s
        """, (nueva_area, int(nuevo_trabajador_id), foto_ruta, ticket_id))

        conexion.commit()
        return jsonify({'exito': True, 'nueva_area': nueva_area}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/taller/ticket/derivar', methods=['POST'])
def derivar_ticket():
    """
    Crea tickets de Tapicería y/o Cojines a partir del fin de Telas.
    BUG FIX: ahora acepta tapicero_id + cojinero_id en un solo POST,
    verifica duplicados y actualiza estado del ticket padre a Terminado.
    """
    data = request.json
    ticket_padre_id = data.get('ticket_padre_id')
    tapicero_id     = data.get('tapicero_id')
    cojinero_id     = data.get('cojinero_id')          # puede ser None
    area_tapiceria  = data.get('area_tapiceria', 'TAPICERIA_SOFAS')

    if not ticket_padre_id or not tapicero_id:
        return jsonify({'error': 'ticket_padre_id y tapicero_id son obligatorios'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Obtener item_id del ticket original
        cursor.execute("SELECT item_id FROM tickets_produccion WHERE id = %s", (ticket_padre_id,))
        res = cursor.fetchone()
        if not res:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        item_id = res[0]

        # Actualizar o crear ticket de Tapicería
        # Si ya existe un ticket Bloqueado de tapicería (creado al momento de la venta),
        # lo actualizamos con el tapicero y lo pasamos a En Proceso.
        cursor.execute("""
            UPDATE tickets_produccion
            SET trabajador_asignado_id = %s,
                estado_ticket = 'En Proceso',
                fecha_inicio = CURRENT_TIMESTAMP
            WHERE item_id = %s AND area_trabajo = %s AND estado_ticket IN ('Bloqueado', 'Pendiente')
        """, (tapicero_id, item_id, area_tapiceria))

        # Si no actualizó ninguna fila (no había ticket previo), insertar uno nuevo
        if cursor.rowcount == 0:
            cursor.execute("""
                INSERT INTO tickets_produccion
                    (item_id, area_trabajo, trabajador_asignado_id, estado_ticket, etapa)
                SELECT %s, %s, %s, 'En Proceso', 2
                WHERE NOT EXISTS (
                    SELECT 1 FROM tickets_produccion
                    WHERE item_id = %s AND area_trabajo = %s AND estado_ticket != 'Terminado'
                )
            """, (item_id, area_tapiceria, tapicero_id, item_id, area_tapiceria))

        # Crear ticket de Cojines si aplica (evitar duplicados)
        if cojinero_id:
            cursor.execute("""
                INSERT INTO tickets_produccion
                    (item_id, area_trabajo, trabajador_asignado_id, estado_ticket, etapa)
                SELECT %s, 'ARMADO_COJINES', %s, 'En Proceso', 2
                WHERE NOT EXISTS (
                    SELECT 1 FROM tickets_produccion
                    WHERE item_id = %s AND area_trabajo = 'ARMADO_COJINES' AND estado_ticket != 'Terminado'
                )
            """, (item_id, cojinero_id, item_id))

        conexion.commit()
        return jsonify({'exito': True}), 201
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)
# ==========================================
# 13. MÓDULO: DESPACHO — ASIGNAR CHOFER
# ==========================================

@app.route('/api/despacho/asignar-chofer', methods=['POST'])
def asignar_chofer_despacho():
    """
    Permite al personal de Despacho asignar un chofer a un ticket de DESPACHO_CENTRAL.
    Solo funciona si el ticket NO está Bloqueado (todas las áreas previas terminadas).
    """
    data      = request.json
    ticket_id = data.get('ticket_id')
    chofer_id = data.get('chofer_id')

    if not ticket_id or not chofer_id:
        return jsonify({'error': 'ticket_id y chofer_id son obligatorios'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Verificar que el ticket existe y es de DESPACHO_CENTRAL
        cursor.execute("""
            SELECT t.id, t.item_id, t.estado_ticket
            FROM tickets_produccion t
            WHERE t.id = %s AND t.area_trabajo = 'DESPACHO_CENTRAL'
        """, (ticket_id,))
        ticket = cursor.fetchone()
        if not ticket:
            return jsonify({'error': 'Ticket de despacho no encontrado'}), 404

        estado_actual = ticket[2]
        item_id       = ticket[1]

        # Verificar que no hay partes pendientes (doble check en backend)
        cursor.execute("""
            SELECT COUNT(*) FROM tickets_produccion
            WHERE item_id = %s
              AND area_trabajo != 'DESPACHO_CENTRAL'
              AND estado_ticket != 'Terminado'
        """, (item_id,))
        pendientes = cursor.fetchone()[0]

        if pendientes > 0:
            return jsonify({'error': f'Aún hay {pendientes} área(s) sin terminar. No se puede despachar.'}), 409

        # Asignar chofer y cambiar estado a En Proceso
        cursor.execute("""
            UPDATE tickets_produccion
            SET trabajador_asignado_id = %s,
                estado_ticket          = 'En Proceso',
                fecha_inicio           = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (chofer_id, ticket_id))

        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Chofer asignado. El despacho está activo.'}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/usuarios/por-area/<string:area>', methods=['GET'])
def obtener_usuarios_por_area(area):
    """
    Devuelve operarios del área indicada + Jefes de Taller y Admins como respaldo.
    Incluye aliases para cubrir nombres históricos distintos que
    representan la misma área (ej: TELAS = CORTE_Y_CONTROL_TELAS).
    Los operarios del área exacta aparecen primero.
    """
    AREA_ALIASES = {
        'CORTE_Y_CONTROL_TELAS': ['CORTE_Y_CONTROL_TELAS', 'TELAS'],
        'TELAS':                  ['TELAS', 'CORTE_Y_CONTROL_TELAS'],
        'TAPICERIA_SOFAS':        ['TAPICERIA_SOFAS', 'TAPICERIA'],
        'TAPICERIA_SILLAS':       ['TAPICERIA_SILLAS', 'TAPICERIA'],
        'ESTRUCTURAS_MUEBLES':    ['ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS', 'CARPINTERIA'],
        'ESTRUCTURAS_SILLAS':     ['ESTRUCTURAS_SILLAS',  'ESTRUCTURAS', 'CARPINTERIA'],
        'ARMADO_COJINES':         ['ARMADO_COJINES', 'COJINES'],
        'PREPARACION_PATAS_ZOCALO': ['PREPARACION_PATAS_ZOCALO', 'PATAS', 'ZOCALO'],
        'TABLEROS_Y_PIEDRAS':     ['TABLEROS_Y_PIEDRAS', 'TABLEROS'],
        'DESPACHO_CENTRAL':       ['DESPACHO_CENTRAL', 'DESPACHO'],
    }
    area_upper = area.upper()
    areas_buscar = AREA_ALIASES.get(area_upper, [area_upper])
    # Añadir también la versión original sin transformar por si acaso
    areas_buscar_set = list(dict.fromkeys([a.upper() for a in areas_buscar] + [area_upper]))

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        placeholders = ",".join(["%s"] * len(areas_buscar_set))
        # Traer operarios del área primero, luego jefes/admins como respaldo
        # Usamos UPPER() en ambos lados para comparación case-insensitive
        cursor.execute(f"""
            SELECT id, nombre, rol, area_asignada,
                CASE
                    WHEN UPPER(COALESCE(area_asignada,'')) IN ({placeholders}) THEN 0
                    ELSE 1
                END AS orden
            FROM usuarios
            WHERE UPPER(COALESCE(area_asignada,'')) IN ({placeholders})
               OR rol IN ('Admin', 'Jefe_Taller', 'JEFE_TALLER')
            ORDER BY orden ASC, nombre ASC;
        """, (*areas_buscar_set, *areas_buscar_set))
        usuarios = [
            {"id": r[0], "nombre": r[1], "rol": r[2], "area": r[3] or ''}
            for r in cursor.fetchall()
        ]
        return jsonify(usuarios), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/despacho/progreso/<int:item_id>', methods=['GET'])
def progreso_despacho(item_id):
    """
    Devuelve el avance de cada área de producción para un item,
    usado por el panel de Despacho para mostrar qué partes faltan.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT area_trabajo, estado_ticket, trabajador_asignado_id,
                   COALESCE(u.nombre, 'Sin asignar') AS trabajador_nombre
            FROM tickets_produccion t
            LEFT JOIN usuarios u ON t.trabajador_asignado_id = u.id
            WHERE t.item_id = %s AND t.area_trabajo != 'DESPACHO_CENTRAL'
            ORDER BY t.etapa ASC, t.id ASC
        """, (item_id,))
        partes = [
            {
                "area":      r[0],
                "estado":    r[1],
                "trabajador": r[3]
            }
            for r in cursor.fetchall()
        ]
        total      = len(partes)
        terminados = sum(1 for p in partes if p['estado'] == 'Terminado')
        return jsonify({
            "partes":    partes,
            "total":     total,
            "terminados": terminados,
            "listo":     total > 0 and terminados == total
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/usuarios/choferes', methods=['GET'])
def obtener_choferes():
    """Lista usuarios con rol Chofer o área DESPACHO."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT id, nombre, rol, area_asignada
            FROM usuarios
            WHERE area_asignada = 'DESPACHO' OR rol = 'Chofer'
            ORDER BY nombre;
        """)
        choferes = [
            {"id": r[0], "nombre": r[1], "rol": r[2], "area": r[3]}
            for r in cursor.fetchall()
        ]
        return jsonify(choferes), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# ==========================================
# 14. MÓDULO: EXPORTAR VENTAS A EXCEL
# ==========================================
@app.route('/api/ventas/exportar', methods=['GET'])
def exportar_ventas_excel():
    """Exporta todas las ventas a un archivo Excel sumando pagos múltiples."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT
                v.codigo_venta,
                v.nombre_cliente,
                v.tipo_comprobante,
                v.dni_cliente,
                v.fecha_emision,
                v.fecha_entrega,
                v.monto_total,
                COALESCE(itm.productos, '') AS productos,
                v.direccion_cliente,
                COALESCE(pg.metodos, 'Sin pago') AS metodo_pago,
                COALESCE(pg.total_pagado, 0) AS monto_adelanto,
                COALESCE(pg.empresas, '---') AS empresa_pago,
                v.fecha_emision AS fecha_registro,
                v.celular_cliente,
                v.moneda,
                v.vendedor_nombre
            FROM ventas v
            LEFT JOIN (
                SELECT venta_id,
                       SUM(monto_bruto) AS total_pagado,
                       STRING_AGG(tipo_pago || ' (' || COALESCE(entidad, '') || ')', ' | ') AS metodos,
                       STRING_AGG(DISTINCT empresa_destino, ' / ') AS empresas
                FROM pagos
                GROUP BY venta_id
            ) pg ON pg.venta_id = v.id
            LEFT JOIN (
                SELECT venta_id, STRING_AGG(producto, ' / ') AS productos
                FROM items_venta
                GROUP BY venta_id
            ) itm ON itm.venta_id = v.id
            ORDER BY v.id DESC;
        """)
        filas = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"

        header_font    = Font(bold=True, color="FFFFFF", size=10)
        header_fill    = PatternFill("solid", fgColor="0F172A")
        center         = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin           = Side(style="thin", color="CBD5E0")
        border         = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "Cód. Venta", "Cliente", "Comprobante", "RUC/DNI/CE",
            "F. Emisión", "F. Entrega", "Monto Total",
            "Producto(s)", "Dirección", "Métodos Pago (Múltiples)",
            "Adelanto Cobrado", "Empresa Receptora",
            "Fecha Registro", "Teléfono", "Moneda", "Vendedor"
        ]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        anchos = [12, 25, 12, 14, 14, 14, 12, 40, 30, 30, 15, 30, 14, 14, 10, 20]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

        fill_par  = PatternFill("solid", fgColor="F8FAFC")
        fill_impar = PatternFill("solid", fgColor="FFFFFF")

        for row_num, f in enumerate(filas, 2):
            fill = fill_par if row_num % 2 == 0 else fill_impar
            valores = [
                f[0], f[1], f[2], f[3], 
                f[4].strftime('%d/%m/%Y') if f[4] else '',
                f[5].strftime('%d/%m/%Y') if f[5] else '',
                float(f[6]) if f[6] else 0,
                f[7], f[8], f[9], 
                float(f[10]) if f[10] else 0,
                f[11], 
                f[12].strftime('%d/%m/%Y') if f[12] else '',
                f[13], f[14], f[15]
            ]
            for col, val in enumerate(valores, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        fecha_hoy = datetime.now().strftime('%Y%m%d_%H%M')
        return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'ventas_innova_{fecha_hoy}.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)

# ==========================================
# EXPORTACIÓN DE VENTAS (FORMATO COMPATIBLE CON EXCEL)
# ==========================================
@app.route('/api/exportar_ventas', methods=['GET'])
def exportar_ventas():
    # 1. Obtenemos las fechas que enviaste desde el calendario HTML
    inicio = request.args.get('inicio')
    fin = request.args.get('fin')
    
    if not inicio or not fin:
        return jsonify({'error': 'Faltan fechas de inicio y fin'}), 400

    # 2. Preparamos el archivo en la memoria del servidor
    si = StringIO()
    cw = csv.writer(si)
    
    # 3. Escribimos las cabeceras (las columnas de tu Excel)
    cw.writerow([
        'Codigo Contrato', 
        'Cliente', 
        'Documento', 
        'Fecha Emision', 
        'Total Venta (S/)', 
        'Total Pagado (S/)', 
        'Saldo (S/)', 
        'Sede'
    ])
    
    # NOTA: Aquí más adelante conectaremos tu base de datos real. 
    # Por ahora enviamos una fila de prueba para validar que la descarga funcione.
    cw.writerow(['INV-0001', 'Cliente de Prueba', '12345678', inicio, '1500.00', '500.00', '1000.00', 'Tienda Principal'])

    # 4. Empaquetamos y enviamos el archivo de texto estructurado
    output = si.getvalue()
    
    # Al enviarlo con extensión .csv, Excel lo abrirá automáticamente con formato de tabla
    nombre_archivo = f"Reporte_Ventas_{inicio}_al_{fin}.csv"
    
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={nombre_archivo}"}
    )

def listar_ventas():
    """Devuelve todas las ventas sumando los pagos múltiples."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT
                v.codigo_venta,
                v.nombre_cliente,
                v.monto_total,
                COALESCE(pg.total_pagado, 0) AS monto_adelanto,
                COALESCE(v.monto_total, 0) - COALESCE(pg.total_pagado, 0) AS saldo,
                COALESCE(v.estado_general, 'Pendiente') AS estado,
                v.fecha_entrega,
                v.vendedor_nombre,
                COALESCE(itm.productos, '') AS productos,
                v.sede
            FROM ventas v
            LEFT JOIN (
                SELECT venta_id, SUM(monto_bruto) AS total_pagado
                FROM pagos
                GROUP BY venta_id
            ) pg ON pg.venta_id = v.id
            LEFT JOIN (
                SELECT venta_id, STRING_AGG(producto, ' / ') AS productos
                FROM items_venta
                GROUP BY venta_id
            ) itm ON itm.venta_id = v.id
            ORDER BY v.id DESC;
        """)
        filas = cursor.fetchall()
        resultado = [{
            'codigo':        f[0],
            'cliente':       f[1],
            'total':         float(f[2]) if f[2] else 0,
            'adelanto':      float(f[3]) if f[3] else 0,
            'saldo':         float(f[4]) if f[4] else 0,
            'estado':        f[5],
            'fecha_entrega': f[6].strftime('%Y-%m-%d') if f[6] else None,
            'vendedor':      f[7],
            'productos':     f[8],
            'sede':          f[9]
        } for f in filas]
        
        return jsonify(resultado), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)
# ==========================================
# MÓDULO: CAMBIO DE PRECIO CON HISTORIAL
# ==========================================

def _crear_tabla_historial_precios(cursor):
    """Crea la tabla si todavía no existe (auto-migración segura)."""
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historial_precios (
            id               SERIAL PRIMARY KEY,
            venta_id         INTEGER       NOT NULL,
            codigo_venta     VARCHAR(50),
            precio_original  NUMERIC(10,2) NOT NULL,
            precio_nuevo     NUMERIC(10,2) NOT NULL,
            motivo           TEXT          NOT NULL,
            vendedor_id      INTEGER,
            vendedor_nombre  VARCHAR(100),
            admin_id         INTEGER,
            admin_nombre     VARCHAR(100),
            estado           VARCHAR(20)   DEFAULT 'Pendiente',
            notas_admin      TEXT,
            fecha_solicitud  TIMESTAMP     DEFAULT NOW(),
            fecha_resolucion TIMESTAMP
        );
    """)


@app.route('/api/ventas/<codigo>/proponer-cambio-precio', methods=['POST'])
def proponer_cambio_precio(codigo):
    """
    Vendedor propone un cambio de precio en una venta activa.
    Body: { precio_nuevo, motivo, vendedor_id, vendedor_nombre }
    """
    data = request.json
    precio_nuevo     = data.get('precio_nuevo')
    motivo           = data.get('motivo', '').strip()
    vendedor_id      = data.get('vendedor_id')
    vendedor_nombre  = data.get('vendedor_nombre', '')

    if not precio_nuevo or not motivo:
        return jsonify({'error': 'precio_nuevo y motivo son obligatorios'}), 400
    if float(precio_nuevo) <= 0:
        return jsonify({'error': 'El precio nuevo debe ser mayor a 0'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _crear_tabla_historial_precios(cursor)

        # Verificar que la venta existe y obtener precio actual
        cursor.execute("SELECT id, monto_total, estado_general FROM ventas WHERE codigo_venta = %s;", (codigo,))
        venta = cursor.fetchone()
        if not venta:
            return jsonify({'error': 'Venta no encontrada'}), 404

        venta_id, precio_original, estado = venta
        if estado in ('Entregado', 'Cancelado'):
            return jsonify({'error': f'No se puede modificar una venta en estado {estado}'}), 400

        # Verificar que no haya ya una solicitud pendiente para esta venta
        cursor.execute("""
            SELECT id FROM historial_precios
            WHERE venta_id = %s AND estado = 'Pendiente';
        """, (venta_id,))
        if cursor.fetchone():
            return jsonify({'error': 'Ya existe una solicitud de cambio de precio pendiente para esta venta'}), 400

        cursor.execute("""
            INSERT INTO historial_precios
                (venta_id, codigo_venta, precio_original, precio_nuevo,
                 motivo, vendedor_id, vendedor_nombre, estado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'Pendiente')
            RETURNING id;
        """, (venta_id, codigo, float(precio_original), float(precio_nuevo),
              motivo, vendedor_id, vendedor_nombre))

        nuevo_id = cursor.fetchone()[0]
        conexion.commit()
        return jsonify({'exito': True, 'id': nuevo_id,
                        'mensaje': 'Solicitud enviada al administrador'}), 201

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/cambios-precio/pendientes', methods=['GET'])
def listar_cambios_precio_pendientes():
    """Admin ve todas las solicitudes de cambio de precio pendientes."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _crear_tabla_historial_precios(cursor)

        cursor.execute("""
            SELECT
                hp.id,
                hp.codigo_venta,
                hp.precio_original,
                hp.precio_nuevo,
                hp.motivo,
                hp.vendedor_nombre,
                hp.fecha_solicitud,
                v.nombre_cliente,
                COALESCE(v.estado_general, 'En Producción') AS estado_venta
            FROM historial_precios hp
            JOIN ventas v ON hp.venta_id = v.id
            WHERE hp.estado = 'Pendiente'
            ORDER BY hp.fecha_solicitud DESC;
        """)
        filas = cursor.fetchall()
        resultado = [{
            'id':               f[0],
            'codigo_venta':     f[1],
            'precio_original':  float(f[2]),
            'precio_nuevo':     float(f[3]),
            'motivo':           f[4],
            'vendedor':         f[5],
            'fecha_solicitud':  f[6].strftime('%d/%m/%Y %H:%M') if f[6] else '',
            'cliente':          f[7],
            'estado_venta':     f[8],
        } for f in filas]
        return jsonify(resultado), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/cambios-precio/<int:cambio_id>/aprobar', methods=['POST'])
def aprobar_cambio_precio(cambio_id):
    """
    Admin aprueba el cambio: actualiza monto_total en ventas y cierra la solicitud.
    Body: { admin_id, admin_nombre }
    """
    data         = request.json
    admin_id     = data.get('admin_id')
    admin_nombre = data.get('admin_nombre', 'Admin')

    try:
        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        cursor.execute("""
            SELECT venta_id, precio_nuevo FROM historial_precios
            WHERE id = %s AND estado = 'Pendiente';
        """, (cambio_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Solicitud no encontrada o ya resuelta'}), 404

        venta_id, precio_nuevo = row

        # 1. Actualizar el total de la venta
        cursor.execute("""
            UPDATE ventas SET monto_total = %s WHERE id = %s;
        """, (precio_nuevo, venta_id))

        # 2. Cerrar la solicitud
        cursor.execute("""
            UPDATE historial_precios
            SET estado = 'Aprobado',
                admin_id = %s,
                admin_nombre = %s,
                fecha_resolucion = NOW()
            WHERE id = %s;
        """, (admin_id, admin_nombre, cambio_id))

        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Precio actualizado con éxito'}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/cambios-precio/<int:cambio_id>/rechazar', methods=['POST'])
def rechazar_cambio_precio(cambio_id):
    """
    Admin rechaza el cambio. El precio original queda intacto.
    Body: { admin_id, admin_nombre, notas_admin }
    """
    data         = request.json
    admin_id     = data.get('admin_id')
    admin_nombre = data.get('admin_nombre', 'Admin')
    notas        = data.get('notas_admin', '').strip()

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute("""
            UPDATE historial_precios
            SET estado = 'Rechazado',
                admin_id = %s,
                admin_nombre = %s,
                notas_admin = %s,
                fecha_resolucion = NOW()
            WHERE id = %s AND estado = 'Pendiente';
        """, (admin_id, admin_nombre, notas, cambio_id))

        if cursor.rowcount == 0:
            return jsonify({'error': 'Solicitud no encontrada o ya resuelta'}), 404

        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Solicitud rechazada'}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@app.route('/api/ventas/<codigo>/historial-precios', methods=['GET'])
def historial_precios_venta(codigo):
    """Devuelve el historial completo de cambios de precio de una venta."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _crear_tabla_historial_precios(cursor)

        cursor.execute("""
            SELECT precio_original, precio_nuevo, motivo, vendedor_nombre,
                   admin_nombre, estado, notas_admin, fecha_solicitud, fecha_resolucion
            FROM historial_precios
            WHERE codigo_venta = %s
            ORDER BY fecha_solicitud DESC;
        """, (codigo,))
        filas = cursor.fetchall()
        resultado = [{
            'precio_original':  float(f[0]),
            'precio_nuevo':     float(f[1]),
            'motivo':           f[2],
            'vendedor':         f[3],
            'admin':            f[4],
            'estado':           f[5],
            'notas_admin':      f[6],
            'fecha_solicitud':  f[7].strftime('%d/%m/%Y %H:%M') if f[7] else '',
            'fecha_resolucion': f[8].strftime('%d/%m/%Y %H:%M') if f[8] else '',
        } for f in filas]
        return jsonify(resultado), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


if __name__ == '__main__':
    app.run(debug=False, port=5000)