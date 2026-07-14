"""
routes_produccion.py — Módulos 8–13: Taller/Kanban, inventario de materiales,
logística externa, recetas BOM, sugerencias de insumos y despacho.
Blueprint: produccion_bp  (sin prefijo de URL)
"""

import json
import math
import re
import uuid
from datetime import datetime
from html import escape as html_escape
from io import BytesIO
from zoneinfo import ZoneInfo
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Blueprint, current_app, jsonify, request, send_file
from database import get_db_connection, release_db_connection, limpiar_foto, notificar_usuario, cloudinary_upload
from auth_middleware import (
    get_usuario_actual,
    requiere_login,
    requiere_rol,
    usuario_puede_operar,
)
from erp_constants import AREA_ALIASES
from workflow_rules import (
    contrato_entregado,
    cotizacion_esta_vigente,
    estructura_disponible_para_tapiceria,
    ids_items_logistica,
    normalizar_fecha_cotizacion,
    normalizar_estado_distribucion,
    normalizar_estado_logistica,
    normalizar_precio_cotizacion,
    tela_puede_distribuirse,
    tela_requiere_comprobante,
    transicion_logistica_permitida,
)

# pip install reportlab==4.2.2
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import mm

produccion_bp = Blueprint('produccion', __name__)

ZONA_HORARIA_LIMA = ZoneInfo('America/Lima')


def _intentar_desbloquear_items(cursor, item_ids):
    """Abre solo tickets cuyas dependencias fisicas ya estan completas."""
    desbloqueados = 0
    for item_id in ids_items_logistica(None, ','.join(map(str, item_ids or []))):
        cursor.execute("""
            SELECT id, area_trabajo
            FROM tickets_produccion
            WHERE item_id = %s
              AND estado_ticket = 'Bloqueado'
              AND area_trabajo IN ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS', 'ARMADO_COJINES')
        """, (item_id,))
        for ticket_id, area in cursor.fetchall():
            if _tela_pendiente_para_item(cursor, item_id):
                continue
            if area in ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS'):
                cursor.execute("""
                    SELECT estado_ticket
                    FROM tickets_produccion
                    WHERE item_id = %s
                      AND area_trabajo IN ('ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS_SILLAS')
                """, (item_id,))
                estados_estructura = [r[0] for r in cursor.fetchall()]
                if not estados_estructura or not all(
                    estructura_disponible_para_tapiceria(e)
                    for e in estados_estructura
                ):
                    continue
            cursor.execute("""
                UPDATE tickets_produccion
                SET estado_ticket = 'En Proceso',
                    fecha_inicio = COALESCE(fecha_inicio, CURRENT_TIMESTAMP)
                WHERE id = %s AND estado_ticket = 'Bloqueado'
            """, (ticket_id,))
            desbloqueados += cursor.rowcount
    return desbloqueados


def _items_asociados_logistica(cursor, logistica_id, item_id, item_ids_extra):
    cursor.execute(
        "SELECT item_id FROM logistica_externa_items WHERE logistica_id = %s ORDER BY item_id",
        (logistica_id,)
    )
    normalizados = [r[0] for r in cursor.fetchall()]
    return normalizados or ids_items_logistica(item_id, item_ids_extra)


def _activar_despacho_si_listo(cursor, venta_id):
    """Activa todos los despachos del contrato solo al completar dependencias."""
    cursor.execute("""
        SELECT COUNT(*)
        FROM tickets_produccion t
        JOIN items_venta i ON i.id = t.item_id
        WHERE i.venta_id = %s AND t.area_trabajo = 'DESPACHO_CENTRAL'
    """, (venta_id,))
    if cursor.fetchone()[0] == 0:
        return None

    cursor.execute("""
        SELECT COUNT(*)
        FROM tickets_produccion t
        JOIN items_venta i ON i.id = t.item_id
        WHERE i.venta_id = %s
          AND t.area_trabajo != 'DESPACHO_CENTRAL'
          AND t.estado_ticket NOT IN ('Terminado', 'Recogido', 'Cancelado')
    """, (venta_id,))
    if cursor.fetchone()[0] > 0:
        return None

    cursor.execute("""
        SELECT COUNT(*)
        FROM logistica_externa
        WHERE venta_id = %s
          AND estado NOT IN ('Recibido', 'Cancelado', 'Rechazado')
    """, (venta_id,))
    if cursor.fetchone()[0] > 0:
        return None

    cursor.execute("""
        SELECT COUNT(*)
        FROM logistica_externa
        WHERE venta_id = %s
          AND (categoria_insumo = 'TELA'
               OR LOWER(COALESCE(insumo_nombre, '')) LIKE '%%tela%%'
               OR LOWER(COALESCE(unidad, '')) IN ('mts', 'metro', 'metros'))
          AND COALESCE(estado_distribucion, '') != 'Distribuido'
          AND estado NOT IN ('Cancelado', 'Rechazado')
    """, (venta_id,))
    if cursor.fetchone()[0] > 0:
        return None

    cursor.execute("""
        UPDATE tickets_produccion
        SET estado_ticket = 'Pendiente'
        WHERE item_id IN (SELECT id FROM items_venta WHERE venta_id = %s)
          AND area_trabajo = 'DESPACHO_CENTRAL'
          AND estado_ticket = 'Bloqueado'
    """, (venta_id,))
    activados = cursor.rowcount
    cursor.execute("""
        UPDATE ventas SET estado_general = 'Listo'
        WHERE id = %s AND COALESCE(estado_general, '') NOT IN ('Entregado', 'Cancelado')
    """, (venta_id,))
    return activados


# FIX (julio 2026): helper compartido para saber si la TELA de un ítem
# específico (no de toda la venta) ya está lista para que tapicería/cojines
# puedan trabajar. Antes esta lógica solo existía como indicador visual en
# obtener_cola_recojo (JOIN por venta_id, no por item_id) y los módulos de Taller se
# limitaba a deshabilitar el botón en el navegador — el endpoint real de
# confirmar recojo no validaba nada de esto en el servidor.
#
# "Tela pendiente" = true si:
#   a) el ítem tiene un ticket interno de CORTE_Y_CONTROL_TELAS que aún no
#      terminó (tela producida internamente), o
#   b) el ítem tiene una fila de logistica_externa de categoria_insumo='TELA'
#      que aún no fue distribuida al tapicero/cojinero.
# Si el ítem no tiene NINGUNA tela asociada (ni ticket interno ni fila de
# logística), no hay nada que esperar y se considera lista.
def _tela_pendiente_para_item(cursor, item_id):
    # FIX (julio 2026 - v2): cuando dos o más ítems de la misma venta usan
    # la misma tela+proveedor (ej. sofá + silla con "Boucle #14"), la
    # consolidación en routes_ventas.py NO crea una fila de logistica_externa
    # por cada ítem — solo una, con item_id apuntando al primero que se
    # procesó. Antes de este fix, el segundo/tercer ítem (ej. la silla)
    # nunca encontraba su propia fila aquí y este semáforo devolvía "no hay
    # tela pendiente" para ellos aunque la tela consolidada todavía no
    # hubiera llegado — dejando avanzar tapicería/recojo de estructura sin
    # esperar el material real. Ahora también se revisa item_ids_extra,
    # donde la consolidación anota los item_id "invitados" a esa misma fila.
    patron_item = f'%,{item_id},%'

    cursor.execute("""
        SELECT EXISTS (
            SELECT 1 FROM tickets_produccion
            WHERE item_id = %s AND area_trabajo = 'CORTE_Y_CONTROL_TELAS'
              AND estado_ticket NOT IN ('Terminado', 'Recogido', 'Cancelado')
        ) OR EXISTS (
            SELECT 1 FROM logistica_externa
            WHERE (
                    COALESCE(categoria_insumo, '') = 'TELA'
                    OR LOWER(COALESCE(insumo_nombre, '')) LIKE '%%tela%%'
                    OR LOWER(COALESCE(unidad, '')) IN ('mts', 'metro', 'metros')
                  )
              AND COALESCE(estado_distribucion, '') != 'Distribuido'
              AND estado NOT IN ('Cancelado', 'Rechazado')
              AND (
                    item_id = %s
                    OR (',' || COALESCE(item_ids_extra, '') || ',') LIKE %s
                    OR EXISTS (
                        SELECT 1 FROM logistica_externa_items lei
                        WHERE lei.logistica_id = logistica_externa.id AND lei.item_id = %s
                    )
              )
        );
    """, (item_id, item_id, patron_item, item_id))
    return bool(cursor.fetchone()[0])


# ==========================================
# 8. TALLER Y KANBAN
# ==========================================

@produccion_bp.route('/api/taller/ticket/<int:id>/finalizar', methods=['POST'])
@requiere_login
def finalizar_ticket(id):
    conexion_permiso = None
    cursor_permiso = None
    try:
        conexion_permiso = get_db_connection()
        cursor_permiso = conexion_permiso.cursor()
        cursor_permiso.execute("""
            SELECT area_trabajo, trabajador_asignado_id, estado_ticket
            FROM tickets_produccion WHERE id = %s
        """, (id,))
        ticket_permiso = cursor_permiso.fetchone()
        if not ticket_permiso:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        if not usuario_puede_operar(ticket_permiso[0], ticket_permiso[1]):
            return jsonify({'error': 'No puedes finalizar un ticket de otra area o persona.'}), 403
        if ticket_permiso[2] in ('Terminado', 'Listo para Recojo', 'Recogido', 'Cancelado'):
            return jsonify({'error': f'El ticket ya esta cerrado ({ticket_permiso[2]}).'}), 409

        if 'foto' not in request.files or request.files['foto'].filename == '':
            return jsonify({'error': 'La foto de evidencia es obligatoria'}), 400

        foto = request.files['foto']
        respuesta_nube = cloudinary_upload(foto, folder="evidencias")
        foto_url_final = respuesta_nube.get('secure_url')

        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        cursor.execute("""
            UPDATE tickets_produccion
            SET estado_ticket = CASE
                    WHEN area_trabajo IN ('ESTRUCTURAS_MUEBLES','ESTRUCTURAS_SILLAS')
                    THEN 'Listo para Recojo'
                    ELSE 'Terminado'
                END,
                foto_evidencia = %s,
                fecha_fin = CURRENT_TIMESTAMP
            WHERE id = %s
              AND estado_ticket NOT IN ('Terminado', 'Listo para Recojo', 'Recogido', 'Cancelado')
            RETURNING item_id, area_trabajo, estado_ticket;
        """, (foto_url_final, id))
        row = cursor.fetchone()
        if not row:
            conexion.rollback()
            return jsonify({'error': 'El ticket fue cerrado por otra sesion. Actualiza la bandeja.'}), 409

        item_id = row[0]
        area_term = row[1]
        desbloqueados = _intentar_desbloquear_items(cursor, [item_id])

        venta_actualizada = False
        if row:
            cursor.execute("SELECT venta_id FROM items_venta WHERE id = %s", (item_id,))
            venta_row = cursor.fetchone()
            if venta_row:
                venta_id_check = venta_row[0]

                # Si el ticket finalizado es DESPACHO_CENTRAL → marcar venta como Entregado
                if area_term == 'DESPACHO_CENTRAL':
                    cursor.execute("""
                        SELECT t.estado_ticket
                        FROM tickets_produccion t
                        JOIN items_venta i ON i.id = t.item_id
                        WHERE i.venta_id = %s
                          AND t.area_trabajo = 'DESPACHO_CENTRAL'
                    """, (venta_id_check,))
                    estados_despacho = [r[0] for r in cursor.fetchall()]
                    if contrato_entregado(estados_despacho):
                        cursor.execute("""
                            UPDATE ventas SET estado_general = 'Entregado'
                            WHERE id = %s AND COALESCE(estado_general,'') != 'Cancelado'
                        """, (venta_id_check,))
                        venta_actualizada = cursor.rowcount > 0
                        if venta_actualizada:
                            from routes_ventas import _liberar_unidades
                            _liberar_unidades(cursor, venta_id_check, 'Vendido', 'Venta')
                    else:
                        cursor.execute("""
                            UPDATE ventas SET estado_general = 'En Despacho'
                            WHERE id = %s AND COALESCE(estado_general,'') NOT IN ('Entregado', 'Cancelado')
                        """, (venta_id_check,))
                else:
                    venta_actualizada = _activar_despacho_si_listo(cursor, venta_id_check) is not None

        conexion.commit()
        es_despacho     = row and row[1] == 'DESPACHO_CENTRAL'
        es_listo_recojo = row and row[2] == 'Listo para Recojo'
        if es_despacho and venta_actualizada:
            msg = 'Entrega confirmada. Todos los items fueron entregados y la venta quedo cerrada.'
        elif es_despacho:
            msg = 'Entrega del item confirmada. El contrato aun tiene otros items por entregar.'
        elif es_listo_recojo:
            msg = '🔴 Estructura lista. Esperando que el chofer confirme el recojo.'
        else:
            msg = 'Ticket finalizado correctamente'
        if not es_despacho and desbloqueados > 0:
            msg += f'. {desbloqueados} ticket(s) de tapicería desbloqueado(s) automáticamente.'
        if not es_despacho and venta_actualizada:
            msg += '. ✅ ¡Producción completa! La venta pasó a estado Listo.'
        return jsonify({'exito': True, 'mensaje': msg, 'desbloqueados': desbloqueados, 'venta_lista': venta_actualizada, 'es_entrega': es_despacho, 'es_listo_recojo': es_listo_recojo}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print(f'Error al finalizar ticket: {e}')
        return jsonify({'error': 'No se pudo finalizar el ticket'}), 500
    finally:
        if cursor_permiso:
            cursor_permiso.close()
        if conexion_permiso:
            release_db_connection(conexion_permiso)
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/taller/cola-recojo', methods=['GET'])
@requiere_login
def obtener_cola_recojo():
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT t.id, t.area_trabajo, i.producto, v.codigo_venta, v.nombre_cliente,
                   COALESCE(u.nombre, 'Sin asignar') AS operario, t.fecha_fin,
                   COALESCE(i.foto_url, '') AS foto_url,
                   COALESCE(t.ticket_details_override, i.color_tela, '') AS especificaciones,
                   t.foto_evidencia, v.direccion_cliente, v.fecha_entrega, t.item_id,
                   COALESCE(ut.nombre, 'Sin asignar') AS tapicero_nombre
            FROM tickets_produccion t
            JOIN items_venta i    ON t.item_id  = i.id
            JOIN ventas v         ON i.venta_id = v.id
            LEFT JOIN usuarios u  ON t.trabajador_asignado_id = u.id
            LEFT JOIN tickets_produccion tap
                ON tap.item_id = t.item_id
               AND tap.area_trabajo IN ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS')
            LEFT JOIN usuarios ut ON tap.trabajador_asignado_id = ut.id
            WHERE t.area_trabajo IN ('ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS_SILLAS')
              AND t.estado_ticket = 'Listo para Recojo'
            ORDER BY t.fecha_fin DESC;
        """)
        raw_estructuras = cursor.fetchall()

        # FIX (julio 2026): antes esto se calculaba en el mismo SELECT con un
        # LEFT JOIN a logistica_externa filtrado por venta_id — eso podía (a)
        # mezclar la tela de OTRA pieza del mismo contrato (venta_id, no
        # item_id) y (b) duplicar la fila del ticket si el item tenía más de
        # una tela asociada (ej. tela principal + tela de cojín). Se calcula
        # ahora por item_id con el mismo helper que usa el gate real de
        # confirmar-recojo, para que el semáforo nunca mienta respecto a lo
        # que el backend realmente va a permitir.
        estructuras = []
        for r in raw_estructuras:
            item_id_r = r[12]
            tela_pendiente = _tela_pendiente_para_item(cursor, item_id_r)
            estructuras.append({
                "ticket_id": r[0], "area": r[1], "producto": r[2], "codigo_venta": r[3],
                "cliente": r[4], "operario": r[5],
                "fecha_fin": r[6].strftime('%d/%m/%Y %H:%M') if r[6] else 'S/F',
                "foto_url": "|".join([limpiar_foto(p) for p in r[7].split('|')]) if r[7] and "|" in r[7] else limpiar_foto(r[7]),
                "especificaciones": r[8] or '',
                "foto_evidencia": r[9] if r[9] else '', "direccion": r[10] or '',
                "fecha_entrega": r[11].strftime('%d/%m/%Y') if r[11] else 'S/F',
                "item_id": item_id_r, "tapicero": r[13],
                "bloqueado_por_telas": tela_pendiente,
                "tela_distribuida": not tela_pendiente,
            })

        # 2. Compras externas (Logística)
        #
        # FIX (julio 2026): existían DOS caminos distintos para marcar un
        # insumo comprado como "listo para que el chofer lo recoja", y este
        # filtro solo reconocía uno de los dos:
        #
        #   a) Manual: el Jefe elige "Listo para Recojo" en el selector de
        #      estado del modal de Logística Externa (ver modules/app/logistica_externa.js,
        #      estadosPosibles) → escribe l.estado = 'Listo para Recojo'.
        #      Este es el que el filtro original SÍ cubría.
        #
        #   b) Automático: al subir el comprobante de pago
        #      (registrar_pago_proveedor) y detectar que el insumo es
        #      ESTRUCTURAL (base/tablero/silla/butaca), el sistema deja
        #      l.estado = 'Pagado' y guarda el "listo" en la columna
        #      l.estado_distribucion, no en l.estado. Ese camino NUNCA
        #      hacía match con este filtro — bases, tableros, sillas
        #      metálicas y butacas/puffs externos pagados por esta vía
        #      jamás llegaban a la cola del chofer.
        #
        # Se cubren ambos caminos. Se excluye TELA porque esa categoría
        # tiene su propia cola (Corte y Control de Telas, ver más abajo y
        # obtener_tickets_taller), no la del chofer de estructuras/compras.
        cursor.execute("""
            SELECT l.id, v.codigo_venta, v.nombre_cliente, l.insumo_nombre, l.sku,
                   COALESCE(p.nombre, l.proveedor_informal, 'Sin proveedor') AS proveedor,
                   COALESCE(p.telefono, '') AS telefono_proveedor,
                   l.url_cotizacion_adjunta, l.notas_proveedor,
                   l.cantidad, l.unidad, l.fecha_entrega_proveedor
            FROM logistica_externa l
            JOIN ventas v ON l.venta_id = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            WHERE (
                    l.estado = 'Listo para Recojo'
                    OR (
                        l.estado IN ('Orden Enviada', 'Confirmado', 'En Tránsito', 'Pagado')
                        AND l.estado_distribucion = 'Listo para Recojo'
                    )
                  )
              AND COALESCE(l.categoria_insumo, 'OTRO') != 'TELA'
            ORDER BY l.id DESC;
        """)
        compras_externas = [{
            "logistica_id": r[0], "codigo_venta": r[1], "cliente": r[2], "insumo": r[3], "sku": r[4],
            "proveedor": r[5], "telefono_proveedor": r[6], "url_cotizacion_adjunta": r[7], "notas_proveedor": r[8],
            "cantidad": float(r[9]) if r[9] else None, "unidad": r[10], "fecha_entrega_proveedor": r[11].strftime('%d/%m/%Y') if r[11] else ''
        } for r in cursor.fetchall()]

        return jsonify({"estructuras": estructuras, "compras_externas": compras_externas}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/taller/ticket/<int:id>/confirmar-recojo', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller', 'Chofer')
def confirmar_recojo_estructura(id):
    """Chofer confirma recojo de estructura: Listo para Recojo → Recogido.
    También intenta desbloquear tapicería si ya están listas las telas."""
    conexion = None
    try:
        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        cursor.execute("""
            SELECT item_id FROM tickets_produccion
            WHERE id = %s AND estado_ticket = 'Listo para Recojo';
        """, (id,))
        row_check = cursor.fetchone()
        if not row_check:
            return jsonify({'error': 'Ticket no encontrado o no está en estado "Listo para Recojo"'}), 400

        cursor.execute("""
            UPDATE tickets_produccion
            SET estado_ticket = 'Recogido'
            WHERE id = %s AND estado_ticket = 'Listo para Recojo'
            RETURNING item_id, area_trabajo;
        """, (id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Ticket no encontrado o no está en estado "Listo para Recojo"'}), 400

        item_id = row[0]

        # El recojo fisico puede ocurrir aunque la tela siga pendiente. Lo que
        # espera a la tela es Tapiceria, no el traslado de la estructura.
        desbloqueados = _intentar_desbloquear_items(cursor, [item_id])

        # ── Verificar si la venta queda completamente lista ──
        cursor.execute("SELECT venta_id FROM items_venta WHERE id = %s", (item_id,))
        venta_row = cursor.fetchone()
        venta_actualizada = False
        if venta_row:
            venta_id_check = venta_row[0]
            venta_actualizada = _activar_despacho_si_listo(cursor, venta_id_check) is not None

        conexion.commit()
        msg = '✅ Recojo confirmado. El ticket pasa a estado Recogido.'
        if desbloqueados > 0:
            msg += f' {desbloqueados} ticket(s) de tapicería desbloqueado(s).'
        if venta_actualizada:
            msg += ' ✅ ¡Producción completa! La venta pasó a estado Listo.'
        return jsonify({'exito': True, 'mensaje': msg, 'desbloqueados': desbloqueados, 'venta_lista': venta_actualizada}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        if conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/<int:logistica_id>/confirmar-recojo-externo', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller', 'Chofer', 'Operario')
def confirmar_recojo_externo(logistica_id):
    """Chofer o Tapicero confirma recojo externo: marca logística como Recibido
    y desbloquea las áreas de producción que correspondan."""
    conexion = None
    try:
        conexion = get_db_connection()
        conexion.autocommit = False
        cursor = conexion.cursor()

        # Leer estado anterior para evitar doble ejecución
        # FIX (julio 2026): antes solo se leía venta_id y el desbloqueo de
        # ahí para abajo se hacía sobre TODOS los tickets 'Bloqueado' de la
        # venta completa — un insumo suelto (ej. la base de UNA butaca)
        # podía desbloquear ARMADO_COJINES o incluso, si por algún motivo
        # estaba en 'Bloqueado', el DESPACHO_CENTRAL de OTRO ítem del mismo
        # contrato que no tenía nada que ver. Ahora se acota por item_id
        # (la pieza puntual de este insumo) y se excluye explícitamente
        # DESPACHO_CENTRAL: ese ticket solo lo debe mover el propio flujo de
        # producción/despacho (asignar_chofer_despacho ya valida todo lo
        # necesario), nunca la llegada suelta de un insumo.
        cursor.execute("""
            SELECT estado, venta_id, item_id, item_ids_extra, operario_id,
                   COALESCE(categoria_insumo, 'OTRO'),
                   COALESCE(estado_distribucion, 'En espera'),
                   COALESCE(tipo_gestion, 'Externo')
            FROM logistica_externa WHERE id = %s FOR UPDATE
        """, (logistica_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Ítem de logística no encontrado'}), 404
        (estado_anterior, venta_id, item_id_logistica, item_ids_extra,
         operario_id, categoria, estado_distribucion, tipo_gestion) = row
        if not usuario_puede_operar('DESPACHO_CENTRAL', operario_id):
            return jsonify({'error': 'No puedes confirmar el recojo asignado a otra persona.'}), 403
        actor = get_usuario_actual()
        es_tela = categoria == 'TELA'

        if tipo_gestion != 'Externo' or es_tela:
            return jsonify({'error': 'Este recojo corresponde solo a compras externas no textiles.'}), 409
        listo_para_recojo = (
            estado_anterior == 'Listo para Recojo'
            or (
                estado_anterior in ('Orden Enviada', 'Confirmado', 'En Tránsito', 'Pagado')
                and estado_distribucion == 'Listo para Recojo'
            )
        )
        if estado_anterior != 'Recibido' and not listo_para_recojo:
            return jsonify({'error': 'El material todavía no está listo para recojo.'}), 409

        if estado_anterior != 'Recibido':
            cursor.execute("""
                UPDATE logistica_externa
                SET estado = 'Recibido',
                    estado_distribucion = CASE WHEN %s THEN 'Recogido' ELSE 'Distribuido' END,
                    fecha_recojo_fisico = NOW(),
                    recogido_por_id = %s
                WHERE id = %s
            """, (es_tela, int(actor['id']), logistica_id))

        item_ids = _items_asociados_logistica(
            cursor, logistica_id, item_id_logistica, item_ids_extra
        )
        desbloqueados = _intentar_desbloquear_items(cursor, item_ids)
        if venta_id:
            _activar_despacho_si_listo(cursor, venta_id)

        conexion.commit()
        return jsonify({
            'exito': True,
            'mensaje': f'✅ Material recogido y recibido. {desbloqueados} ticket(s) desbloqueado(s) para continuar su proceso.',
            'desbloqueados': desbloqueados
        }), 200

    except Exception as e:
        if conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/taller/tickets', methods=['GET'])
@requiere_login
def obtener_tickets_taller():
    area_filtro = request.args.get('area')
    operario_id = request.args.get('operario_id')
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        query = """
            SELECT t.id, i.producto, t.estado_ticket, t.area_trabajo, t.ticket_details_override,
                   t.trabajador_asignado_id, v.codigo_venta, i.color_tela, t.item_id,
                   i.foto_url, t.foto_evidencia, u.nombre, v.id, v.estado_general
            FROM tickets_produccion t
            JOIN items_venta i ON t.item_id  = i.id
            JOIN ventas v      ON i.venta_id = v.id
            LEFT JOIN usuarios u ON t.trabajador_asignado_id = u.id
            WHERE COALESCE(v.estado_general, '') != 'Cancelado'
              AND COALESCE(t.estado_ticket, '') != 'Cancelado'
              AND t.item_id NOT IN (
                -- Excluir items que SOLO tienen tickets de ESTRUCTURAS_SILLAS/TAPICERIA_SILLAS
                -- y ningún ticket de áreas internas reales (tela, cojines, etc.)
                -- Esto indica que la silla/butaca es comprada externamente (logística externa)
                SELECT DISTINCT t2.item_id
                FROM tickets_produccion t2
                WHERE t2.area_trabajo NOT IN ('DESPACHO_CENTRAL')
                GROUP BY t2.item_id
                HAVING
                    -- Todos sus tickets son solo de estructura/tapicería silla
                    bool_and(t2.area_trabajo IN ('ESTRUCTURAS_SILLAS', 'TAPICERIA_SILLAS'))
                    -- Y existe al menos una fila en logística externa activa para esa venta
                    AND EXISTS (
                        SELECT 1
                        FROM logistica_externa le
                        JOIN items_venta i2 ON le.venta_id = i2.venta_id
                        WHERE i2.id = t2.item_id
                          AND le.tipo_gestion = 'Externo'
                          AND le.estado NOT IN ('Recibido', 'Cancelado')
                    )
            )
        """
        params = []
        if area_filtro:
            areas_buscar = AREA_ALIASES.get(area_filtro, [area_filtro])
            placeholders = ','.join(['%s'] * len(areas_buscar))
            query += f' AND t.area_trabajo IN ({placeholders})'
            params.extend(areas_buscar)
        if operario_id:
            query += ' AND t.trabajador_asignado_id = %s'
            params.append(int(operario_id))
        query += " ORDER BY t.etapa ASC, t.id DESC;"

        cursor.execute(query, params)
        raw_rows = cursor.fetchall()

        item_ids_despacho = {row[8] for row in raw_rows if row[3] == 'DESPACHO_CENTRAL'}
        venta_ids_despacho = {row[12] for row in raw_rows if row[3] == 'DESPACHO_CENTRAL'}
        items_incompletos = set()
        if item_ids_despacho:
            placeholders_items = ','.join(['%s'] * len(item_ids_despacho))
            cursor.execute(f"""
                SELECT DISTINCT item_id FROM tickets_produccion
                WHERE item_id IN ({placeholders_items})
                  AND estado_ticket NOT IN ('Terminado', 'Bloqueado') AND area_trabajo != 'DESPACHO_CENTRAL'
            """, tuple(item_ids_despacho))
            items_incompletos = {r[0] for r in cursor.fetchall()}

        ventas_con_logistica_pendiente = set()
        if venta_ids_despacho:
            placeholders_ventas = ','.join(['%s'] * len(venta_ids_despacho))
            cursor.execute(f"""
                SELECT DISTINCT venta_id
                FROM logistica_externa
                WHERE venta_id IN ({placeholders_ventas})
                  AND estado NOT IN ('Recibido', 'Cancelado')
            """, tuple(venta_ids_despacho))
            ventas_con_logistica_pendiente = {r[0] for r in cursor.fetchall()}

        tickets = []
        for row in raw_rows:
            estado = row[2]
            # Compatibilidad con contratos antiguos: si la venta ya fue
            # marcada como Entregada pero su ticket de despacho quedó abierto,
            # no debe seguir apareciendo como pendiente.
            if row[3] == 'DESPACHO_CENTRAL' and row[13] == 'Entregado':
                estado = 'Terminado'
            if (
                row[3] == 'DESPACHO_CENTRAL'
                and row[13] != 'Entregado'
                and (row[8] in items_incompletos or row[12] in ventas_con_logistica_pendiente)
            ):
                estado = 'Bloqueado'

            # La ficha original contiene las notas escritas por el vendedor.
            # El detalle particular del area (SKU, notas internas, etc.) debe
            # complementarla, no reemplazarla.
            ficha_original = (row[7] or '').strip()
            detalle_area = (row[4] or '').strip()
            partes_ficha = [ficha_original]
            if detalle_area and detalle_area not in ficha_original:
                partes_ficha.append(detalle_area)
            especificaciones = '<br>'.join(p for p in partes_ficha if p) or 'Sin notas técnicas'

            tickets.append({
                "id":              row[0],
                "producto":        f"{row[1]} (Ref: {row[6]})",
                "estado":          estado,
                "area":            row[3],
                "trabajador":      row[5],
                "especificaciones": especificaciones,
                "foto":            "|".join([limpiar_foto(p) for p in row[9].split('|')]) if row[9] and "|" in row[9] else limpiar_foto(row[9]),
                "trabajador_nombre": row[11] if row[11] else 'Sin asignar',
                "item_id":         row[8]
            })
            
        # Inyectar elementos de Logística Externa para la cola de Telas
        if not area_filtro or area_filtro in ('TELAS', 'CORTE_Y_CONTROL_TELAS'):
            # FIX (julio 2026): antes este WHERE solo aceptaba
            # estado_distribucion IN ('En Recojo','Recogido','Distribuido').
            # Eso dejaba fuera 'En espera' — el estado que registrar_pago_
            # proveedor() asigna automáticamente a una TELA recién pagada
            # cuando se detecta por SKU en maestro_telas. Como 'En espera'
            # no aparecía aquí, esa tela nunca llegaba a esta bandeja y
            # tampoco existía ningún otro botón en el frontend para
            # moverla — se quedaba pagada pero invisible indefinidamente.
            # Al incluirla, vuelve a pasar por el flujo normal de la
            # bandeja compartida de Telas: cualquier operario del area puede
            # confirmar el recojo sin asignacion individual.
            log_query = """
                SELECT l.id, v.codigo_venta, l.insumo_nombre, l.sku, l.estado_distribucion, v.id,
                       l.operario_id, COALESCE(u.nombre, 'Sin asignar'),
                       COALESCE(l.cantidad, 1)                                   AS cantidad,
                       COALESCE(l.unidad, '')                                    AS unidad,
                       COALESCE(p.nombre, l.proveedor_informal, 'Sin proveedor') AS proveedor,
                       v.nombre_cliente, v.fecha_entrega,
                       COALESCE(l.tipo_gestion, 'Externo'),
                       COALESCE(ur.nombre, ''), COALESCE(ud.nombre, '')
                FROM logistica_externa l
                JOIN ventas v ON l.venta_id = v.id
                LEFT JOIN usuarios u ON l.operario_id = u.id
                LEFT JOIN usuarios ur ON l.recogido_por_id = ur.id
                LEFT JOIN usuarios ud ON l.distribuido_por_id = ud.id
                LEFT JOIN proveedores p ON l.proveedor_id = p.id
                WHERE (l.categoria_insumo = 'TELA' OR LOWER(l.insumo_nombre) LIKE '%%tela%%' OR LOWER(l.unidad) = 'mts')
                  AND COALESCE(l.estado_distribucion, 'En espera')
                      IN ('En espera', 'En Recojo', 'Recogido', 'Distribuido', 'Listo para recojo')
                  AND COALESCE(l.estado, '') NOT IN ('Cancelado', 'Rechazado')
                  AND COALESCE(v.estado_general, '') != 'Cancelado'
            """
            log_params = []
            if operario_id:
                log_query += " AND l.operario_id = %s"
                log_params.append(int(operario_id))
                
            cursor.execute(log_query, log_params)
            log_rows = cursor.fetchall()

            # Para cada venta de estos insumos, averiguar qué tapicero (sofás/sillas)
            # y qué cojinero ya tienen asignado, para que el operario de telas
            # sepa a quién debe entregarle la tela/cojín.
            destinos_por_venta = {}
            venta_ids = {r[5] for r in log_rows}
            if venta_ids:
                placeholders_v = ','.join(['%s'] * len(venta_ids))
                cursor.execute(f"""
                    SELECT i.venta_id, t.area_trabajo, COALESCE(u.nombre, 'Sin asignar')
                    FROM tickets_produccion t
                    JOIN items_venta i ON t.item_id = i.id
                    LEFT JOIN usuarios u ON t.trabajador_asignado_id = u.id
                    WHERE i.venta_id IN ({placeholders_v})
                      AND t.area_trabajo IN ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS', 'ARMADO_COJINES')
                """, tuple(venta_ids))
                for venta_id, area, nombre in cursor.fetchall():
                    d = destinos_por_venta.setdefault(venta_id, {'tapicero': None, 'cojinero': None})
                    if area in ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS'):
                        d['tapicero'] = nombre
                    elif area == 'ARMADO_COJINES':
                        d['cojinero'] = nombre

            for r in log_rows:
                destino = destinos_por_venta.get(r[5], {})
                tickets.append({
                    "id": r[0],
                    "producto": f"TELA EXTERNA: {r[2]}",
                    "estado": normalizar_estado_distribucion(r[4]),
                    "area": "CORTE_Y_CONTROL_TELAS",
                    "trabajador": r[6],
                    "especificaciones": f"Ref: {r[1]} | SKU: {r[3] or 'N/A'}",
                    "foto": "imagenes/sin_foto.jpg",
                    "trabajador_nombre": r[7],
                    "item_id": r[5],
                    "es_logistica": True,
                    "tapicero_destino": destino.get('tapicero'),
                    "cojinero_destino": destino.get('cojinero'),
                    # Campos nuevos para agrupar por contrato en el frontend
                    # (modules/taller/tickets.js: tarjeta de contrato con desglose por línea).
                    "venta_id":       r[5],
                    "codigo_venta":   r[1],
                    "sku":            r[3] or '',
                    "cantidad":       float(r[8]) if r[8] is not None else 1,
                    "unidad":         r[9] or '',
                    "proveedor":      r[10],
                    "cliente":        r[11] or '',
                    "fecha_entrega":  r[12].strftime('%d/%m/%Y') if r[12] else None,
                    "tipo_gestion":   r[13],
                    "recogido_por":   r[14],
                    "distribuido_por": r[15],
                })
                
        return jsonify(tickets), 200
    except Exception:
        current_app.logger.exception("Error al cargar los tickets del taller")
        return jsonify({
            'error': 'No se pudieron cargar los tickets del taller.',
            'codigo': 'TALLER_TICKETS_ERROR',
        }), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/taller/tickets_pendientes', methods=['GET'])
@requiere_login
def obtener_tickets_pendientes():
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT t.id, v.codigo_venta, v.nombre_cliente, i.producto,
                   t.area_trabajo, t.estado_ticket, v.fecha_entrega, i.color_tela
            FROM tickets_produccion t
            JOIN items_venta i ON t.item_id  = i.id
            JOIN ventas v      ON i.venta_id = v.id
            WHERE t.trabajador_asignado_id IS NULL AND t.estado_ticket != 'Terminado'
            ORDER BY v.fecha_entrega ASC;
        """)
        res = [{
            "ticket_id": t[0], "codigo": t[1], "cliente": t[2], "producto": t[3],
            "area": t[4], "estado": t[5],
            "entrega": t[6].strftime('%d/%m/%Y') if t[6] else "S/F",
            "especificaciones": t[7]
        } for t in cursor.fetchall()]
        return jsonify(res), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/taller/asignar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def asignar_maestro_ticket():
    data          = request.json
    ticket_id     = data.get('ticket_id')
    trabajador_id = data.get('trabajador_id')
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT estado_ticket, area_trabajo FROM tickets_produccion WHERE id = %s", (ticket_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        cursor.execute("SELECT rol, area_asignada FROM usuarios WHERE id = %s", (trabajador_id,))
        trabajador = cursor.fetchone()
        if not trabajador:
            return jsonify({'error': 'Trabajador no encontrado'}), 404
        areas_validas = {a.upper() for a in AREA_ALIASES.get(row[1], [row[1]])}
        if trabajador[0] not in ('Admin', 'Jefe_Taller') and str(trabajador[1] or '').upper() not in areas_validas:
            return jsonify({'error': 'El trabajador no pertenece al area de este ticket.'}), 400
        nuevo_estado = row[0] if row[0] == 'Bloqueado' else 'En Proceso'
        cursor.execute("""
            UPDATE tickets_produccion
            SET trabajador_asignado_id = %s,
                estado_ticket = %s,
                fecha_inicio = CASE WHEN %s = 'En Proceso' THEN CURRENT_TIMESTAMP ELSE fecha_inicio END
            WHERE id = %s;
        """, (trabajador_id, nuevo_estado, nuevo_estado, ticket_id))
        conexion.commit()

        # ── NOTIFICACIÓN: avisar al trabajador que tiene un ticket nuevo ──
        # No debe tumbar la respuesta si falla (red caída, correo mal
        # configurado, etc.) — la asignación ya quedó guardada arriba.
        try:
            cursor.execute("""
                SELECT u.nombre, u.email, u.telefono,
                       i.producto, t.area_trabajo, v.codigo_venta
                FROM tickets_produccion t
                JOIN items_venta i  ON t.item_id  = i.id
                JOIN ventas v       ON i.venta_id = v.id
                JOIN usuarios u     ON u.id        = t.trabajador_asignado_id
                WHERE t.id = %s
            """, (ticket_id,))
            info = cursor.fetchone()
            if info:
                nombre, email, telefono, producto, area, codigo_venta = info
                area_legible = (area or '').replace('_', ' ').title()
                notificar_usuario(
                    destinatario_email=email,
                    nombre_destinatario=nombre,
                    asunto=f"Nuevo ticket asignado — {codigo_venta}",
                    mensaje=(
                        f"Hola {nombre},\n\n"
                        f"Se te asignó un nuevo ticket de producción:\n\n"
                        f"  Producto: {producto}\n"
                        f"  Área: {area_legible}\n"
                        f"  Pedido: {codigo_venta}\n\n"
                        f"Ingresa al ERP para ver los detalles.\n\n"
                        f"Innova Möbili — Taller"
                    ),
                    telefono=telefono,
                )
        except Exception as e_notif:
            print(f"⚠️ No se pudo notificar la asignación del ticket {ticket_id}: {e_notif}")
        # ───────────────────────────────────────────────────────────────────

        return jsonify({'exito': True, 'mensaje': 'Maestro asignado correctamente', 'estado': nuevo_estado}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/taller/ticket/<int:ticket_id>/derivar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def derivar_ticket_con_foto(ticket_id):
    nueva_area          = request.form.get('nueva_area')
    nuevo_trabajador_id = request.form.get('nuevo_trabajador_id')
    if not nueva_area or not nuevo_trabajador_id:
        return jsonify({'error': 'nueva_area y nuevo_trabajador_id son obligatorios'}), 400
    foto_ruta = None
    if 'foto' in request.files and request.files['foto'].filename != '':
        respuesta_nube = cloudinary_upload(request.files['foto'], folder="derivaciones")
        foto_ruta = respuesta_nube.get('secure_url')
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT id FROM tickets_produccion WHERE id = %s", (ticket_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Ticket no encontrado'}), 404
        cursor.execute("""
            UPDATE tickets_produccion
            SET area_trabajo = %s, trabajador_asignado_id = %s,
                foto_evidencia = COALESCE(%s, foto_evidencia),
                estado_ticket = 'Pendiente', fecha_inicio = NULL, fecha_fin = NULL
            WHERE id = %s
        """, (nueva_area, int(nuevo_trabajador_id), foto_ruta, ticket_id))
        conexion.commit()
        return jsonify({'exito': True, 'nueva_area': nueva_area}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print(f'Error al derivar ticket con foto: {e}')
        return jsonify({'error': 'No se pudo derivar el ticket'}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/taller/ticket/derivar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def derivar_ticket():
    data            = request.json
    ticket_padre_id = data.get('ticket_padre_id')
    tapicero_id     = data.get('tapicero_id')
    cojinero_id     = data.get('cojinero_id')
    area_tapiceria  = data.get('area_tapiceria', 'TAPICERIA_SOFAS')
    if not ticket_padre_id or not tapicero_id:
        return jsonify({'error': 'ticket_padre_id y tapicero_id son obligatorios'}), 400
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT item_id FROM tickets_produccion WHERE id = %s", (ticket_padre_id,))
        res = cursor.fetchone()
        if not res:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        item_id = res[0]

        cursor.execute("""
            UPDATE tickets_produccion
            SET trabajador_asignado_id = %s, estado_ticket = 'Bloqueado', fecha_inicio = NULL
            WHERE item_id = %s AND area_trabajo = %s AND estado_ticket IN ('Bloqueado', 'Pendiente')
        """, (tapicero_id, item_id, area_tapiceria))

        if cursor.rowcount == 0:
            cursor.execute("""
                INSERT INTO tickets_produccion (item_id, area_trabajo, trabajador_asignado_id, estado_ticket, etapa)
                SELECT %s, %s, %s, 'Bloqueado', 2
                WHERE NOT EXISTS (
                    SELECT 1 FROM tickets_produccion
                    WHERE item_id = %s AND area_trabajo = %s AND estado_ticket != 'Terminado'
                )
            """, (item_id, area_tapiceria, tapicero_id, item_id, area_tapiceria))

        # Cojines: como solo hay un cojinero registrado, se asigna automáticamente
        # (el jefe no necesita seleccionarlo). Si el item ya tenía un ticket de
        # ARMADO_COJINES (creado al registrar la venta porque el mueble lleva cojines),
        # se le asigna directo. Si el jefe envió cojinero_id explícito y no existía
        # ticket, se crea uno nuevo (compatibilidad con flujo anterior).
        cojinero_final = cojinero_id
        if not cojinero_final:
            cursor.execute("""
                SELECT id FROM usuarios
                WHERE area_asignada = 'ARMADO_COJINES' AND rol = 'Operario'
                LIMIT 1
            """)
            row_cojinero = cursor.fetchone()
            if row_cojinero:
                cojinero_final = row_cojinero[0]

        if cojinero_final:
            cursor.execute("""
                UPDATE tickets_produccion
                SET trabajador_asignado_id = %s, estado_ticket = 'Bloqueado', fecha_inicio = NULL
                WHERE item_id = %s AND area_trabajo = 'ARMADO_COJINES' AND estado_ticket IN ('Bloqueado', 'Pendiente')
            """, (cojinero_final, item_id))

            if cursor.rowcount == 0 and cojinero_id:
                cursor.execute("""
                    INSERT INTO tickets_produccion (item_id, area_trabajo, trabajador_asignado_id, estado_ticket, etapa)
                    SELECT %s, 'ARMADO_COJINES', %s, 'Bloqueado', 2
                    WHERE NOT EXISTS (
                        SELECT 1 FROM tickets_produccion
                        WHERE item_id = %s AND area_trabajo = 'ARMADO_COJINES' AND estado_ticket != 'Terminado'
                    )
                """, (item_id, cojinero_final, item_id))

        conexion.commit()
        return jsonify({'exito': True}), 201
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

@produccion_bp.route('/api/taller/stats', methods=['GET'])
@requiere_login
def obtener_taller_stats():
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM tickets_produccion WHERE estado_ticket = 'Pendiente'")
        pendientes = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM tickets_produccion WHERE estado_ticket = 'En Proceso'")
        en_proceso = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM tickets_produccion WHERE estado_ticket IN ('Pendiente', 'En Proceso', 'Bloqueado')")
        activos = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM ventas WHERE estado_general = 'Listo'")
        ventas_listas = cursor.fetchone()[0]
        
        return jsonify({
            'pendientes': pendientes,
            'en_proceso': en_proceso,
            'activos': activos,
            'ventas_listas': ventas_listas
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

@produccion_bp.route('/api/taller/ordenes', methods=['GET'])
@requiere_login
def obtener_ordenes_produccion():
    """
    Julio 2026 — se eliminó el N+1 de este endpoint (era el más grave de
    todo el sistema, según la auditoría de rendimiento): antes hacía
    1 query por venta (items_venta) + 1 por venta (logistica_externa) +
    1 por CADA ítem de cada venta (tickets_produccion). Con 50 ventas
    activas y 3 ítems promedio, eso eran ~250 queries por cada carga de
    esta vista.

    Ahora son 3 queries fijas sin importar cuántas ventas/ítems haya:
      1. Ventas (activas + últimas 30 entregadas, igual que antes)
      2. TODOS los items_venta + TODA la logística de tela de esas ventas,
         con WHERE venta_id IN (...)
      3. TODOS los tickets de TODOS esos items, con WHERE item_id IN (...)
    y el armado por venta/item se hace en Python con diccionarios, no con
    más queries.

    Tope de seguridad: 'activas' ahora tiene LIMIT 150 (antes no tenía
    ningún límite y crecía sin techo con cada venta no entregada/cancelada).
    Es un parche, no paginación real — si este número se vuelve chico,
    conviene aplicarle el mismo patrón de paginación server-side que ya
    tienen Mis Pedidos y Entregados.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # 0) Conteo real de ventas activas — sirve para detectar si el
        #    LIMIT 150 de abajo está cortando pedidos de verdad. Es una
        #    query barata (COUNT sobre el mismo WHERE) y evita que el
        #    corte sea silencioso.
        cursor.execute("""
            SELECT COUNT(*) FROM ventas
            WHERE estado_general NOT IN ('Entregado', 'Cancelado')
        """)
        total_activas_real = cursor.fetchone()[0] or 0
        truncado = total_activas_real > 150

        # 1) Ventas: activas (tope 150) + últimas 30 entregadas
        cursor.execute("""
            SELECT v.id, v.codigo_venta, v.nombre_cliente, v.fecha_entrega,
                   v.vendedor_nombre, v.sede, v.estado_general
            FROM ventas v
            WHERE v.estado_general NOT IN ('Entregado', 'Cancelado')
            ORDER BY v.fecha_entrega ASC NULLS LAST
            LIMIT 150
        """)
        ventas = cursor.fetchall()

        if not ventas:
            respuesta = jsonify([])
            if truncado:
                respuesta.headers['X-Ordenes-Truncado']      = 'true'
                respuesta.headers['X-Ordenes-Activas-Total'] = str(total_activas_real)
            return respuesta, 200

        venta_ids = [v[0] for v in ventas]

        # 2) TODOS los items_venta de estas ventas, en una sola query
        cursor.execute("""
            SELECT id, venta_id, producto, foto_url
            FROM items_venta
            WHERE venta_id = ANY(%s)
            ORDER BY venta_id, id
        """, (venta_ids,))
        items_por_venta = {}
        item_ids = []
        for item_id, venta_id, producto, foto_url in cursor.fetchall():
            items_por_venta.setdefault(venta_id, []).append((item_id, producto, foto_url))
            item_ids.append(item_id)

        # 3) TODA la logística de tela de estas ventas, en una sola query
        #    (el subquery de tapicero_destino sigue correlacionado por fila,
        #    pero eso lo resuelve Postgres en UNA sola pasada, no es un
        #    query por venta desde Python)
        cursor.execute("""
            SELECT l.venta_id, l.id, l.insumo_nombre, l.estado_distribucion,
                   COALESCE(ud.nombre, ur.nombre, u.nombre, 'Sin asignar') AS operario_nombre,
                   COALESCE(ur.nombre, '') AS recogido_por,
                   COALESCE(ud.nombre, '') AS distribuido_por,
                   l.item_id, COALESCE(l.item_ids_extra, ''),
                   COALESCE(
                       (SELECT ARRAY_AGG(lei.item_id ORDER BY lei.item_id)
                        FROM logistica_externa_items lei
                        WHERE lei.logistica_id = l.id),
                       ARRAY[]::INTEGER[]
                   ) AS item_ids_normalizados,
                   (SELECT COALESCE(u2.nombre, 'Sin asignar')
                    FROM tickets_produccion tp2
                    JOIN items_venta iv2 ON tp2.item_id = iv2.id
                    LEFT JOIN usuarios u2 ON tp2.trabajador_asignado_id = u2.id
                    WHERE iv2.venta_id = l.venta_id
                      AND tp2.area_trabajo IN ('TAPICERIA_SOFAS','TAPICERIA_SILLAS')
                    LIMIT 1) AS tapicero_destino
            FROM logistica_externa l
            LEFT JOIN usuarios u ON l.operario_id = u.id
            LEFT JOIN usuarios ur ON l.recogido_por_id = ur.id
            LEFT JOIN usuarios ud ON l.distribuido_por_id = ud.id
            WHERE l.venta_id = ANY(%s)
              AND (l.categoria_insumo = 'TELA'
                   OR LOWER(l.insumo_nombre) LIKE '%%tela%%'
                   OR LOWER(COALESCE(l.unidad, '')) IN ('mts', 'metro', 'metros'))
              AND l.estado_distribucion IS NOT NULL
            ORDER BY l.venta_id, l.id
        """, (venta_ids,))
        logistica_por_venta = {}
        for (venta_id, log_id, insumo_nombre, estado_dist, operario_nombre,
             recogido_por, distribuido_por, item_id_logistica, item_ids_extra,
             item_ids_normalizados, tapicero_destino) in cursor.fetchall():
            logistica_por_venta.setdefault(venta_id, []).append(
                (log_id, insumo_nombre, estado_dist, operario_nombre,
                 recogido_por, distribuido_por, item_id_logistica, item_ids_extra,
                 item_ids_normalizados, tapicero_destino)
            )

        # 4) TODOS los tickets de TODOS los items, en una sola query.
        #    DISTINCT ON (item_id, area_trabajo) reproduce exactamente el
        #    comportamiento anterior (un ticket por área por item, el más
        #    antiguo de esa área) pero para todos los items a la vez.
        tickets_por_item = {}
        if item_ids:
            cursor.execute("""
                SELECT DISTINCT ON (item_id, area_trabajo)
                       item_id, id, area_trabajo, estado_ticket,
                       COALESCE((SELECT nombre FROM usuarios WHERE id = trabajador_asignado_id), 'Sin asignar')
                FROM tickets_produccion
                WHERE item_id = ANY(%s)
                ORDER BY item_id, area_trabajo, id ASC
            """, (item_ids,))
            for item_id, t_id, area, estado_ticket, trabajador in cursor.fetchall():
                tickets_por_item.setdefault(item_id, []).append((t_id, area, estado_ticket, trabajador))

        # 5) Armar el resultado por venta, igual que antes, pero sin
        #    disparar ninguna query dentro de este loop.
        resultado = []
        for v in ventas:
            venta_id = v[0]
            items = items_por_venta.get(venta_id, [])
            logistica_telas = logistica_por_venta.get(venta_id, [])

            items_list = []
            tickets_term = 0
            tickets_total = 0

            for item_id, producto, foto_url in items:
                tickets = tickets_por_item.get(item_id, [])

                tickets_list = []
                for t_id, area, estado_ticket, trabajador in tickets:
                    tickets_total += 1
                    if estado_ticket in ('Terminado', 'Listo para Recojo', 'Recogido', 'Cancelado'):
                        tickets_term += 1
                    tickets_list.append({
                        'id': t_id,
                        'area': area,
                        'estado': estado_ticket,
                        'trabajador': trabajador,
                        'es_logistica': False,
                    })

                items_list.append({
                    'id': item_id,
                    'producto': producto,
                    'foto': limpiar_foto(foto_url.split('|')[0] if foto_url else ''),
                    'tickets': tickets_list,
                })

            # Mostrar la tela bajo cada item asociado, pero contar la operación
            # una sola vez en el progreso del contrato.
            for lg in logistica_telas:
                (log_id, insumo_nombre, estado_dist, operario_nombre,
                 recogido_por, distribuido_por, item_id_logistica, item_ids_extra,
                 item_ids_normalizados, tapicero_destino) = lg
                estado_display = {
                    'En espera':   'Pendiente en Telas',
                    'En Recojo':   'En Recojo',
                    'Recogido':    'Recogido',
                    'Distribuido': 'Distribuido',
                }.get(estado_dist or '', estado_dist or 'Pendiente')
                tickets_total += 1
                if estado_dist == 'Distribuido':
                    tickets_term += 1
                entrada_logistica = {
                    'id': log_id,
                    'area': 'CORTE_Y_CONTROL_TELAS',
                    'estado': estado_display,
                    'trabajador': operario_nombre,
                    'recogido_por': recogido_por,
                    'distribuido_por': distribuido_por,
                    'tapicero_destino': tapicero_destino,
                    'insumo_nombre': insumo_nombre,
                    'es_logistica': True,
                }
                ids_asociados = (
                    list(item_ids_normalizados)
                    if item_ids_normalizados
                    else ids_items_logistica(item_id_logistica, item_ids_extra)
                )
                items_destino = [
                    item for item in items_list if item['id'] in ids_asociados
                ]
                if not items_destino and items_list:
                    items_destino = [items_list[0]]
                if items_destino:
                    for item_destino in items_destino:
                        item_destino['tickets'].append(dict(entrada_logistica))
                else:
                    items_list.append({
                        'id': None,
                        'producto': f'TELA EXTERNA: {insumo_nombre}',
                        'foto': limpiar_foto(''),
                        'tickets': [entrada_logistica],
                    })

            progreso = round((tickets_term / tickets_total * 100)) if tickets_total > 0 else 0
            resultado.append({
                'id': venta_id, 'codigo': v[1], 'codigo_venta': v[1], 'cliente': v[2],
                'fecha_entrega': v[3].strftime('%d/%m/%Y') if v[3] else 'S/F',
                'vendedor': v[4], 'sede': v[5], 'estado': v[6], 'estado_general': v[6],
                'progreso': progreso, 'tickets_term': tickets_term,
                'tickets_total': tickets_total, 'items': items_list,
            })

        respuesta = jsonify(resultado)
        if truncado:
            # No cambia la forma del JSON (sigue siendo un array plano, el
            # frontend hace `_opTodos = data` directo) — el aviso va en
            # headers para no romper esa asignación en busqueda_filtros.js.
            respuesta.headers['X-Ordenes-Truncado']      = 'true'
            respuesta.headers['X-Ordenes-Activas-Total'] = str(total_activas_real)
        return respuesta, 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# ==========================================
# 9. INVENTARIO DE MATERIALES
# ==========================================

@produccion_bp.route('/api/taller/fichatecnica-skus', methods=['GET'])
@requiere_login
def obtener_fotos_skus():
    skus_param = request.args.get('skus', '')
    if not skus_param:
        return jsonify([]), 200
    skus = [s.strip().upper() for s in skus_param.split(',') if s.strip()]
    if not skus:
        return jsonify([]), 200
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        placeholders = ','.join(['%s'] * len(skus))
        cursor.execute(f"""
            SELECT sku, CONCAT(coleccion, ' - ', color) AS nombre, foto_url, 'tela' AS tipo
            FROM maestro_telas WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados = [{'sku': r[0], 'nombre': r[1], 'foto_url': limpiar_foto(r[2]), 'tipo': r[3]} for r in cursor.fetchall()]
        
        cursor.execute(f"""
            SELECT sku, nombre_diseno AS nombre, foto_url, 'cojin' AS tipo
            FROM maestro_disenos_cojin WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados += [{'sku': r[0], 'nombre': r[1], 'foto_url': limpiar_foto(r[2]), 'tipo': r[3]} for r in cursor.fetchall()]
        
        cursor.execute(f"""
            SELECT sku, CONCAT(modelo, ' - ', color) AS nombre, foto_url, 'base' AS tipo
            FROM maestro_bases WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados += [{'sku': r[0], 'nombre': r[1], 'foto_url': limpiar_foto(r[2]), 'tipo': r[3]} for r in cursor.fetchall()]
        
        cursor.execute(f"""
            SELECT sku, CONCAT(modelo, ' - ', color) AS nombre, foto_url, 'base-comedor' AS tipo
            FROM maestro_bases_comedor WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados += [{'sku': r[0], 'nombre': r[1], 'foto_url': limpiar_foto(r[2]), 'tipo': r[3]} for r in cursor.fetchall()]
        
        cursor.execute(f"""
            SELECT sku, CONCAT(nombre_modelo, ' - ', color_veta) AS nombre, foto_url, 'tablero' AS tipo
            FROM maestro_tableros WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados += [{'sku': r[0], 'nombre': r[1], 'foto_url': limpiar_foto(r[2]), 'tipo': r[3]} for r in cursor.fetchall()]
        
        cursor.execute(f"""
            SELECT sku, CONCAT(modelo, ' - ', color_estructura) AS nombre, foto_url, 'silla' AS tipo
            FROM maestro_sillas WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados += [{'sku': r[0], 'nombre': r[1], 'foto_url': limpiar_foto(r[2]), 'tipo': r[3]} for r in cursor.fetchall()]
        
        cursor.execute(f"""
            SELECT sku, CONCAT(modelo, ' - ', color_estructura) AS nombre, foto_url, 'butaca' AS tipo
            FROM maestro_butacas WHERE UPPER(sku) IN ({placeholders})
        """, skus)
        resultados += [{'sku': r[0], 'nombre': r[1], 'foto_url': limpiar_foto(r[2]), 'tipo': r[3]} for r in cursor.fetchall()]
        
        return jsonify(resultados), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

@produccion_bp.route('/api/taller/ticket/<int:ticket_id>/nota', methods=['POST'])
@requiere_login
def agregar_nota_ticket(ticket_id):
    data = request.json or {}
    nota = data.get('nota')
    actor = get_usuario_actual()
    usuario = actor.get('nombre', 'Usuario')
    
    if not nota:
        return jsonify({'error': 'La nota es obligatoria'}), 400
        
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT area_trabajo, trabajador_asignado_id FROM tickets_produccion WHERE id = %s", (ticket_id,))
        ticket = cursor.fetchone()
        if not ticket:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        if not usuario_puede_operar(ticket[0], ticket[1]):
            return jsonify({'error': 'No puedes editar notas de otro ticket.'}), 403
        nota_formateada = f"\n\n[NOTA {datetime.now().strftime('%d/%m %H:%M')} - {usuario}]: {nota}"
        cursor.execute("UPDATE tickets_produccion SET ticket_details_override = COALESCE(ticket_details_override, '') || %s WHERE id = %s", (nota_formateada, ticket_id))
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Nota agregada correctamente'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/taller/inventario', methods=['GET'])
@requiere_login
def obtener_inventario():
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT id, nombre_modelo AS nombre, COALESCE(estado,'Disponible'), 'TABLERO' AS cat, origen_produccion FROM maestro_tableros
            UNION ALL SELECT id, color, COALESCE(estado,'Disponible'), 'TELA', origen_produccion FROM maestro_telas
            UNION ALL SELECT id, nombre_diseno, COALESCE(estado,'Disponible'), 'COJIN', origen_produccion FROM maestro_disenos_cojin
            UNION ALL SELECT id, modelo, COALESCE(estado,'Disponible'), 'BASE', origen_produccion FROM maestro_bases
            UNION ALL SELECT id, modelo, COALESCE(estado,'Disponible'), 'BASE-COMEDOR', origen_produccion FROM maestro_bases_comedor
            UNION ALL SELECT id, modelo, COALESCE(estado,'Disponible'), 'SILLA', origen_produccion FROM maestro_sillas
            UNION ALL SELECT id, modelo, COALESCE(estado,'Disponible'), 'BUTACA', origen_produccion FROM maestro_butacas
            ORDER BY cat, nombre;
        """)
        insumos = [
            {"id": r[0], "nombre": r[1], "estado": r[2], "categoria": r[3]}
            for r in cursor.fetchall()
        ]
        return jsonify(insumos), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/inventario/actualizar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def actualizar_estado_inventario():
    data         = request.json
    item_id      = data.get('id')
    categoria    = data.get('categoria')
    nuevo_estado = data.get('estado')
    tablas_permitidas = {
        'TELA': 'maestro_telas', 'COJIN': 'maestro_disenos_cojin',
        'BASE': 'maestro_bases', 'BASE-COMEDOR': 'maestro_bases_comedor',
        'TABLERO': 'maestro_tableros', 'SILLA': 'maestro_sillas', 'BUTACA': 'maestro_butacas',
    }
    tabla = tablas_permitidas.get(categoria)
    if not tabla:
        return jsonify({'error': 'Categoría no válida'}), 400
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute(f"UPDATE {tabla} SET estado = %s WHERE id = %s;", (nuevo_estado, item_id))
        conexion.commit()
        return jsonify({'exito': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# 10. LOGÍSTICA EXTERNA
# ==========================================

@produccion_bp.route('/api/logistica', methods=['GET'])
@requiere_login
def obtener_logistica():
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute("""
            SELECT l.id, v.codigo_venta, l.insumo_nombre, l.sku,
                   COALESCE(p.nombre, 'Sin asignar') AS proveedor,
                   COALESCE(p.correo, '')            AS correo_proveedor,
                   l.precio_cotizado, l.fecha_entrega_proveedor, l.estado,
                   l.token_usado, l.notas_proveedor, l.url_comprobante_pago,
                   COALESCE(l.cantidad, 1)           AS cantidad,
                   COALESCE(l.unidad, '')            AS unidad,
                   COALESCE(l.tipo_gestion, 'Externo') AS tipo_gestion,
                   l.proveedor_id,
                   COALESCE(l.proveedor_informal, '') AS proveedor_informal,
                   l.url_cotizacion_adjunta,
                   -- Foto 1: la del insumo del maestro, encontrada por SKU
                   -- (la que sale del buscador inteligente al elegir la
                   -- parte del catálogo al armar el pedido).
                   --
                   -- FIX (julio 2026): piezas "✨ PINTEREST" del configurador
                   -- (materiales.js → abrirModalPinterest) usan un SKU temporal
                   -- REQ-PIN-{id} mientras la sugerencia espera aprobación del
                   -- Admin. Ese SKU nunca existe en las tablas maestro_*, así
                   -- que antes esta tarjeta salía sin foto aunque el vendedor
                   -- sí hubiera subido la imagen de referencia — esa foto vive
                   -- en sugerencias_insumos.foto_ref. Se agrega como primera
                   -- opción del COALESCE para recuperarla por el id embebido
                   -- en el SKU (REQ-PIN-<id>).
                   COALESCE(
                       (SELECT foto_ref FROM sugerencias_insumos
                          WHERE l.sku LIKE 'REQ-PIN-%'
                            AND id::text = split_part(l.sku, '-', 3)
                          LIMIT 1),
                       (SELECT foto_url FROM maestro_telas        WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_tableros      WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_bases         WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_bases_comedor WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_sillas        WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_butacas       WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_disenos_cojin WHERE sku = l.sku LIMIT 1)
                   ) AS foto_maestro,
                   -- Foto 2: la foto propia del ítem de venta que generó este
                   -- requerimiento (la que el vendedor sube aparte). Se
                   -- empareja por item_id cuando existe; si el registro es
                   -- viejo (de antes de este fix) y no tiene item_id, cae al
                   -- fallback anterior (primer ítem de la venta) solo para
                   -- no perder la foto que ya se venía mostrando.
                   COALESCE(
                       (SELECT i2.foto_url FROM items_venta i2 WHERE i2.id = l.item_id LIMIT 1),
                       (SELECT i2.foto_url FROM items_venta i2 WHERE i2.venta_id = l.venta_id AND l.item_id IS NULL LIMIT 1)
                   ) AS foto_item,
                   -- Nombre del mueble al que pertenece este requerimiento (ej.
                   -- "Silla Comedor Luna", "Sofá Curvo 3 Cuerpos"). Mismo
                   -- emparejamiento por item_id que foto_item, con el mismo
                   -- fallback para registros viejos sin item_id. Sin esto, la
                   -- tarjeta de "Requerimientos" mostraba la foto del mueble
                   -- pero nunca decía a cuál mueble pertenecía.
                   COALESCE(
                       (SELECT i2.producto FROM items_venta i2 WHERE i2.id = l.item_id LIMIT 1),
                       (SELECT i2.producto FROM items_venta i2 WHERE i2.venta_id = l.venta_id AND l.item_id IS NULL LIMIT 1)
                   ) AS producto_item,
                   -- Ficha/notas del item: aqui suelen vivir medidas, notas de
                   -- casqueria y textos largos del configurador. El frontend
                   -- las usa para mostrar referencias y sugerir metros de tela
                   -- cuando el vendedor los escribio en notas.
                   COALESCE(
                       (SELECT COALESCE(i2.color_tela, '') FROM items_venta i2 WHERE i2.id = l.item_id LIMIT 1),
                       (SELECT COALESCE(i2.color_tela, '') FROM items_venta i2 WHERE i2.venta_id = l.venta_id AND l.item_id IS NULL LIMIT 1),
                       ''
                   ) AS especificaciones_item,
                   -- Detalles descriptivos del insumo según su tipo
                   COALESCE(
                       (SELECT CONCAT(coleccion, ' · ', color) FROM maestro_telas WHERE sku = l.sku LIMIT 1),
                       (SELECT CONCAT(nombre_modelo, ' · ', color_veta, CASE WHEN acabado != '' THEN CONCAT(' · ', acabado) ELSE '' END) FROM maestro_tableros WHERE sku = l.sku LIMIT 1),
                       (SELECT CONCAT(modelo, ' · ', color) FROM maestro_bases WHERE sku = l.sku LIMIT 1),
                       (SELECT CONCAT(modelo, ' · ', color) FROM maestro_bases_comedor WHERE sku = l.sku LIMIT 1),
                       (SELECT CONCAT(modelo, ' · ', color_estructura) FROM maestro_sillas WHERE sku = l.sku LIMIT 1),
                       (SELECT CONCAT(modelo, ' · ', color_estructura) FROM maestro_butacas WHERE sku = l.sku LIMIT 1),
                       (SELECT nombre_diseno FROM maestro_disenos_cojin WHERE sku = l.sku LIMIT 1)
                   ) AS detalle_insumo,
                   COALESCE(l.categoria_insumo, 'OTRO')  AS categoria_insumo,
                   l.estado_distribucion,
                   COALESCE(p.telefono, '')           AS telefono_proveedor
            FROM logistica_externa l
            JOIN ventas v           ON l.venta_id    = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            ORDER BY l.estado ASC, l.id DESC;
        """)
        items = []
        for r in cursor.fetchall():
            # NOTA: se insertaron producto_item y especificaciones_item entre
            # foto_item y detalle_insumo, por eso los índices desde foto_maestro
            # en adelante se corrieron respecto a la version original.
            foto_maestro = limpiar_foto(r[18]) if r[18] else ""
            foto_item    = limpiar_foto(r[19]) if r[19] else ""
            # Lista de fotos sin duplicar (si ambas apuntan a la misma URL,
            # el frontend debe mostrar una sola, no un carrusel de 2 iguales)
            fotos = [f for f in [foto_maestro, foto_item] if f]
            if len(fotos) == 2 and fotos[0] == fotos[1]:
                fotos = fotos[:1]
            items.append({
                "id": r[0], "codigo_venta": r[1], "insumo": r[2], "sku": r[3],
                "proveedor": r[4], "correo_proveedor": r[5],
                "precio_cotizado": float(r[6]) if r[6] else None,
                "fecha_entrega_proveedor": r[7].strftime('%d/%m/%Y') if r[7] else None,
                "estado": r[8],
                "token_usado": r[9], "notas_proveedor": r[10],
                "url_comprobante_pago":    r[11],
                "cantidad":                float(r[12]) if r[12] else 1,
                "unidad":                  r[13],
                "tipo_gestion":            r[14],
                "proveedor_id":            r[15],
                "proveedor_informal":      r[16] or "",
                "foto_maestro":            foto_maestro,
                "foto_item":               foto_item,
                "fotos":                   fotos,
                # Se mantiene foto_url por compatibilidad con cualquier otro
                # lugar del frontend que aún lo lea como foto única.
                "foto_url":                fotos[0] if fotos else "",
                "producto_item":           r[20] or "",
                "especificaciones_item":   r[21] or "",
                "detalle_insumo":          r[22] or "",
                "url_cotizacion_adjunta":  r[17] if len(r) > 17 else None,
                "categoria_insumo":        r[23] if len(r) > 23 else 'OTRO',
                "estado_distribucion":     r[24] if len(r) > 24 else None,
                "telefono_proveedor":      r[25] if len(r) > 25 else "",
            })
        return jsonify(items), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/telas-por-contrato', methods=['GET'])
@requiere_login
def obtener_telas_por_contrato():
    """
    Bandeja operativa del area Telas.

    Devuelve una tarjeta por contrato y dentro agrupa todas las telas que ese
    contrato necesita comprar, recoger o distribuir. Esta es la base para dejar
    de manejar telas sueltas por WhatsApp.
    """
    try:
        if not usuario_puede_operar('CORTE_Y_CONTROL_TELAS'):
            return jsonify({'error': 'Esta bandeja pertenece al area de Telas.'}), 403
        operario_id = request.args.get('operario_id')
        operario_id = int(operario_id) if operario_id else None
        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT
                v.id AS venta_id,
                v.codigo_venta,
                v.nombre_cliente,
                v.estado_general,
                v.fecha_entrega,
                l.id AS logistica_id,
                l.item_id,
                COALESCE(i.producto, '') AS producto_item,
                l.insumo_nombre,
                l.sku,
                COALESCE(l.cantidad, 1) AS cantidad,
                COALESCE(l.unidad, '') AS unidad,
                COALESCE(l.estado, 'Pendiente') AS estado,
                COALESCE(l.estado_distribucion, 'En espera') AS estado_distribucion,
                COALESCE(l.tipo_gestion, 'Externo') AS tipo_gestion,
                COALESCE(p.nombre, 'Sin asignar') AS proveedor,
                l.proveedor_id,
                l.precio_cotizado,
                l.fecha_entrega_proveedor,
                l.url_comprobante_pago,
                l.item_ids_extra,
                l.operario_id,
                COALESCE(ud.nombre, ur.nombre, u.nombre, 'Sin asignar') AS operario_nombre,
                COALESCE(ur.nombre, '') AS recogido_por,
                COALESCE(ud.nombre, '') AS distribuido_por
            FROM logistica_externa l
            JOIN ventas v ON l.venta_id = v.id
            LEFT JOIN items_venta i ON i.id = l.item_id
            LEFT JOIN proveedores p ON p.id = l.proveedor_id
            LEFT JOIN usuarios u ON u.id = l.operario_id
            LEFT JOIN usuarios ur ON ur.id = l.recogido_por_id
            LEFT JOIN usuarios ud ON ud.id = l.distribuido_por_id
            WHERE COALESCE(v.estado_general, '') != 'Cancelado'
              AND (
                    COALESCE(l.categoria_insumo, '') = 'TELA'
                    OR LOWER(COALESCE(l.insumo_nombre, '')) LIKE '%%tela%%'
                    OR LOWER(COALESCE(l.unidad, '')) IN ('mts', 'metro', 'metros')
              )
              AND (%s IS NULL OR l.operario_id = %s)
            ORDER BY v.fecha_entrega ASC NULLS LAST, v.id ASC, l.id ASC;
        """, (operario_id, operario_id))

        rows = cursor.fetchall()
        destinos_por_venta = {}
        venta_ids = {r[0] for r in rows}
        if venta_ids:
            placeholders = ','.join(['%s'] * len(venta_ids))
            cursor.execute(f"""
                SELECT i.venta_id, t.area_trabajo, COALESCE(u.nombre, 'Sin asignar')
                FROM tickets_produccion t
                JOIN items_venta i ON t.item_id = i.id
                LEFT JOIN usuarios u ON t.trabajador_asignado_id = u.id
                WHERE i.venta_id IN ({placeholders})
                  AND t.area_trabajo IN ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS', 'ARMADO_COJINES')
            """, tuple(venta_ids))
            for venta_id, area, nombre in cursor.fetchall():
                destino = destinos_por_venta.setdefault(venta_id, {"tapicero": None, "cojinero": None})
                if area in ('TAPICERIA_SOFAS', 'TAPICERIA_SILLAS'):
                    destino["tapicero"] = nombre
                elif area == 'ARMADO_COJINES':
                    destino["cojinero"] = nombre

        tarjetas = {}
        for r in rows:
            venta_id = r[0]
            destino = destinos_por_venta.get(venta_id, {})
            tarjeta = tarjetas.setdefault(venta_id, {
                "venta_id": venta_id,
                "codigo_venta": r[1],
                "cliente": r[2],
                "estado_contrato": r[3],
                "fecha_entrega": r[4].strftime('%d/%m/%Y') if r[4] else None,
                "tapicero_destino": destino.get("tapicero"),
                "cojinero_destino": destino.get("cojinero"),
                "total_telas": 0,
                "pendientes": 0,
                "en_recojo": 0,
                "recogidas": 0,
                "distribuidas": 0,
                "lineas": [],
            })

            estado_dist = normalizar_estado_distribucion(r[13])
            tarjeta["total_telas"] += 1
            if estado_dist == "Distribuido":
                tarjeta["distribuidas"] += 1
            elif estado_dist == "Recogido":
                tarjeta["recogidas"] += 1
            elif estado_dist == "En Recojo":
                tarjeta["en_recojo"] += 1
            else:
                tarjeta["pendientes"] += 1

            tarjeta["lineas"].append({
                "logistica_id": r[5],
                "item_id": r[6],
                "producto_item": r[7],
                "insumo": r[8],
                "sku": r[9],
                "cantidad": float(r[10]) if r[10] is not None else 1,
                "unidad": r[11],
                "estado": r[12],
                "estado_distribucion": estado_dist,
                "tipo_gestion": r[14],
                "proveedor": r[15],
                "proveedor_id": r[16],
                "precio_cotizado": float(r[17]) if r[17] is not None else None,
                "fecha_entrega_proveedor": r[18].strftime('%d/%m/%Y') if r[18] else None,
                "url_comprobante_pago": r[19],
                "item_ids_extra": r[20] or "",
                "operario_id": r[21],
                "operario_nombre": r[22],
                "recogido_por": r[23],
                "distribuido_por": r[24],
            })

        return jsonify(list(tarjetas.values())), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/actualizar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def actualizar_logistica():
    data                    = request.get_json(silent=True) or {}
    actualiza_proveedor     = 'proveedor_id' in data
    actualiza_informal      = 'proveedor_informal' in data
    actualiza_notas         = 'notas_proveedor' in data
    logistica_id            = data.get('id')
    proveedor_id            = data.get('proveedor_id')
    precio_cotizado         = data.get('precio_cotizado')
    fecha_entrega_proveedor = data.get('fecha_entrega_proveedor')
    tipo_gestion            = data.get('tipo_gestion')
    cantidad                = data.get('cantidad')
    unidad                  = data.get('unidad')
    proveedor_informal      = data.get('proveedor_informal')
    notas_proveedor         = data.get('notas_proveedor')           # ← respuesta WA
    url_cotizacion_adjunta  = data.get('url_cotizacion_adjunta')    # ← foto/PDF del proveedor
    estado                  = data.get('estado')                    # ← cambio de etapa del flujo
    try:
        logistica_id = int(logistica_id)
    except (TypeError, ValueError):
        return jsonify({'error': 'id es obligatorio'}), 400
    if logistica_id <= 0:
        return jsonify({'error': 'id es obligatorio'}), 400

    if proveedor_id not in (None, ''):
        try:
            proveedor_id = int(proveedor_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'proveedor_id no es válido'}), 400
        if proveedor_id <= 0:
            return jsonify({'error': 'proveedor_id no es válido'}), 400
    else:
        proveedor_id = None

    if tipo_gestion is not None:
        tipo_gestion = str(tipo_gestion).strip()
        if tipo_gestion not in {'Externo', 'Interno', 'Informal'}:
            return jsonify({'error': 'tipo_gestion no es valido'}), 400

    def numero_opcional(valor, nombre, minimo, estricto=False):
        if valor is None or valor == '':
            return None
        if isinstance(valor, bool):
            raise ValueError(f'{nombre} no es valido')
        numero = float(valor)
        if not math.isfinite(numero) or numero < minimo or (estricto and numero == minimo):
            comparacion = 'mayor que' if estricto else 'mayor o igual que'
            raise ValueError(f'{nombre} debe ser {comparacion} {minimo}')
        return numero

    try:
        precio_cotizado = numero_opcional(precio_cotizado, 'precio_cotizado', 0, estricto=True)
        cantidad = numero_opcional(cantidad, 'cantidad', 0, estricto=True)
    except (TypeError, ValueError) as exc:
        return jsonify({'error': str(exc)}), 400
    if precio_cotizado is not None and precio_cotizado > 99999999.99:
        return jsonify({'error': 'precio_cotizado supera el límite permitido'}), 400
    if cantidad is not None and cantidad > 9999999999.99:
        return jsonify({'error': 'cantidad supera el límite permitido'}), 400

    if fecha_entrega_proveedor not in (None, ''):
        try:
            fecha_entrega_proveedor = normalizar_fecha_cotizacion(
                fecha_entrega_proveedor,
                fecha_minima=datetime.now(ZONA_HORARIA_LIMA).date(),
            )
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
    else:
        fecha_entrega_proveedor = None

    limites_texto = {
        'unidad': (unidad, 30),
        'proveedor_informal': (proveedor_informal, 200),
        'notas_proveedor': (notas_proveedor, 2000),
        'url_cotizacion_adjunta': (url_cotizacion_adjunta, 2000),
    }
    for nombre, (valor, limite) in limites_texto.items():
        if valor is not None and (not isinstance(valor, str) or len(valor) > limite):
            return jsonify({'error': f'{nombre} no es válido o supera {limite} caracteres'}), 400
    if url_cotizacion_adjunta and not re.match(r'^https?://', url_cotizacion_adjunta, re.I):
        return jsonify({'error': 'url_cotizacion_adjunta debe usar http o https'}), 400
    if estado is not None:
        if not isinstance(estado, str) or len(estado.strip()) > 30:
            return jsonify({'error': 'estado no es válido'}), 400
        estado = normalizar_estado_logistica(estado)
        if estado == 'Orden Enviada':
            return jsonify({'error': 'Genera la Orden de Compra para avanzar a Orden Enviada'}), 409
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute("""
            SELECT estado, COALESCE(tipo_gestion, 'Externo'),
                   COALESCE(estado_distribucion, 'En espera'),
                   url_comprobante_pago, precio_cotizado,
                   fecha_entrega_proveedor, proveedor_id
            FROM logistica_externa
            WHERE id = %s
            FOR UPDATE
        """, (logistica_id,))
        estado_anterior = cursor.fetchone()
        if not estado_anterior:
            conexion.rollback()
            return jsonify({'error': 'Ítem de logística no encontrado'}), 404
        tipo_gestion_final_previsto = tipo_gestion or estado_anterior[1]
        if tipo_gestion is not None and tipo_gestion != estado_anterior[1]:
            if normalizar_estado_logistica(estado_anterior[0]) != 'Pendiente':
                conexion.rollback()
                return jsonify({
                    'error': 'El tipo de gestión solo puede cambiar mientras el insumo está pendiente.'
                }), 409
        if estado is not None and not transicion_logistica_permitida(
            estado_anterior[0], estado, tipo_gestion_final_previsto
        ):
            conexion.rollback()
            return jsonify({
                'error': f'No se permite pasar de {estado_anterior[0]} a {estado}'
            }), 409
        if estado == 'Cotizacion Enviada' and not (proveedor_id or estado_anterior[6]):
            conexion.rollback()
            return jsonify({'error': 'Asigna un proveedor antes de enviar la cotización'}), 409
        if estado == 'Cotizado':
            precio_final = precio_cotizado if precio_cotizado is not None else estado_anterior[4]
            fecha_final = fecha_entrega_proveedor or estado_anterior[5]
            if not precio_final or not fecha_final:
                conexion.rollback()
                return jsonify({'error': 'Precio unitario y fecha son obligatorios para cotizar'}), 409
        if estado == 'Pagado' and not estado_anterior[3]:
            conexion.rollback()
            return jsonify({'error': 'Registra el comprobante antes de marcar como Pagado'}), 409
        if normalizar_estado_distribucion(estado_anterior[2]) == 'Distribuido':
            cambia_estado = estado is not None and estado != estado_anterior[0]
            cambia_gestion = tipo_gestion is not None and tipo_gestion != estado_anterior[1]
            if cambia_estado or cambia_gestion:
                conexion.rollback()
                return jsonify({
                    'error': 'Un insumo distribuido no puede cambiar de estado ni de tipo de gestión.'
                }), 409

        # FIX (julio 2026): antes, marcar tipo_gestion = 'Interno' en este
        # modal solo actualizaba esa columna — el registro se quedaba
        # "Pendiente" para siempre esperando una cotización o un recojo que
        # nunca iban a llegar (porque el taller lo fabrica/consigue por su
        # cuenta). La UI ya prometía esto (texto del modal en modules/app/logistica_externa.js:
        # "Marca como Recibido cuando esté listo... los tickets se
        # desbloquearán automáticamente") pero el backend nunca lo hacía.
        #
        # El desbloqueo debe dispararse cuando el insumo QUEDA en
        # tipo_gestion='Interno' Y estado='Recibido' — no apenas se elige
        # "Interno" (todavía no está listo en ese momento). Por eso se lee
        # el resultado YA aplicado el COALESCE (RETURNING), para que
        # funcione tanto si ambos campos llegan juntos en un solo POST,
        # como si "Interno" se eligió antes y "Recibido" se marca después
        # desde el selector de etapa final (que solo manda `estado`).
        #
        # FIX-2 (julio 2026): el bloque anterior desbloqueaba Tapicería
        # inmediatamente para CUALQUIER insumo Interno, incluida la TELA.
        # Eso saltaba por completo la bandeja compartida de Telas (donde
        # todos los operarios del área ven el insumo en 'En espera' y
        # confirman la distribución) — la tela pasaba de 'Pendiente' a
        # 'Distribuido' sin que ningún operario de Telas la viera nunca.
        # Ahora: si la categoría del insumo es 'TELA', Interno+Recibido solo
        # la deja en 'En espera' (igual que una tela comprada a proveedor)
        # para que aparezca en la bandeja de Telas; el desbloqueo real de
        # Tapicería/Cojines ocurre recién cuando un operario confirma la
        # distribución vía /api/logistica/<id>/confirmar-distribucion (ya
        # existente). Para insumos que NO son tela (estructurales/otros, sin
        # bandeja compartida equivalente) se mantiene el desbloqueo inmediato.
        cursor.execute("""
            UPDATE logistica_externa
            SET proveedor_id             = CASE WHEN %s THEN %s ELSE proveedor_id END,
                precio_cotizado          = COALESCE(%s, precio_cotizado),
                fecha_entrega_proveedor  = COALESCE(%s::date, fecha_entrega_proveedor),
                tipo_gestion             = COALESCE(%s, tipo_gestion),
                cantidad                 = COALESCE(%s, cantidad),
                unidad                   = COALESCE(%s, unidad),
                proveedor_informal       = CASE WHEN %s THEN %s ELSE proveedor_informal END,
                notas_proveedor          = CASE WHEN %s THEN %s ELSE notas_proveedor END,
                url_cotizacion_adjunta   = COALESCE(%s, url_cotizacion_adjunta),
                estado                   = COALESCE(%s, estado)
            WHERE id = %s
            RETURNING venta_id, tipo_gestion, estado, COALESCE(categoria_insumo, 'OTRO'),
                      item_id, item_ids_extra, COALESCE(estado_distribucion, 'En espera');
        """, (actualiza_proveedor, proveedor_id, precio_cotizado, fecha_entrega_proveedor,
              tipo_gestion, cantidad, unidad, actualiza_informal, proveedor_informal,
              actualiza_notas, notas_proveedor, url_cotizacion_adjunta,
              estado,
              logistica_id))

        row = cursor.fetchone()
        if not row:
            conexion.rollback()
            return jsonify({'error': 'Ítem de logística no encontrado'}), 404
        (venta_id, tipo_gestion_final, estado_final, categoria_insumo_final,
         item_id_logistica, item_ids_extra, estado_distribucion_final) = row

        gestion_directa = tipo_gestion_final in ('Interno', 'Informal')
        marcar_directo = gestion_directa and estado_final == 'Recibido'
        directo_es_tela = marcar_directo and categoria_insumo_final == 'TELA'
        directo_no_tela = marcar_directo and categoria_insumo_final != 'TELA'
        tela_ya_distribuida = (
            directo_es_tela
            and normalizar_estado_distribucion(estado_distribucion_final) == 'Distribuido'
        )

        if directo_es_tela and not tela_ya_distribuida:
            # Entra a la bandeja compartida de Telas, tal cual una tela
            # comprada a proveedor — todavía NO se desbloquea nada aquí.
            actor_id = int(get_usuario_actual()['id'])
            cursor.execute("""
                UPDATE logistica_externa
                SET estado_distribucion = 'En espera',
                    recogido_por_id = COALESCE(recogido_por_id, %s),
                    fecha_recojo_fisico = COALESCE(fecha_recojo_fisico, NOW())
                WHERE id = %s
            """, (actor_id, logistica_id))
        elif directo_no_tela:
            cursor.execute("""
                UPDATE logistica_externa
                SET estado_distribucion = 'Distribuido'
                WHERE id = %s
            """, (logistica_id,))

        # Desbloqueo inmediato de Tapicería/Cojines: solo para insumos de
        # gestión directa que NO son tela (no tienen bandeja compartida a la que
        # esperar). La tela Interna se desbloquea después, cuando un
        # operario confirma la distribución desde la bandeja de Telas.
        item_ids_afectados = _items_asociados_logistica(
            cursor, logistica_id, item_id_logistica, item_ids_extra
        )
        desbloqueados = (
            _intentar_desbloquear_items(cursor, item_ids_afectados)
            if directo_no_tela and venta_id else 0
        )
        if venta_id:
            _activar_despacho_si_listo(cursor, venta_id)

        conexion.commit()

        mensaje = 'Guardado correctamente'
        if tela_ya_distribuida:
            mensaje = ('Guardado correctamente. La tela ya estaba distribuida y '
                       'se conservo ese estado.')
        elif directo_es_tela:
            mensaje = ('Tela de gestión directa enviada a la bandeja de Telas. '
                       'Cuando un operario del área confirme la distribución, '
                       'se desbloqueará Tapicería/Cojines.')
        elif directo_no_tela:
            mensaje = (f'Insumo de gestión directa marcado como recibido. '
                       f'{desbloqueados} ticket(s) de Tapicería/Cojines desbloqueado(s).')
        return jsonify({'exito': True, 'mensaje': mensaje, 'desbloqueados': desbloqueados}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print(f'Error al actualizar logística: {e}')
        return jsonify({'error': 'No se pudo actualizar el insumo de logística'}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/<int:logistica_id>/enviar-al-taller', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def enviar_al_taller(logistica_id):
    """
    Flujo 'Informal': el jefe de taller consiguió el material por su cuenta
    y quiere notificar al área de despacho que ya puede armar el pedido.

    Marca la fila de logistica_externa como 'Recibido' (tipo_gestion = 'Informal')
    y desbloquea los tickets_produccion relacionados, igual que cuando llega
    un material externo formal.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Verificar que existe
        cursor.execute(
            """SELECT id, venta_id, insumo_nombre, item_id, item_ids_extra,
                      COALESCE(categoria_insumo, 'OTRO'), estado,
                      COALESCE(tipo_gestion, 'Externo')
                 FROM logistica_externa WHERE id = %s FOR UPDATE""",
            (logistica_id,)
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Ítem de logística no encontrado'}), 404

        venta_id      = row[1]
        insumo_nombre = row[2]
        item_ids = _items_asociados_logistica(cursor, logistica_id, row[3], row[4])
        es_tela = row[5] == 'TELA' or 'tela' in (insumo_nombre or '').lower()
        if normalizar_estado_logistica(row[6]) != 'Pendiente' or row[7] != 'Informal':
            return jsonify({
                'error': 'Solo un insumo informal pendiente puede enviarse directamente al taller.'
            }), 409

        # Marcar como recibido/informal y conservar quién lo ingresó al taller.
        actor_id = int(get_usuario_actual()['id'])
        cursor.execute("""
            UPDATE logistica_externa
            SET estado       = 'Recibido',
                tipo_gestion = 'Informal',
                recogido_por_id = COALESCE(recogido_por_id, %s),
                fecha_recojo_fisico = COALESCE(fecha_recojo_fisico, NOW()),
                estado_distribucion = CASE
                    WHEN %s AND estado_distribucion = 'Distribuido' THEN 'Distribuido'
                    WHEN %s THEN 'En espera'
                    ELSE 'Distribuido'
                END
            WHERE id = %s
              AND estado IN ('POR_PEDIR', 'Por Pedir', 'Pendiente')
              AND tipo_gestion = 'Informal';
        """, (actor_id, es_tela, es_tela, logistica_id))
        if cursor.rowcount != 1:
            conexion.rollback()
            return jsonify({'error': 'El insumo cambió de estado. Actualiza la bandeja.'}), 409

        # Una tela debe pasar por la bandeja de distribucion. Para otros
        # insumos solo se revisan los items relacionados, nunca todo el pedido.
        desbloqueados = 0 if es_tela else _intentar_desbloquear_items(cursor, item_ids)
        if venta_id:
            _activar_despacho_si_listo(cursor, venta_id)

        conexion.commit()
        return jsonify({
            'exito':        True,
            'mensaje':      (
                f'"{insumo_nombre}" enviado a la bandeja de Telas; falta distribuirlo.'
                if es_tela else
                f'"{insumo_nombre}" marcado como recibido. {desbloqueados} ticket(s) desbloqueado(s).'
            ),
            'desbloqueados': desbloqueados,
        }), 200

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print(f'Error al enviar insumo directo al taller: {e}')
        return jsonify({'error': 'No se pudo enviar el insumo al taller'}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# NUEVOS ENDPOINTS LOGÍSTICA (RECOJO / DISTRIBUCIÓN)
# ==========================================

@produccion_bp.route('/api/logistica/<int:id>/confirmar-recojo', methods=['POST'])
@requiere_login
def logistica_confirmar_recojo(id):
    conexion = None
    cursor = None
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT operario_id, COALESCE(categoria_insumo, 'OTRO'),
                   COALESCE(estado_distribucion, 'En espera'), estado,
                   COALESCE(tipo_gestion, 'Externo'), url_comprobante_pago
            FROM logistica_externa WHERE id = %s FOR UPDATE
        """, (id,))
        fila = cursor.fetchone()
        if not fila:
            return jsonify({'error': 'Item de logistica no encontrado'}), 404
        if not usuario_puede_operar('CORTE_Y_CONTROL_TELAS', fila[0]):
            return jsonify({'error': 'Solo el operario asignado, Telas o jefatura puede confirmar el recojo.'}), 403
        if fila[1] != 'TELA':
            return jsonify({'error': 'Este control de recojo es exclusivo para telas.'}), 400
        estado_distribucion = normalizar_estado_distribucion(fila[2])
        if estado_distribucion == 'Distribuido' or fila[3] in ('Cancelado', 'Rechazado'):
            return jsonify({'error': 'La tela ya fue distribuida o esta cancelada.'}), 409
        if estado_distribucion == 'Recogido':
            return jsonify({'exito': True, 'mensaje': 'La tela ya estaba recogida.'}), 200
        if not tela_requiere_comprobante(fila[4]):
            return jsonify({
                'error': 'Esta tela es interna o informal: entregala directamente a Tapiceria.'
            }), 409
        if estado_distribucion not in ('En espera', 'En Recojo'):
            return jsonify({'error': 'La tela no esta disponible para recojo.'}), 409

        url_comprobante = None
        if request.content_type and 'multipart/form-data' in request.content_type:
            if 'comprobante' in request.files and request.files['comprobante'].filename:
                comprobante = request.files['comprobante']
                if not (
                    (comprobante.mimetype or '').startswith('image/')
                    or comprobante.mimetype == 'application/pdf'
                ):
                    return jsonify({'error': 'El comprobante debe ser una imagen o PDF'}), 415
                comprobante.stream.seek(0, 2)
                tamano = comprobante.stream.tell()
                comprobante.stream.seek(0)
                if tamano > 8 * 1024 * 1024:
                    return jsonify({'error': 'El comprobante no puede superar 8 MB'}), 413
                res = cloudinary_upload(comprobante, folder='pagos_proveedores', max_width=1600)
                url_comprobante = res.get('secure_url')
        else:
            data = request.get_json(silent=True) or {}
            url_comprobante = data.get('comprobante_url')

        if url_comprobante and not re.match(r'^https?://', str(url_comprobante), re.I):
            return jsonify({'error': 'La URL del comprobante no es válida'}), 400

        if not (url_comprobante or fila[5]):
            return jsonify({'error': 'El comprobante de pago o recojo es obligatorio.'}), 400
            
        actor = get_usuario_actual()
        cursor.execute("""
            UPDATE logistica_externa
            SET estado_distribucion = 'Recogido',
                fecha_recojo_fisico = NOW(),
                recogido_por_id = %s,
                url_comprobante_pago = COALESCE(%s, url_comprobante_pago)
            WHERE id = %s
        """, (int(actor['id']), url_comprobante, id))
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Recojo confirmado. Tela en taller.'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print(f'Error al confirmar recojo de tela: {e}')
        return jsonify({'error': 'No se pudo confirmar el recojo de la tela'}), 500
    finally:
        if cursor: cursor.close()
        if conexion: release_db_connection(conexion)

@produccion_bp.route('/api/logistica/<int:id>/confirmar-distribucion', methods=['POST'])
@requiere_login
def logistica_confirmar_distribucion(id):
    conexion = None
    cursor = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT venta_id, item_id, item_ids_extra, operario_id,
                   COALESCE(categoria_insumo, 'OTRO'),
                   COALESCE(estado_distribucion, 'En espera'),
                   COALESCE(tipo_gestion, 'Externo'), estado
            FROM logistica_externa WHERE id = %s FOR UPDATE
        """, (id,))
        fila = cursor.fetchone()
        if not fila:
            return jsonify({'error': 'No encontrado'}), 404
        if not usuario_puede_operar('CORTE_Y_CONTROL_TELAS', fila[3]):
            return jsonify({'error': 'Solo el operario asignado, Telas o jefatura puede distribuir.'}), 403
        if fila[4] != 'TELA':
            return jsonify({'error': 'Este control de distribucion es exclusivo para telas.'}), 400
        estado_distribucion = normalizar_estado_distribucion(fila[5])
        if estado_distribucion == 'Distribuido':
            return jsonify({'exito': True, 'desbloqueados': 0, 'mensaje': 'La tela ya estaba distribuida.'}), 200
        if fila[7] in ('Cancelado', 'Rechazado'):
            return jsonify({'error': 'No se puede distribuir una tela cancelada.'}), 409
        if not tela_puede_distribuirse(fila[6], estado_distribucion):
            return jsonify({'error': 'Primero confirma el recojo fisico de la tela.'}), 409

        actor = get_usuario_actual()
        cursor.execute("""
            UPDATE logistica_externa
            SET estado = 'Recibido',
                estado_distribucion = 'Distribuido',
                distribuido_por_id = %s,
                fecha_distribucion = NOW()
            WHERE id = %s
        """, (int(actor['id']), id))

        venta_id = fila[0]
        item_ids = _items_asociados_logistica(cursor, id, fila[1], fila[2])
        desbloqueados = _intentar_desbloquear_items(cursor, item_ids)
        _activar_despacho_si_listo(cursor, venta_id)

        conexion.commit()
        return jsonify({'exito': True, 'desbloqueados': desbloqueados}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        print(f'Error al distribuir tela: {e}')
        return jsonify({'error': 'No se pudo confirmar la distribución de la tela'}), 500
    finally:
        if cursor: cursor.close()
        if conexion: release_db_connection(conexion)


@produccion_bp.route('/api/logistica/telas/confirmar-recojo-lote', methods=['POST'])
@requiere_login
def confirmar_recojo_telas_lote():
    conexion = None
    cursor = None
    try:
        ids_recibidos = json.loads(request.form.get('ids') or '[]')
        if not isinstance(ids_recibidos, list):
            raise ValueError('El campo ids debe ser una lista.')
        ids = list(dict.fromkeys(int(i) for i in ids_recibidos))
        if not ids or len(ids) > 50:
            return jsonify({'error': 'Selecciona entre 1 y 50 telas.'}), 400
        if 'comprobante' not in request.files or not request.files['comprobante'].filename:
            return jsonify({'error': 'El comprobante es obligatorio.'}), 400

        conexion = get_db_connection()
        cursor = conexion.cursor()
        ids_actualizar = []
        for logistica_id in ids:
            cursor.execute("""
                SELECT id, operario_id, COALESCE(categoria_insumo, 'OTRO'),
                       COALESCE(estado_distribucion, 'En espera'), estado,
                       COALESCE(tipo_gestion, 'Externo')
                FROM logistica_externa WHERE id = %s FOR UPDATE
            """, (logistica_id,))
            fila = cursor.fetchone()
            if not fila:
                raise ValueError(f'La linea {logistica_id} ya no existe.')
            if not usuario_puede_operar('CORTE_Y_CONTROL_TELAS', fila[1]):
                return jsonify({'error': f'No tienes permiso sobre la linea {logistica_id}.'}), 403
            estado_distribucion = normalizar_estado_distribucion(fila[3])
            if fila[2] != 'TELA' or estado_distribucion == 'Distribuido' or fila[4] in ('Cancelado', 'Rechazado'):
                raise ValueError(f'La linea {logistica_id} no esta disponible para recojo.')
            if not tela_requiere_comprobante(fila[5]):
                raise ValueError(
                    f'La linea {logistica_id} es interna o informal y se distribuye directamente.'
                )
            if estado_distribucion not in ('En espera', 'En Recojo', 'Recogido'):
                raise ValueError(f'La linea {logistica_id} no esta disponible para recojo.')
            if estado_distribucion != 'Recogido':
                ids_actualizar.append(logistica_id)

        if not ids_actualizar:
            conexion.commit()
            return jsonify({'exito': True, 'actualizados': 0}), 200

        subida = cloudinary_upload(
            request.files['comprobante'], folder='pagos_proveedores', max_width=1600
        )
        url = subida.get('secure_url')
        actor = get_usuario_actual()
        cursor.execute("""
            UPDATE logistica_externa
            SET estado_distribucion = 'Recogido',
                fecha_recojo_fisico = NOW(),
                recogido_por_id = %s,
                url_comprobante_pago = COALESCE(%s, url_comprobante_pago)
            WHERE id = ANY(%s)
        """, (int(actor['id']), url, ids_actualizar))
        actualizados = cursor.rowcount
        conexion.commit()
        return jsonify({'exito': True, 'actualizados': actualizados}), 200
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        if conexion:
            conexion.rollback()
        return jsonify({'error': str(exc)}), 400
    except Exception:
        if conexion:
            conexion.rollback()
        return jsonify({'error': 'No se pudo confirmar el recojo del lote.'}), 500
    finally:
        if cursor:
            cursor.close()
        if conexion:
            release_db_connection(conexion)


@produccion_bp.route('/api/logistica/telas/confirmar-distribucion-lote', methods=['POST'])
@requiere_login
def confirmar_distribucion_telas_lote():
    conexion = None
    cursor = None
    try:
        body = request.get_json(silent=True) or {}
        ids_recibidos = body.get('ids') or []
        if not isinstance(ids_recibidos, list):
            raise ValueError('El campo ids debe ser una lista.')
        ids = list(dict.fromkeys(int(i) for i in ids_recibidos))
        if not ids or len(ids) > 50:
            return jsonify({'error': 'Selecciona entre 1 y 50 telas.'}), 400

        conexion = get_db_connection()
        cursor = conexion.cursor()
        item_ids = set()
        venta_ids = set()
        filas_validas = []
        for logistica_id in ids:
            cursor.execute("""
                SELECT id, venta_id, item_id, item_ids_extra, operario_id,
                       COALESCE(categoria_insumo, 'OTRO'),
                       COALESCE(estado_distribucion, 'En espera'),
                       COALESCE(tipo_gestion, 'Externo'), estado
                FROM logistica_externa WHERE id = %s FOR UPDATE
            """, (logistica_id,))
            fila = cursor.fetchone()
            if not fila:
                raise ValueError(f'La linea {logistica_id} ya no existe.')
            if not usuario_puede_operar('CORTE_Y_CONTROL_TELAS', fila[4]):
                return jsonify({'error': f'No tienes permiso sobre la linea {logistica_id}.'}), 403
            if fila[5] != 'TELA':
                raise ValueError(f'La linea {logistica_id} no corresponde a una tela.')
            estado_distribucion = normalizar_estado_distribucion(fila[6])
            if estado_distribucion == 'Distribuido':
                venta_ids.add(fila[1])
                item_ids.update(
                    _items_asociados_logistica(cursor, logistica_id, fila[2], fila[3])
                )
                continue
            if fila[8] in ('Cancelado', 'Rechazado'):
                raise ValueError(f'La linea {logistica_id} esta cancelada.')
            if not tela_puede_distribuirse(fila[7], estado_distribucion):
                raise ValueError(f'La linea {logistica_id} aun no fue recogida.')
            filas_validas.append(logistica_id)
            venta_ids.add(fila[1])
            item_ids.update(_items_asociados_logistica(cursor, logistica_id, fila[2], fila[3]))

        actor = get_usuario_actual()
        if filas_validas:
            cursor.execute("""
                UPDATE logistica_externa
                SET estado = 'Recibido',
                    estado_distribucion = 'Distribuido',
                    distribuido_por_id = %s,
                    fecha_distribucion = NOW()
                WHERE id = ANY(%s)
            """, (int(actor['id']), filas_validas))
        desbloqueados = _intentar_desbloquear_items(cursor, item_ids)
        for venta_id in venta_ids:
            _activar_despacho_si_listo(cursor, venta_id)
        conexion.commit()
        return jsonify({
            'exito': True,
            'actualizados': len(filas_validas),
            'desbloqueados': desbloqueados,
        }), 200
    except (TypeError, ValueError) as exc:
        if conexion:
            conexion.rollback()
        return jsonify({'error': str(exc)}), 400
    except Exception:
        if conexion:
            conexion.rollback()
        return jsonify({'error': 'No se pudo distribuir el lote de telas.'}), 500
    finally:
        if cursor:
            cursor.close()
        if conexion:
            release_db_connection(conexion)

@produccion_bp.route('/api/logistica/<int:id>/asignar-operario', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def asignar_operario_logistica(id):
    data = request.json
    operario_id = data.get('trabajador_id')
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("""
            UPDATE logistica_externa
            SET operario_id = %s,
                estado_distribucion = CASE
                    WHEN estado_distribucion IS NULL OR estado_distribucion NOT IN ('En Recojo','Recogido','Distribuido')
                    THEN 'En Recojo'
                    ELSE estado_distribucion
                END
            WHERE id = %s
        """, (operario_id, id))
        conexion.commit()

        # ── NOTIFICACIÓN: avisar al operario que tiene un recojo asignado ──
        # No tumba la respuesta si falla — la asignación ya quedó guardada.
        try:
            cursor.execute("""
                SELECT u.nombre, u.email, u.telefono,
                       le.insumo_nombre, le.sku, v.codigo_venta, c.nombre AS proveedor
                FROM logistica_externa le
                JOIN usuarios u         ON u.id            = le.operario_id
                LEFT JOIN items_venta i ON le.item_id      = i.id
                LEFT JOIN ventas v      ON i.venta_id      = v.id
                LEFT JOIN proveedores c ON le.proveedor_id = c.id
                WHERE le.id = %s
            """, (id,))
            info = cursor.fetchone()
            if info:
                nombre, email, telefono, insumo, sku, codigo_venta, proveedor = info
                notificar_usuario(
                    destinatario_email=email,
                    nombre_destinatario=nombre,
                    asunto=f"Nuevo recojo asignado — {codigo_venta or 'Logística'}",
                    mensaje=(
                        f"Hola {nombre},\n\n"
                        f"Se te asignó un recojo de material externo:\n\n"
                        f"  Insumo:    {insumo or '—'}\n"
                        f"  SKU:       {sku or '—'}\n"
                        f"  Proveedor: {proveedor or '—'}\n"
                        f"  Pedido:    {codigo_venta or '—'}\n\n"
                        f"Ingresa al ERP para ver los detalles y confirmar el recojo.\n\n"
                        f"Innova Möbili — Taller"
                    ),
                    telefono=telefono,
                )
        except Exception as e_notif:
            print(f"⚠️ No se pudo notificar la asignación logística {id}: {e_notif}")
        # ────────────────────────────────────────────────────────────────────

        return jsonify({'exito': True}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# ==========================================
# 11. RECETAS DE MUEBLES (BOM)
# ==========================================

@produccion_bp.route('/api/recetas/<int:producto_id>', methods=['GET'])
@requiere_login
def obtener_receta(producto_id):
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT r.id, i.nombre_insumo, r.insumo_id, r.cantidad_necesaria, i.cantidad_actual
            FROM recetas_muebles r
            JOIN inventario_insumos i ON r.insumo_id = i.id
            WHERE r.producto_id = %s ORDER BY i.nombre_insumo;
        """, (producto_id,))
        receta = [{
            "id": r[0], "nombre_insumo": r[1], "insumo_id": r[2],
            "cantidad_necesaria": r[3], "stock_actual": r[4]
        } for r in cursor.fetchall()]
        return jsonify(receta), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/recetas/nueva', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def agregar_ingrediente_receta():
    data               = request.json
    producto_id        = data.get('producto_id')
    insumo_id          = data.get('insumo_id')
    cantidad_necesaria = data.get('cantidad_necesaria', 1)
    if not producto_id or not insumo_id:
        return jsonify({'error': 'producto_id e insumo_id son obligatorios'}), 400
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            INSERT INTO recetas_muebles (producto_id, insumo_id, cantidad_necesaria)
            VALUES (%s,%s,%s) RETURNING id;
        """, (producto_id, insumo_id, cantidad_necesaria))
        nuevo_id = cursor.fetchone()[0]
        conexion.commit()
        return jsonify({'exito': True, 'id': nuevo_id}), 201
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

@produccion_bp.route('/api/inventario/etiquetas-disponibles', methods=['POST'])
@requiere_login
def obtener_etiquetas_disponibles():
    datos = request.get_json(silent=True) or {}
    items = datos.get('items', [])
    por_cantidad = datos.get('por_cantidad', False)
    
    etiquetas = []
    omitidos = []
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        for item in items:
            # Productos enteros y piezas comparten sku_maestro. Las piezas se
            # distinguen por su identidad dimensional, no por tener un SKU.
            es_pieza = all(
                campo in item
                for campo in ('forma', 'largo_cm', 'ancho_cm', 'alto_cm')
            )
            
            if es_pieza:
                # Filtrar stock_piezas por sus medidas exactas
                query = """
                    SELECT codigo_barra, (SELECT nombre FROM sedes WHERE id = sede_id), sku_maestro
                    FROM stock_piezas
                    WHERE sku_maestro = %(sku)s AND forma = %(forma)s
                      AND COALESCE(largo_cm, 0) = %(largo)s AND COALESCE(ancho_cm, 0) = %(ancho)s AND COALESCE(alto_cm, 0) = %(alto)s
                      AND estado = 'Disponible'
                    ORDER BY id
                """
                cursor.execute(query, {
                    'sku': item.get('sku_maestro'),
                    'forma': item.get('forma'),
                    'largo': float(item.get('largo_cm') or 0),
                    'ancho': float(item.get('ancho_cm') or 0),
                    'alto': float(item.get('alto_cm') or 0)
                })
            else:
                # Filtrar stock_productos enteros
                if item.get('catalogo_id'):
                    query = """
                        SELECT sp.codigo_barra,
                               (SELECT nombre FROM sedes WHERE id = sp.sede_id),
                               COALESCE(NULLIF(sp.sku_maestro, ''), cp.sku_maestro, '')
                        FROM stock_productos sp
                        LEFT JOIN catalogo_productos cp ON cp.id = sp.catalogo_id
                        WHERE sp.catalogo_id = %(cat_id)s AND sp.estado = 'Disponible'
                        ORDER BY sp.id
                    """
                    cursor.execute(query, {'cat_id': item.get('catalogo_id')})
                else:
                    # Los modelos registrados directamente en Inventario no
                    # siempre están enlazados a catalogo_productos. La vista
                    # los agrupa por nombre + observaciones, así que la etiqueta
                    # debe usar exactamente la misma identidad.
                    query = """
                        SELECT sp.codigo_barra,
                               (SELECT nombre FROM sedes WHERE id = sp.sede_id),
                               COALESCE(sp.sku_maestro, '')
                        FROM stock_productos sp
                        WHERE LOWER(sp.nombre_modelo) = LOWER(%(nombre)s)
                          AND COALESCE(sp.observaciones, '') = %(observaciones)s
                          AND sp.estado = 'Disponible'
                        ORDER BY sp.id
                    """
                    cursor.execute(query, {
                        'nombre': item.get('nombre_modelo') or '',
                        'observaciones': item.get('observaciones') or ''
                    })
            
            filas = cursor.fetchall()
            nombre_etiqueta = item.get('nombreConMedida') or item.get('nombre_modelo')
            if not filas:
                omitidos.append(nombre_etiqueta or item.get('sku_maestro') or 'Modelo sin nombre')
                continue
            
            if por_cantidad:
                # Imprimir todas las unidades físicas disponibles
                for r in filas:
                    etiquetas.append({
                        'codigo': r[0], 'nombre': nombre_etiqueta,
                        'sede': r[1] or 'Tienda',
                        'sku_maestro': r[2] if len(r) > 2 else (item.get('sku_maestro') or '')
                    })
            else:
                # Imprimir solo 1 por modelo (la primera unidad física que encuentre)
                etiquetas.append({
                    'codigo': filas[0][0], 'nombre': nombre_etiqueta,
                    'sede': filas[0][1] or 'Tienda',
                    'sku_maestro': filas[0][2] if len(filas[0]) > 2 else (item.get('sku_maestro') or '')
                })
                    
        return jsonify({'exito': True, 'etiquetas': etiquetas, 'omitidos': omitidos}), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close()
            release_db_connection(conexion)


# ==========================================
# 12. SUGERENCIAS DE INSUMOS
# ==========================================

@produccion_bp.route('/api/sugerencias', methods=['POST'])
@requiere_login
def guardar_sugerencia():
    try:
        nombre     = request.form.get('nombre')
        tipo       = request.form.get('tipo')
        usuario_id = request.form.get('usuario_id')
        datos_json = request.form.get('datos_json')
        if not nombre or not tipo:
            return jsonify({'error': 'El nombre y tipo de insumo son obligatorios'}), 400
        foto_ruta = request.form.get('foto_ref') or "imagenes/sin_foto.jpg"
        if 'foto' in request.files and request.files['foto'].filename != '':
            respuesta_nube = cloudinary_upload(request.files['foto'], folder="sugerencias")
            foto_ruta = respuesta_nube.get('secure_url')
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            INSERT INTO sugerencias_insumos (nombre, tipo, foto_ref, usuario_id, datos_json, estado)
            VALUES (%s, %s, %s, %s, %s, 'Pendiente') RETURNING id;
        """, (nombre, tipo, foto_ruta, usuario_id, datos_json))
        sugerencia_id = cursor.fetchone()[0]
        conexion.commit()
        return jsonify({'exito': True, 'id': sugerencia_id, 'mensaje': 'Sugerencia enviada al Gestor de Aprobación'}), 201
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/sugerencias', methods=['GET'])
@requiere_login
def obtener_sugerencias():
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT s.id, s.nombre, s.tipo, s.foto_ref,
                   COALESCE(u.nombre, 'Vendedor') AS vendedor, s.datos_json, s.estado
            FROM sugerencias_insumos s
            LEFT JOIN usuarios u ON s.usuario_id = u.id
            WHERE s.estado = 'Pendiente'
            ORDER BY s.fecha_registro DESC;
        """)
        resultado = [{
            "id": r[0], "nombre": r[1], "tipo": r[2], "foto_url": r[3],
            "vendedor": r[4], "datos_json": r[5], "estado": r[6]
        } for r in cursor.fetchall()]
        return jsonify(resultado), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/sugerencias/aprobar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def aprobar_sugerencia_insumo():
    data          = request.json
    sugerencia_id = data.get('sugerencia_id')
    origen        = data.get('origen', 'Externo')
    campo1        = data.get('campo1', '')
    campo2        = data.get('campo2', '')
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT tipo, foto_ref, datos_json FROM sugerencias_insumos WHERE id = %s;", (sugerencia_id,))
        sug = cursor.fetchone()
        if not sug:
            return jsonify({'error': 'Sugerencia no encontrada'}), 404
        tipo_material, foto_ruta, datos_raw = sug
        datos = json.loads(datos_raw) if datos_raw else {}
        nuevo_sku = ""

        def val(key, fallback):
            return datos.get(key) if datos.get(key) else fallback

        inserts = {
            'tela':        ("maestro_telas", "TEL", "(sku, proveedor, coleccion, color, foto_url, origen_produccion, estado, proveedor_id) VALUES (%s,%s,%s,%s,%s,%s,'Disponible',%s)", lambda d: (d.get('proveedor'), val('coleccion', campo1), val('color', campo2), foto_ruta, origen, d.get('proveedor_id'))),
            'cojin':       ("maestro_disenos_cojin", "COJ", "(sku, nombre_diseno, tipo_tela, foto_url, origen_produccion, estado) VALUES (%s,%s,%s,%s,%s,'Disponible')", lambda d: (val('nombre_diseno', campo1), val('tipo_tela', campo2), foto_ruta, origen)),
            'base':        ("maestro_bases", "BAS", "(sku, tipo, material, modelo, color, medida_altura, foto_url, origen_produccion, estado) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'Disponible')", lambda d: (d.get('tipo', 'Patas'), val('material', campo2), val('modelo', campo1), d.get('color'), d.get('medida_altura'), foto_ruta, origen)),
            'tablero':     ("maestro_tableros", "TAB", "(sku, material_base, nombre_modelo, color_veta, acabado, foto_url, origen_produccion, estado) VALUES (%s,%s,%s,%s,%s,%s,%s,'Disponible')", lambda d: (val('material_base', campo2), val('nombre_modelo', campo1), d.get('color_veta'), d.get('acabado'), foto_ruta, origen)),
            'base-comedor':("maestro_bases_comedor", "BAC", "(sku, material, modelo, color, foto_url, origen_produccion, estado) VALUES (%s,%s,%s,%s,%s,%s,'Disponible')", lambda d: (val('material', campo2), val('modelo', campo1), d.get('color'), foto_ruta, origen)),
            'silla':       ("maestro_sillas", "SIL", "(sku, material, modelo, color_estructura, foto_url, origen_produccion, estado) VALUES (%s,%s,%s,%s,%s,%s,'Disponible')", lambda d: (val('material', campo2), val('modelo', campo1), d.get('color_estructura'), foto_ruta, origen)),
            'butaca':      ("maestro_butacas", "BUT", "(sku, material, modelo, color_estructura, foto_url, origen_produccion, estado) VALUES (%s,%s,%s,%s,%s,%s,'Disponible')", lambda d: (val('material', campo2), val('modelo', campo1), d.get('color_estructura'), foto_ruta, origen)),
        }
        if tipo_material not in inserts:
            return jsonify({'error': 'Tipo de material no reconocido'}), 400
        tabla, prefijo, cols_sql, params_fn = inserts[tipo_material]
        cursor.execute(f"SELECT COALESCE(MAX(id), 0) FROM {tabla}")
        nuevo_sku = f"{prefijo}-{str(cursor.fetchone()[0]+1).zfill(4)}"
        params = (nuevo_sku,) + params_fn(datos)
        cursor.execute(f"INSERT INTO {tabla} {cols_sql}", params)
        cursor.execute("UPDATE sugerencias_insumos SET estado = 'Aprobado' WHERE id = %s;", (sugerencia_id,))
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': f'Insumo oficializado con éxito. Nuevo SKU: {nuevo_sku}'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)

@produccion_bp.route('/api/sugerencias/rechazar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def rechazar_sugerencia_insumo():
    data          = request.json
    sugerencia_id = data.get('sugerencia_id')
    motivo        = (data.get('motivo') or '').strip()
    if not sugerencia_id:
        return jsonify({'error': 'sugerencia_id es obligatorio'}), 400
    if not motivo:
        return jsonify({'error': 'El motivo de rechazo es obligatorio'}), 400
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            UPDATE sugerencias_insumos
            SET estado = 'Rechazado', motivo_rechazo = %s
            WHERE id = %s;
        """, (motivo, sugerencia_id))
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Sugerencia de insumo rechazada.'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)


# ==========================================
# 13. DESPACHO — ASIGNAR CHOFER Y PROGRESO
# ==========================================

def _validar_contrato_listo_para_despacho(cursor, ticket_id):
    """Valida que todo el contrato esté completo antes de activar despacho."""
    cursor.execute("""
        SELECT t.id, t.item_id, t.estado_ticket, i.venta_id
        FROM tickets_produccion t
        JOIN items_venta i ON t.item_id = i.id
        WHERE t.id = %s AND t.area_trabajo = 'DESPACHO_CENTRAL'
    """, (ticket_id,))
    ticket = cursor.fetchone()
    if not ticket:
        return None, {'error': 'Ticket de despacho no encontrado'}, 404
    if ticket[2] != 'Pendiente':
        return ticket, {
            'error': f'El despacho no esta disponible para asignacion (estado: {ticket[2]}).'
        }, 409

    venta_id = ticket[3]
    cursor.execute("""
        SELECT COUNT(*)
        FROM tickets_produccion t
        JOIN items_venta i ON t.item_id = i.id
        WHERE i.venta_id = %s
          AND t.area_trabajo != 'DESPACHO_CENTRAL'
          AND t.estado_ticket NOT IN ('Terminado', 'Recogido', 'Cancelado')
    """, (venta_id,))
    pendientes_taller = cursor.fetchone()[0]
    if pendientes_taller > 0:
        return ticket, {
            'error': f'Aún hay {pendientes_taller} área(s) del contrato sin terminar. No se puede despachar.'
        }, 409

    cursor.execute("""
        SELECT COUNT(*)
        FROM logistica_externa
        WHERE venta_id = %s
          AND estado NOT IN ('Recibido', 'Cancelado', 'Rechazado')
    """, (venta_id,))
    pendientes_logistica = cursor.fetchone()[0]
    if pendientes_logistica > 0:
        return ticket, {
            'error': f'Aún hay {pendientes_logistica} insumo(s) de logística externa del contrato sin recibir. '
                     f'No se puede despachar.'
        }, 409

    cursor.execute("""
        SELECT COUNT(*)
        FROM logistica_externa
        WHERE venta_id = %s
          AND (categoria_insumo = 'TELA'
               OR LOWER(COALESCE(insumo_nombre, '')) LIKE '%%tela%%'
               OR LOWER(COALESCE(unidad, '')) IN ('mts', 'metro', 'metros'))
          AND COALESCE(estado_distribucion, '') NOT IN ('Distribuido')
          AND estado NOT IN ('Cancelado', 'Rechazado')
    """, (venta_id,))
    pendientes_telas = cursor.fetchone()[0]
    if pendientes_telas > 0:
        return ticket, {
            'error': f'Aún hay {pendientes_telas} tela(s) sin distribuir al área correspondiente. '
                     f'Márcalas como Distribuido desde la bandeja de Telas antes de despachar.'
        }, 409

    return ticket, None, None

@produccion_bp.route('/api/despacho/asignar-chofer', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def asignar_chofer_despacho():
    data      = request.json
    ticket_id = data.get('ticket_id')
    chofer_id = data.get('chofer_id')
    if not ticket_id or not chofer_id:
        return jsonify({'error': 'ticket_id y chofer_id son obligatorios'}), 400
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _, error_payload, status = _validar_contrato_listo_para_despacho(cursor, ticket_id)
        if error_payload:
            return jsonify(error_payload), status

        cursor.execute("""
            SELECT id
            FROM usuarios
            WHERE id = %s AND rol = 'Chofer' AND COALESCE(estado, TRUE) = TRUE
            FOR SHARE
        """, (chofer_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'El chofer seleccionado no existe o no esta activo.'}), 400

        cursor.execute("""
            UPDATE tickets_produccion
            SET trabajador_asignado_id = %s, estado_ticket = 'En Proceso', fecha_inicio = CURRENT_TIMESTAMP
            WHERE id = %s
              AND estado_ticket = 'Pendiente'
              AND trabajador_asignado_id IS NULL
        """, (chofer_id, ticket_id))
        if cursor.rowcount != 1:
            conexion.rollback()
            return jsonify({'error': 'El despacho ya fue asignado. Actualiza la cola.'}), 409
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Chofer asignado. El despacho está activo.'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/pendientes-por-proveedor', methods=['GET'])
@requiere_login
def logistica_pendientes_por_proveedor():
    """Agrupa filas pendientes de logística externa por proveedor."""
    conexion = None
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()

        # Detectar qué tablas maestro_* existen realmente en la BD
        cursor.execute("""
            SELECT tablename FROM pg_tables
            WHERE schemaname = 'public'
              AND tablename LIKE 'maestro_%'
              AND tablename IN (
                  'maestro_telas','maestro_bases','maestro_tableros',
                  'maestro_bases_comedor','maestro_sillas',
                  'maestro_butacas','maestro_disenos_cojin'
              );
        """)
        tablas_existentes = [r[0] for r in cursor.fetchall()]

        if tablas_existentes:
            subqueries = " ,\n                    ".join(
                f"(SELECT foto_url FROM {t} WHERE sku = l.sku LIMIT 1)"
                for t in tablas_existentes
            )
            foto_expr = f"COALESCE(\n                    {subqueries}\n                )"
        else:
            foto_expr = "NULL"

        # Detectar si las columnas cantidad/unidad ya existen (migración puede no haberse corrido)
        cursor.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'logistica_externa'
              AND column_name IN ('cantidad', 'unidad');
        """)
        cols_existentes = {r[0] for r in cursor.fetchall()}
        cantidad_expr = "l.cantidad" if 'cantidad' in cols_existentes else "NULL"
        unidad_expr   = "l.unidad"   if 'unidad'   in cols_existentes else "NULL"

        cursor.execute(f"""
            SELECT
                p.id AS proveedor_id,
                p.nombre AS proveedor_nombre,
                p.telefono,
                l.id, l.insumo_nombre, l.sku,
                {cantidad_expr} AS cantidad,
                {unidad_expr}   AS unidad,
                {foto_expr}     AS foto_url,
                v.codigo_venta
            FROM logistica_externa l
            JOIN ventas v ON l.venta_id = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            WHERE l.estado IN ('POR_PEDIR', 'Pendiente', 'Por Pedir')
              AND l.proveedor_id IS NOT NULL
            ORDER BY p.id, l.id
        """)
        rows = cursor.fetchall()

        grupos = {}
        for r in rows:
            pid = r[0]
            if pid not in grupos:
                grupos[pid] = {
                    'proveedor_id':     pid,
                    'proveedor_nombre': r[1],
                    'telefono':         r[2],
                    'items': []
                }
            grupos[pid]['items'].append({
                'logistica_id':  r[3],
                'insumo_nombre': r[4],
                'sku':           r[5] or '',
                'cantidad':      float(r[6]) if r[6] else None,
                'unidad':        r[7] or '',
                'foto_url':      r[8] or '',
                'codigo_venta':  r[9],
            })
        return jsonify(list(grupos.values())), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/crear-lote-cotizacion', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def crear_lote_cotizacion():
    conexion = None
    cursor = None
    try:
        datos = request.get_json(silent=True) or {}
        try:
            proveedor_id = int(datos.get('proveedor_id'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Debes seleccionar un proveedor válido'}), 400

        items = datos.get('items')
        if not isinstance(items, list) or not 1 <= len(items) <= 50:
            return jsonify({'error': 'Selecciona entre 1 y 50 materiales'}), 400

        ids_logistica = []
        for item in items:
            if not isinstance(item, dict):
                return jsonify({'error': 'La lista de materiales no es válida'}), 400
            try:
                logistica_id = int(item.get('logistica_id'))
            except (TypeError, ValueError):
                return jsonify({'error': 'Uno de los materiales no tiene un identificador válido'}), 400
            if logistica_id <= 0 or logistica_id in ids_logistica:
                return jsonify({'error': 'La selección contiene materiales inválidos o repetidos'}), 400
            ids_logistica.append(logistica_id)

        conexion = get_db_connection()
        cursor = conexion.cursor()

        cursor.execute("SELECT nombre, telefono FROM proveedores WHERE id = %s", (proveedor_id,))
        prov = cursor.fetchone()
        if not prov:
            return jsonify({'error': 'Proveedor no encontrado'}), 404
        nombre_proveedor, telefono = prov
        if not telefono:
            return jsonify({'error': 'El proveedor no tiene teléfono/WhatsApp registrado'}), 400

        cursor.execute("""
            SELECT tablename
            FROM pg_tables
            WHERE schemaname = 'public'
              AND tablename IN (
                  'maestro_telas', 'maestro_bases', 'maestro_tableros',
                  'maestro_bases_comedor', 'maestro_sillas',
                  'maestro_butacas', 'maestro_disenos_cojin'
              )
        """)
        tablas_foto = [r[0] for r in cursor.fetchall()]
        if tablas_foto:
            candidatos_foto = ', '.join(
                f"(SELECT foto_url FROM {tabla} WHERE sku = l.sku LIMIT 1)"
                for tabla in tablas_foto
            )
            foto_expr = f"COALESCE({candidatos_foto})"
        else:
            foto_expr = "NULL"

        cursor.execute(f"""
            SELECT l.id, l.sku, l.insumo_nombre, l.cantidad, l.unidad,
                   {foto_expr} AS foto_url
            FROM logistica_externa l
            WHERE l.id = ANY(%s)
              AND l.proveedor_id = %s
              AND l.estado IN ('POR_PEDIR', 'Pendiente', 'Por Pedir')
            FOR UPDATE OF l
        """, (ids_logistica, proveedor_id))
        filas_por_id = {r[0]: r for r in cursor.fetchall()}
        faltantes = [item_id for item_id in ids_logistica if item_id not in filas_por_id]
        if faltantes:
            return jsonify({
                'error': (
                    'Uno o más materiales ya cambiaron de estado, no pertenecen '
                    'al proveedor o dejaron de estar disponibles. Actualiza la bandeja.'
                )
            }), 409

        items_autorizados = []
        for item_id in ids_logistica:
            fila = filas_por_id[item_id]
            foto_url = limpiar_foto(fila[5]) if fila[5] else ''
            items_autorizados.append({
                'logistica_id': fila[0],
                'sku': fila[1] or '',
                'insumo_nombre': fila[2] or 'Material sin nombre',
                'cantidad': float(fila[3]) if fila[3] is not None else None,
                'unidad': fila[4] or '',
                'foto_url': foto_url,
            })

        token = uuid.uuid4().hex

        cursor.execute("""
            INSERT INTO cotizaciones_lote (token, proveedor_id, estado, fecha_envio)
            VALUES (%s, %s, 'Pendiente', NOW()) RETURNING id
        """, (token, proveedor_id))
        lote_id = cursor.fetchone()[0]

        for item in items_autorizados:
            cursor.execute("""
                INSERT INTO cotizacion_lote_items
                    (lote_id, logistica_externa_id, sku, insumo_nombre, cantidad, unidad, foto_url)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (lote_id, item.get('logistica_id'), item.get('sku'), item.get('insumo_nombre'),
                  item.get('cantidad'), item.get('unidad'), item.get('foto_url')))

            cursor.execute("""
                UPDATE logistica_externa
                SET estado = 'Cotizacion Enviada', fecha_envio_cotizacion = NOW()
                WHERE id = %s AND estado IN ('POR_PEDIR', 'Pendiente', 'Por Pedir')
            """, (item['logistica_id'],))
            if cursor.rowcount != 1:
                raise RuntimeError('Un material cambió de estado durante la cotización')

        conexion.commit()

        link = f"{request.url_root.rstrip('/')}/cotizar.html?lote={token}"
        return jsonify({
            'exito': True, 'token': token, 'link': link,
            'telefono': telefono, 'nombre_proveedor': nombre_proveedor or 'Proveedor',
            'items': items_autorizados
        }), 200
    except Exception as e:
        if conexion:
            conexion.rollback()
        print(f"Error al crear lote de cotización: {e}")
        return jsonify({'error': 'No se pudo crear el lote de cotización'}), 500
    finally:
        if conexion:
            if cursor:
                cursor.close()
            release_db_connection(conexion)


@produccion_bp.route('/api/cotizar-lote/<token>', methods=['GET', 'POST'])
def cotizar_lote(token):
    if not re.fullmatch(r'[a-f0-9]{32}', token or ''):
        return jsonify({'error': 'Token inválido'}), 404

    conexion = None
    cursor = None
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()

        bloqueo = ' FOR UPDATE' if request.method == 'POST' else ''
        cursor.execute(f"""
            SELECT id, proveedor_id, token_usado, estado, fecha_envio
            FROM cotizaciones_lote
            WHERE token = %s{bloqueo}
        """, (token,))
        lote = cursor.fetchone()
        if not lote:
            return jsonify({'error': 'Token inválido'}), 404
        lote_id, proveedor_id, token_usado, estado, fecha_envio = lote

        if token_usado:
            if request.method == 'GET':
                return jsonify({'ya_respondido': True}), 200
            return jsonify({'error': 'Esta cotización ya fue enviada'}), 409
        if not cotizacion_esta_vigente(
            fecha_envio, ahora=datetime.now(ZONA_HORARIA_LIMA)
        ):
            return jsonify({'error': 'Este enlace de cotización expiró'}), 410

        if request.method == 'GET':
            cursor.execute("SELECT nombre FROM proveedores WHERE id = %s", (proveedor_id,))
            prov = cursor.fetchone()
            cursor.execute("""
                SELECT id, sku, insumo_nombre, cantidad, unidad, foto_url
                FROM cotizacion_lote_items
                WHERE lote_id = %s
                ORDER BY id
            """, (lote_id,))
            items = [{
                'id': r[0],
                'sku': r[1] or '',
                'insumo_nombre': r[2] or 'Material sin nombre',
                'cantidad': float(r[3]) if r[3] is not None else None,
                'unidad': r[4] or '',
                'foto_url': limpiar_foto(r[5]) if r[5] else '',
            } for r in cursor.fetchall()]
            if not items:
                return jsonify({'error': 'Esta cotización no contiene materiales'}), 409
            return jsonify({'proveedor': prov[0] if prov else '', 'items': items}), 200

        body = request.get_json(silent=True) or {}
        respuestas = body.get('respuestas')
        if not isinstance(respuestas, list) or not respuestas:
            return jsonify({'error': 'Debes cotizar todos los materiales del lote'}), 400

        cursor.execute("""
            SELECT id, logistica_externa_id
            FROM cotizacion_lote_items
            WHERE lote_id = %s
            ORDER BY id
            FOR UPDATE
        """, (lote_id,))
        filas_items = cursor.fetchall()
        ids_esperados = {r[0] for r in filas_items}
        respuestas_validadas = {}
        try:
            for respuesta in respuestas:
                if not isinstance(respuesta, dict):
                    raise ValueError('Una respuesta del lote no es válida')
                item_id = int(respuesta.get('item_id'))
                if item_id in respuestas_validadas:
                    raise ValueError('Hay materiales repetidos en la respuesta')
                notas = respuesta.get('notas') or ''
                if not isinstance(notas, str) or len(notas) > 2000:
                    raise ValueError('Las observaciones no pueden superar 2000 caracteres')
                respuestas_validadas[item_id] = {
                    'precio': normalizar_precio_cotizacion(respuesta.get('precio')),
                    'fecha': normalizar_fecha_cotizacion(
                        respuesta.get('fecha_entrega'),
                        fecha_minima=datetime.now(ZONA_HORARIA_LIMA).date(),
                    ),
                    'notas': notas.strip(),
                }
        except (TypeError, ValueError) as error:
            return jsonify({'error': str(error)}), 400

        if set(respuestas_validadas) != ids_esperados:
            return jsonify({'error': 'Debes responder una vez cada material del lote'}), 400

        for item_id, logistica_id in filas_items:
            respuesta = respuestas_validadas[item_id]
            cursor.execute("""
                UPDATE cotizacion_lote_items
                SET precio_cotizado = %s,
                    fecha_entrega_proveedor = %s,
                    notas_item = %s,
                    respondido = TRUE
                WHERE id = %s AND lote_id = %s
            """, (
                respuesta['precio'], respuesta['fecha'], respuesta['notas'],
                item_id, lote_id,
            ))
            cursor.execute("""
                UPDATE logistica_externa
                SET precio_cotizado = %s,
                    fecha_entrega_proveedor = %s,
                    notas_proveedor = %s,
                    estado = 'Cotizado',
                    token_usado = TRUE,
                    fecha_respuesta_proveedor = NOW()
                WHERE id = %s AND estado = 'Cotizacion Enviada'
            """, (
                respuesta['precio'], respuesta['fecha'], respuesta['notas'],
                logistica_id,
            ))
            if cursor.rowcount != 1:
                conexion.rollback()
                return jsonify({
                    'error': 'Uno de los materiales cambió de estado. Actualiza la cotización con Innova.'
                }), 409

        cursor.execute("""
            UPDATE cotizaciones_lote
            SET token_usado = TRUE, estado = 'Respondido', fecha_respuesta = NOW()
            WHERE id = %s AND token_usado = FALSE
        """, (lote_id,))
        if cursor.rowcount != 1:
            conexion.rollback()
            return jsonify({'error': 'Esta cotización ya fue enviada'}), 409
        conexion.commit()
        return jsonify({'exito': True}), 200
    except Exception as e:
        if conexion:
            conexion.rollback()
        print(f"Error en cotización por lote: {e}")
        return jsonify({'error': 'No se pudo procesar la cotización'}), 500
    finally:
        if conexion:
            if cursor:
                cursor.close()
            release_db_connection(conexion)


@produccion_bp.route('/api/despacho/progreso/<int:item_id>', methods=['GET'])
@requiere_login
def progreso_despacho(item_id):
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT area_trabajo, estado_ticket, trabajador_asignado_id,
                   COALESCE(u.nombre, 'Sin asignar') AS trabajador_nombre
            FROM tickets_produccion t
            LEFT JOIN usuarios u ON t.trabajador_asignado_id = u.id
            WHERE t.item_id = %s AND t.area_trabajo != 'DESPACHO_CENTRAL'
            ORDER BY t.etapa ASC, t.id ASC
        """, (item_id,))
        partes = [{"area": r[0], "estado": r[1], "trabajador": r[3]} for r in cursor.fetchall()]
        total      = len(partes)
        terminados = sum(
            1 for p in partes
            if p['estado'] in ('Terminado', 'Recogido', 'Cancelado')
        )
        return jsonify({"partes": partes, "total": total, "terminados": terminados, "listo": total > 0 and terminados == total}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/despacho/ficha-chofer/<int:item_id>', methods=['GET'])
@requiere_login
def ficha_chofer(item_id):
    """
    Devuelve todo lo que el chofer necesita ver antes de hacer la entrega:
    - Datos del cliente (nombre, dirección, teléfono)
    - Financiero (total, adelanto, saldo)
    - Producto y especificaciones
    - Fotos de evidencia de TODAS las áreas de producción terminadas
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Datos de la venta y del ítem
        cursor.execute("""
            SELECT
                i.producto,
                COALESCE(
                    (SELECT ticket_details_override FROM tickets_produccion
                     WHERE item_id = i.id AND area_trabajo = 'DESPACHO_CENTRAL' LIMIT 1),
                    i.color_tela, i.detalles, ''
                ) AS especificaciones,
                i.foto_url,
                v.nombre_cliente,
                COALESCE(v.celular_cliente, '') AS telefono,
                COALESCE(v.direccion_cliente, '') AS direccion,
                COALESCE(v.monto_total, 0)    AS total,
                COALESCE(v.monto_adelanto, 0) AS adelanto,
                v.codigo_venta,
                v.sede,
                TO_CHAR(v.fecha_entrega, 'DD/MM/YYYY') AS fecha_entrega
            FROM items_venta i
            JOIN ventas v ON i.venta_id = v.id
            WHERE i.id = %s;
        """, (item_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Ítem no encontrado'}), 404

        total    = float(row[6])
        adelanto = float(row[7])

        # Fotos de evidencia de todas las áreas terminadas (excepto DESPACHO_CENTRAL)
        cursor.execute("""
            SELECT
                t.area_trabajo,
                t.foto_evidencia,
                COALESCE(u.nombre, 'Sin asignar') AS operario,
                TO_CHAR(t.fecha_fin, 'DD/MM HH24:MI') AS fecha_fin
            FROM tickets_produccion t
            LEFT JOIN usuarios u ON t.trabajador_asignado_id = u.id
            WHERE t.item_id = %s
              AND t.area_trabajo != 'DESPACHO_CENTRAL'
              AND t.estado_ticket IN ('Terminado', 'Listo para Recojo', 'Recogido')
              AND t.foto_evidencia IS NOT NULL
            ORDER BY t.fecha_fin ASC;
        """, (item_id,))

        NOMBRES_AREA = {
            'CORTE_Y_CONTROL_TELAS':    'Telas',
            'TAPICERIA_SOFAS':          'Tapicería Sofás',
            'TAPICERIA_SILLAS':         'Tapicería Sillas',
            'ESTRUCTURAS_MUEBLES':      'Estructura de Sofa',
            'ESTRUCTURAS_SILLAS':       'Estructura de Sillas',
            'ARMADO_COJINES':           'Cojines',
            'PREPARACION_PATAS_ZOCALO': 'Patas y Zócalo',
            'TABLEROS_Y_PIEDRAS':       'Tableros',
        }

        evidencias = []
        for r in cursor.fetchall():
            evidencias.append({
                'area':       NOMBRES_AREA.get(r[0], r[0].replace('_', ' ').title()),
                'foto':       r[1],
                'operario':   r[2],
                'fecha_fin':  r[3] or '',
            })

        return jsonify({
            'producto':        row[0],
            'especificaciones': row[1],
            'foto_producto':   limpiar_foto(row[2]),
            'cliente':         row[3],
            'telefono':        row[4],
            'direccion':       row[5],
            'total':           total,
            'adelanto':        adelanto,
            'saldo':           max(0, total - adelanto),
            'codigo_venta':    row[8],
            'sede':            row[9] or '',
            'fecha_entrega':   row[10] or 'Por coordinar',
            'evidencias':      evidencias,
        }), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# ==========================================
# DESPACHO — HISTORIAL DE ENTREGADOS
# ==========================================

@produccion_bp.route('/api/despacho/entregados/filtros', methods=['GET'])
@requiere_login
def despacho_entregados_filtros():
    """
    Devuelve las sedes y choferes distintos que aparecen en el historial de
    entregados, para llenar los <select> del frontend.

    Antes esto se sacaba de _entTodos (el array completo ya cargado en
    memoria). Con paginación server-side ya no tenemos todo cargado, así
    que este endpoint chico cubre solo lo necesario para los filtros —
    no trae los datos completos de cada entrega.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT DISTINCT v.sede
            FROM tickets_produccion t
            JOIN items_venta i ON t.item_id  = i.id
            JOIN ventas v      ON i.venta_id = v.id
            WHERE t.area_trabajo = 'DESPACHO_CENTRAL' AND t.estado_ticket = 'Terminado'
              AND v.sede IS NOT NULL AND v.sede != ''
            ORDER BY v.sede;
        """)
        sedes = [r[0] for r in cursor.fetchall()]

        cursor.execute("""
            SELECT DISTINCT COALESCE(u.nombre, 'Sin asignar')
            FROM tickets_produccion t
            LEFT JOIN usuarios u ON t.trabajador_asignado_id = u.id
            WHERE t.area_trabajo = 'DESPACHO_CENTRAL' AND t.estado_ticket = 'Terminado'
            ORDER BY 1;
        """)
        choferes = [r[0] for r in cursor.fetchall()]

        return jsonify({'sedes': sedes, 'choferes': choferes}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/despacho/entregados', methods=['GET'])
@requiere_login
def despacho_entregados():
    """
    Devuelve los tickets DESPACHO_CENTRAL ya Terminados (= entregados), paginados.

    Antes tenía un LIMIT 200 fijo sin paginación real (un parche para que no
    se descontrolara el tamaño de la respuesta, pero pasadas las 200 entregas
    las más viejas simplemente dejaban de aparecer). Julio 2026: paginación
    server-side de verdad + los filtros que antes se hacían en memoria
    (_entTodos en busqueda_filtros.js) ahora se hacen en el query.

    Query params opcionales:
        page      (default 1)
        per_page  (default 20, tope 100)
        chofer_id filtra por chofer asignado (uso: historial del chofer)
        q         texto libre: busca en código de venta, cliente o producto
        sede      filtra por sede exacta
        chofer    filtra por nombre de chofer exacto (uso: vista Admin)
    """
    from database import paginar

    chofer_id     = request.args.get('chofer_id')
    page          = request.args.get('page', 1)
    per_page      = request.args.get('per_page', 20)
    q             = (request.args.get('q') or '').strip()
    sede_filtro   = (request.args.get('sede') or '').strip()
    chofer_nombre = (request.args.get('chofer') or '').strip()

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        condiciones = ["t.area_trabajo = 'DESPACHO_CENTRAL'", "t.estado_ticket = 'Terminado'"]
        params = []

        if chofer_id:
            condiciones.append("t.trabajador_asignado_id = %s")
            params.append(int(chofer_id))

        if q:
            condiciones.append("(v.codigo_venta ILIKE %s OR v.nombre_cliente ILIKE %s OR i.producto ILIKE %s)")
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

        if sede_filtro:
            condiciones.append("v.sede = %s")
            params.append(sede_filtro)

        if chofer_nombre:
            condiciones.append("COALESCE(u.nombre, 'Sin asignar') = %s")
            params.append(chofer_nombre)

        query = f"""
            SELECT
                t.id, i.producto, v.codigo_venta, v.nombre_cliente,
                COALESCE(u.nombre, 'Sin asignar') AS chofer,
                t.fecha_fin,
                COALESCE(i.foto_url, '') AS foto_url,
                COALESCE(t.ticket_details_override, i.color_tela, '') AS especificaciones,
                t.foto_evidencia,
                v.direccion_cliente, v.fecha_entrega,
                t.item_id,
                COALESCE(v.monto_total, 0)    AS total,
                COALESCE(v.monto_adelanto, 0) AS adelanto,
                v.sede
            FROM tickets_produccion t
            JOIN items_venta i    ON t.item_id  = i.id
            JOIN ventas v         ON i.venta_id = v.id
            LEFT JOIN usuarios u  ON t.trabajador_asignado_id = u.id
            WHERE {' AND '.join(condiciones)}
            ORDER BY t.fecha_fin DESC
        """

        filas, total, total_pages = paginar(cursor, query, params, page=page, per_page=per_page)

        resultado = []
        for r in filas:
            total_venta = float(r[12])
            adelanto    = float(r[13])
            resultado.append({
                'ticket_id':      r[0],
                'producto':       r[1],
                'codigo_venta':   r[2],
                'cliente':        r[3],
                'chofer':         r[4],
                'fecha_entrega_real': r[5].strftime('%d/%m/%Y %H:%M') if r[5] else '—',
                'foto_url':       limpiar_foto(r[6]),
                'especificaciones': r[7] or '',
                'foto_evidencia': r[8] if r[8] else '',
                'direccion':      r[9] or '',
                'fecha_entrega_pactada': r[10].strftime('%d/%m/%Y') if r[10] else '—',
                'item_id':        r[11],
                'total':          total_venta,
                'adelanto':       adelanto,
                'saldo':          max(0, total_venta - adelanto),
                'sede':           r[14] or '',
            })

        return jsonify({
            'items':       resultado,
            'total':       total,
            'page':        min(max(1, int(page or 1)), total_pages),
            'total_pages': total_pages
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# ──────────────────────────────────────────────────────────
# LOGÍSTICA EXTERNA — endpoints nuevos
# ──────────────────────────────────────────────────────────

@produccion_bp.route('/api/logistica/<int:id>/enviar-cotizacion', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def enviar_cotizacion_proveedor(id):
    """Genera token y devuelve el link + datos para abrir WhatsApp desde el frontend."""
    conexion = None
    cursor = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT l.insumo_nombre, l.sku, v.codigo_venta,
                   p.nombre, p.telefono,
                   -- busca foto del insumo en los maestros por SKU
                   COALESCE(
                       (SELECT foto_url FROM maestro_telas          WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_bases           WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_tableros        WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_bases_comedor   WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_sillas          WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_butacas         WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM maestro_disenos_cojin   WHERE sku = l.sku LIMIT 1),
                       (SELECT foto_url FROM items_venta             WHERE venta_id = l.venta_id LIMIT 1)
                   ) AS foto_url,
                   l.estado
            FROM logistica_externa l
            JOIN ventas v           ON l.venta_id    = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            WHERE l.id = %s
            FOR UPDATE OF l
        """, (id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Registro no encontrado'}), 404

        insumo, sku, codigo_venta, nombre_proveedor, telefono_proveedor, foto_url, estado = row

        if estado not in ('POR_PEDIR', 'Pendiente', 'Por Pedir', 'Cotizacion Enviada'):
            return jsonify({
                'error': f'No se puede enviar una cotización cuando el material está en estado {estado}'
            }), 409

        if not telefono_proveedor:
            return jsonify({'error': 'El proveedor no tiene teléfono/WhatsApp registrado'}), 400

        token = uuid.uuid4().hex
        link  = f"{request.url_root.rstrip('/')}/cotizar.html?token={token}"
        foto_url_limpia = limpiar_foto(foto_url.split('|')[0]) if foto_url else ''

        cursor.execute("""
            UPDATE logistica_externa
            SET token_respuesta            = %s,
                token_usado                = FALSE,
                fecha_envio_cotizacion     = NOW(),
                estado                     = 'Cotizacion Enviada'
            WHERE id = %s
        """, (token, id))
        conexion.commit()

        return jsonify({
            'exito':             True,
            'link':              link,
            'telefono':          telefono_proveedor,
            'nombre_proveedor':  nombre_proveedor or 'Proveedor',
            'insumo':            insumo,
            'sku':               sku or '',
            'codigo_venta':      codigo_venta,
            'foto_url':          foto_url_limpia,
        }), 200

    except Exception as e:
        if conexion:
            conexion.rollback()
        print(f"Error al enviar cotización: {e}")
        return jsonify({'error': 'No se pudo generar el enlace de cotización'}), 500
    finally:
        if conexion:
            if cursor:
                cursor.close()
            release_db_connection(conexion)


@produccion_bp.route('/api/cotizar/<string:token>', methods=['GET'])
def ver_formulario_cotizacion(token):
    """Endpoint PÚBLICO — el proveedor abre el link desde su celular."""
    if not re.fullmatch(r'[a-f0-9]{32}', token or ''):
        return jsonify({'error': 'Token inválido o expirado'}), 404

    conexion = None
    cursor = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT l.id, l.insumo_nombre, l.sku, l.token_usado,
                   v.codigo_venta, COALESCE(p.nombre,'Sin nombre') AS proveedor,
                   l.fecha_envio_cotizacion, l.estado
            FROM logistica_externa l
            JOIN ventas v           ON l.venta_id    = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            WHERE l.token_respuesta = %s
        """, (token,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Token inválido o expirado'}), 404
        if row[3]:   # token_usado
            return jsonify({'ya_respondido': True}), 200
        if not cotizacion_esta_vigente(
            row[6], ahora=datetime.now(ZONA_HORARIA_LIMA)
        ):
            return jsonify({'error': 'Este enlace de cotización expiró'}), 410
        if row[7] != 'Cotizacion Enviada':
            return jsonify({'error': 'Esta cotización ya no está pendiente'}), 409
        return jsonify({
            'id': row[0], 'insumo': row[1], 'sku': row[2] or '',
            'codigo_venta': row[4], 'proveedor': row[5]
        }), 200
    except Exception as e:
        print(f"Error al abrir cotización: {e}")
        return jsonify({'error': 'No se pudo abrir la cotización'}), 500
    finally:
        if conexion:
            if cursor:
                cursor.close()
            release_db_connection(conexion)


@produccion_bp.route('/api/cotizar/<string:token>', methods=['POST'])
def responder_cotizacion(token):
    """Endpoint PÚBLICO — el proveedor envía su precio y fecha."""
    if not re.fullmatch(r'[a-f0-9]{32}', token or ''):
        return jsonify({'error': 'Token inválido o expirado'}), 404

    data = request.get_json(silent=True) or {}
    notas = data.get('notas') or ''
    try:
        precio = normalizar_precio_cotizacion(data.get('precio'))
        fecha = normalizar_fecha_cotizacion(
            data.get('fecha_entrega'),
            fecha_minima=datetime.now(ZONA_HORARIA_LIMA).date(),
        )
        if not isinstance(notas, str) or len(notas) > 2000:
            raise ValueError('Las observaciones no pueden superar 2000 caracteres')
        notas = notas.strip()
    except ValueError as error:
        return jsonify({'error': str(error)}), 400

    conexion = None
    cursor = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT id, token_usado, fecha_envio_cotizacion, estado
            FROM logistica_externa
            WHERE token_respuesta = %s
            FOR UPDATE
        """, (token,))
        cotizacion = cursor.fetchone()
        if not cotizacion:
            return jsonify({'error': 'Token inválido o expirado'}), 404
        logistica_id, token_usado, fecha_envio, estado = cotizacion
        if token_usado:
            return jsonify({'error': 'Esta cotización ya fue enviada'}), 409
        if not cotizacion_esta_vigente(
            fecha_envio, ahora=datetime.now(ZONA_HORARIA_LIMA)
        ):
            return jsonify({'error': 'Este enlace de cotización expiró'}), 410
        if estado != 'Cotizacion Enviada':
            return jsonify({'error': 'Esta cotización ya no está pendiente'}), 409

        cursor.execute("""
            UPDATE logistica_externa
            SET precio_cotizado           = %s,
                fecha_entrega_proveedor   = %s,
                notas_proveedor           = %s,
                token_usado               = TRUE,
                fecha_respuesta_proveedor = NOW(),
                estado = 'Cotizado'
            WHERE id = %s
              AND token_usado = FALSE
              AND estado = 'Cotizacion Enviada'
        """, (precio, fecha, notas, logistica_id))
        if cursor.rowcount == 0:
            conexion.rollback()
            return jsonify({'error': 'Esta cotización ya fue enviada'}), 409
        conexion.commit()
        return jsonify({'exito': True}), 200
    except Exception as e:
        if conexion:
            conexion.rollback()
        print(f"Error al responder cotización: {e}")
        return jsonify({'error': 'No se pudo guardar la cotización'}), 500
    finally:
        if conexion:
            if cursor:
                cursor.close()
            release_db_connection(conexion)

@produccion_bp.route('/api/despacho/cola-general', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller', 'Chofer')
def obtener_cola_despacho_general():
    """
    Devuelve todos los tickets de DESPACHO_CENTRAL que están 'Pendiente'
    (listos para entregar) pero aún no tienen un chofer asignado.
    Usado por la nueva vista "Cola de Despacho" del chofer.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT
                t.id,
                v.codigo_venta,
                v.nombre_cliente,
                v.direccion_cliente,
                v.celular_cliente,
                v.fecha_entrega,
                i.producto,
                i.foto_url
            FROM tickets_produccion t
            JOIN items_venta i ON t.item_id = i.id
            JOIN ventas v      ON i.venta_id = v.id
            WHERE t.area_trabajo = 'DESPACHO_CENTRAL'
              AND t.estado_ticket = 'Pendiente'
              AND t.trabajador_asignado_id IS NULL
              AND COALESCE(v.estado_general, '') NOT IN ('Entregado', 'Cancelado')
              AND NOT EXISTS (
                  SELECT 1
                  FROM tickets_produccion tp
                  JOIN items_venta ip ON tp.item_id = ip.id
                  WHERE ip.venta_id = v.id
                    AND tp.area_trabajo != 'DESPACHO_CENTRAL'
                    AND tp.estado_ticket NOT IN ('Terminado', 'Recogido', 'Cancelado')
              )
              AND NOT EXISTS (
                  SELECT 1
                  FROM logistica_externa le
                  WHERE le.venta_id = v.id
                    AND le.estado NOT IN ('Recibido', 'Cancelado', 'Rechazado')
              )
              AND NOT EXISTS (
                  SELECT 1
                  FROM logistica_externa le
                  WHERE le.venta_id = v.id
                    AND (le.categoria_insumo = 'TELA'
                         OR LOWER(COALESCE(le.insumo_nombre, '')) LIKE '%%tela%%'
                         OR LOWER(COALESCE(le.unidad, '')) IN ('mts', 'metro', 'metros'))
                    AND COALESCE(le.estado_distribucion, '') != 'Distribuido'
                    AND le.estado NOT IN ('Cancelado', 'Rechazado')
              )
            ORDER BY v.fecha_entrega ASC NULLS LAST, v.id ASC;
        """)
        return jsonify([{
            'ticket_id':   r[0], 'codigo_venta': r[1], 'cliente': r[2],
            'direccion':   r[3], 'telefono': r[4],
            'fecha_entrega': r[5].strftime('%d/%m/%Y') if r[5] else 'S/F',
            'producto':    r[6], 'foto_url': limpiar_foto(r[7]),
        } for r in cursor.fetchall()]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/despacho/auto-asignar/<int:ticket_id>', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller', 'Chofer')
def auto_asignar_despacho(ticket_id):
    """
    Permite a un chofer (o admin) asignarse a sí mismo un ticket de despacho
    de la cola general, moviéndolo a su bandeja de "Mis Entregas".
    """
    from auth_middleware import get_usuario_actual
    try:
        usuario   = get_usuario_actual()
        chofer_id = usuario.get('id')

        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _, error_payload, status = _validar_contrato_listo_para_despacho(cursor, ticket_id)
        if error_payload:
            return jsonify(error_payload), status

        cursor.execute("""
            UPDATE tickets_produccion
            SET trabajador_asignado_id = %s, estado_ticket = 'En Proceso', fecha_inicio = NOW()
            WHERE id = %s AND area_trabajo = 'DESPACHO_CENTRAL' AND estado_ticket = 'Pendiente';
        """, (chofer_id, ticket_id))
        if cursor.rowcount == 0:
            conexion.rollback()
            return jsonify({'error': 'El despacho ya fue tomado o no está pendiente.'}), 409
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Te has asignado la entrega. Ahora aparecerá en "Mis Entregas".'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

@produccion_bp.route('/api/logistica/<int:id>/generar-orden', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def generar_orden_compra(id):
    """Genera PDF de Orden de Compra con diseño corporativo y lo sube a Cloudinary."""
    data = request.get_json(silent=True) or {}
    recoger_primero = data.get('recoger_primero', False)
    if not isinstance(recoger_primero, bool):
        return jsonify({'error': 'recoger_primero debe ser verdadero o falso'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute("""
            SELECT l.insumo_nombre, l.sku, l.precio_cotizado,
                   l.fecha_entrega_proveedor, l.notas_proveedor,
                   v.codigo_venta, COALESCE(p.nombre,'Sin proveedor') AS prov,
                   COALESCE(p.telefono,'') AS tel_prov,
                   COALESCE(p.correo,'')  AS correo_prov,
                   COALESCE(l.cantidad, 1) AS cantidad,
                   COALESCE(l.unidad, 'und') AS unidad,
                   COALESCE(l.proveedor_informal,'') AS prov_informal,
                   l.estado
            FROM logistica_externa l
            JOIN ventas v           ON l.venta_id    = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            WHERE l.id = %s
            FOR UPDATE OF l
        """, (id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Registro no encontrado'}), 404

        (insumo, sku, precio, fecha_entrega, notas,
         cod_venta, proveedor, tel_prov, correo_prov,
         cantidad, unidad, prov_informal, estado_actual) = row

        if estado_actual != 'Cotizado':
            return jsonify({
                'error': f'La Orden de Compra solo se genera desde Cotizado (estado actual: {estado_actual})'
            }), 409
        precio_unit = float(precio) if precio is not None else 0.0
        cantidad_num = float(cantidad) if cantidad is not None else 0.0
        if not math.isfinite(precio_unit) or precio_unit <= 0:
            return jsonify({'error': 'La cotización debe tener un precio unitario válido'}), 409
        if not math.isfinite(cantidad_num) or cantidad_num <= 0:
            return jsonify({'error': 'La cantidad del material debe ser mayor que cero'}), 409
        if not fecha_entrega:
            return jsonify({'error': 'La cotización debe tener una fecha de entrega'}), 409
        cantidad_str = f'{cantidad_num:.2f}'.rstrip('0').rstrip('.')

        # Número de OC — intentar obtener el próximo de la secuencia
        cursor.execute("""
            SELECT EXISTS(
                SELECT 1 FROM pg_proc WHERE proname = 'generar_numero_oc'
            )
        """)
        if cursor.fetchone()[0]:
            cursor.execute("SELECT generar_numero_oc()")
            numero_oc = cursor.fetchone()[0]
        else:
            # Serializa el fallback para que dos OC simultáneas no repitan número.
            cursor.execute("SELECT pg_advisory_xact_lock(hashtext('ordenes_compra_numero'))")
            cursor.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM ordenes_compra_seq")
            seq_num   = cursor.fetchone()[0]
            numero_oc = f"OC-{seq_num:04d}"

        fecha_emision = datetime.now().strftime('%d/%m/%Y')
        fecha_entrega_str = fecha_entrega.strftime('%d/%m/%Y') if fecha_entrega else 'Por confirmar'
        subtotal     = precio_unit * cantidad_num
        igv          = round(subtotal * 0.18, 2)
        total        = round(subtotal + igv, 2)
        nombre_prov_display = prov_informal if prov_informal else proveedor

        # ── Colores corporativos ────────────────────────────────────────
        COLOR_OSCURO = rl_colors.HexColor('#0f172a')
        COLOR_DORADO = rl_colors.HexColor('#c9a84c')
        COLOR_GRIS   = rl_colors.HexColor('#f8fafc')
        COLOR_BORDE  = rl_colors.HexColor('#e2e8f0')
        COLOR_TEXTO  = rl_colors.HexColor('#374151')

        # ── Canvas ─────────────────────────────────────────────────────
        buf  = BytesIO()
        c    = rl_canvas.Canvas(buf, pagesize=A4)
        w, h = A4
        margin_x = 40 * mm

        def draw_text(text, x, y, font='Helvetica', size=10, color=COLOR_TEXTO):
            c.setFont(font, size)
            c.setFillColor(color)
            c.drawString(x, y, str(text))

        def draw_rect(x, y, width, height, fill=None, stroke=None, radius=4):
            c.setLineWidth(0.5)
            if fill:
                c.setFillColor(fill)
            if stroke:
                c.setStrokeColor(stroke)
            else:
                c.setStrokeColor(rl_colors.white)
            c.roundRect(x, y, width, height, radius, fill=1 if fill else 0, stroke=1 if stroke else 0)

        # ── HEADER: banda oscura ───────────────────────────────────────
        draw_rect(0, h - 52*mm, w, 52*mm, fill=COLOR_OSCURO)

        # Nombre empresa
        c.setFont('Helvetica-Bold', 22)
        c.setFillColor(rl_colors.white)
        c.drawString(margin_x, h - 18*mm, 'INNOVA')
        c.setFillColor(COLOR_DORADO)
        c.drawString(margin_x + 68, h - 18*mm, 'MÖBILI')

        # Subtítulo empresa
        c.setFont('Helvetica', 8)
        c.setFillColor(rl_colors.HexColor('#94a3b8'))
        c.drawString(margin_x, h - 24*mm, 'Muebles de diseño a medida')

        # Título OC (derecha)
        c.setFont('Helvetica-Bold', 18)
        c.setFillColor(COLOR_DORADO)
        c.drawRightString(w - margin_x, h - 18*mm, 'ORDEN DE COMPRA')

        # Número OC
        c.setFont('Helvetica', 10)
        c.setFillColor(rl_colors.white)
        c.drawRightString(w - margin_x, h - 24*mm, f'N° {numero_oc}')

        # Fecha y ref (segunda línea header)
        c.setFont('Helvetica', 9)
        c.setFillColor(rl_colors.HexColor('#cbd5e1'))
        c.drawString(margin_x, h - 32*mm, f'Fecha de emisión: {fecha_emision}')
        c.drawRightString(w - margin_x, h - 32*mm, f'Ref. pedido: {cod_venta}')

        # Línea separadora dorada
        c.setStrokeColor(COLOR_DORADO)
        c.setLineWidth(1.5)
        c.line(margin_x, h - 38*mm, w - margin_x, h - 38*mm)

        # ── BLOQUE: Proveedor + Condiciones ───────────────────────────
        y_bloque = h - 64*mm
        col_w    = (w - 2 * margin_x - 8*mm) / 2

        # Caja Proveedor
        draw_rect(margin_x, y_bloque - 28*mm, col_w, 32*mm, fill=COLOR_GRIS, stroke=COLOR_BORDE)
        draw_text('PROVEEDOR', margin_x + 4*mm, y_bloque + 1*mm, 'Helvetica-Bold', 8, COLOR_DORADO)
        draw_text(nombre_prov_display[:40], margin_x + 4*mm, y_bloque - 6*mm, 'Helvetica-Bold', 11, COLOR_OSCURO)
        if tel_prov:
            draw_text(f'Tel: {tel_prov}', margin_x + 4*mm, y_bloque - 12*mm, 'Helvetica', 9, COLOR_TEXTO)
        if correo_prov:
            draw_text(correo_prov[:38], margin_x + 4*mm, y_bloque - 18*mm, 'Helvetica', 9, COLOR_TEXTO)

        # Caja Condiciones
        x_cond = margin_x + col_w + 8*mm
        draw_rect(x_cond, y_bloque - 28*mm, col_w, 32*mm, fill=COLOR_GRIS, stroke=COLOR_BORDE)
        draw_text('CONDICIONES', x_cond + 4*mm, y_bloque + 1*mm, 'Helvetica-Bold', 8, COLOR_DORADO)
        draw_text(f'Entrega pactada: {fecha_entrega_str}', x_cond + 4*mm, y_bloque - 6*mm, 'Helvetica', 9, COLOR_TEXTO)
        draw_text('Moneda: Soles (PEN)', x_cond + 4*mm, y_bloque - 12*mm, 'Helvetica', 9, COLOR_TEXTO)
        draw_text('Pago: Contra entrega', x_cond + 4*mm, y_bloque - 18*mm, 'Helvetica', 9, COLOR_TEXTO)

        # ── TABLA DE ÍTEMS ────────────────────────────────────────────
        y_tabla = y_bloque - 36*mm

        # Encabezado tabla
        draw_rect(margin_x, y_tabla - 8*mm, w - 2*margin_x, 8*mm, fill=COLOR_OSCURO)
        headers = [('DESCRIPCIÓN', margin_x + 3*mm),
                   ('CANT.',       margin_x + 95*mm),
                   ('P. UNIT.',    margin_x + 120*mm),
                   ('SUBTOTAL',    margin_x + 145*mm)]
        for txt, xh in headers:
            draw_text(txt, xh, y_tabla - 5*mm, 'Helvetica-Bold', 8, rl_colors.white)

        # Fila de datos
        y_fila = y_tabla - 18*mm
        draw_rect(margin_x, y_fila, w - 2*margin_x, 12*mm, fill=COLOR_GRIS, stroke=COLOR_BORDE)
        # Nombre truncado si es muy largo
        nombre_corto = insumo[:48] if insumo else '—'
        draw_text(nombre_corto, margin_x + 3*mm, y_fila + 4*mm, 'Helvetica', 9, COLOR_TEXTO)
        draw_text(f'{cantidad_str} {unidad}', margin_x + 95*mm, y_fila + 4*mm, 'Helvetica', 9, COLOR_TEXTO)
        draw_text(f'S/ {precio_unit:.2f}', margin_x + 120*mm, y_fila + 4*mm, 'Helvetica', 9, COLOR_TEXTO)
        draw_text(f'S/ {subtotal:.2f}',    margin_x + 145*mm, y_fila + 4*mm, 'Helvetica-Bold', 9, COLOR_OSCURO)

        # ── TOTALES ───────────────────────────────────────────────────
        y_tot = y_fila - 6*mm
        x_tot = margin_x + 120*mm
        tot_w = w - margin_x - x_tot

        def fila_total(label, valor, y, bold=False, highlight=False):
            if highlight:
                draw_rect(x_tot - 2*mm, y - 2*mm, tot_w + 2*mm, 8*mm, fill=COLOR_OSCURO)
                draw_text(label, x_tot, y + 2*mm, 'Helvetica-Bold', 9, rl_colors.white)
                draw_text(valor, x_tot + tot_w - 3*mm, y + 2*mm, 'Helvetica-Bold', 10, COLOR_DORADO)
            else:
                c.setStrokeColor(COLOR_BORDE)
                c.setLineWidth(0.3)
                c.line(x_tot, y - 1*mm, x_tot + tot_w, y - 1*mm)
                draw_text(label, x_tot, y + 2*mm, 'Helvetica-Bold' if bold else 'Helvetica', 9, COLOR_TEXTO)
                draw_text(valor, x_tot + tot_w - 3*mm, y + 2*mm, 'Helvetica-Bold' if bold else 'Helvetica', 9, COLOR_TEXTO)

        fila_total('Subtotal:',  f'S/ {subtotal:.2f}', y_tot)
        fila_total('IGV (18%):', f'S/ {igv:.2f}',      y_tot - 9*mm)
        fila_total('TOTAL:',     f'S/ {total:.2f}',    y_tot - 20*mm, bold=True, highlight=True)

        # ── NOTAS ─────────────────────────────────────────────────────
        if notas:
            y_notas = y_fila - 42*mm
            draw_rect(margin_x, y_notas - 14*mm, w - 2*margin_x, 18*mm, fill=rl_colors.HexColor('#fffbeb'), stroke=rl_colors.HexColor('#fde047'))
            draw_text('OBSERVACIONES', margin_x + 4*mm, y_notas + 1*mm, 'Helvetica-Bold', 8, rl_colors.HexColor('#854d0e'))
            # Truncar notas a 2 líneas de ~90 chars
            linea1 = notas[:90]
            linea2 = notas[90:180] if len(notas) > 90 else ''
            draw_text(linea1, margin_x + 4*mm, y_notas - 6*mm, 'Helvetica', 8, COLOR_TEXTO)
            if linea2:
                draw_text(linea2, margin_x + 4*mm, y_notas - 12*mm, 'Helvetica', 8, COLOR_TEXTO)

        # ── PIE DE PÁGINA ─────────────────────────────────────────────
        draw_rect(0, 0, w, 14*mm, fill=COLOR_OSCURO)
        c.setFont('Helvetica', 7)
        c.setFillColor(rl_colors.HexColor('#64748b'))
        c.drawCentredString(w / 2, 6*mm, 'Innova Möbili — Este documento es una orden de compra oficial.')
        c.drawCentredString(w / 2, 3*mm, f'Generado el {fecha_emision} · Ref. {cod_venta} · OC {numero_oc}')

        c.save()
        buf.seek(0)

        # ── Guardar PDF en base de datos (BYTEA) ─────────────────────
        # Evita depender de Cloudinary para servir el PDF.
        # El endpoint /pdf-oc regenera el PDF desde los datos y lo sirve directo.
        pdf_bytes = buf.getvalue()

        # El PDF, su número y el cambio de etapa se confirman juntos.
        cursor.execute("""
            SELECT id
            FROM ordenes_compra_seq
            WHERE logistica_id = %s
            ORDER BY id DESC
            LIMIT 1
            FOR UPDATE
        """, (id,))
        existing = cursor.fetchone()
        if existing:
            cursor.execute("""
                UPDATE ordenes_compra_seq
                SET numero_oc = %s, pdf_bytes = %s
                WHERE id = %s
            """, (numero_oc, pdf_bytes, existing[0]))
        else:
            cursor.execute("""
                INSERT INTO ordenes_compra_seq (logistica_id, numero_oc, pdf_bytes)
                VALUES (%s, %s, %s)
            """, (id, numero_oc, pdf_bytes))

        cat_insumo = 'OTRO'
        if sku:
            cursor.execute("SELECT 1 FROM maestro_telas WHERE sku = %s LIMIT 1", (sku,))
            if cursor.fetchone():
                cat_insumo = 'TELA'
            else:
                cursor.execute("""
                    SELECT 1 FROM (
                        SELECT sku FROM maestro_tableros WHERE sku = %s UNION ALL
                        SELECT sku FROM maestro_bases WHERE sku = %s UNION ALL
                        SELECT sku FROM maestro_bases_comedor WHERE sku = %s UNION ALL
                        SELECT sku FROM maestro_sillas WHERE sku = %s UNION ALL
                        SELECT sku FROM maestro_butacas WHERE sku = %s
                    ) t LIMIT 1
                """, (sku, sku, sku, sku, sku))
                if cursor.fetchone():
                    cat_insumo = 'ESTRUCTURAL'

        if cat_insumo == 'OTRO' and (
            'tela' in (insumo or '').lower()
            or (unidad or '').lower() in ('mts', 'mt', 'metro', 'metros')
        ):
            cat_insumo = 'TELA'

        estado_dist = None
        if recoger_primero:
            if cat_insumo == 'TELA':
                estado_dist = 'En Recojo'
            elif cat_insumo == 'ESTRUCTURAL':
                estado_dist = 'Listo para Recojo'

        token_orden_compra = uuid.uuid4().hex

        cursor.execute("""
            UPDATE logistica_externa
            SET estado = 'Orden Enviada',
                categoria_insumo = %s,
                estado_distribucion = COALESCE(%s, estado_distribucion),
                token_orden_compra = %s
            WHERE id = %s AND estado = 'Cotizado'
        """, (cat_insumo, estado_dist, token_orden_compra, id))
        if cursor.rowcount != 1:
            raise RuntimeError('La cotización cambió de estado durante la generación de la OC')
        conexion.commit()

        return jsonify({
            'exito':      True,
            'numero_oc':  numero_oc,
            'proveedor':  nombre_prov_display,
            'telefono':   tel_prov,
            'url_pdf': (
                f"{request.url_root.rstrip('/')}/api/logistica/orden-publica/"
                f"{token_orden_compra}"
            ),
        }), 200


    except Exception:
        if 'conexion' in locals() and conexion:
            conexion.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'No se pudo generar la Orden de Compra'}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/<int:id>/registrar-pago', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def registrar_pago_proveedor(id):
    """Sube voucher de pago a Cloudinary y actualiza estado.
    Detecta automáticamente si el insumo es TELA o ESTRUCTURAL
    y asigna estado_distribucion correspondiente.
    """
    if 'comprobante' not in request.files or not request.files['comprobante'].filename:
        return jsonify({'error': 'Campo comprobante es obligatorio'}), 400
    archivo = request.files['comprobante']
    if not (
        (archivo.mimetype or '').startswith('image/')
        or archivo.mimetype == 'application/pdf'
    ):
        return jsonify({'error': 'El comprobante debe ser una imagen o PDF'}), 400
    archivo.stream.seek(0, 2)
    tamano_archivo = archivo.stream.tell()
    archivo.stream.seek(0)
    if tamano_archivo > 8 * 1024 * 1024:
        return jsonify({'error': 'El comprobante no puede superar 8 MB'}), 413
    conexion = None
    cursor = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Bloquea la compra mientras se valida y registra el pago.
        # FIX (julio 2026 - v4): antes esta función siempre recalculaba
        # estado_distribucion desde cero (p.ej. 'En espera' para TELA) y lo
        # sobrescribía sin mirar el valor previo. Eso rompía el caso real de
        # "proveedor de confianza": si el material ya se recogió con
        # generar-orden(recoger_primero=True) — pasando por 'En Recojo' →
        # 'Recogido' → incluso 'Distribuido' — y el pago (nota de pedido)
        # recién se registra días después, este UPDATE retrocedía el estado
        # a 'En espera' como si la tela nunca hubiera llegado. Eso podía
        # reabrir el semáforo de _tela_pendiente_para_item() para un ítem
        # que ya estaba resuelto, o hacer reaparecer en la cola de recojo
        # algo que el chofer ya había entregado.
        cursor.execute("""
            SELECT sku, estado_distribucion, estado,
                   COALESCE(tipo_gestion, 'Externo'),
                   COALESCE(categoria_insumo, 'OTRO'),
                   insumo_nombre, unidad, venta_id
            FROM logistica_externa
            WHERE id = %s
            FOR UPDATE
        """, (id,))
        row_sku = cursor.fetchone()
        if not row_sku:
            return jsonify({'error': 'Registro de logística no encontrado'}), 404
        sku = (row_sku[0] or '').strip()
        estado_distribucion_previo = row_sku[1]
        estado_previo = row_sku[2]
        tipo_gestion = row_sku[3]
        categoria_previa = row_sku[4]
        insumo_nombre = row_sku[5] or ''
        unidad = row_sku[6] or ''
        venta_id = row_sku[7]
        estados_pagables = {
            'Orden Enviada', 'Confirmado', 'En Tránsito', 'Pagado',
            'Listo para Recojo', 'Recibido',
        }
        if tipo_gestion != 'Externo' or estado_previo not in estados_pagables:
            return jsonify({
                'error': f'No se puede registrar un pago desde el estado {estado_previo}'
            }), 409

        resp = cloudinary_upload(archivo, folder='pagos_proveedores', max_width=1600)
        url_voucher = resp.get('secure_url')
        if not url_voucher:
            raise RuntimeError('Cloudinary no devolvió la URL del comprobante')

        # Detectar categoría cruzando con tablas maestro
        categoria_insumo   = categoria_previa if categoria_previa in {'TELA', 'ESTRUCTURAL'} else 'OTRO'
        estado_recien_pagado = None

        if categoria_insumo == 'OTRO' and sku:
            cursor.execute("SELECT 1 FROM maestro_telas WHERE sku = %s LIMIT 1", (sku,))
            if cursor.fetchone():
                categoria_insumo     = 'TELA'
                estado_recien_pagado = 'En espera'   # disponible en la bandeja compartida de Telas
            else:
                cursor.execute("""
                    SELECT 1 FROM (
                        SELECT sku FROM maestro_tableros      WHERE sku = %s
                        UNION ALL
                        SELECT sku FROM maestro_bases         WHERE sku = %s
                        UNION ALL
                        SELECT sku FROM maestro_bases_comedor WHERE sku = %s
                        UNION ALL
                        SELECT sku FROM maestro_sillas        WHERE sku = %s
                        UNION ALL
                        SELECT sku FROM maestro_butacas       WHERE sku = %s
                    ) t LIMIT 1
                """, (sku, sku, sku, sku, sku))
                if cursor.fetchone():
                    categoria_insumo     = 'ESTRUCTURAL'
                    estado_recien_pagado = 'Listo para Recojo'   # va directo a cola

        if categoria_insumo == 'OTRO' and (
            'tela' in insumo_nombre.lower()
            or unidad.lower() in ('mts', 'mt', 'metro', 'metros')
        ):
            categoria_insumo = 'TELA'
            estado_recien_pagado = 'En espera'
        elif categoria_insumo == 'TELA':
            estado_recien_pagado = 'En espera'
        elif categoria_insumo == 'ESTRUCTURAL':
            estado_recien_pagado = 'Listo para Recojo'

        # Si el material YA avanzó más allá de "recién pagado" (porque se
        # recogió de confianza antes de pagar), no retrocedemos su estado —
        # solo se registra el pago. Si todavía no tenía avance (NULL o
        # 'En espera'), sí aplicamos el valor recién calculado.
        ESTADOS_AVANZADOS = ('En Recojo', 'Recogido', 'Distribuido')
        if estado_distribucion_previo in ESTADOS_AVANZADOS:
            estado_distribucion = estado_distribucion_previo
        else:
            estado_distribucion = estado_recien_pagado

        estado_despues_pago = (
            'Recibido'
            if estado_previo == 'Recibido' or estado_distribucion_previo == 'Distribuido'
            else 'Pagado'
        )

        cursor.execute("""
            UPDATE logistica_externa
            SET url_comprobante_pago  = %s,
                fecha_pago            = NOW(),
                estado                = %s,
                categoria_insumo      = %s,
                estado_distribucion   = %s
            WHERE id = %s
        """, (url_voucher, estado_despues_pago, categoria_insumo, estado_distribucion, id))
        if venta_id and estado_despues_pago == 'Recibido':
            _activar_despacho_si_listo(cursor, venta_id)

        conexion.commit()
        return jsonify({
            'exito':               True,
            'url':                 url_voucher,
            'categoria_insumo':    categoria_insumo,
            'estado_distribucion': estado_distribucion,
            'estado':              estado_despues_pago,
        }), 200
    except Exception as e:
        if conexion:
            conexion.rollback()
        print(f"Error al registrar pago de proveedor: {e}")
        return jsonify({'error': 'No se pudo registrar el comprobante de pago'}), 500
    finally:
        if conexion:
            if cursor:
                cursor.close()
            release_db_connection(conexion)

@produccion_bp.route('/api/logistica/resumen', methods=['GET'])
@requiere_login
def resumen_logistica():
    """
    Devuelve resumen de movimientos de logística externa para un rango de fechas.
    Query params:
      desde  — fecha inicio  YYYY-MM-DD  (default: lunes de esta semana)
      hasta  — fecha fin     YYYY-MM-DD  (default: hoy)
    """
    from datetime import date, timedelta
    hoy = date.today()
    dia_semana = hoy.weekday()           # 0=lun … 6=dom
    lunes = hoy - timedelta(days=dia_semana)

    desde_str = request.args.get('desde', lunes.isoformat())
    hasta_str = request.args.get('hasta', hoy.isoformat())

    try:
        desde = date.fromisoformat(desde_str)
        hasta = date.fromisoformat(hasta_str)
    except ValueError:
        return jsonify({'error': 'Formato de fecha inválido. Usa YYYY-MM-DD.'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # ── Todos los movimientos del período ─────────────────────────────────
        cursor.execute("""
            SELECT
                l.id,
                v.codigo_venta,
                l.insumo_nombre,
                l.sku,
                COALESCE(p.nombre, l.proveedor_informal, 'Sin asignar') AS proveedor,
                l.precio_cotizado,
                l.cantidad,
                l.unidad,
                l.estado,
                COALESCE(l.categoria_insumo, 'OTRO')  AS categoria_insumo,
                l.estado_distribucion,
                TO_CHAR(l.fecha_pago,               'DD/MM/YYYY') AS fecha_pago,
                TO_CHAR(l.fecha_entrega_proveedor,  'DD/MM/YYYY') AS fecha_entrega_est,
                TO_CHAR(l.fecha_recojo_fisico,      'DD/MM/YYYY') AS fecha_recojo,
                l.tipo_gestion,
                l.url_comprobante_pago
            FROM logistica_externa l
            JOIN  ventas    v ON l.venta_id     = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            WHERE
                (
                    (l.fecha_pago          IS NOT NULL AND l.fecha_pago::date          BETWEEN %s AND %s)
                 OR (l.fecha_recojo_fisico IS NOT NULL AND l.fecha_recojo_fisico::date BETWEEN %s AND %s)
                )
            ORDER BY COALESCE(l.fecha_pago, l.fecha_recojo_fisico) DESC;
        """, (desde, hasta, desde, hasta))

        rows = cursor.fetchall()

        movimientos = []
        for r in rows:
            precio    = float(r[5] or 0)
            cantidad  = float(r[6] or 1)
            subtotal  = precio * cantidad
            movimientos.append({
                'id':              r[0],
                'codigo_venta':    r[1],
                'insumo':          r[2],
                'sku':             r[3] or '',
                'proveedor':       r[4],
                'precio_unit':     precio,
                'cantidad':        cantidad,
                'unidad':          r[7] or '',
                'subtotal':        subtotal,
                'estado':          r[8] or '',
                'categoria':       r[9],
                'estado_dist':     r[10] or '',
                'fecha_pago':      r[11] or '',
                'fecha_entrega_est': r[12] or '',
                'fecha_recojo':    r[13] or '',
                'tipo_gestion':    r[14] or 'Externo',
                'tiene_comprobante': bool(r[15]),
                'comprobante_url': r[15] or '',
            })

        # ── Estadísticas agregadas ─────────────────────────────────────────────
        total_pagado     = sum(m['subtotal'] for m in movimientos if m['estado'] == 'Pagado')
        total_recibido   = sum(m['subtotal'] for m in movimientos if m['fecha_recojo'])
        por_categoria    = {}
        por_proveedor    = {}
        por_dia          = {}

        for m in movimientos:
            cat  = m['categoria']
            prov = m['proveedor']
            dia  = m['fecha_pago'] or m['fecha_recojo'] or ''

            por_categoria.setdefault(cat,  {'cantidad': 0, 'total': 0.0})
            por_categoria[cat]['cantidad'] += 1
            por_categoria[cat]['total']    += m['subtotal']

            por_proveedor.setdefault(prov, {'cantidad': 0, 'total': 0.0})
            por_proveedor[prov]['cantidad'] += 1
            por_proveedor[prov]['total']    += m['subtotal']

            if dia:
                por_dia.setdefault(dia, {'cantidad': 0, 'total': 0.0})
                por_dia[dia]['cantidad'] += 1
                por_dia[dia]['total']    += m['subtotal']

        # ── Gastos sueltos del período ─────────────────────────────────────────
        gastos_periodo = []
        total_gastos   = 0.0
        cursor.execute("""
            SELECT id, concepto, monto, categoria, proveedor_nombre, fecha_gasto
            FROM gastos_logistica
            WHERE fecha_gasto BETWEEN %s AND %s
            ORDER BY fecha_gasto DESC;
        """, (desde, hasta))
        for gr in cursor.fetchall():
            m = float(gr[2] or 0)
            total_gastos += m
            gastos_periodo.append({
                'id': gr[0], 'concepto': gr[1], 'monto': m,
                'categoria': gr[3], 'proveedor': gr[4] or '',
                'fecha': gr[5].strftime('%d/%m/%Y') if gr[5] else '',
            })
            cat = gr[3] or 'Otro'
            por_categoria.setdefault(cat, {'cantidad': 0, 'total': 0.0})
            por_categoria[cat]['cantidad'] += 1
            por_categoria[cat]['total']    += m

        return jsonify({
            'desde':            desde.strftime('%d/%m/%Y'),
            'hasta':            hasta.strftime('%d/%m/%Y'),
            'total_registros':  len(movimientos),
            'total_pagado':     total_pagado,
            'total_recibido':   total_recibido,
            'total_gastos':     total_gastos,
            'total_periodo':    total_recibido + total_gastos,
            'por_categoria':    por_categoria,
            'por_proveedor':    por_proveedor,
            'por_dia':          por_dia,
            'movimientos':      movimientos,
            'gastos_logistica': gastos_periodo,
        }), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/<int:id>/estado-distribucion', methods=['PATCH'])
@requiere_rol('Admin', 'Jefe_Taller')
def actualizar_estado_distribucion(id):
    """El operario de telas cambia el estado de distribución de un ítem pagado.
    Solo aplica a ítems con categoria_insumo = 'TELA'.
    Body: { "estado": "En espera" | "En Recojo" }
    """
    data = request.get_json(silent=True) or {}
    estado_solicitado = str(data.get('estado') or '').strip()
    if not estado_solicitado:
        return jsonify({'error': 'estado es obligatorio'}), 400
    estado = normalizar_estado_distribucion(estado_solicitado)
    estados_validos = ('En espera', 'En Recojo')
    if estado not in estados_validos:
        return jsonify({'error': f'Estado debe ser uno de: {estados_validos}'}), 400
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT categoria_insumo, estado, COALESCE(estado_distribucion, 'En espera')
            FROM logistica_externa WHERE id = %s FOR UPDATE
        """, (id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Ítem no encontrado'}), 404
        if row[0] != 'TELA':
            return jsonify({'error': 'Este control solo aplica a telas'}), 400
        if row[1] != 'Pagado':
            return jsonify({'error': 'Solo se puede cambiar la distribución de ítems pagados'}), 409
        if normalizar_estado_distribucion(row[2]) == 'Distribuido':
            return jsonify({'error': 'Una tela distribuida no puede volver a pendiente'}), 409
        cursor.execute("""
            UPDATE logistica_externa
            SET estado_distribucion = %s
            WHERE id = %s
        """, (estado, id))
        conexion.commit()
        return jsonify({'exito': True, 'estado_distribucion': estado}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/<int:id>/pdf-oc', methods=['GET'])
@requiere_login
def servir_pdf_oc(id):
    """
    Genera y sirve la Orden de Compra como HTML con diseño corporativo.
    El browser abre el HTML en ventana nueva y puede imprimirlo como PDF.
    Elimina la dependencia de Cloudinary y problemas de Content-Type.
    """
    from flask import make_response
    conexion = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Obtener datos del requerimiento
        cursor.execute("""
            SELECT l.insumo_nombre, l.sku, l.precio_cotizado,
                   l.fecha_entrega_proveedor, l.notas_proveedor,
                   v.codigo_venta, COALESCE(p.nombre,'Sin proveedor') AS prov,
                   COALESCE(p.telefono,'') AS tel_prov,
                   COALESCE(p.correo,'')  AS correo_prov,
                   COALESCE(l.cantidad, 1) AS cantidad,
                   COALESCE(l.unidad, 'und') AS unidad,
                   COALESCE(l.proveedor_informal,'') AS prov_informal
            FROM logistica_externa l
            JOIN ventas v           ON l.venta_id    = v.id
            LEFT JOIN proveedores p ON l.proveedor_id = p.id
            WHERE l.id = %s
        """, (id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Registro no encontrado'}), 404

        (insumo, sku, precio, fecha_entrega, notas,
         cod_venta, proveedor, tel_prov, correo_prov,
         cantidad, unidad, prov_informal) = row

        # Obtener número de OC
        numero_oc = f'OC-{id:04d}'
        try:
            cursor.execute("""
                SELECT numero_oc FROM ordenes_compra_seq
                WHERE logistica_id = %s ORDER BY id DESC LIMIT 1
            """, (id,))
            oc_row = cursor.fetchone()
            if oc_row and oc_row[0]:
                numero_oc = oc_row[0]
        except Exception:
            conexion.rollback()

        from datetime import datetime as dt
        fecha_emision     = dt.now().strftime('%d/%m/%Y')
        fecha_entrega_str = fecha_entrega.strftime('%d/%m/%Y') if fecha_entrega else 'Por confirmar'
        precio_unit = float(precio) if precio else 0.0
        cantidad_num = float(cantidad) if cantidad is not None else 0.0
        cantidad_str = f'{cantidad_num:.2f}'.rstrip('0').rstrip('.')
        subtotal    = precio_unit * cantidad_num
        igv         = round(subtotal * 0.18, 2)
        total       = round(subtotal + igv, 2)
        numero_oc_html = html_escape(str(numero_oc))
        cod_venta_html = html_escape(str(cod_venta or 'S/C'))
        nombre_prov = html_escape(str(prov_informal or proveedor or 'Sin proveedor'))
        tel_prov_html = html_escape(str(tel_prov or ''))
        correo_prov_html = html_escape(str(correo_prov or ''))
        insumo_html = html_escape(str(insumo or '—'))
        unidad_html = html_escape(str(unidad or 'und'))
        notas_html = (
            f'<div class="notas-box"><b>Observaciones:</b> {html_escape(str(notas))}</div>'
            if notas else ''
        )
        logo_url = html_escape(
            f"{request.url_root.rstrip('/')}/imagenes/Logo3.png",
            quote=True,
        )

        html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Orden de Compra {numero_oc_html}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@300;400;600;700&family=Jost:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Jost', Arial, sans-serif;
    background: #f5f0e8;
    padding: 30px 16px;
    color: #2c1f0e;
  }}
  .page {{
    width: 210mm;
    margin: auto;
    background: #fff;
    overflow: hidden;
    box-shadow: 0 12px 48px rgba(44,31,14,0.18);
  }}

  /* ── HEADER ── */
  .header {{
    background: #1a120b;
    padding: 28px 40px 22px;
    position: relative;
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
  }}
  .header-left {{
    display: flex;
    align-items: center;
    gap: 18px;
  }}
  /* Logo */
  .logo-mark {{
    height: 80px;
    width: auto;
    flex-shrink: 0;
    filter: drop-shadow(0px 4px 6px rgba(0,0,0,0.25));
  }}
  .brand-text {{}}
  .brand-name {{
    font-family: 'Cormorant Garamond', serif;
    font-size: 22px;
    font-weight: 700;
    color: #f5f0e8;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    line-height: 1;
  }}
  .brand-name em {{
    color: #c9a84c;
    font-style: italic;
  }}
  .brand-sub {{
    font-family: 'Jost', sans-serif;
    font-size: 9px;
    font-weight: 300;
    color: #8a7560;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    margin-top: 5px;
  }}
  .header-right {{
    text-align: right;
  }}
  .oc-label {{
    font-family: 'Jost', sans-serif;
    font-size: 9px;
    font-weight: 600;
    letter-spacing: 0.35em;
    text-transform: uppercase;
    color: #8a7560;
    margin-bottom: 6px;
  }}
  .oc-title {{
    font-family: 'Cormorant Garamond', serif;
    font-size: 26px;
    font-weight: 600;
    color: #c9a84c;
    letter-spacing: 0.06em;
    line-height: 1;
  }}
  .oc-num {{
    font-family: 'Jost', sans-serif;
    font-size: 12px;
    font-weight: 300;
    color: #8a7560;
    letter-spacing: 0.1em;
    margin-top: 4px;
  }}

  /* Línea dorada separadora */
  .header-divider {{
    border: none;
    border-top: 1px solid rgba(201,168,76,0.35);
    margin: 18px 40px 0;
  }}
  .header-meta {{
    background: #1a120b;
    display: flex;
    justify-content: space-between;
    padding: 10px 40px 20px;
    font-family: 'Jost', sans-serif;
    font-size: 10px;
    font-weight: 300;
    letter-spacing: 0.1em;
    color: #8a7560;
  }}
  .header-meta b {{ color: #e8dcc8; font-weight: 500; }}

  /* ── BODY ── */
  .body {{ padding: 32px 40px; }}

  /* CAJAS INFO */
  .info-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
    margin-bottom: 32px;
  }}
  .info-box {{
    border: 1px solid #e8dcc8;
    border-radius: 4px;
    padding: 16px 20px;
    background: #fdfaf5;
  }}
  .info-box-label {{
    font-family: 'Jost', sans-serif;
    font-size: 8px;
    font-weight: 600;
    color: #c9a84c;
    text-transform: uppercase;
    letter-spacing: 0.3em;
    margin-bottom: 10px;
    padding-bottom: 8px;
    border-bottom: 1px solid #e8dcc8;
  }}
  .info-box-nombre {{
    font-family: 'Cormorant Garamond', serif;
    font-size: 18px;
    font-weight: 600;
    color: #2c1f0e;
    margin-bottom: 8px;
    line-height: 1.2;
  }}
  .info-box-row {{
    font-family: 'Jost', sans-serif;
    font-size: 11px;
    font-weight: 300;
    color: #8a7560;
    margin-bottom: 4px;
    letter-spacing: 0.04em;
  }}
  .info-box-row b {{ font-weight: 500; color: #2c1f0e; }}

  /* TABLA */
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 24px; }}
  thead tr {{
    background: #2c1f0e;
  }}
  thead th {{
    font-family: 'Jost', sans-serif;
    color: #e8dcc8;
    font-size: 9px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.2em;
    padding: 12px 14px;
    text-align: left;
  }}
  thead th:last-child {{ text-align: right; }}
  tbody tr {{ background: #fdfaf5; }}
  tbody tr:nth-child(even) {{ background: #f5f0e8; }}
  tbody td {{
    padding: 16px 14px;
    font-family: 'Jost', sans-serif;
    font-size: 12px;
    font-weight: 400;
    color: #2c1f0e;
    border-bottom: 1px solid #e8dcc8;
    vertical-align: middle;
  }}
  tbody td:last-child {{
    text-align: right;
    font-weight: 600;
    color: #2c1f0e;
    font-family: 'Cormorant Garamond', serif;
    font-size: 15px;
    white-space: nowrap;
  }}
  .td-sku {{
    font-family: 'Jost', sans-serif;
    font-size: 10px;
    font-weight: 300;
    color: #b07d3a;
    letter-spacing: 0.08em;
  }}

  /* TOTALES */
  .totales {{ display: flex; justify-content: flex-end; margin-bottom: 24px; }}
  .totales-inner {{ width: 260px; }}
  .tot-row {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 9px 0;
    border-bottom: 1px solid #e8dcc8;
    font-family: 'Jost', sans-serif;
    font-size: 11px;
    font-weight: 300;
    color: #8a7560;
    letter-spacing: 0.06em;
  }}
  .tot-row.final {{
    background: #2c1f0e;
    color: #f5f0e8;
    padding: 12px 16px;
    margin-top: 8px;
    border: none;
    font-weight: 600;
    font-size: 12px;
    letter-spacing: 0.12em;
    text-transform: uppercase;
  }}
  .tot-row.final span:last-child {{
    font-family: 'Cormorant Garamond', serif;
    font-size: 20px;
    font-weight: 700;
    color: #c9a84c;
    letter-spacing: 0.05em;
  }}

  /* NOTAS */
  .notas-box {{
    background: rgba(201,168,76,0.06);
    border-left: 3px solid #c9a84c;
    padding: 12px 18px;
    font-family: 'Jost', sans-serif;
    font-size: 11px;
    font-weight: 300;
    color: #8a7560;
    margin-bottom: 28px;
    letter-spacing: 0.04em;
  }}
  .notas-box b {{ color: #2c1f0e; font-weight: 500; }}

  /* PIE */
  .footer {{
    background: #1a120b;
    padding: 16px 40px;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 16px;
  }}
  .footer-logo {{
    opacity: 0.5;
    display: flex;
    align-items: center;
    gap: 8px;
  }}
  .footer-logo span {{
    font-family: 'Cormorant Garamond', serif;
    font-size: 11px;
    font-weight: 300;
    color: #8a7560;
    letter-spacing: 0.2em;
    text-transform: uppercase;
  }}
  .footer-sep {{ color: #8a7560; opacity: 0.3; }}
  .footer-text {{
    font-family: 'Jost', sans-serif;
    font-size: 9px;
    font-weight: 300;
    color: #8a7560;
    letter-spacing: 0.12em;
  }}

  /* BOTÓN IMPRIMIR */
  .print-bar {{ text-align: center; margin-bottom: 24px; }}
  .btn-print {{
    background: #2c1f0e;
    color: #c9a84c;
    border: none;
    padding: 13px 36px;
    font-family: 'Jost', sans-serif;
    font-size: 13px;
    font-weight: 500;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    cursor: pointer;
    transition: background 0.2s;
  }}
  .btn-print:hover {{ background: #1a120b; }}

  @media print {{
    body {{ background: #fff; padding: 0; }}
    .page {{ box-shadow: none; width: 100%; }}
    .print-bar {{ display: none; }}
  }}
</style>
</head>
<body>
<div class="print-bar">
  <button class="btn-print" onclick="window.print()">🖨️ &nbsp;Imprimir / Guardar como PDF</button>
</div>
<div class="page">

  <!-- HEADER -->
  <div class="header">
    <div class="header-left">
      <!-- Logo real (mismo archivo que usa el contrato de venta) -->
      <img src="{logo_url}" class="logo-mark"
           onerror="this.style.display='none'">
      <div class="brand-text">
        <div class="brand-name">Innova <em>Möbili</em></div>
        <div class="brand-sub">Muebles de diseño a medida &nbsp;·&nbsp; RUC 20600768175</div>
      </div>
    </div>
    <div class="header-right">
      <div class="oc-label">Documento comercial</div>
      <div class="oc-title">Orden de Compra</div>
      <div class="oc-num">N° {numero_oc_html}</div>
    </div>
  </div>
  <hr class="header-divider">
  <div class="header-meta">
    <span>Fecha de emisión: <b>{fecha_emision}</b></span>
    <span>Ref. pedido: <b>{cod_venta_html}</b></span>
  </div>

  <!-- BODY -->
  <div class="body">
    <div class="info-grid">
      <div class="info-box">
        <div class="info-box-label">Proveedor</div>
        <div class="info-box-nombre">{nombre_prov}</div>
        {'<div class="info-box-row">📞 &nbsp;' + tel_prov_html + '</div>' if tel_prov_html else ''}
        {'<div class="info-box-row">✉ &nbsp;' + correo_prov_html + '</div>' if correo_prov_html else ''}
      </div>
      <div class="info-box">
        <div class="info-box-label">Condiciones</div>
        <div class="info-box-row"><b>Entrega pactada:</b> &nbsp;{fecha_entrega_str}</div>
        <div class="info-box-row"><b>Moneda:</b> &nbsp;Soles (PEN)</div>
        <div class="info-box-row"><b>Pago:</b> &nbsp;Contra entrega</div>
      </div>
    </div>

    <table>
      <thead>
        <tr>
          <th>Descripción</th>
          <th>Cant.</th>
          <th>P. Unit.</th>
          <th>Subtotal</th>
        </tr>
      </thead>
      <tbody>
        <tr>
          <td>{insumo_html}</td>
          <td>{cantidad_str} {unidad_html}</td>
          <td>S/ {precio_unit:.2f}</td>
          <td>S/ {subtotal:.2f}</td>
        </tr>
      </tbody>
    </table>

    <div class="totales">
      <div class="totales-inner">
        <div class="tot-row"><span>Subtotal</span><span>S/ {subtotal:.2f}</span></div>
        <div class="tot-row"><span>IGV (18%)</span><span>S/ {igv:.2f}</span></div>
        <div class="tot-row final"><span>Total</span><span>S/ {total:.2f}</span></div>
      </div>
    </div>

    {notas_html}
  </div>

  <!-- PIE -->
  <div class="footer">
    <div class="footer-logo">
      <svg width="14" height="14" viewBox="0 0 56 56" fill="none">
        <polygon points="28,4 52,28 28,52 4,28" fill="#c9a84c" opacity="0.6"/>
      </svg>
      <span>Innova Möbili</span>
    </div>
    <span class="footer-sep">·</span>
    <span class="footer-text">Documento generado el {fecha_emision} &nbsp;·&nbsp; Ref. {cod_venta_html} &nbsp;·&nbsp; {numero_oc_html}</span>
  </div>

</div>
</body>
</html>"""

        resp = make_response(html)
        resp.headers['Content-Type'] = 'text/html; charset=utf-8'
        resp.headers['Cache-Control'] = 'no-store'
        return resp

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/orden-publica/<token>', methods=['GET'])
def servir_pdf_oc_publico(token):
    """Permite al proveedor abrir una OC sin exponer IDs ni requerir sesión ERP."""
    if not re.fullmatch(r'[a-f0-9]{32}', token or ''):
        return jsonify({'error': 'Enlace de Orden de Compra inválido'}), 404

    conexion = None
    cursor = None
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT id
            FROM logistica_externa
            WHERE token_orden_compra = %s
              AND estado NOT IN ('Pendiente', 'POR_PEDIR', 'Por Pedir', 'Cotizacion Enviada')
        """, (token,))
        fila = cursor.fetchone()
        if not fila:
            return jsonify({'error': 'Enlace de Orden de Compra inválido'}), 404
        logistica_id = fila[0]
    except Exception as error:
        print(f"Error al resolver Orden de Compra pública: {error}")
        return jsonify({'error': 'No se pudo abrir la Orden de Compra'}), 500
    finally:
        if conexion:
            if cursor:
                cursor.close()
            release_db_connection(conexion)

    return servir_pdf_oc.__wrapped__(logistica_id)


# ─── STOCK ESTRUCTURAS SOFÁ ───────────────────────────────────────────────────

@produccion_bp.route('/api/stock-estructuras', methods=['GET'])
@requiere_login
def listar_stock_estructuras():
    # Filtros opcionales
    desde       = (request.args.get('desde') or '').strip()
    hasta       = (request.args.get('hasta') or '').strip()
    carpintero  = (request.args.get('carpintero') or '').strip()
    estado_pago = (request.args.get('pago') or '').strip()   # 'pagado' | 'pendiente'
    try:
        limit = min(max(int(request.args.get('limit', 500)), 1), 1000)
    except (TypeError, ValueError):
        limit = 500

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        conditions, params = [], []
        if desde:
            conditions.append("fecha_registro::date >= %s"); params.append(desde)
        if hasta:
            conditions.append("fecha_registro::date <= %s"); params.append(hasta)
        if carpintero:
            conditions.append("COALESCE(carpintero_nombre,'') ILIKE %s")
            params.append(f'%{carpintero}%')
        if estado_pago == 'pagado':
            conditions.append("COALESCE(pagado, FALSE) = TRUE")
        elif estado_pago == 'pendiente':
            conditions.append("COALESCE(pagado, FALSE) = FALSE")

        where_sql = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

        cursor.execute(f"""
            SELECT id, nombre_modelo, ancho, profundidad, alto,
                   medida_estandar, foto_url, tipo, cantidad, estado,
                   ticket_id, TO_CHAR(fecha_registro AT TIME ZONE 'America/Lima','DD/MM/YYYY'), COALESCE(precio, 0),
                   COALESCE(modelo_base, ''), COALESCE(chofer_nombre, ''),
                   COALESCE(tipo_base, ''),
                   COALESCE(medida_base::numeric, 0),
                   COALESCE(medida_base_estandar, FALSE),
                   COALESCE(pagado, FALSE),
                   TO_CHAR(fecha_entrega_chofer AT TIME ZONE 'America/Lima', 'DD/MM/YYYY HH24:MI'),
                   COALESCE(carpintero_nombre, ''),
                   COALESCE(es_antiguo, FALSE),
                   medida_brazo,
                   COALESCE(es_juego_completo, TRUE),
                   COALESCE(completado_por, ''),
                   TO_CHAR(fecha_completado, 'DD/MM/YYYY'),
                   COALESCE(comentario_parte, ''),
                   COALESCE(foto_completado_url, ''),
                   COALESCE(foto_entrega_url, ''),
                   COALESCE(comentario_entrega, '')
            FROM stock_estructuras_sofa
            {where_sql}
            ORDER BY fecha_registro DESC
            LIMIT %s
        """, [*params, limit])
        rows = cursor.fetchall()
        return jsonify([{
            'id': r[0], 'nombre_modelo': r[1],
            'ancho': float(r[2] or 0), 'profundidad': float(r[3] or 0), 'alto': float(r[4] or 0),
            'medida_estandar': r[5], 'foto_url': r[6], 'tipo': r[7],
            'cantidad': r[8], 'estado': r[9], 'ticket_id': r[10], 'fecha': r[11],
            'precio': float(r[12] or 0),
            'modelo_base': r[13], 'chofer_nombre': r[14],
            'tipo_base': r[15],
            'medida_base': float(r[16] or 0),
            'medida_base_estandar': r[17],
            'pagado': bool(r[18]),
            'fecha_entrega_chofer': r[19] or '',
            'carpintero_nombre': r[20] or '',
            'es_antiguo': bool(r[21]),
            'medida_brazo': float(r[22]) if r[22] is not None else None,
            'es_juego_completo': bool(r[23]),
            'completado_por': r[24] or '',
            'fecha_completado': r[25] or '',
            'comentario_parte': r[26] or '',
            'foto_completado_url': r[27] or '',
            'foto_entrega_url': r[28] or '',
            'comentario_entrega': r[29] or '',
        } for r in rows]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/stock-estructuras/exportar', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller')
def exportar_stock_estructuras_excel():
    """
    Descarga un Excel con dos hojas separadas: 'Estructuras Nuevas' y
    'Estructuras Antiguas' (campo es_antiguo), usando los mismos filtros
    opcionales que el listado normal (desde, hasta, carpintero, pago).
    """
    desde       = (request.args.get('desde') or '').strip()
    hasta       = (request.args.get('hasta') or '').strip()
    carpintero  = (request.args.get('carpintero') or '').strip()
    estado_pago = (request.args.get('pago') or '').strip()

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        conditions, params = [], []
        if desde:
            conditions.append("fecha_registro::date >= %s"); params.append(desde)
        if hasta:
            conditions.append("fecha_registro::date <= %s"); params.append(hasta)
        if carpintero:
            conditions.append("COALESCE(carpintero_nombre,'') ILIKE %s")
            params.append(f'%{carpintero}%')
        if estado_pago == 'pagado':
            conditions.append("COALESCE(pagado, FALSE) = TRUE")
        elif estado_pago == 'pendiente':
            conditions.append("COALESCE(pagado, FALSE) = FALSE")

        where_sql = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

        cursor.execute(f"""
            SELECT nombre_modelo, tipo, cantidad, estado,
                   COALESCE(modelo_base, ''), COALESCE(tipo_base, ''),
                   ancho, profundidad, alto, medida_brazo,
                   COALESCE(precio, 0), COALESCE(carpintero_nombre, ''),
                   COALESCE(chofer_nombre, ''), COALESCE(pagado, FALSE),
                   TO_CHAR(fecha_registro AT TIME ZONE 'America/Lima', 'DD/MM/YYYY'),
                   TO_CHAR(fecha_entrega_chofer AT TIME ZONE 'America/Lima', 'DD/MM/YYYY HH24:MI'),
                   COALESCE(es_antiguo, FALSE)
            FROM stock_estructuras_sofa
            {where_sql}
            ORDER BY COALESCE(es_antiguo, FALSE) ASC, fecha_registro DESC
        """, params)
        filas = cursor.fetchall()

        nuevas   = [f for f in filas if not f[16]]
        antiguas = [f for f in filas if f[16]]

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="0F172A")
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin        = Side(style="thin", color="CBD5E0")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill_par    = PatternFill("solid", fgColor="F8FAFC")
        fill_impar  = PatternFill("solid", fgColor="FFFFFF")
        fill_entregado_si = PatternFill("solid", fgColor="DCFCE7")  # verde clarito
        fill_entregado_no  = PatternFill("solid", fgColor="FEF3C7")  # ambar clarito

        headers = [
            "Modelo", "Tipo", "Cantidad", "Estado",
            "Antiguo", "Entregado",
            "Modelo Base", "Tipo Base",
            "Ancho", "Profundidad", "Alto", "Medida Brazo",
            "Precio (S/)", "Carpintero", "Chofer", "Pagado",
            "Fecha Registro", "Fecha Entrega Chofer"
        ]
        anchos = [22, 14, 10, 16, 10, 12, 18, 14, 10, 12, 10, 13, 12, 22, 20, 10, 14, 20]

        def _llenar_hoja(ws, datos):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = center; cell.border = border
            for col, ancho in enumerate(anchos, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
            for row_num, f in enumerate(datos, 2):
                fill = fill_par if row_num % 2 == 0 else fill_impar
                entregado = (f[3] or '').strip().lower() == 'entregado'
                valores = [
                    f[0], f[1], f[2], f[3],
                    'Sí' if f[16] else 'No',
                    'Sí' if entregado else 'No',
                    f[4], f[5],
                    float(f[6] or 0), float(f[7] or 0), float(f[8] or 0),
                    float(f[9]) if f[9] is not None else '',
                    float(f[10] or 0), f[11], f[12],
                    'Sí' if f[13] else 'No',
                    f[14] or '', f[15] or ''
                ]
                for col, val in enumerate(valores, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if col == 6:  # columna "Entregado" resaltada
                        cell.fill = fill_entregado_si if entregado else fill_entregado_no
                    else:
                        cell.fill = fill
            ws.freeze_panes = "A2"

        ws_nuevas = wb.active
        ws_nuevas.title = "Estructuras Nuevas"
        _llenar_hoja(ws_nuevas, nuevas)

        ws_antiguas = wb.create_sheet("Estructuras Antiguas")
        _llenar_hoja(ws_antiguas, antiguas)

        buffer = BytesIO()
        wb.save(buffer); buffer.seek(0)
        fecha_hoy = datetime.now().strftime('%Y%m%d_%H%M')
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'estructuras_innova_{fecha_hoy}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/stock-estructuras', methods=['POST'])
@requiere_login
def registrar_stock_estructura():
    if not usuario_puede_operar('ESTRUCTURAS_MUEBLES'):
        return jsonify({'error': 'Solo Estructuras o jefatura puede registrar este stock.'}), 403
    try:
        nombre              = request.form.get('nombre_modelo')
        modelo_base         = request.form.get('modelo_base', '')
        ancho               = request.form.get('ancho') or 0
        profundidad         = request.form.get('profundidad') or 0
        alto                = request.form.get('alto') or 0
        medida_estandar     = request.form.get('medida_estandar') == 'true'
        tipo                = request.form.get('tipo', 'estructura')
        cantidad            = int(request.form.get('cantidad', 1))
        precio              = float(request.form.get('precio') or 0)
        # A8: Campos nuevos para pata/zócalo
        tipo_base           = request.form.get('tipo_base', '')  # 'patas', 'zocalo', o vacío
        medida_base         = request.form.get('medida_base') or None
        medida_base_estandar = request.form.get('medida_base_estandar') == 'true'
        es_antiguo          = request.form.get('es_antiguo') == 'true'
        medida_brazo        = request.form.get('medida_brazo') or None
        es_juego_completo   = request.form.get('es_juego_completo') != 'false'

        if not (nombre or '').strip():
            return jsonify({'error': 'El nombre del modelo es obligatorio.'}), 400
        if cantidad < 1 or cantidad > 50:
            return jsonify({'error': 'La cantidad debe estar entre 1 y 50.'}), 400
        if precio < 0:
            return jsonify({'error': 'El precio no puede ser negativo.'}), 400

        # Guardar quién registró la estructura (el carpintero que la creó)
        carpintero_nombre = get_usuario_actual().get('nombre', '')

        # Validar que si hay tipo_base, también hay medida_base
        if tipo_base and not medida_base:
            return jsonify({'error': 'Si selecciona un tipo de base, debe ingresar la medida'}), 400

        foto_url = None
        if 'foto' in request.files and request.files['foto'].filename:
            res = cloudinary_upload(request.files['foto'], folder='stock_estructuras')
            foto_url = res.get('secure_url')

        conexion = get_db_connection()
        cursor   = conexion.cursor()

        if cantidad > 1:
            new_ids = []
            for _ in range(cantidad):
                cursor.execute("""
                    INSERT INTO stock_estructuras_sofa
                        (nombre_modelo, modelo_base, ancho, profundidad, alto,
                         medida_estandar, foto_url, tipo, cantidad, precio,
                         tipo_base, medida_base, medida_base_estandar, carpintero_nombre,
                         es_antiguo, medida_brazo, es_juego_completo)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,1,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
                """, (nombre, modelo_base, ancho, profundidad, alto,
                      medida_estandar, foto_url, tipo, precio,
                      tipo_base, medida_base, medida_base_estandar, carpintero_nombre or None,
                      es_antiguo, medida_brazo, es_juego_completo))
                new_ids.append(cursor.fetchone()[0])
            new_id = new_ids[0]
        else:
            cursor.execute("""
                INSERT INTO stock_estructuras_sofa
                    (nombre_modelo, modelo_base, ancho, profundidad, alto,
                     medida_estandar, foto_url, tipo, cantidad, precio,
                     tipo_base, medida_base, medida_base_estandar, carpintero_nombre,
                     es_antiguo, medida_brazo, es_juego_completo)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
            """, (nombre, modelo_base, ancho, profundidad, alto,
                  medida_estandar, foto_url, tipo, cantidad, precio,
                  tipo_base, medida_base, medida_base_estandar, carpintero_nombre or None,
                  es_antiguo, medida_brazo, es_juego_completo))
            new_id = cursor.fetchone()[0]

        conexion.commit()
        return jsonify({'exito': True, 'id': new_id}), 201
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

@produccion_bp.route('/api/stock-estructuras/<int:stock_id>/usar', methods=['POST'])
@requiere_login
def usar_stock_estructura(stock_id):
    """Marca una estructura como entregada, la vincula al ticket y termina el ticket."""
    data      = request.get_json(silent=True) or {}
    ticket_id = data.get('ticket_id')
    if not ticket_id:
        return jsonify({'error': 'El ticket_id es obligatorio.'}), 400
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT item_id, area_trabajo, trabajador_asignado_id, estado_ticket
            FROM tickets_produccion
            WHERE id = %s
            FOR UPDATE
        """, (ticket_id,))
        ticket = cursor.fetchone()
        if not ticket:
            return jsonify({'error': 'Ticket no encontrado.'}), 404
        if ticket[1] not in ('ESTRUCTURAS_MUEBLES', 'ESTRUCTURAS_SILLAS'):
            return jsonify({'error': 'La estructura solo puede aplicarse a un ticket de estructuras.'}), 400
        if not usuario_puede_operar(ticket[1], ticket[2]):
            return jsonify({'error': 'No puedes usar stock en un ticket de otra persona o área.'}), 403
        if ticket[3] in ('Terminado', 'Listo para Recojo', 'Recogido', 'Cancelado'):
            return jsonify({'error': f'El ticket ya está cerrado ({ticket[3]}).'}), 409

        cursor.execute("""
            SELECT estado, ticket_id
            FROM stock_estructuras_sofa
            WHERE id = %s
            FOR UPDATE
        """, (stock_id,))
        stock = cursor.fetchone()
        if not stock:
            return jsonify({'error': 'Estructura de stock no encontrada.'}), 404
        if (stock[0] or '').lower() != 'disponible':
            return jsonify({'error': 'La estructura ya no está disponible.'}), 409
        cursor.execute(
            "SELECT 1 FROM stock_estructuras_sofa WHERE ticket_id = %s LIMIT 1",
            (ticket_id,),
        )
        if cursor.fetchone():
            return jsonify({'error': 'Este ticket ya tiene una estructura de stock vinculada.'}), 409

        cursor.execute("""
            UPDATE stock_estructuras_sofa
            SET estado = 'entregado', ticket_id = %s
            WHERE id = %s AND LOWER(COALESCE(estado, '')) = 'disponible'
            RETURNING id
        """, (ticket_id, stock_id))
        if not cursor.fetchone():
            return jsonify({'error': 'La estructura fue tomada por otra operación.'}), 409
        cursor.execute("""
            UPDATE tickets_produccion
            SET estado_ticket = 'Terminado', fecha_fin = CURRENT_TIMESTAMP
            WHERE id = %s
              AND estado_ticket NOT IN ('Terminado', 'Listo para Recojo', 'Recogido', 'Cancelado')
        """, (ticket_id,))
        desbloqueados = _intentar_desbloquear_items(cursor, [ticket[0]])
        cursor.execute("SELECT venta_id FROM items_venta WHERE id = %s", (ticket[0],))
        venta = cursor.fetchone()
        if venta:
            _activar_despacho_si_listo(cursor, venta[0])
        conexion.commit()
        return jsonify({'exito': True, 'desbloqueados': desbloqueados}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# ==========================================
# GESTOR DE MODELOS DE SOFÁ (CUSTOM)
# ==========================================

@produccion_bp.route('/api/sofa-modelos', methods=['GET'])
@requiere_login
def get_sofa_modelos():
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("SELECT id, key, label, medidas, foto_url FROM sofa_modelos_custom ORDER BY label ASC")
        modelos = [{'id': r[0], 'key': r[1], 'label': r[2], 'medidas': r[3], 'foto': r[4]} for r in cursor.fetchall()]
        return jsonify(modelos), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close()
            release_db_connection(conexion)

@produccion_bp.route('/api/sofa-modelos', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def add_sofa_modelo():
    from auth_middleware import get_usuario_actual
    import time
    try:
        label = request.form.get('label')
        medidas = request.form.get('medidas')
        if not label or not medidas:
            return jsonify({'error': 'Nombre y tipo de medidas son obligatorios'}), 400
        if 'foto' not in request.files or not request.files['foto'].filename:
            return jsonify({'error': 'La foto es obligatoria'}), 400

        foto_url = None
        try:
            res = cloudinary_upload(request.files['foto'], folder='modelos_sofa')
            foto_url = res.get('secure_url')
        except Exception as e:
            return jsonify({'error': f'Error al subir la foto: {e}'}), 500
        
        if not foto_url:
            return jsonify({'error': 'No se pudo obtener la URL de la foto subida'}), 500

        key = 'custom_' + str(int(time.time()))
        user_id = get_usuario_actual().get('id')

        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("""
            INSERT INTO sofa_modelos_custom (key, label, medidas, foto_url, created_by)
            VALUES (%s, %s, %s, %s, %s) RETURNING id, key, label, medidas, foto_url
        """, (key, label, medidas, foto_url, user_id))
        nuevo = cursor.fetchone()
        conexion.commit()

        return jsonify({
            'exito': True,
            'modelo': {'id': nuevo[0], 'key': nuevo[1], 'label': nuevo[2], 'medidas': nuevo[3], 'foto': nuevo[4]}
        }), 201

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close()
            release_db_connection(conexion)

@produccion_bp.route('/api/sofa-modelos/<int:modelo_id>', methods=['DELETE'])
@requiere_rol('Admin', 'Jefe_Taller')
def delete_sofa_modelo(modelo_id):
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("DELETE FROM sofa_modelos_custom WHERE id = %s RETURNING id", (modelo_id,))
        deleted = cursor.fetchone()
        conexion.commit()
        if not deleted:
            return jsonify({'error': 'Modelo no encontrado'}), 404
        return jsonify({'exito': True}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close()
            release_db_connection(conexion)

@produccion_bp.route('/api/sofa-modelos/<int:modelo_id>/foto', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def update_sofa_modelo_foto(modelo_id):
    if 'foto' not in request.files or not request.files['foto'].filename:
        return jsonify({'error': 'La foto es obligatoria'}), 400
    
    try:
        foto_url = None
        try:
            res = cloudinary_upload(request.files['foto'], folder='modelos_sofa')
            foto_url = res.get('secure_url')
        except Exception as e:
            return jsonify({'error': f'Error al subir la foto: {e}'}), 500
        
        if not foto_url:
            return jsonify({'error': 'No se pudo obtener la URL de la foto subida'}), 500

        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute("UPDATE sofa_modelos_custom SET foto_url = %s WHERE id = %s RETURNING foto_url", (foto_url, modelo_id))
        updated = cursor.fetchone()
        conexion.commit()
        if not updated:
            return jsonify({'error': 'Modelo no encontrado'}), 404
        return jsonify({'exito': True, 'foto_url': updated[0]}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close()
            release_db_connection(conexion)

# =============================================================================
# INNOVA MÖBILI — routes_produccion.py  PATCH
# Reemplazar SOLO la función sugerir_estructura() (~línea 2039).
# El resto del archivo no se toca.
# =============================================================================


@produccion_bp.route('/api/stock-estructuras/sugerir', methods=['GET'])
@requiere_login
def sugerir_estructura():
    """
    Devuelve estructuras disponibles ordenadas: primero por modelo exacto,
    luego estándar, luego por medidas similares (±15 cm).

    Query params:
      ancho, profundidad, alto  — medidas del ticket (float, default 0)
      modelo_base               — modelo del sofá (str, opcional)
      solo_estandar             — si 'true', ignora medidas y devuelve
                                  solo tipo='estructura' con medida_estandar=TRUE
    """
    try:
        ancho        = float(request.args.get('ancho', 0))
        profundidad  = float(request.args.get('profundidad', 0))
        alto         = float(request.args.get('alto', 0))
        modelo_base  = request.args.get('modelo_base', '').strip()
        notas        = request.args.get('notas', '').strip()[:1200]
        solo_estandar = request.args.get('solo_estandar', 'false').lower() == 'true'
        margen       = 15  # cm

        palabras_omitidas = {
            'para', 'con', 'sin', 'del', 'las', 'los', 'una', 'uno', 'medida',
            'medidas', 'estructura', 'estructuras', 'modelo', 'notas', 'nota',
            'cliente', 'confirmar', 'final', 'sku', 'ref'
        }
        palabras_notas = [
            p for p in re.findall(r'[a-záéíóúñ0-9]+', notas.lower())
            if len(p) >= 4 and p not in palabras_omitidas and not p.isdigit()
        ][:8]
        patrones_notas = [f'%{p}%' for p in dict.fromkeys(palabras_notas)]

        conexion = get_db_connection()
        cursor   = conexion.cursor()

        if solo_estandar:
            # ── Caso C: solo estructuras estándar, sin filtro de medidas ──────
            cursor.execute("""
                SELECT id, nombre_modelo, ancho, profundidad, alto,
                       medida_estandar, foto_url, tipo, cantidad,
                       COALESCE(modelo_base, '')
                FROM stock_estructuras_sofa
                WHERE estado = 'disponible'
                  AND medida_estandar = TRUE
                  AND tipo = 'estructura'
                  AND COALESCE(es_antiguo, FALSE) = FALSE
                ORDER BY nombre_modelo ASC
                LIMIT 8
            """)
        else:
            # Determinar si debemos aplicar el filtro de medidas similares.
            # Si los tres valores son 0, NO aplicar (evita "ancho <= 15" que
            # devuelve basura).
            hay_medidas = (ancho > 0 or profundidad > 0 or alto > 0)

            if hay_medidas:
                medidas_clause = """
                    OR (
                        ABS(ancho       - %(ancho)s)       <= %(margen)s AND
                        ABS(profundidad - %(profundidad)s) <= %(margen)s AND
                        (%(alto)s = 0 OR ABS(alto - %(alto)s) <= %(margen)s)
                    )
                """
            else:
                # Sin medidas: no incluir cláusula de distancia
                medidas_clause = ""

            notas_clause = ""
            if patrones_notas:
                notas_clause = """
                    OR LOWER(CONCAT_WS(' ', nombre_modelo, modelo_base, tipo_base,
                                             medida_base, comentario_parte, comentario_entrega))
                       LIKE ANY (%(patrones_notas)s)
                """

            sql = f"""
                SELECT id, nombre_modelo, ancho, profundidad, alto,
                       medida_estandar, foto_url, tipo, cantidad,
                       COALESCE(modelo_base, '')
                FROM stock_estructuras_sofa
                WHERE estado = 'disponible'
                  AND COALESCE(es_antiguo, FALSE) = FALSE
                  AND (
                    medida_estandar = TRUE
                    OR (%(modelo_base)s != '' AND LOWER(modelo_base) = LOWER(%(modelo_base)s))
                    {medidas_clause}
                    {notas_clause}
                  )
                ORDER BY
                    CASE WHEN %(modelo_base)s != ''
                              AND LOWER(modelo_base) = LOWER(%(modelo_base)s)
                         THEN 0 ELSE 1 END ASC,
                    medida_estandar DESC,
                    ABS(ancho - %(ancho)s) ASC
                LIMIT 8
            """
            cursor.execute(sql, {
                'ancho':        ancho,
                'profundidad':  profundidad,
                'alto':         alto,
                'modelo_base':  modelo_base,
                'margen':       margen,
                'patrones_notas': patrones_notas,
            })

        rows = cursor.fetchall()
        return jsonify([{
            'id':             r[0],
            'nombre_modelo':  r[1],
            'ancho':          float(r[2] or 0),
            'profundidad':    float(r[3] or 0),
            'alto':           float(r[4] or 0),
            'medida_estandar': r[5],
            'foto_url':       r[6],
            'tipo':           r[7],
            'cantidad':       r[8],
            'modelo_base':    r[9],
        } for r in rows]), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/stock-estructuras/<int:stock_id>/entregar', methods=['PATCH'])
@requiere_login
def entregar_estructura(stock_id):
    """
    El carpintero marca una estructura como entregada al chofer.
    Recibe: { "chofer_nombre": "Juan Quispe" }
    Registra quién la recogió para historial.
    """
    data          = request.get_json() or {}
    chofer_nombre = (data.get('chofer_nombre') or '').strip()
    foto_entrega_url   = (data.get('foto_entrega_url') or '').strip()
    comentario_entrega = (data.get('comentario_entrega') or '').strip()

    if not usuario_puede_operar('ESTRUCTURAS_MUEBLES'):
        return jsonify({'error': 'Solo Estructuras o jefatura puede registrar la entrega.'}), 403

    if not chofer_nombre:
        return jsonify({'error': 'Debes indicar el nombre del chofer.'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Verificar que la estructura existe y está disponible
        cursor.execute(
            "SELECT id, estado FROM stock_estructuras_sofa WHERE id = %s FOR UPDATE;",
            (stock_id,)
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Estructura no encontrada.'}), 404
        if row[1] != 'disponible':
            return jsonify({'error': 'Esta estructura ya fue entregada o no está disponible.'}), 409

        cursor.execute("""
            UPDATE stock_estructuras_sofa
            SET estado = 'entregado', chofer_nombre = %s,
                fecha_entrega_chofer = NOW(),
                foto_entrega_url = %s,
                comentario_entrega = %s
            WHERE id = %s AND estado = 'disponible'
        """, (chofer_nombre, foto_entrega_url or None, comentario_entrega or None, stock_id))
        if cursor.rowcount != 1:
            return jsonify({'error': 'La estructura cambió de estado. Actualiza la lista.'}), 409
        conexion.commit()
        return jsonify({'exito': True, 'chofer_nombre': chofer_nombre}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion:
            conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# A10: Completar parte de estructura
@produccion_bp.route('/api/stock-estructuras/<int:stock_id>/completar-parte', methods=['PATCH'])
@requiere_login
def completar_parte_estructura(stock_id):
    if not usuario_puede_operar('ESTRUCTURAS_MUEBLES'):
        return jsonify({'error': 'Solo Estructuras o jefatura puede completar una parte.'}), 403
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()

        completado_por = get_usuario_actual().get('nombre', '')
        fecha_completado = request.form.get('fecha_completado', '').strip()
        comentario_parte = request.form.get('comentario_parte', '').strip()
        
        foto_url = None
        if 'foto' in request.files and request.files['foto'].filename:
            res = cloudinary_upload(request.files['foto'], folder='stock_estructuras')
            foto_url = res.get('secure_url')

        cursor.execute("""
            UPDATE stock_estructuras_sofa
            SET completado_por = %s,
                fecha_completado = %s::date,
                comentario_parte = %s,
                foto_completado_url = COALESCE(%s, foto_completado_url)
            WHERE id = %s
        """, (completado_por, fecha_completado if fecha_completado else None, comentario_parte, foto_url, stock_id))
        conexion.commit()
        return jsonify({'exito': True}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# A9: Toggle pago y edición de estructura de sofá ─────────────────────────────

@produccion_bp.route('/api/stock-estructuras/<int:stock_id>/pago', methods=['PATCH'])
@requiere_rol('Admin', 'Jefe_Taller')
def toggle_pago_estructura(stock_id):
    """Alterna el estado de pago de una estructura."""
    data   = request.get_json() or {}
    pagado = bool(data.get('pagado', False))

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute(
            "UPDATE stock_estructuras_sofa SET pagado = %s WHERE id = %s RETURNING id;",
            (pagado, stock_id)
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Estructura no encontrada.'}), 404

        conexion.commit()
        return jsonify({'exito': True, 'pagado': pagado}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion:
            conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/stock-estructuras/<int:stock_id>/editar', methods=['PATCH'])
@requiere_rol('Admin', 'Jefe_Taller')
def editar_estructura(stock_id):
    """
    Edita los campos de una estructura existente.
    Acepta multipart/form-data (puede incluir nueva foto) o JSON.
    Campos actualizables: nombre_modelo, modelo_base, ancho, profundidad, alto,
    medida_estandar, tipo_base, medida_base, medida_base_estandar, precio, cantidad.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        cursor.execute(
            "SELECT id FROM stock_estructuras_sofa WHERE id = %s;",
            (stock_id,)
        )
        if not cursor.fetchone():
            return jsonify({'error': 'Estructura no encontrada.'}), 404

        # Leer campos (soporta form-data o JSON)
        if request.content_type and 'multipart' in request.content_type:
            get = lambda k, d=None: request.form.get(k, d)
        else:
            body = request.get_json() or {}
            get = lambda k, d=None: body.get(k, d)

        nombre          = get('nombre_modelo')
        modelo_base     = get('modelo_base', '')
        ancho           = get('ancho')
        profundidad     = get('profundidad')
        alto            = get('alto')
        medida_estandar = get('medida_estandar')
        tipo_base       = get('tipo_base', '')
        medida_base     = get('medida_base')
        medida_base_est = get('medida_base_estandar')
        precio          = get('precio')
        cantidad        = get('cantidad')
        es_antiguo      = get('es_antiguo')
        medida_brazo    = get('medida_brazo')

        # Foto opcional
        foto_url = None
        if request.files and 'foto' in request.files and request.files['foto'].filename:
            res      = cloudinary_upload(request.files['foto'], folder='stock_estructuras')
            foto_url = res.get('secure_url')

        # Construir SET dinámico solo con campos enviados
        sets, vals = [], []
        if nombre          is not None: sets.append('nombre_modelo = %s');       vals.append(nombre)
        if modelo_base     is not None: sets.append('modelo_base = %s');         vals.append(modelo_base)
        if ancho           is not None: sets.append('ancho = %s');               vals.append(float(ancho) if ancho else 0)
        if profundidad     is not None: sets.append('profundidad = %s');         vals.append(float(profundidad) if profundidad else 0)
        if alto            is not None: sets.append('alto = %s');                vals.append(float(alto) if alto else 0)
        if medida_estandar is not None: sets.append('medida_estandar = %s');     vals.append(str(medida_estandar).lower() == 'true')
        if tipo_base       is not None: sets.append('tipo_base = %s');           vals.append(tipo_base)
        if medida_base     is not None: sets.append('medida_base = %s');         vals.append(float(medida_base) if medida_base else None)
        if medida_base_est is not None: sets.append('medida_base_estandar = %s');vals.append(str(medida_base_est).lower() == 'true')
        if precio          is not None: sets.append('precio = %s');              vals.append(float(precio) if precio else 0)
        if cantidad        is not None: sets.append('cantidad = %s');            vals.append(int(cantidad) if cantidad else 1)
        if es_antiguo      is not None: sets.append('es_antiguo = %s');          vals.append(str(es_antiguo).lower() == 'true')
        if medida_brazo    is not None: sets.append('medida_brazo = %s');        vals.append(float(medida_brazo) if medida_brazo else None)
        if foto_url        is not None: sets.append('foto_url = %s');            vals.append(foto_url)

        if not sets:
            return jsonify({'error': 'No se enviaron campos para actualizar.'}), 400

        vals.append(stock_id)
        cursor.execute(
            f"UPDATE stock_estructuras_sofa SET {', '.join(sets)} WHERE id = %s;",
            vals
        )
        conexion.commit()
        return jsonify({'exito': True}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion:
            conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)

# ══════════════════════════════════════════════════════════════════════════════
# DELETE: ELIMINAR ESTRUCTURA DE STOCK (solo Admin)
# ══════════════════════════════════════════════════════════════════════════════

@produccion_bp.route('/api/stock-estructuras/<int:stock_id>', methods=['DELETE'])
@requiere_rol('Admin')
def eliminar_stock_estructura(stock_id):
    """
    Elimina permanentemente un registro de stock_estructuras_sofa.
    Solo el Admin puede hacerlo.
    No elimina si la estructura ya fue usada en un ticket (estado='entregado' y ticket_id no nulo),
    para preservar la trazabilidad.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        # Verificar que existe y que no está vinculada a un ticket activo
        cursor.execute(
            "SELECT id, estado, ticket_id FROM stock_estructuras_sofa WHERE id = %s;",
            (stock_id,)
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Estructura no encontrada.'}), 404

        estado    = row[1]
        ticket_id = row[2]

        if estado == 'entregado' and ticket_id:
            return jsonify({
                'error': f'No se puede eliminar: esta estructura ya fue usada en el ticket #{ticket_id}. '
                         f'Eliminar rompería la trazabilidad de producción.'
            }), 400

        cursor.execute("DELETE FROM stock_estructuras_sofa WHERE id = %s;", (stock_id,))
        conexion.commit()
        return jsonify({'exito': True, 'eliminado': stock_id}), 200

    except Exception as e:
        if 'conexion' in locals() and conexion:
            conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ══════════════════════════════════════════════════════════════════════════════
# A10: HISTORIAL DE PAGOS A CARPINTEROS
# ══════════════════════════════════════════════════════════════════════════════

@produccion_bp.route('/api/stock-estructuras/cerrar-pago-semanal', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def cerrar_pago_semanal():
    """
    Cierra el pago semanal de un carpintero:
      1. Busca todas sus estructuras con estado='entregado' y pagado=FALSE
         dentro del rango [semana_inicio, semana_fin].
      2. Suma el monto total.
      3. Marca esas estructuras como pagado=TRUE.
      4. Inserta un registro en pagos_carpinteros.
    Body: { carpintero_nombre, semana_inicio (YYYY-MM-DD), semana_fin (YYYY-MM-DD), notas }
    """
    data           = request.get_json() or {}
    carpintero     = (data.get('carpintero_nombre') or '').strip()
    semana_inicio  = (data.get('semana_inicio') or '').strip()
    semana_fin     = (data.get('semana_fin') or '').strip()
    notas          = (data.get('notas') or '').strip()
    voucher_url    = (data.get('voucher_url') or '').strip() or None
    registrado_por = get_usuario_actual().get('nombre', 'Sistema')

    if not carpintero or not semana_inicio or not semana_fin:
        return jsonify({'error': 'carpintero_nombre, semana_inicio y semana_fin son obligatorios.'}), 400

    try:
        from datetime import date as _date
        _date.fromisoformat(semana_inicio)
        _date.fromisoformat(semana_fin)
    except ValueError:
        return jsonify({'error': 'Fechas inválidas. Usa formato YYYY-MM-DD.'}), 400

    try:
        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        # Buscar estructuras pendientes — primero por carpintero_nombre, luego fallback a chofer_nombre
        cursor.execute("""
            SELECT id, precio, nombre_modelo, cantidad, fecha_entrega_chofer
            FROM stock_estructuras_sofa
            WHERE estado = 'entregado'
              AND COALESCE(pagado, FALSE) = FALSE
              AND COALESCE(carpintero_nombre, chofer_nombre) = %s
              AND fecha_entrega_chofer::date BETWEEN %s AND %s
            ORDER BY fecha_entrega_chofer ASC
            FOR UPDATE;
        """, (carpintero, semana_inicio, semana_fin))
        filas = cursor.fetchall()

        if not filas:
            return jsonify({
                'error': f'No hay estructuras entregadas y pendientes de pago para '
                         f'"{carpintero}" entre {semana_inicio} y {semana_fin}.'
            }), 404

        ids_estructuras = [r[0] for r in filas]
        monto_total     = sum(float(r[1] or 0) * int(r[3] or 1) for r in filas)
        detalle         = [
            {
                'id':           r[0],
                'nombre_modelo': r[2] or 'Sin modelo',
                'cantidad':     r[3] or 1,
                'precio':       float(r[1] or 0),
                'subtotal':     float(r[1] or 0) * int(r[3] or 1),
                'fecha_entrega': r[4].strftime('%d/%m/%Y') if r[4] else '',
            }
            for r in filas
        ]

        # Marcar como pagadas
        placeholders = ','.join(['%s'] * len(ids_estructuras))
        cursor.execute(
            f"UPDATE stock_estructuras_sofa SET pagado = TRUE WHERE id IN ({placeholders});",
            ids_estructuras
        )

        # Insertar registro de pago
        cursor.execute("""
            INSERT INTO pagos_carpinteros
                (carpintero_nombre, semana_inicio, semana_fin, estructuras_ids,
                 cantidad_estructuras, monto_total, registrado_por, notas, voucher_url)
            VALUES (%s, %s, %s, %s::jsonb, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (
            carpintero, semana_inicio, semana_fin,
            json.dumps(ids_estructuras),
            len(ids_estructuras), monto_total,
            registrado_por, notas or None, voucher_url
        ))
        pago_id = cursor.fetchone()[0]
        conexion.commit()

        return jsonify({
            'exito':               True,
            'pago_id':             pago_id,
            'carpintero':          carpintero,
            'estructuras_pagadas': len(ids_estructuras),
            'monto_total':         monto_total,
            'detalle':             detalle,
        }), 201

    except Exception as e:
        if 'conexion' in locals() and conexion:
            conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/stock-estructuras/historial-pagos', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller')
def historial_pagos_carpinteros():
    """
    GET /api/stock-estructuras/historial-pagos?carpintero=X&semana=YYYY-MM-DD
    semana: cualquier fecha dentro de la semana (filtra semana_inicio <= fecha <= semana_fin).
    Sin filtros devuelve los últimos 100 registros.
    """
    carpintero = (request.args.get('carpintero') or '').strip()
    semana     = (request.args.get('semana') or '').strip()

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        conditions, params = [], []
        if carpintero:
            conditions.append("carpintero_nombre ILIKE %s")
            params.append(f'%{carpintero}%')
        if semana:
            conditions.append("semana_inicio <= %s AND semana_fin >= %s")
            params.extend([semana, semana])

        where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
        cursor.execute(f"""
            SELECT id, carpintero_nombre, semana_inicio, semana_fin,
                   estructuras_ids, cantidad_estructuras, monto_total,
                   registrado_por, fecha_pago, notas, voucher_url
            FROM pagos_carpinteros
            {where}
            ORDER BY fecha_pago DESC
            LIMIT 100;
        """, params)

        rows = cursor.fetchall()
        return jsonify([
            {
                'id':                    r[0],
                'carpintero':            r[1],
                'semana_inicio':         r[2].strftime('%d/%m/%Y') if r[2] else '',
                'semana_fin':            r[3].strftime('%d/%m/%Y') if r[3] else '',
                'estructuras_ids':       r[4] if r[4] else [],
                'cantidad_estructuras':  r[5],
                'monto_total':           float(r[6] or 0),
                'registrado_por':        r[7] or '',
                'fecha_pago':            r[8].strftime('%d/%m/%Y %H:%M') if r[8] else '',
                'notas':                 r[9] or '',
                'voucher_url':           r[10] or '',
            }
            for r in rows
        ]), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/stock-estructuras/carpinteros', methods=['GET'])
@requiere_login
def listar_carpinteros():
    """Devuelve nombres únicos de carpinteros que tienen estructuras registradas."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT DISTINCT COALESCE(carpintero_nombre, chofer_nombre) AS nombre
            FROM stock_estructuras_sofa
            WHERE COALESCE(carpintero_nombre, chofer_nombre) IS NOT NULL
              AND COALESCE(carpintero_nombre, chofer_nombre) != ''
            ORDER BY nombre;
        """)
        return jsonify([r[0] for r in cursor.fetchall()]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ══════════════════════════════════════════════════════════════════════════════
# B1: GASTOS DE LOGÍSTICA / FLETE
# ══════════════════════════════════════════════════════════════════════════════

@produccion_bp.route('/api/logistica/gasto', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller', 'Chofer')
def registrar_gasto_logistica():
    """
    Registra un gasto suelto de logística.
    Body: { concepto, monto, categoria, proveedor_nombre, fecha_gasto (YYYY-MM-DD), notas }
    """
    data             = request.get_json() or {}
    concepto         = (data.get('concepto') or '').strip()
    monto            = data.get('monto')
    categoria        = (data.get('categoria') or 'Otro').strip()
    proveedor_nombre = (data.get('proveedor_nombre') or '').strip() or None
    fecha_gasto      = (data.get('fecha_gasto') or '').strip()
    notas            = (data.get('notas') or '').strip() or None
    registrado_por = get_usuario_actual().get('nombre', 'Sistema')

    if not concepto:
        return jsonify({'error': 'El concepto es obligatorio.'}), 400
    if monto is None:
        return jsonify({'error': 'El monto es obligatorio.'}), 400
    if categoria not in ('Flete', 'Transporte', 'Compra directa', 'Otro'):
        return jsonify({'error': 'Categoría inválida. Usa: Flete, Transporte, Compra directa, Otro'}), 400
    try:
        monto = float(monto)
    except (TypeError, ValueError):
        return jsonify({'error': 'El monto debe ser un número.'}), 400
    if monto <= 0:
        return jsonify({'error': 'El monto debe ser mayor a cero.'}), 400

    if fecha_gasto:
        try:
            from datetime import date as _date
            _date.fromisoformat(fecha_gasto)
        except ValueError:
            return jsonify({'error': 'fecha_gasto inválida. Usa YYYY-MM-DD.'}), 400
    else:
        from datetime import date as _date
        fecha_gasto = _date.today().isoformat()

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            INSERT INTO gastos_logistica
                (concepto, monto, categoria, proveedor_nombre, fecha_gasto, registrado_por, notas)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (concepto, monto, categoria, proveedor_nombre, fecha_gasto, registrado_por, notas))
        gasto_id = cursor.fetchone()[0]
        conexion.commit()
        return jsonify({'exito': True, 'id': gasto_id}), 201

    except Exception as e:
        if 'conexion' in locals() and conexion:
            conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@produccion_bp.route('/api/logistica/gastos', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller', 'Chofer')
def listar_gastos_logistica():
    """
    GET /api/logistica/gastos?desde=YYYY-MM-DD&hasta=YYYY-MM-DD&categoria=Flete
    """
    desde_str = request.args.get('desde', '')
    hasta_str = request.args.get('hasta', '')
    categoria = (request.args.get('categoria') or '').strip()

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        conditions, params = [], []
        if desde_str:
            conditions.append("fecha_gasto >= %s"); params.append(desde_str)
        if hasta_str:
            conditions.append("fecha_gasto <= %s"); params.append(hasta_str)
        if categoria:
            conditions.append("categoria = %s"); params.append(categoria)

        where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
        cursor.execute(f"""
            SELECT id, concepto, monto, categoria, proveedor_nombre,
                   fecha_gasto, registrado_por, notas
            FROM gastos_logistica
            {where}
            ORDER BY fecha_gasto DESC, id DESC
            LIMIT 500;
        """, params)

        gastos = [
            {
                'id':               r[0],
                'concepto':         r[1],
                'monto':            float(r[2] or 0),
                'categoria':        r[3],
                'proveedor_nombre': r[4] or '',
                'fecha_gasto':      r[5].strftime('%d/%m/%Y') if r[5] else '',
                'fecha_gasto_iso':  r[5].isoformat() if r[5] else '',
                'registrado_por':   r[6] or '',
                'notas':            r[7] or '',
            }
            for r in cursor.fetchall()
        ]
        return jsonify({
            'gastos': gastos,
            'total':  sum(g['monto'] for g in gastos),
            'count':  len(gastos),
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)
