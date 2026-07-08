"""
routes_ventas.py — Módulos 2, 3 y 14: Ventas, seguimiento, producción y exportación.
Blueprint: ventas_bp  (sin prefijo de URL)

PUENTE INVENTARIO (Mayo 2026)
─────────────────────────────
Cuando se registra una venta con muebles que tienen es_stock=True y un
stock_producto_id / stock_pieza_id, este módulo ahora:
  1. Marca la unidad física como "Reservado"  en stock_productos / stock_piezas
  2. Deja registro en historial_inventario
  3. Vincula el id de la unidad al ítem de venta (items_venta.stock_producto_id)

Al cambiar estado a "Entregado":  Reservado → Vendido
Al anular la venta:               Reservado → Disponible

El carrito (carrito.js) debe enviar stock_producto_id o stock_pieza_id
dentro de cada elemento del array muebles.
"""

import io
import json
import traceback
import openpyxl
from datetime import datetime
from io import BytesIO, StringIO

import cloudinary.uploader
from flask import Blueprint, jsonify, request, send_file
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import get_db_connection, release_db_connection, limpiar_foto
from auth_middleware import requiere_login, requiere_rol
from notification_service import (
    diagnosticar_correo_prueba,
    enviar_resumen_operativo,
    enviar_correo_prueba,
    notificar_contrato_creado,
    notificar_estado_contrato,
    resumen_operativo,
)

ventas_bp = Blueprint('ventas', __name__)

# ─── Migración de esquema (se ejecuta una sola vez por proceso) ───────────────
_schema_listo = False

def _asegurar_columnas_inventario():
    """
    Añade stock_producto_id / stock_pieza_id a items_venta si aún no existen.
    Se ejecuta con autocommit=True para que un ALTER TABLE no envenene
    ninguna transacción abierta en caso de error.
    """
    global _schema_listo
    if _schema_listo:
        return
    conn = None
    cur  = None
    try:
        conn = get_db_connection()
        conn.autocommit = True          # DDL fuera de cualquier transacción
        cur  = conn.cursor()
        cur.execute("ALTER TABLE items_venta ADD COLUMN IF NOT EXISTS stock_producto_id INTEGER;")
        cur.execute("ALTER TABLE items_venta ADD COLUMN IF NOT EXISTS stock_pieza_id INTEGER;")
        cur.execute("ALTER TABLE items_venta ADD COLUMN IF NOT EXISTS es_stock BOOLEAN DEFAULT FALSE;")
        _schema_listo = True
    except Exception as e:
        print(f"⚠️ _asegurar_columnas_inventario: {e}")
        # Aunque falle el ALTER (columna ya existe en versiones <9.6), marcamos
        # igual como listo para no reintentar en cada request.
        _schema_listo = True
    finally:
        if cur:
            cur.close()
        if conn:
            conn.autocommit = False     # Restaurar antes de devolver al pool
            release_db_connection(conn)


# ─── Helper: historial de precios ────────────────────────────────────────────

def _crear_tabla_historial_precios(cursor):
    """Crea la tabla si todavía no existe (auto-migración segura)."""
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historial_precios (
            id               SERIAL PRIMARY KEY,
            venta_id         INTEGER       NOT NULL,
            codigo_venta     VARCHAR(50),
            precio_original  NUMERIC(10,2) NOT NULL,
            precio_nuevo     NUMERIC(10,2) NOT NULL,
            motivo           TEXT          NOT NULL,
            vendedor_id      INTEGER,
            vendedor_nombre  VARCHAR(100),
            admin_id         INTEGER,
            admin_nombre     VARCHAR(100),
            estado           VARCHAR(20)   DEFAULT 'Pendiente',
            notas_admin      TEXT,
            fecha_solicitud  TIMESTAMP     DEFAULT NOW(),
            fecha_resolucion TIMESTAMP
        );
    """)
    # Migración lazy: cada solicitud ahora apunta a UN producto específico
    # del contrato (item_id) y trae un tipo_cambio para saber qué hacer al
    # aprobar: cambiar el precio de ese item, cambiar su tela/material, o
    # agregar un producto nuevo al contrato (con su propio precio).
    cursor.execute("""
        ALTER TABLE historial_precios
            ADD COLUMN IF NOT EXISTS item_id         INTEGER,
            ADD COLUMN IF NOT EXISTS producto_nombre  VARCHAR(200),
            ADD COLUMN IF NOT EXISTS tipo_cambio      VARCHAR(30) DEFAULT 'precio',
            ADD COLUMN IF NOT EXISTS detalle_nuevo    TEXT;
    """)


def _recalcular_total_venta(cursor, venta_id):
    """Suma precio_unitario de todos los items del contrato y actualiza
    ventas.monto_total. Se llama siempre que se agrega/edita un item para
    que el total nunca quede desincronizado."""
    cursor.execute(
        "SELECT COALESCE(SUM(precio_unitario), 0) FROM items_venta WHERE venta_id = %s;",
        (venta_id,)
    )
    nuevo_total = cursor.fetchone()[0]
    cursor.execute("UPDATE ventas SET monto_total = %s WHERE id = %s;", (nuevo_total, venta_id))
    return float(nuevo_total)


# ─── Helper: puente con el inventario real ───────────────────────────────────

class UnidadNoDisponibleError(Exception):
    """
    Se lanza cuando la unidad física elegida para un ítem de venta ya no
    está 'Disponible' (otra venta ya la reservó/vendió). A propósito NO se
    captura dentro de _reservar_unidad — debe subir hasta guardar_venta()
    y abortar TODA la venta (rollback), en vez de guardar un contrato con
    una pieza fantasma sin unidad física real detrás.
    """
    pass


def _reservar_unidad(cursor, venta_id, codigo_venta,
                     stock_prod_id, stock_piez_id,
                     usuario_id, usuario_nombre, item_id):
    """
    Marca una unidad física como 'Reservado' y deja huella en
    historial_inventario. Vincula el id al ítem de venta.

    FIX (julio 2026 - doble venta): esta función tenía 3 problemas reales:
      1) El SELECT no bloqueaba la fila (sin FOR UPDATE) — dos ventas
         concurrentes podían leer el mismo estado 'Disponible' antes de que
         cualquiera de las dos confirmara, y ambas terminaban marcando la
         MISMA unidad física como reservada para dos contratos distintos.
      2) El estado leído nunca se validaba — si la unidad YA estaba
         'Reservado' o 'Vendido' (por otra venta, incluso sin condición de
         carrera), igual se volvía a marcar 'Reservado' y se vinculaba a
         este nuevo ítem, pisando silenciosamente la reserva/venta anterior.
      3) Si algo fallaba, el error se tragaba (SAVEPOINT + continuar) y la
         venta se guardaba igual, con un ítem sin ninguna unidad física
         real detrás — nadie se enteraba hasta que alguien iba a entregar
         un mueble que ya no existía en stock.

    Ahora: 'FOR UPDATE' bloquea la fila hasta que esta transacción termine
    (la segunda venta concurrente espera y luego ve el estado ya
    actualizado), y se valida que el estado sea 'Disponible' antes de
    reservar. Si la unidad ya no está disponible, se lanza
    UnidadNoDisponibleError — NO se captura aquí, sube hasta guardar_venta()
    y aborta toda la venta con un mensaje claro para el vendedor.
    """
    if stock_prod_id:
        tabla, col, tipo = 'stock_productos', 'stock_producto_id', 'producto'
        reg_id = stock_prod_id
    elif stock_piez_id:
        tabla, col, tipo = 'stock_piezas', 'stock_pieza_id', 'pieza'
        reg_id = stock_piez_id
    else:
        return

    cursor.execute("SAVEPOINT reservar_unidad")

    try:
        cursor.execute(
            f"SELECT estado, sede_id, codigo_barra FROM {tabla} WHERE id = %s FOR UPDATE",
            (reg_id,)
        )
        fila = cursor.fetchone()
        if not fila:
            cursor.execute("RELEASE SAVEPOINT reservar_unidad")
            return  # La unidad ya no existe; continuar sin romper la transacción

        estado_ant, sede_id, barcode = fila

        if estado_ant != 'Disponible':
            cursor.execute("ROLLBACK TO SAVEPOINT reservar_unidad")
            cursor.execute("RELEASE SAVEPOINT reservar_unidad")
            raise UnidadNoDisponibleError(
                f"La unidad {barcode or reg_id} ya no está disponible "
                f"(estado actual: {estado_ant}). Es posible que otra venta "
                f"la haya tomado justo ahora — elige otra unidad de stock "
                f"para este ítem y vuelve a intentar."
            )

        # Cambiar a Reservado
        cursor.execute(
            f"UPDATE {tabla} SET estado = 'Reservado', actualizado_en = NOW() WHERE id = %s",
            (reg_id,)
        )

        # Historial
        cursor.execute("""
            INSERT INTO historial_inventario
                (tipo_registro, registro_id, codigo_barra, tipo_evento,
                 sede_origen_id, sede_destino_id, estado_anterior, estado_nuevo,
                 usuario_id, usuario_nombre, venta_id, codigo_venta, notas)
            VALUES (%s,%s,%s,'Reserva',%s,NULL,%s,'Reservado',%s,%s,%s,%s,%s);
        """, (tipo, reg_id, barcode, sede_id, estado_ant,
              usuario_id, usuario_nombre, venta_id, codigo_venta,
              f"Reservado al registrar venta {codigo_venta}"))

        # Vincular al ítem de venta
        cursor.execute(f"UPDATE items_venta SET {col} = %s WHERE id = %s", (reg_id, item_id))

        cursor.execute("RELEASE SAVEPOINT reservar_unidad")

    except UnidadNoDisponibleError:
        raise  # Error de negocio: debe abortar toda la venta, no tragárselo

    except Exception as e:
        # Errores técnicos inesperados (no de negocio) — igual que antes,
        # no rompen la transacción padre por sí solos.
        print(f"⚠️  _reservar_unidad ERROR — tabla={tabla} reg_id={reg_id}")
        print(f"⚠️  Excepción: {type(e).__name__}: {e}")
        print(f"⚠️  Traceback:\n{traceback.format_exc()}")
        cursor.execute("ROLLBACK TO SAVEPOINT reservar_unidad")
        cursor.execute("RELEASE SAVEPOINT reservar_unidad")


def _liberar_unidades(cursor, venta_id, estado_destino, evento_nombre):
    """
    Itera los ítems de la venta y pasa las unidades reservadas al
    estado indicado (Disponible para anulaciones, Vendido para entregas).
    """
    for tabla, col, tipo in [
        ('stock_productos', 'stock_producto_id', 'producto'),
        ('stock_piezas',    'stock_pieza_id',    'pieza'),
    ]:
        try:
            cursor.execute(f"""
                SELECT iv.{col}, sp.estado, sp.sede_id, sp.codigo_barra
                FROM items_venta iv
                JOIN {tabla} sp ON sp.id = iv.{col}
                WHERE iv.venta_id = %s AND iv.{col} IS NOT NULL
                  AND sp.estado = 'Reservado'
            """, (venta_id,))
        except Exception:
            continue  # La columna puede no existir en BD más antiguas

        for reg_id, estado_ant, sede_id, barcode in cursor.fetchall():
            cursor.execute(
                f"UPDATE {tabla} SET estado = %s, actualizado_en = NOW() WHERE id = %s",
                (estado_destino, reg_id)
            )
            cursor.execute("""
                INSERT INTO historial_inventario
                    (tipo_registro, registro_id, codigo_barra, tipo_evento,
                     sede_origen_id, sede_destino_id, estado_anterior, estado_nuevo,
                     usuario_id, usuario_nombre, venta_id, codigo_venta, notas)
                VALUES (%s, %s, %s, %s, %s, NULL, %s, %s, NULL, 'Sistema', %s, NULL, %s)
            """, (tipo, reg_id, barcode, evento_nombre,
                  sede_id, estado_ant, estado_destino,
                  venta_id, f"Cambio automático — {evento_nombre}"))


def _sincronizar_tickets_con_estado_venta(cursor, venta_id, nuevo_estado_venta):
    """
    Cuando el Admin cambia manualmente el estado general de la venta
    (botón "Gestionar Venta" en Reportes/Ventas), los tickets de producción
    de cada pieza (taller, cojines, estructuras, despacho, etc.) NO se
    actualizan solos. Esta función los cierra para mantener coherencia.
    """
    mapa_estados = {
        'Listo':      'Terminado',
        'Despachado': 'Terminado',
        'Entregado':  'Terminado',
    }
    ticket_destino = mapa_estados.get(nuevo_estado_venta)
    if not ticket_destino:
        return

    # Áreas normales de producción (corte, tapicería, estructuras, cojines...)
    cursor.execute("""
        UPDATE tickets_produccion
        SET estado_ticket = %s,
            fecha_fin = COALESCE(fecha_fin, CURRENT_TIMESTAMP)
        WHERE item_id IN (SELECT id FROM items_venta WHERE venta_id = %s)
          AND area_trabajo != 'DESPACHO_CENTRAL'
          AND estado_ticket NOT IN ('Terminado', 'Cancelado', 'Recogido', 'Listo para Recojo')
    """, (ticket_destino, venta_id))

    # Si el pedido ya se Despachó/Entregó, el material externo se marca como recibido.
    cursor.execute("""
        UPDATE logistica_externa
        SET estado = 'Recibido'
        WHERE venta_id = %s AND estado NOT IN ('Recibido', 'Cancelado', 'Rechazado')
    """, (venta_id,))

# ==========================================
# VENTAS — REGISTRO Y LISTADO
# ==========================================

@ventas_bp.route('/api/ventas', methods=['GET', 'POST'])
@requiere_login
def guardar_venta():
    if request.method == 'GET':
        return listar_ventas()

    _asegurar_columnas_inventario()   # ← migración segura al primer uso

    datos = request.json
    _paso_actual = "inicio"
    try:
        print(f"\n{'='*60}")
        print(f"[VENTA] Iniciando registro — código: {datos.get('codigo')}")
        print(f"[VENTA] Payload recibido: {datos}")
        print(f"{'='*60}\n")

        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        lista_pagos_raw = datos.get('pagos', [])
        empresa_pago_resumen = lista_pagos_raw[0].get('empresa', '') if lista_pagos_raw else ''

        _paso_actual = "INSERT ventas"
        print(f"[PASO 1] {_paso_actual}")
        cursor.execute("""
            INSERT INTO ventas (
                codigo_venta, nombre_cliente, dni_cliente, celular_cliente,
                direccion_cliente, vendedor_id, fecha_emision, fecha_entrega,
                monto_total, vendedor_nombre, moneda, tipo_cambio, tipo_comprobante,
                empresa_ruc, empresa_pago, sede, nombre_empresa_cliente
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;
        """, (
            datos['codigo'],          datos['cliente'],
            datos.get('dni'),         datos.get('celular'),
            datos.get('direccion'),   datos.get('vendedor_id'),
            datos['fecha_emision'],   datos.get('fecha_entrega'),
            datos.get('monto_total', 0),
            datos.get('vendedor_nombre'),
            datos.get('moneda', 'PEN'),
            datos.get('tipo_cambio', 1.00),
            datos.get('tipo_comprobante', 'Boleta'),
            datos.get('empresa_ruc'),
            empresa_pago_resumen,
            datos.get('sede', 'Sede Central'),
            datos.get('empresa_cliente', 'Particular')
        ))
        venta_id = cursor.fetchone()[0]
        print(f"[PASO 1] OK — venta_id={venta_id}")

        _paso_actual = "crear_tabla_historial_precios"
        print(f"[PASO 2] {_paso_actual}")
        _crear_tabla_historial_precios(cursor)
        cursor.execute("""
            INSERT INTO historial_precios (
                venta_id, codigo_venta, precio_original, precio_nuevo,
                motivo, vendedor_id, vendedor_nombre, estado, fecha_solicitud, fecha_resolucion
            ) VALUES (%s, %s, %s, %s, 'Precio base de creación del contrato', %s, %s, 'Aprobado', NOW(), NOW());
        """, (
            venta_id,
            datos['codigo'],
            float(datos.get('monto_total', 0)),
            float(datos.get('monto_total', 0)),
            datos.get('vendedor_id'),
            datos.get('vendedor_nombre')
        ))
        print(f"[PASO 2] OK")

        # Pagos múltiples
        _paso_actual = "INSERT pagos"
        print(f"[PASO 3] {_paso_actual} — {len(lista_pagos_raw)} pago(s)")
        total_adelanto = 0
        for idx_p, p in enumerate(lista_pagos_raw):
            monto_bruto = p.get('monto', 0)
            total_adelanto += monto_bruto
            print(f"[PASO 3.{idx_p+1}] Pago: {p}")
            cursor.execute("""
                INSERT INTO pagos (
                    venta_id, tipo_pago, entidad, numero_operacion,
                    monto_bruto, comision_pos, monto_neto, empresa_destino, comprobante_url
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s);
            """, (
                venta_id,
                p.get('tipo'), p.get('entidad'), p.get('operacion'),
                monto_bruto, p.get('comision', 0), p.get('monto_neto', 0),
                p.get('empresa'), p.get('comprobante_url', 'Sin imagen')
            ))
        print(f"[PASO 3] OK — total_adelanto={total_adelanto}")

        if total_adelanto > 0:
            _paso_actual = "UPDATE ventas monto_adelanto"
            cursor.execute(
                "UPDATE ventas SET monto_adelanto = %s WHERE id = %s",
                (total_adelanto, venta_id)
            )
            print(f"[PASO 3b] UPDATE monto_adelanto OK")

        # ── Motor Make-vs-Buy ─────────────────────────────────────────────────
        # FIX (julio 2026): este diccionario existía desde antes pero nunca se
        # usaba en ningún INSERT — por eso en "Logística Externa / Requerimientos"
        # solo se veía el nombre puro del material (ej. "Boucle - #14") sin
        # indicar para qué pieza del mueble era (¿cojín?, ¿butaca?, ¿base?).
        # Ahora se completa para TODAS las claves de componentes (antes solo
        # cubría telas y cojines) y se usa más abajo al armar insumo_nombre.
        SUFIJO_TELA = {
            'tela':              'Tela Principal (Sofá/Silla)',
            'tela-silla':        'Tela para Silla Comedor',
            'tela-butaca':       'Tela para Butaca',
            'tela-cojin':        'Tela para Cojines',
            'cojin-entero':      'Cojín Entero',
            'cojin-diseno':      'Cojín c/Diseño',
            'base':              'Base',
            'tablero':           'Tablero',
            'tablero-centro':    'Tablero Mesa de Centro',
            'silla':             'Sillería',
            'estructura-butaca': 'Estructura de Butaca',
            'base-mesa':         'Base de Mesa',
            'base-centro':       'Base Mesa de Centro',
        }
        mapeo_erp = {
            'tela':              ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'tela-cojin':        ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'tela-silla':        ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'tela-butaca':       ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'cojin-entero':      ('maestro_telas',         'CORTE_Y_CONTROL_TELAS'),
            'cojin-diseno':      ('maestro_disenos_cojin', 'ARMADO_COJINES'),
            'base':              ('maestro_bases',          'PREPARACION_PATAS_ZOCALO'),
            'tablero':           ('maestro_tableros',       'TABLEROS_Y_PIEDRAS'),
            'tablero-centro':    ('maestro_tableros',       'TABLEROS_Y_PIEDRAS'),
            'silla':             ('maestro_sillas',         'ESTRUCTURAS_SILLAS'),
            'estructura-butaca': ('maestro_butacas',        'ESTRUCTURAS_SILLAS'),
            'base-mesa':         ('maestro_bases_comedor',  'TABLEROS_Y_PIEDRAS'),
            'base-centro':       ('maestro_bases_comedor',  'TABLEROS_Y_PIEDRAS'),
        }

        _paso_actual = "loop muebles"
        print(f"[PASO 4] Procesando {len(datos.get('muebles', []))} mueble(s)")
        for idx_m, m in enumerate(datos['muebles']):
            _paso_actual = f"mueble[{idx_m}] INSERT items_venta"
            print(f"[PASO 4.{idx_m+1}] Mueble: {m}")
            # Limpiar foto_url: si viene con prefijo /uploads/ seguido de una URL absoluta, extraer solo la URL real
            foto_url_raw = m.get('foto', '') or ''
            if '/uploads/https://' in foto_url_raw:
                foto_url_raw = foto_url_raw[foto_url_raw.index('/uploads/https://') + len('/uploads/'):]
            elif '/uploads/http://' in foto_url_raw:
                foto_url_raw = foto_url_raw[foto_url_raw.index('/uploads/http://') + len('/uploads/'):]

            cursor.execute("""
                INSERT INTO items_venta (venta_id, producto, color_tela, foto_url, precio_unitario, es_stock)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id;
            """, (venta_id, m.get('tipo'), m.get('tela'), foto_url_raw, m.get('precio'), m.get('es_stock', False)))
            item_id = cursor.fetchone()[0]
            print(f"[PASO 4.{idx_m+1}] item_id={item_id}")

            componentes = m.get('componentes', {})
            areas_internas_creadas = set()

            nombre_lower = m['tipo'].lower()
            # categoria viene del catálogo (Sofá, Butaca, Silla, Sillón, etc.)
            categoria_lower = (m.get('categoria') or '').lower()
            area_estructura = None
            # FIX (julio 2026): "Curvo", "En U" y "Juego de Sala (3-2-1)" — 3 de
            # los 7 modelos base del sistema — no contenían NINGUNA de estas
            # palabras clave en su nombre. Para esos 3 modelos, la detección
            # dependía al 100% de que 'categoria' llegara como 'Sofá' desde el
            # frontend, sin ninguna red de seguridad si por lo que sea no
            # llegaba (JS cacheado viejo, deploy a medias, etc.). Se agregan
            # 'curvo', 'juego de sala' y 'en u' como keywords adicionales.
            if any(p in nombre_lower for p in ['sofá', 'sofa', 'seccional', 'modular', 'multi', 'curvado', 'curvo', 'plantilla', 'juego de sala', 'en u']) \
               or any(p in categoria_lower for p in ['sofá', 'sofa', 'seccional', 'modular', 'sillón', 'sillon']):
                area_estructura = 'ESTRUCTURAS_MUEBLES'
            elif any(p in nombre_lower for p in ['silla', 'butaca', 'sitial', 'puff']) \
                 or any(p in categoria_lower for p in ['silla', 'butaca', 'sitial', 'puff']):
                area_estructura = 'ESTRUCTURAS_SILLAS'

            # FIX (julio 2026 - v2): último respaldo para modelos de Sofá
            # personalizados (creados con la "tuerquita ⚙️ Gestionar" del
            # configurador). Si el mueble trae un componente bajo la clave
            # 'tela' (a diferencia de 'tela-silla' / 'tela-butaca', esta
            # clave la usa ÚNICAMENTE el configurador de Sofá — ver
            # confirmarPersonalizadoSofa en catalogo.js), es un sofá con
            # certeza, sin importar si el nombre es libre ("Cuervo partido
            # en 2", etc.) o si 'categoria' no llegó bien desde el frontend
            # (caché de JS viejo, deploy a medias). Evita que un mueble se
            # quede sin ticket de Estructura/Tapicería.
            if not area_estructura and componentes.get('tela'):
                area_estructura = 'ESTRUCTURAS_MUEBLES'

            # FIX (julio 2026): log de diagnóstico. Si un mueble vuelve a
            # quedarse sin ticket de Estructura/Tapicería, este print en los
            # logs de Render muestra exactamente qué texto llegó como 'tipo'
            # y 'categoria' para ese mueble puntual — así no hay que adivinar
            # si el problema fue el nombre del modelo o que 'categoria' llegó
            # vacía desde el frontend.
            print(f"[DEBUG ÁREA ESTRUCTURA] mueble[{idx_m}] tipo={m.get('tipo')!r} "
                  f"categoria={m.get('categoria')!r} -> area_estructura={area_estructura!r}")

            # Verificar si el componente principal (silla/butaca) es externo ANTES de crear tickets
            # Si es externo va solo a logística, no al taller
            componente_principal_externo = False
            sku_silla = None  # queda disponible más abajo (ej. para el ticket de Corte y Tela)
            if area_estructura == 'ESTRUCTURAS_SILLAS':
                # FIX (julio 2026): el configurador de Butaca Personalizada
                # (confirmarButaca() en catalogo.js) manda el SKU de la
                # estructura bajo la clave 'estructura-butaca' (igual que
                # mapeo_erp más abajo), NO bajo 'butaca' — esa clave nunca
                # existió en ningún payload del frontend. Con el nombre
                # equivocado, sku_silla siempre salía None para butacas
                # personalizadas, este chequeo de "¿es externa?" se saltaba
                # por completo, y el sistema creaba de todos modos los
                # tickets de ESTRUCTURAS_SILLAS/TAPICERIA_SILLAS aunque el
                # maestro dijera origen_produccion = 'Externo'.
                sku_silla = componentes.get('silla') or componentes.get('estructura-butaca')
                if sku_silla:
                    for tabla_ext in ('maestro_sillas', 'maestro_butacas'):
                        try:
                            cursor.execute(
                                f"SELECT origen_produccion FROM {tabla_ext} WHERE sku = %s LIMIT 1",
                                (sku_silla,)
                            )
                            res_origen = cursor.fetchone()
                            if res_origen and res_origen[0] == 'Externo':
                                componente_principal_externo = True
                                break
                        except Exception:
                            pass

            # SKU de la tela principal (si el mueble lleva tela) — se usa para
            # que el ticket de Tapicería y el de Corte de Telas muestren la
            # foto de la tela en la Ficha Técnica (galería de /fichatecnica-skus).
            sku_tela_principal = (
                componentes.get('tela') or componentes.get('tela-silla') or componentes.get('tela-butaca')
            )

            if area_estructura and not componente_principal_externo:
                _paso_actual = f"mueble[{idx_m}] INSERT ticket {area_estructura}"
                print(f"[PASO 4.{idx_m+1}.a] Ticket area_estructura={area_estructura}")
                # Si es una silla/butaca armada internamente, dejamos su SKU en el
                # ticket para que el área de Estructuras vea también la foto de la silla.
                override_estructura = f"SKU: {sku_silla}" if (area_estructura == 'ESTRUCTURAS_SILLAS' and sku_silla) else None
                cursor.execute("""
                    INSERT INTO tickets_produccion (item_id, area_trabajo, estado_ticket, etapa, ticket_details_override)
                    VALUES (%s, %s, 'Pendiente', 1, %s)
                """, (item_id, area_estructura, override_estructura))
                areas_internas_creadas.add(area_estructura)

                area_tap = 'TAPICERIA_SOFAS' if area_estructura == 'ESTRUCTURAS_MUEBLES' else 'TAPICERIA_SILLAS'
                override_tap = f"SKU: {sku_tela_principal}" if sku_tela_principal else None
                cursor.execute("""
                    INSERT INTO tickets_produccion (item_id, area_trabajo, estado_ticket, etapa, ticket_details_override)
                    VALUES (%s, %s, 'Bloqueado', 2, %s)
                """, (item_id, area_tap, override_tap))
                areas_internas_creadas.add(area_tap)

            # ── Acumular detalle de cojines para ticket_details_override ─────
            # Leemos los SKUs de cojines del componente antes de procesar el loop
            detalle_cojines_armado  = []   # para ARMADO_COJINES
            detalle_cojines_corte   = []   # para CORTE_Y_CONTROL_TELAS

            for key_c, sku_c in componentes.items():
                if not sku_c:
                    continue
                if key_c == 'cojin-entero':
                    try:
                        cursor.execute("SAVEPOINT sp_cojin_entero")
                        cursor.execute(
                            "SELECT coleccion, color FROM maestro_telas WHERE sku = %s", (sku_c,)
                        )
                        row_te = cursor.fetchone()
                        nombre_tela = f"{row_te[0]} - {row_te[1]}" if row_te else sku_c
                        cursor.execute("RELEASE SAVEPOINT sp_cojin_entero")
                    except Exception:
                        cursor.execute("ROLLBACK TO SAVEPOINT sp_cojin_entero")
                        cursor.execute("RELEASE SAVEPOINT sp_cojin_entero")
                        nombre_tela = sku_c
                    linea = f"Cojín Entero → [{sku_c}] {nombre_tela}"
                    detalle_cojines_armado.append(linea)
                    detalle_cojines_corte.append(linea)

                elif key_c == 'cojin-diseno':
                    try:
                        cursor.execute("SAVEPOINT sp_cojin_diseno")
                        cursor.execute(
                            "SELECT nombre_diseno, tipo_tela FROM maestro_disenos_cojin WHERE sku = %s", (sku_c,)
                        )
                        row_dc = cursor.fetchone()
                        nombre_dis = row_dc[0] if row_dc else sku_c
                        tipo_tela  = row_dc[1] if row_dc else ''
                        cursor.execute("RELEASE SAVEPOINT sp_cojin_diseno")
                    except Exception:
                        cursor.execute("ROLLBACK TO SAVEPOINT sp_cojin_diseno")
                        cursor.execute("RELEASE SAVEPOINT sp_cojin_diseno")
                        nombre_dis, tipo_tela = sku_c, ''
                    linea = f"Cojín c/Diseño → [{sku_c}] {nombre_dis}" + (f" ({tipo_tela})" if tipo_tela else "")
                    detalle_cojines_armado.append(linea)
                    detalle_cojines_corte.append(linea)

            # ─────────────────────────────────────────────────────────────────

            for key, sku in componentes.items():
                if not sku or key not in mapeo_erp:
                    continue
                tabla, area_destino = mapeo_erp[key]
                _paso_actual = f"mueble[{idx_m}] componente key={key} sku={sku} tabla={tabla}"
                print(f"[PASO 4.{idx_m+1}.c] Componente key={key}, sku={sku}, tabla={tabla}, area={area_destino}")

                if key == 'silla':
                    cursor.execute("SELECT material, modelo, origen_produccion, proveedor_id FROM maestro_sillas WHERE sku = %s", (sku,))
                    res_silla = cursor.fetchone()
                    if res_silla:
                        mat = (res_silla[0] or '').lower()
                        nombre_insumo_silla = res_silla[1] or sku
                        prov_id_silla = res_silla[3]
                        if any(w in mat for w in ['metal', 'acero', 'fierro', 'aluminio']) or res_silla[2] == 'Externo':
                            # FIX (julio 2026): antes no se guardaba cuántas sillas
                            # lleva el comedor — la columna cantidad quedaba NULL y
                            # la vista de Logística Externa mostraba siempre "1"
                            # sin importar si eran 6, 8 o 10 sillas. El frontend
                            # ahora manda componentes.cantidad_silla; se valida y,
                            # si por lo que sea no llega, se asume 1 (nunca 0/None).
                            try:
                                cantidad_silla = int(componentes.get('cantidad_silla') or 1)
                                if cantidad_silla < 1:
                                    cantidad_silla = 1
                            except (TypeError, ValueError):
                                cantidad_silla = 1

                            cursor.execute(
                                "ALTER TABLE logistica_externa ADD COLUMN IF NOT EXISTS cantidad INTEGER;"
                            )
                            # FIX (julio 2026): se agrega el rol del componente
                            # (ej. "Sillería") entre paréntesis al nombre del
                            # insumo, para que en "Requerimientos" se sepa para
                            # qué pieza del pedido es esta compra externa —
                            # antes solo se veía el nombre del modelo de silla,
                            # sin contexto de qué parte del mueble representaba.
                            rol_silla = SUFIJO_TELA.get(key, '')
                            nombre_insumo_silla_ctx = (
                                f"{nombre_insumo_silla} ({rol_silla})" if rol_silla else nombre_insumo_silla
                            )
                            # FIX (julio 2026 - v3): mismo fix que el resto de
                            # componentes — clasificar categoria_insumo desde
                            # la creación (aquí siempre 'ESTRUCTURAL', las
                            # sillas nunca llegan por esta rama como tela).
                            cursor.execute(
                                "ALTER TABLE logistica_externa "
                                "ADD COLUMN IF NOT EXISTS categoria_insumo VARCHAR(30) DEFAULT 'OTRO';"
                            )
                            cursor.execute(
                                """INSERT INTO logistica_externa
                                       (venta_id, item_id, insumo_nombre, sku, estado, tipo_gestion,
                                        proveedor_id, cantidad, categoria_insumo)
                                   VALUES (%s, %s, %s, %s, 'Pendiente', 'Externo', %s, %s, 'ESTRUCTURAL')""",
                                (venta_id, item_id, nombre_insumo_silla_ctx, sku, prov_id_silla, cantidad_silla)
                            )
                            continue

                try:
                    cursor.execute(f"SELECT origen_produccion FROM {tabla} WHERE sku = %s", (sku,))
                    res = cursor.fetchone()
                except Exception:
                    res = None

                if res and res[0] == 'Interno':
                    if area_destino not in areas_internas_creadas:
                        # Determinar detalle a guardar en el ticket.
                        # IMPORTANTE: todas las líneas "SKU: XXXX" que agreguemos aquí
                        # son las que luego /api/taller/fichatecnica-skus usa para
                        # traer la foto del material y mostrarla en la Ficha Técnica
                        # del ticket (galería por área) — ver modules/taller/ficha_asignaciones.js:verFichaTaller.
                        partes_override = []
                        if area_destino == 'ARMADO_COJINES' and detalle_cojines_armado:
                            partes_override.append("COJINERÍA:\n" + "\n".join(detalle_cojines_armado))
                        elif area_destino == 'CORTE_Y_CONTROL_TELAS':
                            if sku_tela_principal:
                                partes_override.append(f"SKU: {sku_tela_principal}")
                            # Si el mueble es una silla/butaca armada internamente, se
                            # incluye también su SKU para que quien corta la tela vea
                            # de una vez la silla + la tela (y calcule cuánta tela usar).
                            if area_estructura == 'ESTRUCTURAS_SILLAS' and sku_silla:
                                partes_override.append(f"SKU: {sku_silla}")
                            if detalle_cojines_corte:
                                partes_override.append("COJINES:\n" + "\n".join(detalle_cojines_corte))
                        else:
                            partes_override.append(f"SKU: {sku}")

                        override_texto = "\n".join(partes_override) if partes_override else None

                        cursor.execute("""
                            INSERT INTO tickets_produccion (item_id, area_trabajo, estado_ticket, etapa, ticket_details_override)
                            VALUES (%s, %s, 'Pendiente', 1, %s)
                        """, (item_id, area_destino, override_texto))
                        areas_internas_creadas.add(area_destino)
                    else:
                        # Ya existe un ticket para esta área (ej: 'tablero' y 'base-mesa'
                        # ambos van a TABLEROS_Y_PIEDRAS) — no duplicamos el ticket,
                        # solo le sumamos el SKU de este componente para que también
                        # aparezca su foto en la Ficha Técnica.
                        try:
                            cursor.execute(
                                "SELECT id, ticket_details_override FROM tickets_produccion "
                                "WHERE item_id = %s AND area_trabajo = %s LIMIT 1",
                                (item_id, area_destino)
                            )
                            row_tk = cursor.fetchone()
                            if row_tk:
                                texto_previo = row_tk[1] or ''
                                nuevo_texto = (texto_previo + "\n" if texto_previo else "") + f"SKU: {sku}"
                                cursor.execute(
                                    "UPDATE tickets_produccion SET ticket_details_override = %s WHERE id = %s",
                                    (nuevo_texto, row_tk[0])
                                )
                        except Exception:
                            pass
                elif res is None or res[0] == 'Externo':
                    # Obtener nombre real del insumo desde la tabla correspondiente
                    try:
                        cursor.execute("SAVEPOINT sp_nombre_insumo")
                        col_nombre = {
                            'maestro_telas':         "CONCAT(coleccion, ' - ', color)",
                            'maestro_bases':         'modelo',
                            'maestro_bases_comedor': 'modelo',
                            'maestro_butacas':       'modelo',
                            'maestro_sillas':        'modelo',
                            'maestro_tableros':      'nombre_modelo',
                            'maestro_disenos_cojin': 'nombre_diseno',
                        }.get(tabla, 'sku')
                        cursor.execute(f"SELECT {col_nombre} FROM {tabla} WHERE sku = %s", (sku,))
                        row_nombre = cursor.fetchone()
                        nombre_insumo_real = row_nombre[0] if row_nombre and row_nombre[0] else sku
                        cursor.execute("RELEASE SAVEPOINT sp_nombre_insumo")
                    except Exception:
                        cursor.execute("ROLLBACK TO SAVEPOINT sp_nombre_insumo")
                        cursor.execute("RELEASE SAVEPOINT sp_nombre_insumo")
                        nombre_insumo_real = sku
    
                    # Intentar obtener proveedor_id desde la tabla del maestro correspondiente
                    prov_id_logistica = None
                    col_prov = 'proveedor_id'  # todas las tablas del maestro ahora tienen esta columna
                    try:
                        cursor.execute(f"SELECT proveedor_id FROM {tabla} WHERE sku = %s", (sku,))
                        row_prov = cursor.fetchone()
                        prov_id_logistica = row_prov[0] if row_prov else None
                    except Exception:
                        pass
    
                    # --- Check for existing logistica entry to consolidate ---
                    # FIX (julio 2026 - v2): antes, al consolidar, el item_id
                    # del segundo/tercer componente (ej. la silla que comparte
                    # tela con el sofá) se perdía por completo — la fila de
                    # logistica_externa solo quedaba asociada al primer
                    # item_id. Eso dejaba a _tela_pendiente_para_item()
                    # (routes_produccion.py) sin forma de saber que ese
                    # segundo ítem también estaba esperando esta tela. Ahora
                    # se guarda en item_ids_extra (lista separada por comas)
                    # para que el semáforo revise también esos item_id.
                    cursor.execute(
                        "ALTER TABLE logistica_externa "
                        "ADD COLUMN IF NOT EXISTS item_ids_extra TEXT DEFAULT NULL;"
                    )
                    cursor.execute("""
                        SELECT id, insumo_nombre, item_id, COALESCE(item_ids_extra, '')
                        FROM logistica_externa
                        WHERE venta_id = %s AND sku = %s AND proveedor_id IS NOT DISTINCT FROM %s
                        LIMIT 1;
                    """, (venta_id, sku, prov_id_logistica))

                    existing_log_row = cursor.fetchone()
                    rol_componente = SUFIJO_TELA.get(key, '')

                    if existing_log_row:
                        existing_log_id, existing_insumo_nombre, existing_item_id, existing_extra = existing_log_row
                        # Append new role to name if not already present
                        if rol_componente and rol_componente.lower() not in existing_insumo_nombre.lower():
                            if '(' in existing_insumo_nombre and existing_insumo_nombre.endswith(')'):
                                # Append inside parenthesis
                                nuevo_nombre = f"{existing_insumo_nombre[:-1]}, {rol_componente})"
                            else:
                                # Add parenthesis
                                nuevo_nombre = f"{existing_insumo_nombre} ({rol_componente})"
                            
                            cursor.execute(
                                "UPDATE logistica_externa SET insumo_nombre = %s WHERE id = %s",
                                (nuevo_nombre, existing_log_id)
                            )

                        # Registrar este item_id como "invitado" de la fila
                        # consolidada, si todavía no está anotado.
                        ids_extra = [x for x in existing_extra.split(',') if x.strip()]
                        item_id_str = str(item_id)
                        if item_id_str != str(existing_item_id) and item_id_str not in ids_extra:
                            ids_extra.append(item_id_str)
                            cursor.execute(
                                "UPDATE logistica_externa SET item_ids_extra = %s WHERE id = %s",
                                (','.join(ids_extra), existing_log_id)
                            )

                        continue # Skip insertion
    
                    nombre_insumo_con_rol = f"{nombre_insumo_real} ({rol_componente})" if rol_componente else nombre_insumo_real
    
                    # FIX (julio 2026 - v3): antes categoria_insumo se quedaba
                    # NULL/'OTRO' hasta que alguien subía el comprobante de pago
                    # (ver registrar_pago_proveedor en routes_produccion.py).
                    # Mientras tanto, la cola de Corte y Tela, el semáforo de
                    # "estructura lista para recojo" y el progreso de Orden de
                    # Pedido no podían saber que esta fila era tela. Ahora se
                    # clasifica de una vez según la tabla maestro de origen.
                    categoria_insumo_ins = 'TELA' if tabla == 'maestro_telas' else 'ESTRUCTURAL'

                    cursor.execute(
                        "ALTER TABLE logistica_externa "
                        "ADD COLUMN IF NOT EXISTS categoria_insumo VARCHAR(30) DEFAULT 'OTRO';"
                    )
                    cursor.execute(
                        """INSERT INTO logistica_externa
                               (venta_id, item_id, insumo_nombre, sku, estado, proveedor_id,
                                tipo_gestion, categoria_insumo)
                           VALUES (%s, %s, %s, %s, 'Pendiente', %s, 'Externo', %s)""",
                        (venta_id, item_id, nombre_insumo_con_rol, sku, prov_id_logistica, categoria_insumo_ins)
                    )

            # Ticket de Despacho
            # Si es stock ya fabricado → Pendiente; si es producción → Bloqueado
            estado_despacho = 'Pendiente' if m.get('es_stock') else 'Bloqueado'
            _paso_actual = f"mueble[{idx_m}] INSERT ticket DESPACHO_CENTRAL"
            print(f"[PASO 4.{idx_m+1}.d] Ticket DESPACHO_CENTRAL estado={estado_despacho}")
            cursor.execute("""
                INSERT INTO tickets_produccion (item_id, area_trabajo, estado_ticket, etapa)
                VALUES (%s, 'DESPACHO_CENTRAL', %s, 99)
            """, (item_id, estado_despacho))

            # Descuento genérico de stock (contador en catálogo)
            if m.get('es_stock') and m.get('catalogo_id'):
                _paso_actual = f"mueble[{idx_m}] UPDATE catalogo_productos stock"
                print(f"[PASO 4.{idx_m+1}.e] UPDATE stock catalogo_id={m['catalogo_id']}")
                cursor.execute("""
                    UPDATE catalogo_productos
                    SET stock_cantidad = GREATEST(0, stock_cantidad - 1),
                        en_stock = CASE WHEN stock_cantidad - 1 <= 0 THEN false ELSE en_stock END
                    WHERE id = %s AND en_stock = true
                """, (m['catalogo_id'],))

            # ── PUENTE INVENTARIO REAL ────────────────────────────────────────
            _paso_actual = f"mueble[{idx_m}] _reservar_unidad stock_producto_id={m.get('stock_producto_id')} stock_pieza_id={m.get('stock_pieza_id')}"
            print(f"[PASO 4.{idx_m+1}.f] _reservar_unidad: {_paso_actual}")
            _reservar_unidad(
                cursor, venta_id, datos['codigo'],
                m.get('stock_producto_id'),
                m.get('stock_pieza_id'),
                datos.get('vendedor_id'),
                datos.get('vendedor_nombre', ''),
                item_id
            )
            print(f"[PASO 4.{idx_m+1}.f] _reservar_unidad OK")
            # ─────────────────────────────────────────────────────────────────

            # Descuento de insumos por receta (solo productos fabricados)
            _paso_actual = f"mueble[{idx_m}] descuento insumos receta"
            cursor.execute(
                "SELECT cp.id FROM catalogo_productos cp WHERE cp.nombre_modelo = %s LIMIT 1;",
                (m['tipo'],)
            )
            prod_row = cursor.fetchone()
            if prod_row and not m.get('es_stock'):
                print(f"[PASO 4.{idx_m+1}.g] UPDATE insumos receta prod_id={prod_row[0]}")
                cursor.execute("""
                    UPDATE inventario_insumos i
                    SET cantidad_actual = GREATEST(0, i.cantidad_actual - r.cantidad_necesaria)
                    FROM recetas_muebles r
                    WHERE r.insumo_id = i.id AND r.producto_id = %s;
                """, (prod_row[0],))

        print(f"[PASO 5] COMMIT")
        conexion.commit()

        try:
            notif = notificar_contrato_creado(
                cursor,
                venta_id,
                datos,
                cantidad_items=len(datos.get('muebles') or []),
            )
            print(f"[NOTIF] Contrato {datos.get('codigo')}: {notif}")
        except Exception as ex_notif:
            print(f"[NOTIF] No se pudo notificar contrato {datos.get('codigo')}: {ex_notif}")

        print(f"[VENTA] ✅ Venta registrada exitosamente — venta_id={venta_id}\n")
        return jsonify({"mensaje": "Venta procesada exitosamente", "id": venta_id}), 201

    except UnidadNoDisponibleError as ex_stock:
        # FIX (julio 2026 - doble venta): antes _reservar_unidad se tragaba
        # este caso en silencio y la venta se guardaba igual con un ítem
        # sin unidad física real. Ahora aborta toda la venta (ya se hizo
        # rollback dentro de _reservar_unidad vía SAVEPOINT, pero
        # confirmamos aquí el rollback de la transacción completa) y le
        # devuelve al vendedor un mensaje claro y accionable, no un 500
        # genérico con traceback técnico.
        if 'conexion' in locals() and conexion: conexion.rollback()
        print(f"\n[VENTA] ⚠️ Unidad no disponible — venta abortada: {ex_stock}\n")
        return jsonify({"error": str(ex_stock)}), 409

    except Exception as ex:
        if 'conexion' in locals() and conexion: conexion.rollback()
        tb = traceback.format_exc()
        print(f"\n{'!'*60}")
        print(f"[ERROR] PASO QUE FALLÓ: {_paso_actual}")
        print(f"[ERROR] EXCEPCIÓN: {type(ex).__name__}: {str(ex)}")
        print(f"[ERROR] TRACEBACK COMPLETO:\n{tb}")
        print(f"{'!'*60}\n")
        error_msg = str(ex)
        if "llave duplicada" in error_msg.lower() or "uniqueviolation" in error_msg.lower() or "duplicate key" in error_msg.lower():
            return jsonify({"error": f"El N° de Contrato ({datos.get('codigo')}) ya está registrado en el sistema. Por favor, utiliza un código diferente."}), 400
        return jsonify({
            "error": error_msg,
            "paso": _paso_actual,
            "detalle": tb
        }), 500
    finally:
        if 'conexion' in locals() and conexion:
            if 'cursor' in locals() and cursor: cursor.close()
            release_db_connection(conexion)


def listar_ventas():
    """Devuelve todas las ventas sumando los pagos múltiples."""
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        q      = (request.args.get('q') or '').strip()
        estado = (request.args.get('estado') or '').strip()
        desde  = (request.args.get('desde') or '').strip()
        hasta  = (request.args.get('hasta') or '').strip()
        try:
            limit = min(max(int(request.args.get('limit', 500)), 1), 1000)
        except (TypeError, ValueError):
            limit = 500

        condiciones = []
        params = []
        if q:
            like = f"%{q.lower()}%"
            condiciones.append("""(
                LOWER(v.codigo_venta) LIKE %s
                OR LOWER(v.nombre_cliente) LIKE %s
                OR LOWER(COALESCE(v.vendedor_nombre, '')) LIKE %s
                OR LOWER(COALESCE(itm.productos, '')) LIKE %s
            )""")
            params.extend([like, like, like, like])
        if estado:
            condiciones.append("COALESCE(v.estado_general, 'Pendiente') = %s")
            params.append(estado)
        if desde:
            condiciones.append("v.fecha_emision::date >= %s")
            params.append(desde)
        if hasta:
            condiciones.append("v.fecha_emision::date <= %s")
            params.append(hasta)

        where_sql = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""
        params.append(limit)

        cursor.execute(f"""
            SELECT
                v.id, v.codigo_venta, v.nombre_cliente, v.monto_total,
                COALESCE(pg.total_pagado, 0) AS monto_adelanto,
                COALESCE(v.monto_total, 0) - COALESCE(pg.total_pagado, 0) AS saldo,
                COALESCE(v.estado_general, 'Pendiente') AS estado,
                v.fecha_entrega, v.vendedor_nombre,
                COALESCE(itm.productos, '') AS productos,
                v.sede, v.fecha_emision
            FROM ventas v
            LEFT JOIN (
                SELECT venta_id, SUM(monto_bruto) AS total_pagado
                FROM pagos GROUP BY venta_id
            ) pg ON pg.venta_id = v.id
            LEFT JOIN (
                SELECT venta_id, STRING_AGG(producto, ' / ') AS productos
                FROM items_venta GROUP BY venta_id
            ) itm ON itm.venta_id = v.id
            {where_sql}
            ORDER BY v.id DESC
            LIMIT %s;
        """, params)
        filas = cursor.fetchall()
        resultado = [{
            'id':            f[0],
            'codigo':        f[1], 'cliente':       f[2],
            'total':         float(f[3]) if f[3] else 0,
            'adelanto':      float(f[4]) if f[4] else 0,
            'saldo':         float(f[5]) if f[5] else 0,
            'estado':        f[6],
            'fecha_entrega': f[7].strftime('%Y-%m-%d') if f[7] else None,
            'vendedor':      f[8], 'productos':     f[9], 'sede': f[10],
            'fecha_emision': f[11].strftime('%Y-%m-%d') if f[11] else None
        } for f in filas]
        return jsonify(resultado), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)


# ==========================================
# GESTIÓN MANUAL DE ESTADOS Y ANULACIÓN
# ==========================================

@ventas_bp.route('/api/ventas/rapidas', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller')
def reporte_ventas_rapidas():
    """Reporte operativo de items vendidos por el flujo de Venta Rapida."""
    desde = (request.args.get('desde') or '').strip()
    hasta = (request.args.get('hasta') or '').strip()
    try:
        limit = min(max(int(request.args.get('limit', 200)), 1), 1000)
    except (TypeError, ValueError):
        limit = 200

    condiciones = ["iv.color_tela ILIKE %s"]
    params = ["%VENTA RAPIDA%"]
    if desde:
        condiciones.append("v.fecha_emision >= %s")
        params.append(desde)
    if hasta:
        condiciones.append("v.fecha_emision <= %s")
        params.append(hasta)

    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        cursor.execute(f"""
            SELECT v.codigo_venta,
                   v.nombre_cliente,
                   COALESCE(v.vendedor_nombre, 'Sin asignar') AS vendedor,
                   COALESCE(v.sede, '') AS sede_venta,
                   v.fecha_emision,
                   COALESCE(v.estado_general, 'Pendiente') AS estado,
                   iv.producto,
                   COALESCE(iv.precio_unitario, 0) AS precio,
                   COALESCE(iv.foto_url, '') AS foto_url,
                   COALESCE(iv.color_tela, '') AS detalles
            FROM items_venta iv
            JOIN ventas v ON v.id = iv.venta_id
            WHERE {' AND '.join(condiciones)}
            ORDER BY v.fecha_emision DESC, v.id DESC, iv.id DESC
            LIMIT %s
        """, params + [limit])
        rows = cursor.fetchall()
        items = [{
            'codigo': r[0],
            'cliente': r[1],
            'vendedor': r[2],
            'sede': r[3],
            'fecha_emision': r[4].strftime('%Y-%m-%d') if r[4] else '',
            'estado': r[5],
            'producto': r[6],
            'precio': float(r[7] or 0),
            'foto_url': limpiar_foto(r[8]),
            'detalles': r[9],
        } for r in rows]
        return jsonify({
            'items': items,
            'total_items': len(items),
            'total_monto': round(sum(i['precio'] for i in items), 2),
            'limit': limit,
        }), 200
    except Exception as ex:
        return jsonify({'error': str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


def _cerrar_db_silencioso(cursor=None, conexion=None):
    try:
        if cursor:
            cursor.close()
    except Exception as ex:
        print(f"[NOTIF] No se pudo cerrar cursor: {ex}")
    try:
        if conexion:
            release_db_connection(conexion)
    except Exception as ex:
        print(f"[NOTIF] No se pudo devolver conexion: {ex}")


@ventas_bp.route('/api/notificaciones/resumen-operativo', methods=['GET', 'POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def notificaciones_resumen_operativo():
    """Consulta o envia por correo el resumen operativo pendiente."""
    enviar = request.method == 'POST' or request.args.get('enviar') == '1'
    try:
        conexion = get_db_connection()
        cursor = conexion.cursor()
        data = enviar_resumen_operativo(cursor) if enviar else resumen_operativo(cursor)
        conexion.commit()
        return jsonify(data), 200
    except Exception as ex:
        if 'conexion' in locals() and conexion:
            conexion.rollback()
        return jsonify({'error': str(ex)}), 500
    finally:
        _cerrar_db_silencioso(
            cursor if 'cursor' in locals() else None,
            conexion if 'conexion' in locals() else None,
        )


@ventas_bp.route('/api/notificaciones/ping', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller')
def notificaciones_ping():
    return jsonify({'ok': True, 'modulo': 'notificaciones'}), 200


@ventas_bp.route('/api/notificaciones/probar-correo', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def notificaciones_probar_correo():
    """Envia un correo simple para validar SMTP en Render."""
    try:
        data = diagnosticar_correo_prueba()
        return jsonify(data), 200
    except Exception as ex:
        return jsonify({'error': str(ex)}), 500


@ventas_bp.route('/api/ventas/<int:venta_id>/estado', methods=['PUT'])
@requiere_login
def cambiar_estado_venta(venta_id):
    nuevo_estado = request.json.get('estado')
    if not nuevo_estado:
        return jsonify({'error': 'El estado es obligatorio'}), 400

    try:
        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        cursor.execute("UPDATE ventas SET estado_general = %s WHERE id = %s", (nuevo_estado, venta_id))

        # Sincronizar tickets y logística con el nuevo estado de la venta.
        _sincronizar_tickets_con_estado_venta(cursor, venta_id, nuevo_estado)

        # Si la venta se marca como Entregada, las unidades de stock físico
        # que estaban 'Reservado' pasan a 'Vendido'.
        if nuevo_estado == 'Entregado':
            _liberar_unidades(cursor, venta_id, 'Vendido', 'Venta')
        # ─────────────────────────────────────────────────────────────────────

        conexion.commit()
        try:
            notif = notificar_estado_contrato(cursor, venta_id, nuevo_estado)
            print(f"[NOTIF] Estado venta {venta_id}: {notif}")
        except Exception as ex_notif:
            print(f"[NOTIF] No se pudo notificar estado de venta {venta_id}: {ex_notif}")
        return jsonify({'exito': True, 'mensaje': f'Estado actualizado a {nuevo_estado}'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            if 'cursor' in locals() and cursor: cursor.close()
            release_db_connection(conexion)


@ventas_bp.route('/api/ventas/<int:venta_id>/anular', methods=['POST'])
@requiere_login
def anular_venta_completa(venta_id):
    try:
        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        # 1. Marcar venta como Cancelado
        cursor.execute("UPDATE ventas SET estado_general = 'Cancelado' WHERE id = %s", (venta_id,))

        # 2. Cancelar Tickets de Producción
        cursor.execute("""
            UPDATE tickets_produccion
            SET estado_ticket = 'Cancelado'
            WHERE item_id IN (SELECT id FROM items_venta WHERE venta_id = %s)
              AND estado_ticket != 'Terminado'
        """, (venta_id,))

        # 3. Cancelar compras en Logística Externa
        cursor.execute("""
            UPDATE logistica_externa
            SET estado = 'Cancelado'
            WHERE venta_id = %s AND estado != 'Recibido'
        """, (venta_id,))

        # 4. Devolver stock genérico (contador en catalogo_productos)
        cursor.execute("""
            SELECT cp.id, iv.es_stock
            FROM catalogo_productos cp
            JOIN items_venta iv ON cp.nombre_modelo = iv.producto
            -- Solo devolver al stock si originalmente se descontó de él
            WHERE iv.venta_id = %s
        """, (venta_id,))
        for p in cursor.fetchall():
            cursor.execute("""
                UPDATE catalogo_productos
                SET stock_cantidad = stock_cantidad + 1, en_stock = true
                WHERE id = %s
            """, (p[0],))

        # 5. ── PUENTE INVENTARIO: liberar unidades físicas reservadas ─────────
        _liberar_unidades(cursor, venta_id, 'Disponible', 'Devolucion')
        # ─────────────────────────────────────────────────────────────────────

        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Venta y procesos de taller anulados con éxito.'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            if 'cursor' in locals() and cursor: cursor.close()
            release_db_connection(conexion)


def _crear_tabla_ventas_eliminadas_log(cursor):
    """
    Auditoría de borrados definitivos. La venta y todo lo que cuelga de
    ella (items, tickets, pagos, logística, historial de precios) se
    borra de verdad — pero acá queda una fotografía de qué se borró,
    quién lo hizo y por qué, para siempre. Sin esto, un borrado en
    cascada es imposible de investigar después.
    """
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ventas_eliminadas_log (
            id                      SERIAL PRIMARY KEY,
            venta_id                INTEGER,
            codigo_venta            VARCHAR(50),
            nombre_cliente          VARCHAR(150),
            monto_total             NUMERIC(10,2),
            vendedor_nombre         VARCHAR(100),
            fecha_emision_original  TIMESTAMP,
            snapshot_items          JSONB,
            motivo                  TEXT NOT NULL,
            eliminado_por_id        INTEGER,
            eliminado_por_nombre    VARCHAR(100),
            fecha_eliminacion       TIMESTAMP DEFAULT NOW()
        );
    """)


@ventas_bp.route('/api/ventas/<int:venta_id>/eliminar-completo', methods=['POST'])
@requiere_rol('Admin')
def eliminar_venta_completa(venta_id):
    """
    Borra la venta y TODO lo relacionado a ella como si nunca hubiera
    existido: items, tickets de taller, pagos, logística externa y
    solicitudes de cambio de precio. A diferencia de /anular (que solo
    cambia el estado a 'Cancelado' y conserva el registro), esto ejecuta
    DELETE real sobre las filas.

    Requiere 'motivo' — se guarda en ventas_eliminadas_log ANTES de borrar
    nada, junto con una foto de los items del contrato, para dejar rastro
    de qué se eliminó y por qué (dato irrecuperable después de esto).

    Solo Admin puede ejecutarlo por ser una acción irreversible.
    """
    data         = request.json or {}
    motivo       = (data.get('motivo') or '').strip()
    admin_id     = data.get('admin_id')
    admin_nombre = data.get('admin_nombre', 'Admin')

    if not motivo:
        return jsonify({'error': 'Debes indicar el motivo de la eliminación'}), 400

    try:
        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()

        cursor.execute("""
            SELECT codigo_venta, nombre_cliente, monto_total, vendedor_nombre, fecha_emision
            FROM ventas WHERE id = %s;
        """, (venta_id,))
        venta = cursor.fetchone()
        if not venta:
            return jsonify({'error': 'Venta no encontrada'}), 404
        codigo_venta, nombre_cliente, monto_total, vendedor_nombre, fecha_emision = venta

        cursor.execute("""
            SELECT producto, color_tela, precio_unitario FROM items_venta WHERE venta_id = %s;
        """, (venta_id,))
        items_snapshot = [
            {'producto': r[0], 'detalles': r[1], 'precio_unitario': float(r[2] or 0)}
            for r in cursor.fetchall()
        ]

        # 1. Dejar rastro ANTES de borrar nada
        _crear_tabla_ventas_eliminadas_log(cursor)
        cursor.execute("""
            INSERT INTO ventas_eliminadas_log
                (venta_id, codigo_venta, nombre_cliente, monto_total, vendedor_nombre,
                 fecha_emision_original, snapshot_items, motivo,
                 eliminado_por_id, eliminado_por_nombre)
            VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s, %s, %s);
        """, (venta_id, codigo_venta, nombre_cliente, monto_total, vendedor_nombre,
              fecha_emision, json.dumps(items_snapshot, default=str), motivo,
              admin_id, admin_nombre))

        # 2. Devolver al inventario cualquier unidad física reservada,
        #    igual que en la anulación — eliminar la venta no debe dejar
        #    muebles "perdidos" marcados como Reservado para siempre.
        _liberar_unidades(cursor, venta_id, 'Disponible', 'Eliminacion definitiva')

        # 3. Devolver stock genérico del catálogo (mismo criterio que /anular)
        cursor.execute("""
            SELECT cp.id FROM catalogo_productos cp
            JOIN items_venta iv ON cp.nombre_modelo = iv.producto
            WHERE iv.venta_id = %s
        """, (venta_id,))
        for p in cursor.fetchall():
            cursor.execute("""
                UPDATE catalogo_productos SET stock_cantidad = stock_cantidad + 1, en_stock = true
                WHERE id = %s
            """, (p[0],))

        # 4. Borrar en cascada — hijos primero, padre al final
        cursor.execute("DELETE FROM historial_inventario WHERE venta_id = %s;", (venta_id,))
        cursor.execute("""
            DELETE FROM tickets_produccion
            WHERE item_id IN (SELECT id FROM items_venta WHERE venta_id = %s);
        """, (venta_id,))
        cursor.execute("DELETE FROM pagos WHERE venta_id = %s;", (venta_id,))

        # FIX: logistica_externa tiene tablas hijas (ordenes_compra_seq vía
        # logistica_id, cotizacion_lote_items vía logistica_externa_id) con FK
        # SIN ON DELETE CASCADE. Si esta venta generó una cotización de tela
        # externa u orden de compra, borrar logistica_externa directamente
        # violaba esa FK y Postgres abortaba toda la transacción con un 500
        # genérico. Hay que borrar esas tablas hijas primero.
        cursor.execute("SELECT id FROM logistica_externa WHERE venta_id = %s;", (venta_id,))
        logistica_ids = [r[0] for r in cursor.fetchall()]
        if logistica_ids:
            cursor.execute(
                "DELETE FROM cotizacion_lote_items WHERE logistica_externa_id = ANY(%s);",
                (logistica_ids,)
            )
            cursor.execute(
                "DELETE FROM ordenes_compra_seq WHERE logistica_id = ANY(%s);",
                (logistica_ids,)
            )
        cursor.execute("DELETE FROM logistica_externa WHERE venta_id = %s;", (venta_id,))

        cursor.execute("DELETE FROM historial_precios WHERE venta_id = %s;", (venta_id,))
        cursor.execute("DELETE FROM items_venta WHERE venta_id = %s;", (venta_id,))
        cursor.execute("DELETE FROM ventas WHERE id = %s;", (venta_id,))

        conexion.commit()
        return jsonify({
            'exito': True,
            'mensaje': f'La venta #{codigo_venta} fue eliminada por completo del sistema.'
        }), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        # DEBUG TEMPORAL (julio 2026): el 500 no mostraba el error real,
        # ni en los logs de Render (que solo registran el access log de
        # gunicorn) ni en el popup del frontend (str(e) a veces viene vacío
        # según el tipo de excepción de psycopg2). Con esto:
        #   1. El traceback completo queda impreso en stdout → aparece en
        #      Render → Logs, buscándolo como "ELIMINAR-COMPLETO ERROR".
        #   2. El JSON de respuesta incluye tipo + mensaje + pgcode/pgerror
        #      si es un error de Postgres, para verlo directo en el SweetAlert.
        print(f"\n{'='*70}\nELIMINAR-COMPLETO ERROR (venta_id={venta_id})\n{'='*70}")
        traceback.print_exc()
        print(f"{'='*70}\n")

        detalle = f"{type(e).__name__}: {e}"
        pgcode  = getattr(e, 'pgcode', None)
        pgerror = getattr(e, 'pgerror', None)
        if pgcode or pgerror:
            detalle += f" | pgcode={pgcode} | pgerror={pgerror}"

        return jsonify({'error': detalle}), 500
    finally:
        if 'conexion' in locals() and conexion:
            if 'cursor' in locals() and cursor: cursor.close()
            release_db_connection(conexion)


# ==========================================
# SEGUIMIENTO Y PRODUCCIÓN
# ==========================================

@ventas_bp.route('/api/mis-ventas/<int:vendedor_id>', methods=['GET'])
@requiere_login
def obtener_mis_ventas(vendedor_id):
    """
    Paginado server-side (julio 2026). Antes traía TODAS las ventas del
    vendedor sin límite y el frontend paginaba/filtraba en memoria
    (_mpTodos en busqueda_filtros.js) — con pocos cientos de pedidos ya
    implicaba bajar un JSON completo en cada apertura de "Mis Pedidos".

    Query params opcionales:
        page      (default 1)
        per_page  (default 20, tope 100 — ver database.paginar)
        q         texto libre: busca en código de venta o nombre de cliente
        estado    filtra por estado_general exacto (mismo valor que ya
                  manda el <select> del frontend, ej. "Entregado")
    """
    from database import paginar

    page     = request.args.get('page', 1)
    per_page = request.args.get('per_page', 20)
    q        = (request.args.get('q') or '').strip()
    estado   = (request.args.get('estado') or '').strip()

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()

        condiciones = ["v.vendedor_id = %s"]
        params      = [vendedor_id]

        if q:
            condiciones.append("(v.codigo_venta ILIKE %s OR v.nombre_cliente ILIKE %s)")
            params.extend([f"%{q}%", f"%{q}%"])

        if estado:
            condiciones.append("COALESCE(v.estado_general, 'Pendiente') = %s")
            params.append(estado)

        query = f"""
            SELECT v.codigo_venta, v.nombre_cliente, v.fecha_entrega,
                   COALESCE(COUNT(t.id), 0) AS total,
                   COALESCE(SUM(CASE WHEN t.estado_ticket = 'Terminado' THEN 1 ELSE 0 END), 0) AS terminados,
                   v.monto_total, COALESCE(v.estado_general, 'Pendiente') AS estado
            FROM ventas v
            LEFT JOIN items_venta i        ON v.id  = i.venta_id
            LEFT JOIN tickets_produccion t ON i.id  = t.item_id
            WHERE {' AND '.join(condiciones)}
            GROUP BY v.id, v.codigo_venta, v.nombre_cliente, v.fecha_entrega, v.monto_total, v.estado_general
            ORDER BY v.id DESC
        """

        filas, total, total_pages = paginar(cursor, query, params, page=page, per_page=per_page)

        res = []
        for v in filas:
            total_tickets = v[3]
            terminados    = v[4]
            porcentaje    = round((terminados / total_tickets * 100), 0) if total_tickets > 0 else 0
            res.append({
                "codigo":      v[0], "cliente":    v[1],
                "entrega":     v[2].strftime('%d/%m/%Y') if v[2] else "S/F",
                "progreso":    porcentaje,
                "monto_total": float(v[5]),
                "estado":      v[6]
            })

        return jsonify({
            "items":       res,
            "total":       total,
            "page":        min(max(1, int(page or 1)), total_pages),
            "total_pages": total_pages
        })
    except Exception as ex:
        print("Error en seguimiento:", ex)
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/produccion/<area>', methods=['GET'])
@requiere_login
def ver_tickets_area(area):
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT t.id, i.producto, i.color_tela, v.codigo_venta,
                   t.estado_ticket, i.foto_url, v.fecha_entrega
            FROM tickets_produccion t
            JOIN items_venta i ON t.item_id  = i.id
            JOIN ventas v      ON i.venta_id = v.id
            WHERE t.area_trabajo = %s AND t.estado_ticket = 'Pendiente'
            ORDER BY v.fecha_entrega ASC;
        """, (area,))
        lista = [{
            "ticket_id":    t[0], "producto":     t[1], "color":        t[2],
            "codigo_venta": t[3], "estado":       t[4], "foto":         t[5],
            "fecha_entrega": t[6].strftime('%d/%m/%Y') if t[6] else "S/F"
        } for t in cursor.fetchall()]
        return jsonify({"area": area, "tareas": lista})
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/pedido/detalle/<codigo>', methods=['GET'])
@requiere_login
def obtener_detalle_pedido(codigo):
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT id, codigo_venta, nombre_cliente, fecha_entrega, empresa_ruc, vendedor_nombre
            FROM ventas WHERE codigo_venta = %s;
        """, (codigo,))
        venta = cursor.fetchone()
        if not venta:
            return jsonify({"error": "Pedido no encontrado"}), 404
        cursor.execute(
            "SELECT producto, color_tela, foto_url, COALESCE(precio_unitario, 0) FROM items_venta WHERE venta_id = %s;",
            (venta[0],)
        )
        items = []
        for i in cursor.fetchall():
            # FIX (julio 2026): foto_url puede traer varias fotos separadas
            # por '|' (las que el vendedor eligió del catálogo/carta al
            # armar el pedido — antes solo se guardaba/mostraba la primera
            # y el resto se perdía). Ahora se devuelven TODAS en `fotos`,
            # y se deja `foto` (la primera) para no romper nada que ya
            # dependa de ese campo como único.
            fotos_item = [f.strip() for f in (i[2] or '').split('|') if f.strip()]
            items.append({
                "producto": i[0], "detalles": i[1],
                "foto":  fotos_item[0] if fotos_item else "",
                "fotos": fotos_item,
                "precio": float(i[3] or 0),
            })

        # Comprobantes de pago subidos al finalizar la venta (uno por cada
        # pago registrado en el checkout: efectivo, POS, transferencia, etc.)
        pagos = []
        try:
            cursor.execute("""
                SELECT tipo_pago, entidad, monto_bruto, comprobante_url,
                       TO_CHAR(fecha_pago, 'DD/MM/YYYY')
                FROM pagos WHERE venta_id = %s ORDER BY id;
            """, (venta[0],))
            pagos = [{
                "tipo":        r[0] or "—",
                "entidad":     r[1] or "—",
                "monto":       float(r[2] or 0),
                "comprobante": r[3] if r[3] and r[3] != 'Sin imagen' else None,
                "fecha":       r[4] or "—",
            } for r in cursor.fetchall()]
        except Exception:
            conexion.rollback()

        return jsonify({
            "codigo":  venta[1], "cliente": venta[2],
            "entrega": venta[3].strftime('%d/%m/%Y') if venta[3] else "S/F",
            "items":   items,
            "pagos":   pagos,
        }), 200
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# CAMBIO DE PRECIO CON HISTORIAL
# ==========================================

@ventas_bp.route('/api/ventas/<codigo>/items-editables', methods=['GET'])
@requiere_login
def items_editables_venta(codigo):
    """
    Lista los productos de un contrato para que el vendedor elija a cuál
    le quiere pedir un cambio (precio, tela/material, o agregar uno nuevo).
    Usado por el paso 1 del modal 'Cambiar precio' en el frontend.
    """
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("SELECT id, monto_total, estado_general FROM ventas WHERE codigo_venta = %s;", (codigo,))
        venta = cursor.fetchone()
        if not venta:
            return jsonify({'error': 'Venta no encontrada'}), 404
        venta_id, monto_total, estado = venta

        cursor.execute("""
            SELECT id, producto, color_tela, foto_url, COALESCE(precio_unitario, 0)
            FROM items_venta WHERE venta_id = %s ORDER BY id;
        """, (venta_id,))
        items = [{
            'id':             r[0],
            'producto':       r[1],
            'detalles':       r[2] or '',
            'foto':           limpiar_foto(r[3]).split('|')[0] if r[3] else 'imagenes/sin_foto.jpg',
            'precio_unitario': float(r[4] or 0),
        } for r in cursor.fetchall()]

        return jsonify({
            'codigo_venta':   codigo,
            'estado_venta':   estado,
            'monto_total':    float(monto_total or 0),
            'items':          items,
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/ventas/<codigo>/proponer-cambio-precio', methods=['POST'])
@requiere_login
def proponer_cambio_precio(codigo):
    """
    tipo_cambio puede ser:
      'precio'          -> cambia el precio_unitario de item_id
      'material'        -> cambia la tela/material (detalle_nuevo) de item_id,
                            opcionalmente tambien el precio si sube el costo
      'nuevo_producto'  -> agrega un producto adicional al contrato
                            (producto_nombre + precio_nuevo), item_id no aplica
    """
    data            = request.json or {}
    tipo_cambio     = data.get('tipo_cambio', 'precio')
    item_id         = data.get('item_id')
    producto_nombre = (data.get('producto_nombre') or '').strip()
    precio_nuevo    = data.get('precio_nuevo')
    detalle_nuevo   = (data.get('detalle_nuevo') or '').strip()
    motivo          = (data.get('motivo') or '').strip()
    vendedor_id     = data.get('vendedor_id')
    vendedor_nombre = data.get('vendedor_nombre', '')

    if tipo_cambio not in ('precio', 'material', 'nuevo_producto'):
        return jsonify({'error': 'tipo_cambio invalido'}), 400
    if not motivo:
        return jsonify({'error': 'El motivo del cambio es obligatorio'}), 400
    if tipo_cambio == 'nuevo_producto' and not producto_nombre:
        return jsonify({'error': 'Debes indicar el nombre del producto nuevo'}), 400
    if tipo_cambio in ('precio', 'material') and not item_id:
        return jsonify({'error': 'Debes elegir a que producto del contrato aplica el cambio'}), 400
    if tipo_cambio == 'material' and not detalle_nuevo:
        return jsonify({'error': 'Describe la tela/material nueva'}), 400
    if tipo_cambio in ('precio', 'nuevo_producto'):
        if precio_nuevo is None or float(precio_nuevo) <= 0:
            return jsonify({'error': 'Ingresa un precio valido'}), 400

    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _crear_tabla_historial_precios(cursor)

        cursor.execute("SELECT id, monto_total, estado_general FROM ventas WHERE codigo_venta = %s;", (codigo,))
        venta = cursor.fetchone()
        if not venta:
            return jsonify({'error': 'Venta no encontrada'}), 404
        venta_id, monto_total_actual, estado = venta
        if estado in ('Entregado', 'Cancelado'):
            return jsonify({'error': f'No se puede modificar una venta en estado {estado}'}), 400

        precio_original = 0.0
        if tipo_cambio in ('precio', 'material'):
            cursor.execute("SELECT producto, precio_unitario FROM items_venta WHERE id = %s AND venta_id = %s;",
                            (item_id, venta_id))
            item_row = cursor.fetchone()
            if not item_row:
                return jsonify({'error': 'El producto seleccionado no pertenece a este contrato'}), 400
            producto_nombre = producto_nombre or item_row[0]
            precio_original = float(item_row[1] or 0)
            if tipo_cambio == 'material' and precio_nuevo is None:
                precio_nuevo = precio_original

        cursor.execute(
            "SELECT id FROM historial_precios WHERE venta_id = %s AND estado = 'Pendiente';",
            (venta_id,)
        )
        if cursor.fetchone():
            return jsonify({'error': 'Ya existe una solicitud pendiente para esta venta. Espera su resolucion antes de enviar otra.'}), 400

        cursor.execute("""
            INSERT INTO historial_precios
                (venta_id, codigo_venta, precio_original, precio_nuevo,
                 motivo, vendedor_id, vendedor_nombre, estado,
                 item_id, producto_nombre, tipo_cambio, detalle_nuevo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s, %s, %s)
            RETURNING id;
        """, (venta_id, codigo, precio_original, float(precio_nuevo or 0),
              motivo, vendedor_id, vendedor_nombre,
              item_id, producto_nombre, tipo_cambio, detalle_nuevo or None))
        nuevo_id = cursor.fetchone()[0]
        conexion.commit()
        return jsonify({'exito': True, 'id': nuevo_id, 'mensaje': 'Solicitud enviada al administrador'}), 201

    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/cambios-precio/pendientes', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller')
def listar_cambios_precio_pendientes():
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _crear_tabla_historial_precios(cursor)
        cursor.execute("""
            SELECT hp.id, hp.codigo_venta, hp.precio_original, hp.precio_nuevo,
                   hp.motivo, hp.vendedor_nombre, hp.fecha_solicitud,
                   v.nombre_cliente, COALESCE(v.estado_general, 'En Producción') AS estado_venta,
                   hp.producto_nombre, COALESCE(hp.tipo_cambio, 'precio'), hp.detalle_nuevo
            FROM historial_precios hp
            JOIN ventas v ON hp.venta_id = v.id
            WHERE hp.estado = 'Pendiente'
            ORDER BY hp.fecha_solicitud DESC;
        """)
        resultado = [{
            'id': f[0], 'codigo_venta': f[1],
            'precio_original': float(f[2]), 'precio_nuevo': float(f[3]),
            'motivo': f[4], 'vendedor': f[5],
            'fecha_solicitud': f[6].strftime('%d/%m/%Y %H:%M') if f[6] else '',
            'cliente': f[7], 'estado_venta': f[8],
            'producto': f[9] or '—', 'tipo_cambio': f[10], 'detalle_nuevo': f[11] or '',
        } for f in cursor.fetchall()]
        return jsonify(resultado), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/cambios-precio/<int:cambio_id>/aprobar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def aprobar_cambio_precio(cambio_id):
    data         = request.json
    admin_id     = data.get('admin_id')
    admin_nombre = data.get('admin_nombre', 'Admin')
    try:
        conexion = get_db_connection()
        conexion.autocommit = False
        cursor   = conexion.cursor()
        cursor.execute("""
            SELECT venta_id, precio_nuevo, item_id, producto_nombre,
                   COALESCE(tipo_cambio, 'precio'), detalle_nuevo, codigo_venta
            FROM historial_precios WHERE id = %s AND estado = 'Pendiente';
        """, (cambio_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Solicitud no encontrada o ya resuelta'}), 404
        venta_id, precio_nuevo, item_id, producto_nombre, tipo_cambio, detalle_nuevo, codigo_venta = row

        if tipo_cambio == 'precio':
            cursor.execute("UPDATE items_venta SET precio_unitario = %s WHERE id = %s AND venta_id = %s;",
                            (precio_nuevo, item_id, venta_id))
        elif tipo_cambio == 'material':
            if detalle_nuevo:
                cursor.execute("UPDATE items_venta SET color_tela = %s WHERE id = %s AND venta_id = %s;",
                                (detalle_nuevo, item_id, venta_id))
            if precio_nuevo is not None:
                cursor.execute("UPDATE items_venta SET precio_unitario = %s WHERE id = %s AND venta_id = %s;",
                                (precio_nuevo, item_id, venta_id))
        elif tipo_cambio == 'nuevo_producto':
            cursor.execute("""
                INSERT INTO items_venta (venta_id, producto, precio_unitario, es_stock)
                VALUES (%s, %s, %s, FALSE);
            """, (venta_id, producto_nombre, precio_nuevo))

        nuevo_total = _recalcular_total_venta(cursor, venta_id)

        cursor.execute("""
            UPDATE historial_precios
            SET estado = 'Aprobado', admin_id = %s, admin_nombre = %s, fecha_resolucion = NOW()
            WHERE id = %s;
        """, (admin_id, admin_nombre, cambio_id))
        conexion.commit()
        return jsonify({
            'exito': True,
            'mensaje': f'Cambio aplicado. Nuevo total del contrato: S/ {nuevo_total:.2f}',
            'monto_total': nuevo_total,
        }), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/cambios-precio/<int:cambio_id>/rechazar', methods=['POST'])
@requiere_rol('Admin', 'Jefe_Taller')
def rechazar_cambio_precio(cambio_id):
    data         = request.json
    admin_id     = data.get('admin_id')
    admin_nombre = data.get('admin_nombre', 'Admin')
    notas        = data.get('notas_admin', '').strip()
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            UPDATE historial_precios
            SET estado = 'Rechazado', admin_id = %s, admin_nombre = %s,
                notas_admin = %s, fecha_resolucion = NOW()
            WHERE id = %s AND estado = 'Pendiente';
        """, (admin_id, admin_nombre, notas, cambio_id))
        if cursor.rowcount == 0:
            return jsonify({'error': 'Solicitud no encontrada o ya resuelta'}), 404
        conexion.commit()
        return jsonify({'exito': True, 'mensaje': 'Solicitud rechazada'}), 200
    except Exception as e:
        if 'conexion' in locals() and conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/ventas/<codigo>/historial-precios', methods=['GET'])
@requiere_login
def historial_precios_venta(codigo):
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _crear_tabla_historial_precios(cursor)
        cursor.execute("""
            SELECT precio_original, precio_nuevo, motivo, vendedor_nombre,
                   admin_nombre, estado, notas_admin, fecha_solicitud, fecha_resolucion,
                   producto_nombre, COALESCE(tipo_cambio, 'precio'), detalle_nuevo
            FROM historial_precios WHERE codigo_venta = %s ORDER BY fecha_solicitud DESC;
        """, (codigo,))
        resultado = [{
            'precio_original':  float(f[0]), 'precio_nuevo': float(f[1]),
            'motivo': f[2], 'vendedor': f[3], 'admin': f[4], 'estado': f[5],
            'notas_admin': f[6],
            'fecha_solicitud':  f[7].strftime('%d/%m/%Y %H:%M') if f[7] else '',
            'fecha_resolucion': f[8].strftime('%d/%m/%Y %H:%M') if f[8] else '',
            'producto': f[9] or '—', 'tipo_cambio': f[10], 'detalle_nuevo': f[11] or '',
        } for f in cursor.fetchall()]
        return jsonify(resultado), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion:
            cursor.close(); release_db_connection(conexion)


# ==========================================
# EXPORTACIÓN DE VENTAS
# ==========================================

@ventas_bp.route('/api/ventas/exportar', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller')
def exportar_ventas_excel():
    try:
        inicio_str = request.args.get('inicio')
        fin_str = request.args.get('fin')

        conexion = get_db_connection()
        cursor   = conexion.cursor()

        params = []
        where_clauses = []
        if inicio_str:
            where_clauses.append("v.fecha_emision >= %s")
            params.append(inicio_str)
        if fin_str:
            where_clauses.append("v.fecha_emision <= %s")
            params.append(fin_str)
        
        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        cursor.execute(f"""
            SELECT v.codigo_venta, v.nombre_cliente, v.tipo_comprobante, v.dni_cliente,
                   v.fecha_emision, v.fecha_entrega, v.monto_total,
                   COALESCE(itm.productos, '') AS productos, v.direccion_cliente,
                   COALESCE(pg.metodos, 'Sin pago') AS metodo_pago,
                   COALESCE(pg.operaciones, '') AS numero_operacion,
                   COALESCE(pg.total_pagado, 0) AS monto_adelanto,
                   COALESCE(pg.empresas, '---') AS empresa_pago,
                   v.fecha_emision AS fecha_registro, v.celular_cliente,
                   v.moneda, v.vendedor_nombre
            FROM ventas v
            LEFT JOIN (
                SELECT venta_id, SUM(monto_bruto) AS total_pagado,
                       STRING_AGG(tipo_pago || ' (' || COALESCE(entidad, '') || ')', ' | ') AS metodos,
                       STRING_AGG(DISTINCT empresa_destino, ' / ') AS empresas,
                       STRING_AGG(NULLIF(numero_operacion, ''), ' | ') AS operaciones
                FROM pagos GROUP BY venta_id
            ) pg ON pg.venta_id = v.id
            LEFT JOIN (
                SELECT venta_id, STRING_AGG(producto, ' / ') AS productos
                FROM items_venta GROUP BY venta_id
            ) itm ON itm.venta_id = v.id
            {where_sql}
            ORDER BY v.fecha_emision DESC;
        """, params)
        filas = cursor.fetchall()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        def _formatear_hoja_excel(ws, datos_hoja):
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill("solid", fgColor="0F172A")
            center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin        = Side(style="thin", color="CBD5E0")
            border      = Border(left=thin, right=thin, top=thin, bottom=thin)
            headers = [
                "Cód. Venta", "Cliente", "Comprobante", "RUC/DNI/CE",
                "F. Emisión", "F. Entrega", "Monto Total",
                "Producto(s)", "Dirección", "Métodos Pago (Múltiples)",
                "N° Operación", "Adelanto Cobrado", "Empresa Receptora",
                "Fecha Registro", "Teléfono", "Moneda", "Vendedor"
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = center; cell.border = border
            anchos = [12, 25, 12, 14, 14, 14, 12, 40, 30, 30, 18, 15, 30, 14, 14, 10, 20]
            for col, ancho in enumerate(anchos, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
            fill_par   = PatternFill("solid", fgColor="F8FAFC")
            fill_impar = PatternFill("solid", fgColor="FFFFFF")
            for row_num, f in enumerate(datos_hoja, 2):
                fill = fill_par if row_num % 2 == 0 else fill_impar
                valores = [
                    f[0], f[1], f[2], f[3],
                    f[4].strftime('%d/%m/%Y') if f[4] else '',
                    f[5].strftime('%d/%m/%Y') if f[5] else '',
                    float(f[6]) if f[6] else 0,
                    f[7], f[8], f[9],
                    f[10],
                    float(f[11]) if f[11] else 0,
                    f[12],
                    f[13].strftime('%d/%m/%Y') if f[13] else '',
                    f[14], f[15], f[16]
                ]
                for col, val in enumerate(valores, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.fill = fill; cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.freeze_panes = "A2"

        semanas = {}
        if inicio_str and fin_str:
            from datetime import timedelta
            inicio_dt = datetime.strptime(inicio_str, '%Y-%m-%d')
            fin_dt = datetime.strptime(fin_str, '%Y-%m-%d')
            if (fin_dt - inicio_dt).days > 6:
                for fila in filas:
                    fecha_emision = fila[4]
                    if fecha_emision:
                        inicio_semana = fecha_emision - timedelta(days=fecha_emision.weekday())
                        semana_key = inicio_semana.strftime('%Y-%m-%d')
                        semanas.setdefault(semana_key, []).append(fila)
        
        if not semanas:
            semanas['Ventas'] = filas

        sorted_semanas = sorted(semanas.items(), key=lambda item: item[0] if item[0] != 'Ventas' else '0000-00-00')

        for i, (semana_key, datos_semana) in enumerate(sorted_semanas):
            if semana_key == 'Ventas':
                titulo_hoja = 'Ventas'
            else:
                from datetime import timedelta
                inicio_sem = datetime.strptime(semana_key, '%Y-%m-%d')
                fin_sem = inicio_sem + timedelta(days=6)
                titulo_hoja = f"Sem {inicio_sem.strftime('%d')}-{fin_sem.strftime('%d %b')}"
            
            ws = wb.create_sheet(title=titulo_hoja, index=i)
            _formatear_hoja_excel(ws, datos_semana)

        buffer = io.BytesIO()
        wb.save(buffer); buffer.seek(0)
        fecha_hoy = datetime.now().strftime('%Y%m%d_%H%M')
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'ventas_innova_{fecha_hoy}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conexion' in locals() and conexion: cursor.close(); release_db_connection(conexion)


# A4: La ruta /api/exportar_ventas (CSV stub) fue eliminada.
# El único endpoint de exportación es /api/ventas/exportar (Excel completo).
# El frontend ya apunta correctamente a /api/ventas/exportar.


# ═══════════════════════════════════════════════════════════════════════════
#  COMISIONES DE VENDEDORES  (v2 — sueldos, descuentos acumulados, aumentos)
#
#  GET  /api/vendedores/comisiones
#       ?desde=YYYY-MM-DD &hasta=YYYY-MM-DD &vendedor=Nombre
#       Devuelve TODOS los vendedores (aunque no hayan vendido nada).
#       Incluye sueldo_base, descuentos, aumentos y saldo_acumulado.
#
#  POST /api/vendedores/ajuste
#       { usuario_id, tipo: 'descuento'|'aumento', monto, motivo, semana_inicio, semana_fin }
#       Registra un descuento o aumento para el vendedor en esa semana.
#
#  POST /api/vendedores/cerrar-semana
#       { usuario_id, semana_inicio, semana_fin, monto_pagado, notas, voucher_url }
#       Cierra la semana: si comision=0 → no se paga sueldo; descuentos pendientes
#       pasan como saldo_acumulado a la siguiente semana.
#
#  Tabla auto-creada: ajustes_sueldo_vendedor
#    id, usuario_id, tipo ('descuento'|'aumento'), monto, motivo,
#    semana_inicio, semana_fin, aplicado (bool), created_at
#
#  Tabla auto-creada: cierres_semanales_vendedor
#    id, usuario_id, semana_inicio, semana_fin, sueldo_base, comision,
#    aumentos, descuentos, saldo_anterior, monto_pagado, notas, voucher_url,
#    created_at
# ═══════════════════════════════════════════════════════════════════════════

SUELDO_BASE_VENDEDOR = 350.0
TASA_COMISION        = 0.03


def _ensure_tablas_vendedor(cursor):
    """Crea las tablas auxiliares si no existen (auto-migración segura)."""
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ajustes_sueldo_vendedor (
            id             SERIAL PRIMARY KEY,
            usuario_id     INTEGER NOT NULL,
            tipo           VARCHAR(20) NOT NULL CHECK (tipo IN ('descuento','aumento')),
            monto          NUMERIC(10,2) NOT NULL CHECK (monto > 0),
            motivo         TEXT,
            semana_inicio  DATE NOT NULL,
            semana_fin     DATE NOT NULL,
            aplicado       BOOLEAN DEFAULT FALSE,
            created_at     TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS cierres_semanales_vendedor (
            id              SERIAL PRIMARY KEY,
            usuario_id      INTEGER NOT NULL,
            semana_inicio   DATE NOT NULL,
            semana_fin      DATE NOT NULL,
            sueldo_base     NUMERIC(10,2) DEFAULT 350,
            comision        NUMERIC(10,2) DEFAULT 0,
            aumentos        NUMERIC(10,2) DEFAULT 0,
            descuentos      NUMERIC(10,2) DEFAULT 0,
            saldo_anterior  NUMERIC(10,2) DEFAULT 0,
            monto_pagado    NUMERIC(10,2) DEFAULT 0,
            notas           TEXT,
            voucher_url     TEXT,
            created_at      TIMESTAMP DEFAULT NOW(),
            UNIQUE (usuario_id, semana_inicio)
        );
    """)


@ventas_bp.route('/api/vendedores/comisiones', methods=['GET'])
@requiere_rol('Admin')
def obtener_comisiones_vendedores():
    desde    = request.args.get('desde', '')
    hasta    = request.args.get('hasta', '')
    vendedor = request.args.get('vendedor', '').strip()

    conexion = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _ensure_tablas_vendedor(cursor)
        conexion.commit()

        # ── 1. Todos los vendedores registrados en el sistema ──────────────
        filtro_nombre = ""
        params_usu    = []
        if vendedor:
            filtro_nombre = "AND LOWER(nombre) = LOWER(%s)"
            params_usu.append(vendedor)

        cursor.execute(f"""
            SELECT id, nombre, area_asignada
            FROM usuarios
            WHERE rol = 'Vendedor'
            {filtro_nombre}
            ORDER BY nombre;
        """, params_usu)
        vendedores_db = cursor.fetchall()   # [(id, nombre, area), ...]

        # ── 2. Ventas del período por vendedor ──────────────────────────────
        # FIX (julio 2026): antes se agrupaba por LOWER(TRIM(vendedor_nombre))
        # y se cruzaba contra usuarios.nombre. Eso fallaba (mostrando "0
        # contratos" para vendedores que SÍ tenían pedidos) apenas había una
        # diferencia mínima entre el nombre guardado en la venta y el nombre
        # actual en `usuarios`: una tilde, mayúscula/minúscula, un espacio de
        # más, o simplemente que un admin editó el nombre del vendedor después
        # de que ya tuviera ventas registradas.
        #
        # `ventas.vendedor_id` se guarda siempre al crear el pedido (viene de
        # usuarioActivo.id en carrito.js) y es un ID numérico que nunca
        # cambia ni tiene errores de tipeo, así que ahora cruzamos por ahí.
        # Se conserva un mapa auxiliar por nombre solo como respaldo para
        # ventas antiguas que pudieran no tener vendedor_id guardado (NULL).
        cond_v  = []
        params_v = []
        if desde:
            cond_v.append("fecha_emision::date >= %s")
            params_v.append(desde)
        if hasta:
            cond_v.append("fecha_emision::date <= %s")
            params_v.append(hasta)
        cond_v.append("COALESCE(estado_general, 'Pendiente') <> 'Cancelado'")
        where_v = ("WHERE " + " AND ".join(cond_v)) if cond_v else ""

        cursor.execute(f"""
            SELECT
                vendedor_id,
                LOWER(TRIM(vendedor_nombre))        AS vnom,
                COUNT(DISTINCT id)                  AS contratos,
                COALESCE(SUM(COALESCE(monto_total, 0)), 0) AS total_ventas
            FROM ventas
            {where_v}
            GROUP BY vendedor_id, LOWER(TRIM(vendedor_nombre));
        """, params_v)
        ventas_map_por_id     = {}   # {vendedor_id: {...}}   ← fuente principal
        ventas_map_por_nombre = {}   # {nombre normalizado: {...}}  ← respaldo
        for vid, vnom, contratos, total in cursor.fetchall():
            dato = {'contratos': int(contratos), 'total_ventas': float(total)}
            if vid is not None:
                # Si ya había datos para ese id (no debería pasar, pero por
                # si acaso), sumamos en vez de pisar.
                acumulado = ventas_map_por_id.get(vid, {'contratos': 0, 'total_ventas': 0.0})
                ventas_map_por_id[vid] = {
                    'contratos':    acumulado['contratos'] + dato['contratos'],
                    'total_ventas': acumulado['total_ventas'] + dato['total_ventas'],
                }
            elif vnom:
                acumulado = ventas_map_por_nombre.get(vnom, {'contratos': 0, 'total_ventas': 0.0})
                ventas_map_por_nombre[vnom] = {
                    'contratos':    acumulado['contratos'] + dato['contratos'],
                    'total_ventas': acumulado['total_ventas'] + dato['total_ventas'],
                }

        # ── 3. Ajustes pendientes (descuentos/aumentos no aplicados) ───────
        cond_a  = ["aplicado = FALSE"]
        params_a = []
        if desde:
            cond_a.append("semana_inicio >= %s")
            params_a.append(desde)
        if hasta:
            cond_a.append("semana_fin <= %s")
            params_a.append(hasta)
        where_a = "WHERE " + " AND ".join(cond_a)

        cursor.execute(f"""
            SELECT usuario_id, tipo, COALESCE(SUM(monto), 0)
            FROM ajustes_sueldo_vendedor
            {where_a}
            GROUP BY usuario_id, tipo;
        """, params_a)
        ajustes_map = {}   # {usuario_id: {'descuento': X, 'aumento': Y}}
        for uid, tipo, monto in cursor.fetchall():
            if uid not in ajustes_map:
                ajustes_map[uid] = {'descuento': 0.0, 'aumento': 0.0}
            ajustes_map[uid][tipo] = float(monto)

        # ── 4. Saldo acumulado de descuentos de semanas anteriores ─────────
        #       (descuentos que no se pudieron descontar porque no hubo venta)
        cursor.execute("""
            SELECT usuario_id,
                   COALESCE(SUM(descuentos - monto_pagado + sueldo_base + comision), 0)
            FROM cierres_semanales_vendedor
            WHERE monto_pagado = 0
            GROUP BY usuario_id;
        """)
        # Mejor calcular saldo acumulado como descuentos no cobrados anteriores
        cursor.execute("""
            SELECT usuario_id,
                   COALESCE(SUM(
                       CASE WHEN monto_pagado = 0
                            THEN descuentos - aumentos
                            ELSE 0 END
                   ), 0) AS saldo_acum
            FROM cierres_semanales_vendedor
            GROUP BY usuario_id;
        """)
        saldo_map = {r[0]: float(r[1]) for r in cursor.fetchall()}

        # ── 5. Armar resultado ─────────────────────────────────────────────
        resultado = []
        for uid, nombre, area in vendedores_db:
            vk    = nombre.lower().strip()
            # Cruce principal por vendedor_id; si ese vendedor no tiene ventas
            # con vendedor_id (por ejemplo, ventas viejas previas a que ese
            # campo se guardara), se busca por nombre normalizado como respaldo.
            vdata = ventas_map_por_id.get(uid) or ventas_map_por_nombre.get(vk) \
                    or {'contratos': 0, 'total_ventas': 0.0}
            ajuste      = ajustes_map.get(uid, {'descuento': 0.0, 'aumento': 0.0})
            saldo_acum  = max(0.0, saldo_map.get(uid, 0.0))

            total_ventas   = vdata['total_ventas']
            comision       = round(total_ventas * TASA_COMISION, 2)
            # FIX (julio 2026): antes esto era `total_ventas > 0`, es decir,
            # el sueldo base dependía de la SUMA EN SOLES de sus contratos.
            # Eso dejaba sin sueldo a vendedores que sí registraron pedidos
            # pero con monto_total en 0 (producto de stock entregado sin
            # precio cargado, contrato aún pendiente de cotización, etc.).
            # Ahora se basa en si hizo o no pedidos (contratos), que es la
            # regla real: "si hace pedidos se le paga, si no hace ninguno no".
            vendio_algo    = vdata['contratos'] > 0

            # Si no vendió nada → no cobra sueldo base ni comisión esta semana
            sueldo_efectivo = SUELDO_BASE_VENDEDOR if vendio_algo else 0.0

            descuentos_semana = ajuste['descuento']
            aumentos_semana   = ajuste['aumento']

            # Descuentos que no se pudieron aplicar antes se suman a los de ahora
            descuentos_total = descuentos_semana + saldo_acum

            bruto = sueldo_efectivo + comision + aumentos_semana
            neto  = max(0.0, round(bruto - descuentos_total, 2))

            # Deuda pendiente (descuentos que superan el bruto → pasan a próxima semana)
            deuda_siguiente = max(0.0, round(descuentos_total - bruto, 2)) if not vendio_algo else 0.0

            resultado.append({
                'usuario_id':         uid,
                'vendedor_nombre':    nombre,
                'sede':               area or '',
                'total_contratos':    vdata['contratos'],
                'total_ventas':       total_ventas,
                'comision':           comision,
                'sueldo_base':        SUELDO_BASE_VENDEDOR,
                'sueldo_efectivo':    sueldo_efectivo,   # 0 si no vendió nada
                'aumentos':           aumentos_semana,
                'descuentos':         descuentos_semana,
                'saldo_acumulado':    saldo_acum,        # descuentos arrastrados
                'descuentos_total':   round(descuentos_total, 2),
                'neto':               neto,
                'deuda_siguiente':    deuda_siguiente,   # pasa a próx. semana
                'vendio':             vendio_algo,
                'hizo_pedido':         vendio_algo,
            })

        # Ordenar: primero los que sí vendieron, luego los que no
        resultado.sort(key=lambda x: (-x['total_ventas'], x['vendedor_nombre']))

        return jsonify({
            'vendedores':      resultado,
            'total_ventas':    round(sum(r['total_ventas']  for r in resultado), 2),
            'total_contratos': sum(r['total_contratos']     for r in resultado),
            'total_comision':  round(sum(r['comision']      for r in resultado), 2),
            'total_sueldos':   round(sum(r['sueldo_efectivo'] for r in resultado), 2),
            'total_neto':      round(sum(r['neto']          for r in resultado), 2),
            'sueldo_base':     SUELDO_BASE_VENDEDOR,
            'tasa':            TASA_COMISION,
        }), 200

    except Exception as e:
        if conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/vendedores/ajuste', methods=['POST'])
@requiere_rol('Admin')
def registrar_ajuste_vendedor():
    """Registra un descuento o aumento para un vendedor en una semana."""
    data = request.get_json() or {}
    usuario_id     = data.get('usuario_id')
    tipo           = data.get('tipo', '').strip()       # 'descuento' | 'aumento'
    monto          = float(data.get('monto', 0) or 0)
    motivo         = data.get('motivo', '').strip()
    semana_inicio  = data.get('semana_inicio', '')
    semana_fin     = data.get('semana_fin', '')

    if not all([usuario_id, tipo in ('descuento', 'aumento'), monto > 0,
                semana_inicio, semana_fin]):
        return jsonify({'error': 'Faltan campos obligatorios o monto inválido'}), 400

    conexion = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _ensure_tablas_vendedor(cursor)

        cursor.execute("""
            INSERT INTO ajustes_sueldo_vendedor
                (usuario_id, tipo, monto, motivo, semana_inicio, semana_fin)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (usuario_id, tipo, monto, motivo, semana_inicio, semana_fin))
        nuevo_id = cursor.fetchone()[0]
        conexion.commit()

        return jsonify({'exito': True, 'id': nuevo_id,
                        'mensaje': f'{tipo.capitalize()} de S/ {monto:.2f} registrado'}), 201
    except Exception as e:
        if conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/vendedores/ajustes/<int:uid>', methods=['GET'])
@requiere_rol('Admin')
def listar_ajustes_vendedor(uid):
    """Lista todos los ajustes (descuentos/aumentos) de un vendedor."""
    conexion = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _ensure_tablas_vendedor(cursor)
        cursor.execute("""
            SELECT id, tipo, monto, motivo, semana_inicio, semana_fin, aplicado,
                   TO_CHAR(created_at, 'DD/MM/YYYY HH24:MI') AS fecha
            FROM ajustes_sueldo_vendedor
            WHERE usuario_id = %s
            ORDER BY created_at DESC
            LIMIT 50;
        """, (uid,))
        ajustes = [{
            'id': r[0], 'tipo': r[1], 'monto': float(r[2]),
            'motivo': r[3] or '', 'semana_inicio': str(r[4]),
            'semana_fin': str(r[5]), 'aplicado': r[6], 'fecha': r[7]
        } for r in cursor.fetchall()]
        return jsonify(ajustes), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/vendedores/ajuste/<int:ajuste_id>', methods=['DELETE'])
@requiere_rol('Admin')
def eliminar_ajuste_vendedor(ajuste_id):
    """Elimina un ajuste (solo si no fue aplicado)."""
    conexion = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        cursor.execute("""
            DELETE FROM ajustes_sueldo_vendedor
            WHERE id = %s AND aplicado = FALSE
            RETURNING id;
        """, (ajuste_id,))
        deleted = cursor.fetchone()
        conexion.commit()
        if not deleted:
            return jsonify({'error': 'Ajuste no encontrado o ya aplicado'}), 404
        return jsonify({'exito': True}), 200
    except Exception as e:
        if conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)


@ventas_bp.route('/api/vendedores/cerrar-semana', methods=['POST'])
@requiere_rol('Admin')
def cerrar_semana_vendedor():
    """
    Cierra la semana de un vendedor:
    - Si no vendió nada → monto_pagado = 0, descuentos se acumulan.
    - Marca los ajustes de ese período como aplicado=TRUE.
    """
    data = request.get_json() or {}
    usuario_id    = data.get('usuario_id')
    semana_inicio = data.get('semana_inicio', '')
    semana_fin    = data.get('semana_fin', '')
    monto_pagado  = float(data.get('monto_pagado', 0) or 0)
    notas         = data.get('notas', '').strip()
    voucher_url   = data.get('voucher_url', '').strip()

    if not all([usuario_id, semana_inicio, semana_fin]):
        return jsonify({'error': 'Faltan campos obligatorios'}), 400

    conexion = None
    try:
        conexion = get_db_connection()
        cursor   = conexion.cursor()
        _ensure_tablas_vendedor(cursor)

        # Ventas del período para este vendedor
        cursor.execute("""
            SELECT nombre FROM usuarios WHERE id = %s;
        """, (usuario_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Vendedor no encontrado'}), 404
        nombre = row[0]

        # FIX (julio 2026): mismo cruce que en /api/vendedores/comisiones —
        # antes se buscaba solo por LOWER(TRIM(vendedor_nombre)) = nombre
        # actual en `usuarios`, lo cual fallaba (0 contratos) si el nombre
        # del vendedor cambió o tiene una diferencia mínima de tildes/mayús-
        # culas/espacios frente al valor guardado al momento de la venta.
        # Ahora se cruza por vendedor_id (fuente principal) y se suma,
        # como respaldo, cualquier venta vieja sin vendedor_id que sí
        # coincida por nombre.
        cursor.execute("""
            SELECT COUNT(*), COALESCE(SUM(COALESCE(monto_total, 0)), 0)
            FROM ventas
            WHERE fecha_emision::date BETWEEN %s AND %s
              AND COALESCE(estado_general, 'Pendiente') <> 'Cancelado'
              AND (
                    vendedor_id = %s
                    OR (vendedor_id IS NULL AND LOWER(TRIM(vendedor_nombre)) = LOWER(TRIM(%s)))
              );
        """, (semana_inicio, semana_fin, usuario_id, nombre))
        total_contratos, total_ventas = cursor.fetchone()
        total_contratos = int(total_contratos)
        total_ventas    = float(total_ventas)
        comision        = round(total_ventas * TASA_COMISION, 2)
        # FIX (julio 2026): mismo criterio que en /api/vendedores/comisiones —
        # el sueldo base se paga si el vendedor HIZO PEDIDOS (contratos),
        # no si la suma de esos pedidos en soles fue mayor a 0. Antes un
        # contrato con monto_total en 0 (stock entregado sin precio cargado,
        # contrato pendiente de cotización, etc.) dejaba al vendedor sin
        # sueldo aunque sí hubiera trabajo/pedido real registrado.
        vendio          = total_contratos > 0
        sueldo_ef       = SUELDO_BASE_VENDEDOR if vendio else 0.0

        # Ajustes del período
        cursor.execute("""
            SELECT tipo, COALESCE(SUM(monto), 0)
            FROM ajustes_sueldo_vendedor
            WHERE usuario_id = %s AND semana_inicio >= %s AND semana_fin <= %s
              AND aplicado = FALSE
            GROUP BY tipo;
        """, (usuario_id, semana_inicio, semana_fin))
        ajustes = {r[0]: float(r[1]) for r in cursor.fetchall()}
        aumentos   = ajustes.get('aumento', 0.0)
        descuentos = ajustes.get('descuento', 0.0)

        # Saldo arrastrado de semanas anteriores sin pagar
        cursor.execute("""
            SELECT COALESCE(SUM(
                CASE WHEN monto_pagado = 0 THEN descuentos - aumentos ELSE 0 END
            ), 0)
            FROM cierres_semanales_vendedor
            WHERE usuario_id = %s;
        """, (usuario_id,))
        saldo_anterior = max(0.0, float(cursor.fetchone()[0]))

        # Registrar cierre
        cursor.execute("""
            INSERT INTO cierres_semanales_vendedor
                (usuario_id, semana_inicio, semana_fin, sueldo_base, comision,
                 aumentos, descuentos, saldo_anterior, monto_pagado, notas, voucher_url)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (usuario_id, semana_inicio) DO UPDATE SET
                monto_pagado = EXCLUDED.monto_pagado,
                notas        = EXCLUDED.notas,
                voucher_url  = EXCLUDED.voucher_url;
        """, (usuario_id, semana_inicio, semana_fin, SUELDO_BASE_VENDEDOR,
              comision, aumentos, descuentos, saldo_anterior, monto_pagado,
              notas, voucher_url))

        # Marcar ajustes como aplicados
        cursor.execute("""
            UPDATE ajustes_sueldo_vendedor
            SET aplicado = TRUE
            WHERE usuario_id = %s AND semana_inicio >= %s AND semana_fin <= %s;
        """, (usuario_id, semana_inicio, semana_fin))

        conexion.commit()
        return jsonify({
            'exito':       True,
            'vendedor':    nombre,
            'vendio':      vendio,
            'comision':    comision,
            'sueldo_base': SUELDO_BASE_VENDEDOR,
            'sueldo_ef':   sueldo_ef,
            'neto':        max(0.0, round(sueldo_ef + comision + aumentos - descuentos - saldo_anterior, 2)),
            'monto_pagado': monto_pagado,
        }), 200

    except Exception as e:
        if conexion: conexion.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conexion:
            cursor.close(); release_db_connection(conexion)
