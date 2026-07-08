"""
routes_inventario.py — Sistema de Inventario Completo
Innova Mobili ERP

Rutas:
  GET  /api/inventario/resumen           → stock agrupado por categoría/modelo/sede
  GET  /api/inventario/piezas/resumen    → piezas agrupadas por sku/medida/sede
  GET  /api/inventario/buscar/<barcode>  → buscar unidad por código de barras
  POST /api/inventario/producto/nuevo    → registrar producto entero
  POST /api/inventario/pieza/nueva       → registrar pieza a medida
  PUT  /api/inventario/<tipo>/<id>/estado → cambiar estado (traslado, venta, baja, etc.)
  GET  /api/inventario/historial/<tipo>/<id> → historial de una unidad
  GET  /api/inventario/historial/sede/<sede_id> → movimientos de una sede
  GET  /api/inventario/exportar          → Excel (.xlsx) completo, productos y piezas juntos
"""

import os
import csv
import io
from datetime import datetime
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Blueprint, request, jsonify, Response, send_file
from database import get_db_connection, release_db_connection
from auth_middleware import requiere_login, requiere_rol

inventario_bp = Blueprint('inventario', __name__)
tz_peru = pytz.timezone('America/Lima')
_schema_fotos_adicionales_listo = False

# Roles autorizados para modificar stock
ROLES_INVENTARIO = ('Admin', 'Jefe_Taller')


def _arg_bool(nombre):
    return str(request.args.get(nombre, '')).strip().lower() in ('1', 'true', 'si', 'yes')


def _arg_limit(default=None, maximo=1000):
    raw = request.args.get('limit')
    if raw in (None, ''):
        return default
    try:
        return min(max(int(raw), 1), maximo)
    except (TypeError, ValueError):
        return default

# Prefijos por categoría para el código de barras
PREFIJOS = {
    'Sofa':           'SOF',
    'Butaca':         'BUT',
    'Silla':          'SIL',
    'Espejo':         'ESP',
    'Cuadro':         'CUA',
    'Cojin':          'COJ',
    'Mesa Centro':    'MEC',
    'Consola':        'CON',
    'Esquinero':      'ESQ',
    'Florero':        'FLO',
    'Manta':          'MAN',
    'tablero':        'TAB',
    'base-comedor':   'BAC',
    'base-consola':   'BCS',
    'base-mesa-centro': 'BMC',
    'silla':            'SIL',
    'butaca':           'BUT',
}


def _conn():
    return get_db_connection()


def _rel(c):
    release_db_connection(c)


def _verificar_rol(rol):
    return rol in ROLES_INVENTARIO


def _asegurar_columna_fotos_adicionales():
    """Añade fotos_adicionales a stock_productos y stock_piezas si no existen."""
    global _schema_fotos_adicionales_listo
    if _schema_fotos_adicionales_listo:
        return
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute("ALTER TABLE stock_productos ADD COLUMN IF NOT EXISTS fotos_adicionales TEXT;")
        cur.execute("ALTER TABLE stock_productos ADD COLUMN IF NOT EXISTS observaciones TEXT;")
        cur.execute("ALTER TABLE stock_piezas ADD COLUMN IF NOT EXISTS fotos_adicionales TEXT;")
        cur.execute("ALTER TABLE stock_piezas ADD COLUMN IF NOT EXISTS foto_url TEXT;")
        # Índices de performance para queries frecuentes
        cur.execute("CREATE INDEX IF NOT EXISTS idx_stock_prod_nombre ON stock_productos (LOWER(nombre_modelo));")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_stock_prod_sede ON stock_productos (sede_id);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_stock_piezas_sku ON stock_piezas (sku_maestro);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_stock_piezas_sede ON stock_piezas (sede_id);")
        _schema_fotos_adicionales_listo = True
    except Exception as e:
        print(f"⚠️  _asegurar_columna_fotos_adicionales: {e}")
        _schema_fotos_adicionales_listo = True # Don't retry
    finally:
        if cur: cur.close()
        if conn:
            conn.autocommit = False
            release_db_connection(conn)


def _generar_codigo(cur, prefijo, tabla_col):
    """Genera el siguiente código de barras único: IM-{PREFIX}-{N:05d}"""
    cur.execute(f"SELECT COUNT(*) FROM {tabla_col};")
    n = cur.fetchone()[0] + 1
    codigo = f"IM-{prefijo}-{str(n).zfill(5)}"
    # Verificar unicidad (si ya existe, incrementar)
    cur.execute(f"SELECT 1 FROM {tabla_col} WHERE codigo_barra = %s", (codigo,))
    while cur.fetchone():
        n += 1
        codigo = f"IM-{prefijo}-{str(n).zfill(5)}"
        cur.execute(f"SELECT 1 FROM {tabla_col} WHERE codigo_barra = %s", (codigo,))
    return codigo


def _registrar_historial(cur, tipo, reg_id, barcode, evento,
                          sede_orig, sede_dest, est_ant, est_nuevo,
                          usuario_id, usuario_nombre, venta_id=None,
                          codigo_venta=None, notas=None):
    cur.execute("""
        INSERT INTO historial_inventario
            (tipo_registro, registro_id, codigo_barra, tipo_evento,
             sede_origen_id, sede_destino_id, estado_anterior, estado_nuevo,
             usuario_id, usuario_nombre, venta_id, codigo_venta, notas)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
    """, (tipo, reg_id, barcode, evento,
          sede_orig, sede_dest, est_ant, est_nuevo,
          usuario_id, usuario_nombre, venta_id, codigo_venta, notas))


# Mapa: categoría de pieza → (tabla_maestro, columna_nombre_modelo)
_MAESTRO_FOTO_MAP = {
    'tablero':          ('maestro_tableros',     'nombre_modelo'),
    'silla':            ('maestro_sillas',        'modelo'),
    'butaca':           ('maestro_butacas',       'modelo'),
    'base-comedor':     ('maestro_bases_comedor', 'modelo'),
    'base-consola':     ('maestro_bases_comedor', 'modelo'),
    'base-mesa-centro': ('maestro_bases_comedor', 'modelo'),
}


def _obtener_foto_maestro(cur, categoria, nombre_modelo):
    """
    Busca la foto_url del maestro correspondiente a una categoría y nombre de modelo.
    Devuelve la URL o URLs (str) o '' si no encuentra nada.
    """
    if not categoria or not nombre_modelo:
        return ''
    entry = _MAESTRO_FOTO_MAP.get(categoria.lower())
    if not entry:
        return ''
    tabla, col = entry
    try:
        cur.execute(
            f"SELECT foto_url FROM {tabla} WHERE LOWER({col}) = LOWER(%s) LIMIT 1",
            (nombre_modelo,)
        )
        row = cur.fetchone()
        if row and row[0]:
            return row[0]
    except Exception:
        pass
    return ''


# ─────────────────────────────────────────────────────────────────────────────
# 1. RESUMEN DE PRODUCTOS ENTEROS (pivot por sede)
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/resumen', methods=['GET'])
@requiere_login
def resumen_productos():
    categoria = request.args.get('categoria', '')
    q         = request.args.get('q', '').strip().lower()
    sede_id   = request.args.get('sede_id', '')
    stock_only = _arg_bool('stock_only')
    limit = _arg_limit()

    where, params = [], []
    if categoria:
        where.append("sp.categoria = %s"); params.append(categoria)
    if q:
        where.append("LOWER(sp.nombre_modelo) LIKE %s"); params.append(f"%{q}%")
    if stock_only:
        where.append("sp.estado = 'Disponible'")
    # FIX-SEDE-FILTRO: sede_id NO va en el WHERE de SQL. Si se filtra ahí,
    # las filas de las DEMÁS tiendas quedan completamente fuera de la
    # consulta desde antes del GROUP BY — y como la foto y el desglose
    # por tienda se calculan agregando TODAS las filas del modelo
    # (MAX(sp.foto_url), sede_stock, etc.), el resultado filtrado pierde
    # la foto (si está en otra sede) y no muestra el stock de las demás
    # tiendas, aunque sí exista. El filtro de tienda se aplica más abajo,
    # en Python, sobre el modelo ya armado con todos sus datos completos.
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        cur.execute("SELECT id, nombre FROM sedes ORDER BY id;")
        sedes = cur.fetchall()

        cur.execute(f"""
            SELECT
                sp.categoria,
                sp.nombre_modelo,
                sp.catalogo_id,
                sp.sede_id,
                se.nombre AS sede_nombre,
                COUNT(*)                                                    AS total,
                COUNT(*) FILTER (WHERE sp.estado = 'Disponible')           AS disponibles,
                COUNT(*) FILTER (WHERE sp.estado = 'Reservado')            AS reservados,
                COUNT(*) FILTER (WHERE sp.estado = 'Vendido')              AS vendidos,
                MAX(sp.foto_url)                                            AS foto_url,
                MAX(cp.foto_url)                                            AS cat_foto_url,
                MAX(cp.fotos_urls)                                          AS cat_fotos_urls,
                MAX(sp.fotos_adicionales)                                   AS stock_fotos_adicionales,
                COALESCE(sp.observaciones, MAX(cp.observaciones))           AS observaciones
            FROM stock_productos sp
            JOIN sedes se ON sp.sede_id = se.id
            LEFT JOIN catalogo_productos cp
                   ON cp.id = sp.catalogo_id
                   OR (sp.catalogo_id IS NULL
                       AND LOWER(cp.nombre_modelo) = LOWER(sp.nombre_modelo))
            {where_sql}
            GROUP BY sp.categoria, sp.nombre_modelo, sp.catalogo_id, sp.observaciones, sp.sede_id, se.nombre
            ORDER BY sp.categoria, sp.nombre_modelo, sp.observaciones, se.nombre;
        """, params)
        rows = cur.fetchall()

        # Agrupar por modelo
        # FIX: como la consulta SQL agrupa (GROUP BY) también por sede, cada
        # fila 'r' representa un modelo EN UNA SOLA sede. Antes, la lista de
        # fotos del modelo se armaba solo con los datos de la PRIMERA fila
        # encontrada (la primera sede en orden alfabético) — si esa sede no
        # tenía foto pero otra sí (p.ej. "Tienda Grande"), la tarjeta del
        # modelo completo se mostraba "Sin foto" aunque sí hubiera fotos
        # reales en otras tiendas. Ahora se recorren las fotos de TODAS las
        # filas (todas las sedes) del mismo modelo antes de decidir cuál usar.
        modelos = {}
        for r in rows:
            observaciones_row = r[13] or ""
            key = (r[0], r[1], observaciones_row)
            if key not in modelos:
                modelos[key] = {
                    "categoria":     r[0],
                    "nombre_modelo": r[1],
                    "catalogo_id":   r[2],
                    "foto_url":      "",
                    "fotos":         [],
                    "_fotos_seen":   set(),
                    "observaciones": observaciones_row,
                    "total":         0,
                    "disponibles":   0,
                    "sede_stock":    {s[1]: {"total":0,"disponibles":0} for s in sedes},
                }

            m = modelos[key]

            # Si esta fila trae un catalogo_id y el modelo aún no tiene uno, usarlo
            if not m["catalogo_id"] and r[2]:
                m["catalogo_id"] = r[2]

            # Acumular fotos de esta sede: catálogo, luego las del registro de stock
            # sp.foto_url (r[9]) es legacy, puede contener la foto del catálogo
            # en el momento del registro
            for photo_str in (r[10] or "", r[11] or "", r[9] or "", r[12] or ""):
                if not photo_str:
                    continue
                for f_url in photo_str.split('|'):
                    f = f_url.strip()
                    if f and f not in m["_fotos_seen"]:
                        m["fotos"].append(f)
                        m["_fotos_seen"].add(f)

            if not m["foto_url"] and m["fotos"]:
                m["foto_url"] = m["fotos"][0]

            m["total"]       += r[5]
            m["disponibles"] += r[6]
            m["sede_stock"][r[4]] = {
                "total":      r[5],
                "disponibles": r[6],
                "reservados":  r[7],
                "vendidos":    r[8],
            }

        # Limpiar el set interno auxiliar (no serializable a JSON)
        for m in modelos.values():
            m.pop("_fotos_seen", None)

        modelos_lista = list(modelos.values())

        # FIX-SEDE-FILTRO: el filtro de tienda se aplica AQUÍ, sobre modelos
        # ya armados con la foto y el desglose completo de todas las sedes.
        # Así, un modelo con stock en "Tienda Grande" pero cuya foto está
        # registrada en "Tienda del Medio" sigue mostrando esa foto y el
        # resto de tiendas donde también tiene stock, en vez de perderlos.
        if sede_id:
            sede_nombre_filtro = next((s[1] for s in sedes if str(s[0]) == str(sede_id)), None)
            if sede_nombre_filtro:
                modelos_lista = [
                    m for m in modelos_lista
                    if m["sede_stock"].get(sede_nombre_filtro, {}).get("total", 0) > 0
                ]

        total_modelos = len(modelos_lista)
        if limit:
            modelos_lista = modelos_lista[:limit]

        return jsonify({
            "sedes":   [s[1] for s in sedes],
            "modelos": modelos_lista,
            "total_modelos": total_modelos,
            "limit": limit,
            "truncated": bool(limit and total_modelos > limit)
        }), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# 2. RESUMEN DE PIEZAS A MEDIDA (pivot por sede)
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/piezas/resumen', methods=['GET'])
@requiere_login
def resumen_piezas():
    categoria = request.args.get('categoria', '')
    q         = request.args.get('q', '').strip().lower()
    stock_only = _arg_bool('stock_only')
    limit = _arg_limit()

    where, params = [], []
    if categoria:
        where.append("sp.categoria = %s"); params.append(categoria)
    if q:
        where.append("(LOWER(sp.nombre_modelo) LIKE %s OR LOWER(sp.material) LIKE %s)")
        params.extend([f"%{q}%", f"%{q}%"])
    if stock_only:
        where.append("sp.estado = 'Disponible'")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        cur.execute("SELECT id, nombre FROM sedes ORDER BY id;")
        sedes = cur.fetchall()

        cur.execute(f"""
            SELECT
                sp.categoria, sp.sku_maestro, sp.nombre_modelo,
                sp.material, sp.color_acabado, sp.forma,
                sp.largo_cm, sp.ancho_cm, sp.alto_cm,
                sp.sede_id, se.nombre AS sede_nombre,
                COUNT(*) FILTER (WHERE sp.estado = 'Disponible') AS disponibles,
                COUNT(*) AS total,
                COALESCE(
                    MAX(mt.foto_url),
                    MAX(ms.foto_url),
                    MAX(mb.foto_url),
                    MAX(mbu.foto_url)
                ) AS foto_maestro,
                MAX(sp.fotos_adicionales) AS fotos_adicionales
            FROM stock_piezas sp
            JOIN sedes se ON sp.sede_id = se.id
            LEFT JOIN maestro_tableros      mt  ON sp.categoria = 'tablero'
                AND (mt.sku = sp.sku_maestro OR LOWER(mt.nombre_modelo) = LOWER(sp.nombre_modelo))
            LEFT JOIN maestro_sillas        ms  ON sp.categoria = 'silla'
                AND (ms.sku = sp.sku_maestro OR LOWER(ms.modelo) = LOWER(sp.nombre_modelo))
            LEFT JOIN maestro_bases_comedor mb  ON sp.categoria IN ('base-comedor','base-consola','base-mesa-centro')
                AND (mb.sku = sp.sku_maestro OR LOWER(mb.modelo) = LOWER(sp.nombre_modelo))
            LEFT JOIN maestro_butacas       mbu ON sp.categoria = 'butaca'
                AND (mbu.sku = sp.sku_maestro OR LOWER(mbu.modelo) = LOWER(sp.nombre_modelo))
            {where_sql}
            GROUP BY sp.categoria, sp.sku_maestro, sp.nombre_modelo,
                     sp.material, sp.color_acabado, sp.forma,
                     sp.largo_cm, sp.ancho_cm, sp.alto_cm,
                     sp.sede_id, se.nombre
            ORDER BY sp.categoria, sp.sku_maestro, sp.forma, se.nombre;
        """, params)
        rows = cur.fetchall()

        grupos = {}
        for r in rows:
            key = (r[1], r[5], r[6], r[7], r[8])  # sku+forma+medidas
            if key not in grupos:
                foto_maestro      = r[13] or ""
                fotos_adicionales = r[14] or ""
                todas_fotos = []
                for f in foto_maestro.split('|'):
                    f = f.strip()
                    if f:
                        todas_fotos.append(f)
                for f in fotos_adicionales.split('|'):
                    f = f.strip()
                    if f and f not in todas_fotos:
                        todas_fotos.append(f)
                grupos[key] = {
                    "categoria":     r[0],
                    "sku_maestro":   r[1],
                    "nombre_modelo": r[2],
                    "material":      r[3],
                    "color_acabado": r[4],
                    "forma":         r[5],
                    "largo_cm":      float(r[6]) if r[6] else None,
                    "ancho_cm":      float(r[7]) if r[7] else None,
                    "alto_cm":       float(r[8]) if r[8] else None,
                    "foto_url":      todas_fotos[0] if todas_fotos else "",
                    "fotos":         todas_fotos,
                    "total":         0,
                    "disponibles":   0,
                    "sede_stock":    {s[1]: 0 for s in sedes},
                }
            grupos[key]["disponibles"] += r[11]
            grupos[key]["total"]       += r[12]
            grupos[key]["sede_stock"][r[10]] = r[11]  # disponibles por sede

        piezas_lista = list(grupos.values())
        total_piezas = len(piezas_lista)
        if limit:
            piezas_lista = piezas_lista[:limit]

        return jsonify({
            "sedes":  [s[1] for s in sedes],
            "piezas": piezas_lista,
            "total_piezas": total_piezas,
            "limit": limit,
            "truncated": bool(limit and total_piezas > limit)
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# 11. UNIDADES DE PIEZAS DISPONIBLES POR SKU (para el picker del carrito)
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/piezas/disponibles/<sku_maestro>', methods=['GET'])
@requiere_login
def unidades_piezas_disponibles_por_sku(sku_maestro):
    sede_id = request.args.get('sede_id', '')
    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        where_extra = ""
        params = [sku_maestro]
        if sede_id:
            where_extra = " AND sp.sede_id = %s"
            params.append(sede_id)

        cur.execute(f"""
            SELECT sp.id, sp.codigo_barra, se.nombre AS sede,
                   sp.material, sp.color_acabado, sp.forma, sp.largo_cm, sp.ancho_cm, sp.alto_cm,
                   sp.costo_ingreso,
                   TO_CHAR(sp.fecha_ingreso, 'DD/MM/YYYY') AS fecha_ingreso
            FROM stock_piezas sp
            JOIN sedes se ON sp.sede_id = se.id
            WHERE sp.sku_maestro = %s
              AND sp.estado = 'Disponible'
              {where_extra}
            ORDER BY se.nombre, sp.fecha_ingreso;
        """, params)

        unidades = []
        for r in cur.fetchall():
            medida = ""
            if r[5] == 'Circular': medida = f"⌀ {r[6]} cm" if r[6] else 'Circular'
            elif r[5] == 'Rectangular':
                l = r[6] if r[6] else '?'
                a = f" x {r[7]}" if r[7] else ''
                h = f" / H:{r[8]}" if r[8] else ''
                medida = f"{l}{a} cm{h}"
            else: medida = f"{r[6]} cm" if r[6] else 'Irregular'

            label_parts = [r[1], r[2], medida]
            if r[3]: label_parts.append(r[3])
            if r[4]: label_parts.append(r[4])

            unidades.append({
                'id':            r[0],
                'codigo_barra':  r[1],
                'sede':          r[2],
                'label':         ' — '.join(label_parts),
            })
        return jsonify(unidades), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# 3. BUSCAR POR CÓDIGO DE BARRAS
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/buscar/<barcode>', methods=['GET'])
@requiere_login
def buscar_por_barcode(barcode):
    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        # 1. Buscar en stock_productos (JOIN al catálogo para fotos del modelo maestro)
        cur.execute("""
            SELECT sp.id, sp.codigo_barra, sp.nombre_modelo, sp.categoria,
                   sp.color_tela, sp.acabado, sp.estado, se.nombre AS sede,
                   sp.foto_url, sp.costo_ingreso, sp.precio_venta,
                   sp.fecha_ingreso, 'producto' AS tipo, sp.fotos_adicionales,
                   COALESCE(cp.foto_url, '')    AS cat_foto_url,
                   COALESCE(cp.fotos_urls, '')  AS cat_fotos_urls
            FROM stock_productos sp
            LEFT JOIN sedes se ON sp.sede_id = se.id
            LEFT JOIN catalogo_productos cp
                   ON cp.id = COALESCE(sp.catalogo_id,
                        (SELECT id FROM catalogo_productos
                         WHERE LOWER(nombre_modelo) = LOWER(sp.nombre_modelo)
                         LIMIT 1))
            WHERE sp.codigo_barra = %s;
        """, (barcode,))
        row = cur.fetchone()
        if row:
            # Fotos: catálogo primero, luego stock (pipe-sep), luego adicionales
            todas_fotos = []
            seen_fotos = set()

            def add_photos(photo_str):
                if not photo_str: return
                for f_url in photo_str.split('|'):
                    f = f_url.strip()
                    if f and f not in seen_fotos:
                        todas_fotos.append(f)
                        seen_fotos.add(f)

            add_photos(row[14]) # cat_foto_url
            add_photos(row[15]) # cat_fotos_urls
            add_photos(row[8])  # sp.foto_url (legacy)
            add_photos(row[13]) # sp.fotos_adicionales

            return jsonify({
                "tipo":              "producto",
                "id":                row[0],
                "codigo_barra":      row[1],
                "nombre_modelo":     row[2],
                "categoria":         row[3],
                "color_tela":        row[4],
                "acabado":           row[5],
                "estado":            row[6],
                "sede":              row[7],
                "foto_url":          todas_fotos[0] if todas_fotos else "",
                "fotos":             todas_fotos,
                "costo_ingreso":     float(row[9]) if row[9] else None,
                "precio_venta":      float(row[10]) if row[10] else None,
                "fecha_ingreso":     row[11].strftime('%d/%m/%Y') if row[11] else None,
                "fotos_adicionales": row[13] or "",
            }), 200

        # 2. Buscar en stock_piezas
        cur.execute("""
            SELECT sp.id, sp.codigo_barra, sp.nombre_modelo, sp.categoria,
                   sp.material, sp.color_acabado, sp.estado, se.nombre AS sede, sp.forma,
                   sp.largo_cm, sp.ancho_cm, sp.alto_cm, sp.costo_ingreso,
                   sp.fecha_ingreso, 'pieza' AS tipo, sp.fotos_adicionales, sp.sku_maestro
            FROM stock_piezas sp
            LEFT JOIN sedes se ON sp.sede_id = se.id
            WHERE sp.codigo_barra = %s;
        """, (barcode,))
        row = cur.fetchone()
        if row:
            fotos_adicionales_str = row[15] or ""
            sku_maestro_pieza     = row[16] or ""
            categoria_pieza       = (row[3] or '').lower()
            nombre_pieza          = row[2] or ''

            # Buscar foto del maestro según categoría (query separada, sin CASE)
            foto_maestro = ''
            try:
                tabla_maestro, col_nombre = '', ''
                if categoria_pieza == 'tablero': tabla_maestro, col_nombre = 'maestro_tableros', 'nombre_modelo'
                elif categoria_pieza == 'silla': tabla_maestro, col_nombre = 'maestro_sillas', 'modelo'
                elif categoria_pieza == 'butaca': tabla_maestro, col_nombre = 'maestro_butacas', 'modelo'
                elif 'base' in categoria_pieza: tabla_maestro, col_nombre = 'maestro_bases_comedor', 'modelo'

                if tabla_maestro:
                    cur.execute(
                        f"SELECT foto_url FROM {tabla_maestro} "
                        f"WHERE sku = %s OR LOWER({col_nombre}) = LOWER(%s) LIMIT 1",
                        (sku_maestro_pieza, nombre_pieza)
                    )
                    r_foto = cur.fetchone()
                    if r_foto and r_foto[0]:
                        foto_maestro = r_foto[0]
            except Exception:
                pass

            # Maestro primero, luego fotos adicionales subidas al registrar
            todas_fotos = []
            seen_fotos = set()
            def add_photos(photo_str):
                if not photo_str: return
                for f_url in photo_str.split('|'):
                    f = f_url.strip()
                    if f and f not in seen_fotos:
                        todas_fotos.append(f)
                        seen_fotos.add(f)

            add_photos(foto_maestro)
            add_photos(fotos_adicionales_str)

            return jsonify({
                "tipo":              "pieza",
                "id":                row[0],
                "codigo_barra":      row[1],
                "nombre_modelo":     row[2],
                "categoria":         row[3],
                "material":          row[4],
                "color_acabado":     row[5],
                "estado":            row[6],
                "sede":              row[7],
                "forma":             row[8],
                "largo_cm":          float(row[9]) if row[9] else None,
                "ancho_cm":          float(row[10]) if row[10] else None,
                "alto_cm":           float(row[11]) if row[11] else None,
                "costo_ingreso":     float(row[12]) if row[12] else None,
                "fecha_ingreso":     row[13].strftime('%d/%m/%Y') if row[13] else None,
                "foto_url":          todas_fotos[0] if todas_fotos else "",
                "fotos":             todas_fotos,
                "fotos_adicionales": fotos_adicionales_str,
            }), 200

        # 3. ← NUEVO: Buscar en stock_unidades
        cur.execute("""
            SELECT su.id, su.codigo_barra, su.nombre_modelo, su.categoria,
                   su.color_tela, su.acabado, su.estado, se.nombre AS sede,
                   COALESCE(cp.foto_url, '') AS foto_url, su.costo_ingreso, su.fecha_ingreso,
                   'unidad' AS tipo
            FROM stock_unidades su
            LEFT JOIN sedes se ON su.sede_id = se.id
            LEFT JOIN catalogo_productos cp
                   ON LOWER(cp.nombre_modelo) = LOWER(su.nombre_modelo)
            WHERE su.codigo_barra = %s;
        """, (barcode,))
        row = cur.fetchone()
        if row:
            return jsonify({
                "tipo": "producto",   # tratarlo como producto para que el frontend lo muestre igual
                "id": row[0], "codigo_barra": row[1],
                "nombre_modelo": row[2], "categoria": row[3],
                "color_tela": row[4], "acabado": row[5], "estado": row[6],
                "sede": row[7], "foto_url": row[8] or "",
                "costo_ingreso": float(row[9]) if row[9] else None,
                "precio_venta": None,
                "fecha_ingreso": row[10].strftime('%d/%m/%Y') if row[10] else None,
            }), 200

        return jsonify({'error': 'Código no encontrado'}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: cur.close(); _rel(conn)
# ─────────────────────────────────────────────────────────────────────────────
# 4. REGISTRAR PRODUCTO ENTERO
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/producto/nuevo', methods=['POST'])
@requiere_login
def registrar_producto():
    _asegurar_columna_fotos_adicionales()
    data = request.json or {}

    if not _verificar_rol(data.get('usuario_rol', '')):
        return jsonify({'error': 'Sin permisos. Solo Admin o Jefe de Taller.'}), 403

    required = ['nombre_modelo', 'categoria', 'sede_id', 'usuario_id']
    missing  = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({'error': f'Campos faltantes: {", ".join(missing)}'}), 400

    cantidad = max(1, min(50, int(data.get('cantidad', 1))))

    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        unidades_creadas = []
        for _ in range(cantidad):
            prefijo = PREFIJOS.get(data['categoria'], 'PRD')
            barcode = _generar_codigo(cur, prefijo, 'stock_productos')

            cur.execute("""
                INSERT INTO stock_productos
                    (catalogo_id, nombre_modelo, categoria, codigo_barra,
                     color_tela, acabado, observaciones, foto_url, fotos_adicionales,
                     sede_id, estado, costo_ingreso, precio_venta, creado_por)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Disponible',%s,%s,%s)
                RETURNING id;
            """, (
                data.get('catalogo_id'),
                data['nombre_modelo'],
                data['categoria'],
                barcode,
                data.get('color_tela'),
                data.get('acabado'),
                data.get('observaciones'),
                data.get('foto_url'),
                data.get('fotos_adicionales'),
                data['sede_id'],
                data.get('costo_ingreso'),
                data.get('precio_venta'),
                data['usuario_id'],
            ))
            nuevo_id = cur.fetchone()[0]

            _registrar_historial(
                cur, 'producto', nuevo_id, barcode,
                'Ingreso', None, data['sede_id'],
                None, 'Disponible',
                data['usuario_id'], data.get('usuario_nombre', ''),
                notas=f"Ingreso inicial. {data.get('observaciones', '')}"
            )
            unidades_creadas.append({'id': nuevo_id, 'codigo_barra': barcode})

        conn.commit()
        # Compatibilidad: si solo se registró 1, devolver también codigo_barra directo
        resp = {'exito': True, 'unidades': unidades_creadas}
        if len(unidades_creadas) == 1:
            resp['id'] = unidades_creadas[0]['id']
            resp['codigo_barra'] = unidades_creadas[0]['codigo_barra']
        return jsonify(resp), 201

    except Exception as e:
        if conn: conn.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# 5. REGISTRAR PIEZA A MEDIDA
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/pieza/nueva', methods=['POST'])
@requiere_login
def registrar_pieza():
    _asegurar_columna_fotos_adicionales()
    data = request.json or {}

    if not _verificar_rol(data.get('usuario_rol', '')):
        return jsonify({'error': 'Sin permisos. Solo Admin o Jefe de Taller.'}), 403

    required = ['sku_maestro', 'nombre_modelo', 'categoria', 'forma', 'sede_id', 'usuario_id']
    missing  = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({'error': f'Campos faltantes: {", ".join(missing)}'}), 400

    # Puede registrarse más de 1 unidad a la vez (cantidad)
    cantidad = max(1, int(data.get('cantidad', 1)))

    conn = None
    try:
        conn = _conn(); cur = conn.cursor()
        prefijo  = PREFIJOS.get(data['categoria'], 'PIE')
        generados = []

        for _ in range(cantidad):
            barcode = _generar_codigo(cur, prefijo, 'stock_piezas')
            cur.execute("""
                INSERT INTO stock_piezas
                    (categoria, sku_maestro, nombre_modelo, material, color_acabado,
                    codigo_barra, forma, largo_cm, ancho_cm, alto_cm, fotos_adicionales,
                    foto_url, sede_id, estado, costo_ingreso, proveedor, usuario_ingreso_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Disponible',%s,%s,%s)
                RETURNING id;
            """, (
                data['categoria'],
                data['sku_maestro'],
                data['nombre_modelo'],
                data.get('material'),
                data.get('color_acabado'),
                barcode,
                data['forma'],
                data.get('largo_cm'),
                data.get('ancho_cm'),
                data.get('alto_cm'),
                data.get('fotos_adicionales'),
                data.get('foto_url'),          # foto del maestro de materiales
                data['sede_id'],
                data.get('costo_ingreso'),
                data.get('proveedor'),
                data['usuario_id'],
            ))
            nuevo_id = cur.fetchone()[0]
            _registrar_historial(
                cur, 'pieza', nuevo_id, barcode,
                'Ingreso', None, data['sede_id'],
                None, 'Disponible',
                data['usuario_id'], data.get('usuario_nombre', ''),
                notas=data.get('notas', 'Ingreso inicial')
            )
            generados.append({'id': nuevo_id, 'codigo_barra': barcode})

        conn.commit()
        return jsonify({'exito': True, 'unidades': generados}), 201

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# 6. CAMBIAR ESTADO (Traslado, Venta, Baja, Reserva, etc.)
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/<tipo>/<int:reg_id>/estado', methods=['PUT'])
@requiere_login
def cambiar_estado(tipo, reg_id):
    """
    tipo: 'producto' o 'pieza'
    Body JSON:
    {
      "estado_nuevo":   "En Traslado",
      "sede_destino_id": 3,          (requerido si es traslado)
      "tipo_evento":    "Traslado",
      "usuario_id":     5,
      "usuario_rol":    "Admin",
      "usuario_nombre": "Carlos",
      "notas":          "Traslado para exhibición",
      "venta_id":       null,
      "codigo_venta":   null
    }
    """
    if tipo not in ('producto', 'pieza'):
        return jsonify({'error': 'tipo debe ser producto o pieza'}), 400

    data = request.json or {}
    if not _verificar_rol(data.get('usuario_rol', '')):
        return jsonify({'error': 'Sin permisos. Solo Admin o Jefe de Taller.'}), 403

    tabla = 'stock_productos' if tipo == 'producto' else 'stock_piezas'
    conn  = None
    try:
        conn = _conn(); cur = conn.cursor()

        cur.execute(
            f"SELECT estado, sede_id, codigo_barra FROM {tabla} WHERE id = %s",
            (reg_id,)
        )
        row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Registro no encontrado'}), 404

        estado_ant, sede_orig, barcode = row
        estado_nuevo   = data.get('estado_nuevo', estado_ant)
        sede_destino   = data.get('sede_destino_id')
        nueva_sede_id  = sede_destino if sede_destino else sede_orig

        # Actualizar estado y sede si hay traslado
        cur.execute(f"""
            UPDATE {tabla}
               SET estado = %s, sede_id = %s, actualizado_en = NOW()
             WHERE id = %s;
        """, (estado_nuevo, nueva_sede_id, reg_id))

        _registrar_historial(
            cur, tipo, reg_id, barcode,
            data.get('tipo_evento', 'Ajuste'),
            sede_orig, sede_destino,
            estado_ant, estado_nuevo,
            data['usuario_id'], data.get('usuario_nombre', ''),
            data.get('venta_id'), data.get('codigo_venta'),
            data.get('notas')
        )

        conn.commit()
        return jsonify({'exito': True, 'estado_nuevo': estado_nuevo}), 200

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# 7. HISTORIAL DE UNA UNIDAD
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/historial/<tipo>/<int:reg_id>', methods=['GET'])
@requiere_login
def historial_unidad(tipo, reg_id):
    conn = None
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            SELECT h.tipo_evento, h.estado_anterior, h.estado_nuevo,
                   so.nombre AS sede_origen, sd.nombre AS sede_destino,
                   h.usuario_nombre, h.codigo_venta, h.notas,
                   TO_CHAR(h.fecha AT TIME ZONE 'America/Lima', 'DD/MM/YYYY HH24:MI') AS fecha
            FROM historial_inventario h
            LEFT JOIN sedes so ON h.sede_origen_id  = so.id
            LEFT JOIN sedes sd ON h.sede_destino_id = sd.id
            WHERE h.tipo_registro = %s AND h.registro_id = %s
            ORDER BY h.fecha DESC;
        """, (tipo, reg_id))
        rows = cur.fetchall()
        return jsonify([{
            "evento":        r[0], "estado_ant": r[1], "estado_nuevo": r[2],
            "sede_origen":   r[3], "sede_destino": r[4],
            "usuario":       r[5], "venta":       r[6],
            "notas":         r[7], "fecha":        r[8],
        } for r in rows]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# 8. MOVIMIENTOS DE UNA SEDE
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/historial/sede/<int:sede_id>', methods=['GET'])
@requiere_login
def historial_sede(sede_id):
    limite = int(request.args.get('limite', 50))
    conn   = None
    try:
        conn = _conn(); cur = conn.cursor()
        cur.execute("""
            SELECT h.tipo_registro, h.codigo_barra, h.tipo_evento,
                   h.estado_anterior, h.estado_nuevo,
                   so.nombre AS sede_origen, sd.nombre AS sede_destino,
                   h.usuario_nombre, h.codigo_venta, h.notas,
                   TO_CHAR(h.fecha AT TIME ZONE 'America/Lima', 'DD/MM/YYYY HH24:MI') AS fecha
            FROM historial_inventario h
            LEFT JOIN sedes so ON h.sede_origen_id  = so.id
            LEFT JOIN sedes sd ON h.sede_destino_id = sd.id
            WHERE h.sede_origen_id = %s OR h.sede_destino_id = %s
            ORDER BY h.fecha DESC
            LIMIT %s;
        """, (sede_id, sede_id, limite))
        rows = cur.fetchall()
        return jsonify([{
            "tipo": r[0], "codigo_barra": r[1], "evento": r[2],
            "estado_ant": r[3], "estado_nuevo": r[4],
            "sede_origen": r[5], "sede_destino": r[6],
            "usuario": r[7], "venta": r[8], "notas": r[9], "fecha": r[10],
        } for r in rows]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# 9. EXPORTAR EXCEL COMPLETO (productos + piezas, todo en una sola tabla)
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/exportar', methods=['GET'])
@requiere_rol('Admin', 'Jefe_Taller')
def exportar_inventario():
    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        # Productos enteros + Piezas a medida, unidos en un solo conjunto
        # (mismo orden de columnas para ambos, por eso los NULL de relleno)
        cur.execute("""
            SELECT 'Producto', sp.categoria, sp.nombre_modelo, sp.codigo_barra,
                   sp.color_tela, sp.acabado, sp.estado, se.nombre,
                   sp.costo_ingreso, sp.precio_venta,
                   sp.fecha_ingreso, sp.observaciones,
                   NULL, NULL, NULL, NULL, NULL
            FROM stock_productos sp JOIN sedes se ON sp.sede_id = se.id
            UNION ALL
            SELECT 'Pieza', sp.categoria, sp.nombre_modelo, sp.codigo_barra,
                   sp.material, sp.color_acabado, sp.estado, se.nombre,
                   sp.costo_ingreso, NULL,
                   sp.fecha_ingreso, NULL,
                   sp.forma, sp.largo_cm, sp.ancho_cm, sp.alto_cm,
                   sp.proveedor
            FROM stock_piezas sp JOIN sedes se ON sp.sede_id = se.id
            ORDER BY 1, 2, 3;
        """)
        filas = cur.fetchall()

        # ── Construir el Excel ──────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventario'

        headers = [
            'Tipo', 'Categoría', 'Modelo', 'Código Barras',
            'Color/Material', 'Acabado/Color', 'Estado', 'Sede',
            'Costo Ingreso', 'Precio Venta', 'Fecha Ingreso', 'Observaciones',
            'Forma', 'Largo/Diám (cm)', 'Ancho (cm)', 'Alto (cm)', 'Proveedor'
        ]

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="0F172A")
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin        = Side(style="thin", color="CBD5E0")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center; cell.border = border

        anchos = [10, 14, 22, 16, 16, 16, 14, 16, 13, 13, 14, 28, 12, 14, 12, 12, 20]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

        fill_par   = PatternFill("solid", fgColor="F8FAFC")
        fill_impar = PatternFill("solid", fgColor="FFFFFF")

        for row_num, f in enumerate(filas, 2):
            fill = fill_par if row_num % 2 == 0 else fill_impar
            valores = [
                f[0], f[1], f[2], f[3],
                f[4], f[5], f[6], f[7],
                float(f[8]) if f[8] is not None else None,
                float(f[9]) if f[9] is not None else None,
                f[10].strftime('%d/%m/%Y') if f[10] else '',
                f[11],
                f[12],
                float(f[13]) if f[13] is not None else None,
                float(f[14]) if f[14] is not None else None,
                float(f[15]) if f[15] is not None else None,
                f[16],
            ]
            for col, val in enumerate(valores, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill = fill; cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:Q{max(len(filas) + 1, 1)}"

        buffer = io.BytesIO()
        wb.save(buffer); buffer.seek(0)

        fecha = datetime.now(tz_peru).strftime('%Y%m%d_%H%M')
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'inventario_{fecha}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)

# ─────────────────────────────────────────────────────────────────────────────
# 10. UNIDADES DISPONIBLES POR CATÁLOGO (para el picker del carrito)
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/disponibles/<int:catalogo_id>', methods=['GET'])
@requiere_login
def unidades_disponibles_por_catalogo(catalogo_id):
    """
    Devuelve todas las unidades de stock_productos con estado='Disponible'
    que pertenecen al producto de catálogo indicado.

    Usado por addStockItemToCart() en catalogo.js para que el vendedor
    pueda elegir la pieza física exacta que va a vender.

    Query param opcional:  ?sede_id=3  → filtrar por tienda
    """
    sede_id = request.args.get('sede_id', '')
    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        where_extra = ""
        params = [catalogo_id]
        if sede_id:
            where_extra = " AND sp.sede_id = %s"
            params.append(sede_id)

        cur.execute(f"""
            SELECT sp.id, sp.codigo_barra, se.nombre AS sede,
                   sp.color_tela, sp.acabado, sp.observaciones,
                   sp.costo_ingreso, sp.precio_venta,
                   TO_CHAR(sp.fecha_ingreso, 'DD/MM/YYYY') AS fecha_ingreso
            FROM stock_productos sp
            JOIN sedes se ON sp.sede_id = se.id
            WHERE sp.catalogo_id = %s
              AND sp.estado = 'Disponible'
              {where_extra}
            ORDER BY se.nombre, sp.fecha_ingreso;
        """, params)

        unidades = []
        for r in cur.fetchall():
            # Etiqueta descriptiva para el picker
            label_parts = [r[1], r[2]]                   # código + sede
            if r[3]: label_parts.append(r[3])            # color/tela
            if r[4]: label_parts.append(r[4])            # acabado
            if r[5]: label_parts.append(r[5])            # observaciones
            unidades.append({
                'id':            r[0],
                'codigo_barra':  r[1],
                'sede':          r[2],
                'color_tela':    r[3] or '',
                'acabado':       r[4] or '',
                'observaciones': r[5] or '',
                'costo_ingreso': float(r[6]) if r[6] else None,
                'precio_venta':  float(r[7]) if r[7] else None,
                'fecha_ingreso': r[8] or '',
                'label':         ' — '.join(label_parts),
            })

        return jsonify(unidades), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)

# ─────────────────────────────────────────────────────────────────────────────
# ELIMINAR ITEM (Pruebas)
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/<tipo>/<int:reg_id>', methods=['DELETE'])
@requiere_rol('Admin')
def eliminar_item_inventario(tipo, reg_id):
    if tipo not in ('producto', 'pieza'):
        return jsonify({'error': 'tipo debe ser producto o pieza'}), 400

    tabla = 'stock_productos' if tipo == 'producto' else 'stock_piezas'
    conn = None
    try:
        conn = _conn()
        cur = conn.cursor()

        # Borrar del historial primero para evitar problemas de relación
        cur.execute("DELETE FROM historial_inventario WHERE tipo_registro = %s AND registro_id = %s", (tipo, reg_id))
        
        # Eliminar el registro de stock físico
        cur.execute(f"DELETE FROM {tabla} WHERE id = %s RETURNING id", (reg_id,))
        if not cur.fetchone():
            return jsonify({'error': 'Registro no encontrado'}), 404

        conn.commit()
        return jsonify({'exito': True}), 200
    except Exception as e:
        if conn: conn.rollback()
        err_msg = str(e)
        if "foreign key" in err_msg.lower() or "llave foránea" in err_msg.lower():
            err_msg = "No se puede eliminar porque este registro ya está vinculado a una venta o proceso activo."
        return jsonify({'error': err_msg}), 500
    finally:
        if conn: _rel(conn)

# ─────────────────────────────────────────────────────────────────────────────
# LISTAR UNIDADES DE UN MODELO (Para eliminar pruebas/gestión)
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# NOTA (limpieza julio 2026): este archivo tenía una segunda definición de
# POST /api/inventario/etiquetas-disponibles que colisionaba con la de
# routes_produccion.py (misma ruta y método registradas en dos blueprints).
# Como produccion_bp se registra antes que inventario_bp en app.py, la versión
# de aquí NUNCA se ejecutaba (Flask/Werkzeug usa la primera regla que matchea).
# Se eliminó por ser además menos precisa: matcheaba solo por nombre_modelo,
# mientras que la de routes_produccion.py matchea por catalogo_id (productos
# enteros) o por sku_maestro+forma+medidas exactas (piezas a medida), que es
# lo que realmente envía inventario.js. La versión activa vive en:
#   routes_produccion.py → obtener_etiquetas_disponibles()
# ─────────────────────────────────────────────────────────────────────────────

@inventario_bp.route('/api/inventario/unidades-modelo', methods=['GET'])
@requiere_login
def unidades_modelo():
    tipo = request.args.get('tipo')
    nombre_modelo = request.args.get('modelo', '')
    
    conn = None
    try:
        conn = _conn()
        cur = conn.cursor()
        
        tabla = 'stock_productos' if tipo == 'producto' else 'stock_piezas'
        
        cur.execute(f"""
            SELECT sp.id, sp.codigo_barra, se.nombre AS sede, sp.estado, sp.fecha_ingreso
            FROM {tabla} sp
            LEFT JOIN sedes se ON sp.sede_id = se.id
            WHERE sp.nombre_modelo = %s
            ORDER BY sp.fecha_ingreso DESC
        """, (nombre_modelo,))
            
        unidades = [{
            'id': r[0], 'codigo_barra': r[1],
            'sede': r[2] or 'Sin sede', 'estado': r[3],
            'fecha_ingreso': r[4].strftime('%d/%m/%Y') if r[4] else ''
        } for r in cur.fetchall()]
            
        return jsonify(unidades), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: _rel(conn)

# ─────────────────────────────────────────────────────────────────────────────
# VENTA DIRECTA DESDE STOCK EN TIENDA
# ─────────────────────────────────────────────────────────────────────────────

def _asegurar_tabla_ventas_tienda(cur):
    """Crea la tabla ventas_tienda si no existe (migración lazy)."""
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ventas_tienda (
            id              SERIAL PRIMARY KEY,
            fecha           TIMESTAMP DEFAULT NOW(),
            usuario_id      INTEGER,
            usuario_nombre  VARCHAR(120),
            tipo_registro   VARCHAR(20) DEFAULT 'producto',  -- 'producto' o 'pieza'
            registro_id     INTEGER NOT NULL,                -- id en stock_productos o stock_piezas
            codigo_barra    VARCHAR(80),
            nombre_producto VARCHAR(200),
            categoria       VARCHAR(80),
            foto_url        TEXT,
            sede_nombre     VARCHAR(120),
            precio_venta    NUMERIC(10,2) NOT NULL,
            observaciones   TEXT
        );
    """)


@inventario_bp.route('/api/inventario/venta-directa', methods=['POST'])
@requiere_login
def venta_directa_tienda():
    """
    Registra una venta directa de un producto de stock de tienda.
    Marca la unidad como 'Vendido' y guarda el registro en ventas_tienda.

    Body JSON:
    {
        "tipo":            "producto" | "pieza",
        "registro_id":     123,
        "precio_venta":    350.00,
        "usuario_id":      5,
        "usuario_nombre":  "Rommel",
        "nombre_producto": "Cojin peluche",
        "categoria":       "Cojin",
        "foto_url":        "https://...",
        "sede_nombre":     "Tienda de Plaza Vea",
        "codigo_barra":    "COJ-0023",
        "observaciones":   ""
    }
    """
    data = request.json or {}

    tipo         = data.get('tipo', 'producto')
    registro_id  = data.get('registro_id')
    precio_venta = data.get('precio_venta')

    if not registro_id or precio_venta is None:
        return jsonify({'error': 'Faltan campos obligatorios: registro_id, precio_venta'}), 400
    if tipo not in ('producto', 'pieza'):
        return jsonify({'error': 'tipo debe ser producto o pieza'}), 400
    try:
        precio_venta = float(precio_venta)
        if precio_venta < 0:
            raise ValueError()
    except (ValueError, TypeError):
        return jsonify({'error': 'precio_venta inválido'}), 400

    tabla = 'stock_productos' if tipo == 'producto' else 'stock_piezas'
    conn  = None
    try:
        conn = _conn()
        cur  = conn.cursor()

        # Verificar que la unidad existe y está disponible
        cur.execute(
            f"SELECT estado, sede_id, codigo_barra FROM {tabla} WHERE id = %s",
            (registro_id,)
        )
        row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Unidad no encontrada'}), 404
        if row[0] != 'Disponible':
            return jsonify({'error': f'La unidad ya no está disponible (estado: {row[0]})'}), 409

        estado_ant  = row[0]
        sede_orig   = row[1]
        codigo_barra = data.get('codigo_barra') or row[2]

        # Marcar como Vendido en la tabla de stock
        cur.execute(
            f"UPDATE {tabla} SET estado = 'Vendido', actualizado_en = NOW() WHERE id = %s",
            (registro_id,)
        )

        # Registrar en historial_inventario (para trazabilidad)
        _registrar_historial(
            cur, tipo, registro_id, codigo_barra,
            'Venta Directa',
            sede_orig, None,
            estado_ant, 'Vendido',
            data.get('usuario_id'), data.get('usuario_nombre', ''),
            None, None,
            f"Venta directa S/ {precio_venta:.2f}"
        )

        # Guardar en ventas_tienda
        _asegurar_tabla_ventas_tienda(cur)
        cur.execute("""
            INSERT INTO ventas_tienda
                (usuario_id, usuario_nombre, tipo_registro, registro_id,
                 codigo_barra, nombre_producto, categoria, foto_url,
                 sede_nombre, precio_venta, observaciones)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id;
        """, (
            data.get('usuario_id'),
            data.get('usuario_nombre', ''),
            tipo,
            registro_id,
            codigo_barra,
            data.get('nombre_producto', ''),
            data.get('categoria', ''),
            data.get('foto_url', ''),
            data.get('sede_nombre', ''),
            precio_venta,
            data.get('observaciones', '')
        ))
        venta_id = cur.fetchone()[0]
        conn.commit()

        return jsonify({'exito': True, 'venta_tienda_id': venta_id}), 200

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


# ─────────────────────────────────────────────────────────────────────────────
# AJUSTAR CANTIDAD DE STOCK (agregar o quitar unidades de un modelo/sede)
# ─────────────────────────────────────────────────────────────────────────────
@inventario_bp.route('/api/inventario/stock-producto/cantidad', methods=['PATCH'])
@requiere_login
def ajustar_cantidad_stock():
    """
    Ajusta la cantidad disponible de un modelo en una sede.
    Si la nueva cantidad es mayor → inserta las unidades faltantes.
    Si es menor → elimina las unidades 'Disponible' sobrantes (las más recientes).
    No toca unidades Vendidas, Reservadas o en Traslado.

    Body JSON:
    {
        "nombre_modelo": "Sofá Roma",
        "categoria":     "Sofa",
        "catalogo_id":   12,          (puede ser null)
        "sede_id":       3,
        "cantidad_nueva": 5,
        "usuario_id":    7,
        "usuario_nombre": "Carlos"
    }
    """
    _asegurar_columna_fotos_adicionales()
    data = request.json or {}
    required = ['nombre_modelo', 'categoria', 'sede_id', 'cantidad_nueva', 'usuario_id']
    missing  = [f for f in required if data.get(f) is None]
    if missing:
        return jsonify({'error': f'Campos faltantes: {", ".join(missing)}'}), 400

    cantidad_nueva = int(data['cantidad_nueva'])
    if cantidad_nueva < 0:
        return jsonify({'error': 'La cantidad no puede ser negativa'}), 400

    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        # Contar unidades Disponibles actuales del modelo en esa sede
        cur.execute("""
            SELECT id FROM stock_productos
            WHERE LOWER(nombre_modelo) = LOWER(%s)
              AND sede_id = %s
              AND estado = 'Disponible'
            ORDER BY id ASC;
        """, (data['nombre_modelo'], data['sede_id']))
        disponibles = [r[0] for r in cur.fetchall()]
        cantidad_actual = len(disponibles)

        diferencia = cantidad_nueva - cantidad_actual

        if diferencia > 0:
            # Agregar unidades nuevas
            prefijo = PREFIJOS.get(data['categoria'], 'PRD')

            # ── Resolver observaciones para las unidades nuevas ──────────────
            # FIX-DUPLICADO (red de seguridad backend): /api/inventario/resumen
            # agrupa las tarjetas por (categoria, nombre_modelo, observaciones).
            # Si las unidades nuevas se crean con observaciones distintas (o NULL)
            # a las del resto del modelo, aparecen como una tarjeta aparte —
            # duplicada, sin foto y sin ver el stock que ya existía. El frontend
            # ya manda las observaciones correctas, pero por si algún llamador
            # las omite, heredamos las de cualquier unidad hermana existente
            # del mismo modelo (misma lógica que ya se usa para la foto abajo).
            observaciones_nuevas = data.get('observaciones')
            if not observaciones_nuevas:
                cur.execute("""
                    SELECT observaciones FROM stock_productos
                    WHERE LOWER(nombre_modelo) = LOWER(%s)
                      AND categoria = %s
                      AND observaciones IS NOT NULL AND observaciones != ''
                    ORDER BY id ASC LIMIT 1;
                """, (data['nombre_modelo'], data['categoria']))
                row_obs = cur.fetchone()
                if row_obs:
                    observaciones_nuevas = row_obs[0]

            # ── Resolver la foto para las unidades nuevas ────────────────────
            # FIX: antes solo se buscaba la foto en catalogo_productos usando
            # catalogo_id. Pero la mayoría de productos registrados directo
            # en inventario (botón "Registrar...") NO tienen catalogo_id, y
            # aunque lo tuvieran, esa ficha del catálogo puede no tener foto.
            # Resultado: las unidades nuevas creadas en otra tienda quedaban
            # con foto_url = NULL y aparecían "sin imagen".
            #
            # Ahora se busca en 3 pasos, en orden:
            #   1) Foto del catálogo, si el modelo viene de la carta.
            #   2) Foto de CUALQUIER unidad existente de este mismo modelo
            #      (en cualquier sede) que sí tenga foto — la fuente más
            #      confiable, porque es la imagen real que ya se venía usando.
            #   3) Coincidencia por nombre en catalogo_productos (igual que
            #      hace /api/inventario/resumen), por si el modelo tiene
            #      ficha en la carta pero no quedó enlazado por catalogo_id.
            foto_url = None
            fotos_adicionales = None

            if data.get('catalogo_id'):
                cur.execute("SELECT foto_url FROM catalogo_productos WHERE id = %s", (data['catalogo_id'],))
                row = cur.fetchone()
                if row and row[0]:
                    foto_url = row[0]

            if not foto_url:
                # FIX (foto en fotos_adicionales): antes este query exigía
                # foto_url IS NOT NULL, así que un modelo sin catalogo_id cuya
                # única foto viviera en fotos_adicionales (subida con "Tomar
                # foto" / "Subir fotos" al registrarlo) quedaba totalmente
                # descartado — sus unidades hermanas tienen foto_url = NULL,
                # aunque sí tengan foto en fotos_adicionales. Resultado: las
                # unidades nuevas creadas en otra tienda (o más stock en la
                # misma) se creaban sin foto, aunque el modelo sí tuviera una
                # (visible en la tarjeta de /api/inventario/resumen, que sí
                # mezcla fotos_adicionales). Ahora el WHERE también matchea
                # filas cuya foto vive únicamente en fotos_adicionales.
                cur.execute("""
                    SELECT foto_url, fotos_adicionales FROM stock_productos
                    WHERE LOWER(nombre_modelo) = LOWER(%s)
                      AND categoria = %s
                      AND (
                          (foto_url IS NOT NULL AND foto_url != '')
                          OR (fotos_adicionales IS NOT NULL AND fotos_adicionales != '')
                      )
                    ORDER BY id ASC LIMIT 1;
                """, (data['nombre_modelo'], data['categoria']))
                row = cur.fetchone()
                if row:
                    foto_url = row[0] or None
                    fotos_adicionales = row[1] or None
                    # Si la foto real estaba solo en fotos_adicionales, usar
                    # la primera de esa lista como foto_url de las unidades
                    # nuevas, para que el carousel de detalle siempre tenga
                    # algo en la posición principal.
                    if not foto_url and fotos_adicionales:
                        foto_url = fotos_adicionales.split('|')[0].strip()

            if not foto_url:
                # Igual que el paso 1, pero por nombre en vez de catalogo_id.
                # Se incluye fotos_urls como respaldo por si el registro del
                # catálogo tiene su foto principal (foto_url) vacía pero sí
                # fotos extra guardadas en fotos_urls.
                cur.execute("""
                    SELECT foto_url, fotos_urls FROM catalogo_productos
                    WHERE LOWER(nombre_modelo) = LOWER(%s) LIMIT 1;
                """, (data['nombre_modelo'],))
                row = cur.fetchone()
                if row:
                    if row[0]:
                        foto_url = row[0]
                    elif row[1]:
                        foto_url = row[1].split('|')[0].strip()

            for _ in range(diferencia):
                barcode = _generar_codigo(cur, prefijo, 'stock_productos')
                cur.execute(
                    """
                    INSERT INTO stock_productos
                        (catalogo_id, nombre_modelo, categoria, codigo_barra,
                         observaciones, foto_url, fotos_adicionales, sede_id, estado, creado_por)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Disponible', %s)
                    RETURNING id;
                    """, (
                    data.get('catalogo_id'),
                    data['nombre_modelo'],
                    data['categoria'],
                    barcode,
                    observaciones_nuevas,
                    foto_url,
                    fotos_adicionales,
                    data['sede_id'],
                    data['usuario_id'],
                ))
                nuevo_id = cur.fetchone()[0]
                _registrar_historial(
                    cur, 'producto', nuevo_id, barcode,
                    'Ajuste', None, data['sede_id'],
                    None, 'Disponible',
                    data['usuario_id'], data.get('usuario_nombre', ''),
                    notas=f'Ajuste: {cantidad_actual} → {cantidad_nueva} uds'
                )

        elif diferencia < 0:
            # Eliminar las unidades Disponibles más recientes (las últimas ingresadas)
            ids_a_eliminar = disponibles[diferencia:]  # slice desde el final
            for reg_id in ids_a_eliminar:
                cur.execute("SELECT codigo_barra, sede_id FROM stock_productos WHERE id = %s", (reg_id,))
                row = cur.fetchone()
                barcode, sede_orig = row if row else (None, None)
                _registrar_historial(
                    cur, 'producto', reg_id, barcode,
                    'Ajuste', sede_orig, None,
                    'Disponible', 'Eliminado',
                    data['usuario_id'], data.get('usuario_nombre', ''),
                    notas=f'Ajuste: {cantidad_actual} → {cantidad_nueva} uds'
                )
                cur.execute("DELETE FROM stock_productos WHERE id = %s", (reg_id,))

        conn.commit()
        return jsonify({
            'exito':           True,
            'cantidad_anterior': cantidad_actual,
            'cantidad_nueva':    cantidad_nueva,
            'diferencia':        diferencia,
            'mensaje':          f'Stock actualizado: {cantidad_actual} → {cantidad_nueva} unidades'
        }), 200

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


@inventario_bp.route('/api/inventario/producto/editar', methods=['PUT'])
@requiere_login
def editar_producto_inventario():
    """
    Edita los datos (nombre, categoría, observaciones) de un modelo agrupado
    en el inventario por tiendas. Como /api/inventario/resumen agrupa las
    unidades de stock_productos por (categoria, nombre_modelo, observaciones),
    aquí identificamos el grupo original con esos 3 campos y actualizamos
    TODAS las unidades que coincidan, en todas las sedes.

    Si el modelo viene de un producto de catálogo (catalogo_id), también se
    actualiza catalogo_productos para que el cambio se refleje en la carta.

    Body JSON:
    {
        "categoria":          "Sofa",          (categoría actual, para ubicar el grupo)
        "nombre_modelo":      "Sofá Roma",      (nombre actual)
        "observaciones":      "",               (observaciones actuales, puede ser "")
        "catalogo_id":        12,               (opcional)

        "nuevo_nombre":         "Sofá Roma XL",
        "nueva_categoria":      "Sofa",
        "nuevas_observaciones": "Edición 2026",
        "nuevo_precio":         1500.00,        (opcional, solo si hay catalogo_id)
        "nueva_foto_url":       "https://res.cloudinary.com/...",  (opcional — ver más abajo)

        "usuario_id":    7,
        "usuario_nombre": "Carlos"
    }

    FIX (julio 2026): antes no existía forma de subir/cambiar la foto de un
    modelo ya registrado en inventario — si se creó sin foto (p.ej. desde
    "Registrar..." sin adjuntar imagen), quedaba "Sin foto" para siempre,
    sin importar en cuántas tiendas se le sumara stock después. Ahora, si
    el body trae 'nueva_foto_url' (subida antes con /api/upload-foto desde
    el frontend), se aplica a TODAS las unidades del modelo en TODAS las
    sedes, y también a catalogo_productos si el modelo está enlazado a la
    carta (catalogo_id).
    """
    data = request.json or {}

    required = ['categoria', 'nombre_modelo', 'nuevo_nombre', 'nueva_categoria', 'usuario_id']
    missing  = [f for f in required if data.get(f) in (None, '')]
    if missing:
        return jsonify({'error': f'Campos faltantes: {", ".join(missing)}'}), 400

    nuevo_nombre    = data['nuevo_nombre'].strip()
    nueva_categoria = data['nueva_categoria'].strip()
    if not nuevo_nombre:
        return jsonify({'error': 'El nombre del modelo no puede estar vacío'}), 400

    obs_actuales  = data.get('observaciones') or ''
    nuevas_obs    = data.get('nuevas_observaciones', obs_actuales) or ''
    nueva_foto_url = (data.get('nueva_foto_url') or '').strip()

    conn = None
    try:
        conn = _conn(); cur = conn.cursor()

        if nueva_foto_url:
            cur.execute("""
                UPDATE stock_productos
                   SET nombre_modelo = %s,
                       categoria      = %s,
                       observaciones  = %s,
                       foto_url       = %s
                 WHERE LOWER(nombre_modelo) = LOWER(%s)
                   AND categoria = %s
                   AND COALESCE(observaciones, '') = %s;
            """, (
                nuevo_nombre, nueva_categoria, nuevas_obs, nueva_foto_url,
                data['nombre_modelo'], data['categoria'], obs_actuales
            ))
        else:
            cur.execute("""
                UPDATE stock_productos
                   SET nombre_modelo = %s,
                       categoria      = %s,
                       observaciones  = %s
                 WHERE LOWER(nombre_modelo) = LOWER(%s)
                   AND categoria = %s
                   AND COALESCE(observaciones, '') = %s;
            """, (
                nuevo_nombre, nueva_categoria, nuevas_obs,
                data['nombre_modelo'], data['categoria'], obs_actuales
            ))
        unidades_afectadas = cur.rowcount

        if data.get('catalogo_id'):
            campos  = ["nombre_modelo = %s", "categoria = %s", "observaciones = %s"]
            valores = [nuevo_nombre, nueva_categoria, nuevas_obs]
            if data.get('nuevo_precio') not in (None, ''):
                campos.append("precio_base = %s")
                valores.append(float(data['nuevo_precio']))
            if nueva_foto_url:
                campos.append("foto_url = %s")
                valores.append(nueva_foto_url)
            valores.append(data['catalogo_id'])
            cur.execute(f"""
                UPDATE catalogo_productos
                   SET {', '.join(campos)}
                 WHERE id = %s;
            """, valores)

        conn.commit()
        return jsonify({
            'exito': True,
            'unidades_afectadas': unidades_afectadas,
            'mensaje': f'Producto actualizado ({unidades_afectadas} unidad(es) en inventario)'
        }), 200

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)


@inventario_bp.route('/api/inventario/ventas-tienda', methods=['GET'])
@requiere_login
def listar_ventas_tienda():
    """
    Lista las ventas directas registradas desde Stock en Tienda.
    Params opcionales: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD&sede=nombre
    """
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    sede  = request.args.get('sede', '')

    conn = None
    try:
        conn = _conn()
        cur  = conn.cursor()
        _asegurar_tabla_ventas_tienda(cur)
        conn.commit()

        conds  = []
        params = []
        if desde:
            conds.append("DATE(fecha) >= %s"); params.append(desde)
        if hasta:
            conds.append("DATE(fecha) <= %s"); params.append(hasta)
        if sede:
            conds.append("sede_nombre ILIKE %s"); params.append(f'%{sede}%')

        where = ('WHERE ' + ' AND '.join(conds)) if conds else ''

        cur.execute(f"""
            SELECT id, TO_CHAR(fecha AT TIME ZONE 'America/Lima', 'DD/MM/YYYY HH24:MI') AS fecha,
                   usuario_nombre, nombre_producto, categoria,
                   foto_url, sede_nombre, precio_venta, observaciones, codigo_barra
            FROM ventas_tienda
            {where}
            ORDER BY fecha DESC
            LIMIT 500;
        """, params)

        resultado = [{
            'id':              r[0],
            'fecha':           r[1],
            'vendedor':        r[2],
            'producto':        r[3],
            'categoria':       r[4],
            'foto_url':        r[5] or '',
            'sede':            r[6],
            'precio_venta':    float(r[7]),
            'observaciones':   r[8] or '',
            'codigo_barra':    r[9] or '',
        } for r in cur.fetchall()]

        return jsonify(resultado), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _rel(conn)
